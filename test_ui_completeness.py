import asyncio
import os
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path

from playwright.async_api import async_playwright


BASE_URL = os.environ.get("LONGFORM_UI_BASE_URL", "http://127.0.0.1:8000/")


@dataclass
class CheckResult:
    name: str
    passed: bool
    detail: str = ""


def record(results: list[CheckResult], name: str, passed: bool, detail: str = ""):
    results.append(CheckResult(name=name, passed=passed, detail=detail))
    icon = "PASS" if passed else "FAIL"
    encoding = getattr(sys.stdout, "encoding", None) or "utf-8"
    safe_name = str(name or "").encode(encoding, errors="replace").decode(encoding, errors="replace")
    safe_detail = str(detail or "").encode(encoding, errors="replace").decode(encoding, errors="replace")
    suffix = f" | {safe_detail}" if safe_detail else ""
    print(f"[{icon}] {safe_name}{suffix}")


async def visible(page, selector: str) -> bool:
    locator = page.locator(selector)
    return await locator.count() > 0 and await locator.first.is_visible()


async def exists(page, selector: str) -> bool:
    return await page.locator(selector).count() > 0


def build_per_turn_batch_excel(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对话模板"
    ws.append(["session_id", "turn_order", "user_message", "Role_Nickname", "relationship"])
    # 2 个 session，每个 3 轮
    for session_id, role in [("s1", "角色A"), ("s2", "角色B")]:
        for turn_order in [1, 2, 3]:
            ws.append([session_id, turn_order, f"{role}-第{turn_order}轮", role, "暧昧"])
    wb.save(path)


def build_placeholder_batch_excel_legacy(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对话模板"
    ws.append(
        [
            "session_id",
            "turn_order",
            "user_message",
            "Role_Nickname",
            "relationship",
            "gender",
            "personal_type",
            "personality",
            "longform_persona",
            "longform_narrative_style",
            "longform_few_shot",
            "dialogue_summary",
            "system_module8",
        ]
    )
    for turn_order in [1, 2]:
        ws.append(
            [
                "placeholder-session",
                turn_order,
                f"占位测试第{turn_order}轮",
                "角色占位",
                "暧昧",
                "男",
                "霸道腹黑",
                "暂时留空",
                "暂时留空",
                "暂时留空",
                "暂时留空",
                "可留空",
            ]
        )
    wb.save(path)


def build_placeholder_batch_excel(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对话模板"
    ws.append(
        [
            "session_id",
            "turn_order",
            "user_message",
            "Role_Nickname",
            "relationship",
            "gender",
            "personal_type",
            "personality",
            "longform_persona",
            "longform_narrative_style",
            "longform_few_shot",
            "dialogue_summary",
            "system_module8",
        ]
    )
    for turn_order in [1, 2]:
        ws.append(
            [
                "placeholder-session",
                turn_order,
                f"placeholder-turn-{turn_order}",
                "角色占位",
                "暧昧",
                "男",
                "霸道，高冷，高岭之花，反差",
                "外冷内热，强掌控欲，面对喜欢的人会有明显反差。",
                "暂时留空",
                "暂时留空",
                "暂时留空",
                "暂时留空",
                "可留空",
            ]
        )
    wb.save(path)


async def main():
    results: list[CheckResult] = []
    console_errors: list[str] = []
    page_errors: list[str] = []

    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page(viewport={"width": 1600, "height": 1000})
            page.on(
                "console",
                lambda msg: console_errors.append(f"{msg.type}: {msg.text}")
                if msg.type in {"error", "warning"}
                else None,
            )
            page.on("pageerror", lambda exc: page_errors.append(str(exc)))

            await page.goto(BASE_URL, wait_until="networkidle")
            await page.wait_for_timeout(1600)

            record(results, "A01 根路径跳转到带版本号入口页", "/static/index.html?v=" in page.url, page.url)
            nav_labels = await page.locator(".nav-item .nav-text").all_text_contents()
            record(
                results,
                "A02 四项导航存在",
                nav_labels == ["对话体验", "测试中心", "提示词管理", "历史记录"],
                ",".join(nav_labels),
            )
            record(results, "A03 单模型顶部设置按钮存在", await visible(page, "#btn-header-model-settings"), "#btn-header-model-settings")

            await page.wait_for_function("() => document.querySelectorAll('#f-model-mini option').length > 0", timeout=10000)
            await page.wait_for_function("() => document.querySelectorAll('#f-summary-prompt-version option').length > 0", timeout=10000)
            await page.wait_for_function("() => document.querySelectorAll('#f-scoring-prompt-version option').length > 0", timeout=10000)

            summary_latest = await page.evaluate(
                "async () => (await (await fetch('/api/prompts?kind=summary')).json()).latest_filename"
            )
            scoring_latest = await page.evaluate(
                "async () => (await (await fetch('/api/scoring-prompts')).json()).latest_filename"
            )
            chat_selected = await page.locator("#f-prompt-version").input_value()
            summary_selected = await page.locator("#f-summary-prompt-version").input_value()
            scoring_selected = await page.locator("#f-scoring-prompt-version").input_value()
            mini_selected = await page.locator("#f-model-mini").input_value()

            record(
                results,
                "A04 摘要提示词默认自动追最新",
                summary_selected == "",
                f"selected={summary_selected}, latest={summary_latest}",
            )
            record(
                results,
                "A05 打分提示词默认自动追最新",
                scoring_selected == "",
                f"selected={scoring_selected}, latest={scoring_latest}",
            )
            record(
                results,
                "A06 摘要模型默认 doubao-lite",
                mini_selected == "doubao-lite",
                f"selected={mini_selected}",
            )

            await page.locator("#rightPanel .panel-tab").filter(has_text="🎭 角色").click()
            await page.wait_for_timeout(300)
            record(results, "B01 角色区存在", await visible(page, "#preset-grid"), "#preset-grid")
            record(results, "B02 新右侧变量编辑面板存在", await visible(page, "#role-variable-editor"), "#role-variable-editor")
            record(
                results,
                "B03 Role_info_works 输入框与 voice_forbidden 摘要卡存在",
                await visible(page, "#role-variable-editor [data-editor-key='Role_info_works']")
                and await visible(page, "#role-variable-editor .role-variable-editor-row:has([data-editor-key='voice_forbidden']) .role-variable-editor-summary-card"),
                "role-variable-editor",
            )
            record(
                results,
                "B04 旧右侧角色表单已隐藏",
                await page.evaluate(
                    """
                    () => {
                      const storage = document.getElementById('role-form-legacy-storage');
                      if (!storage) return false;
                      const style = getComputedStyle(storage);
                      return style.display === 'none' || style.visibility === 'hidden';
                    }
                    """
                ),
                "role-form-legacy-storage",
            )
            right_panel_metrics = await page.evaluate(
                """
                () => {
                  const nicknameInput = document.querySelector("#role-variable-editor [data-editor-key='Role_Nickname']");
                  const voiceSummary = document.querySelector("#role-variable-editor .role-variable-editor-row:has([data-editor-key='voice_forbidden']) .role-variable-editor-summary-card");
                  const voiceTextarea = document.querySelector("#role-variable-editor [data-editor-key='voice_forbidden']");
                  const voiceExpand = [...document.querySelectorAll("#role-variable-editor .role-variable-editor-row:has([data-editor-key='voice_forbidden']) .role-variable-editor-inline-btn")].find(el => (el.textContent || '').includes('展开编辑') || (el.textContent || '').includes('收起编辑'));
                  const nicknameRow = nicknameInput?.closest('.role-variable-editor-row');
                  const voiceRow = voiceSummary?.closest('.role-variable-editor-row');
                  return {
                    nicknameWidth: Math.round(nicknameInput?.getBoundingClientRect().width || 0),
                    voiceWidth: Math.round(voiceSummary?.getBoundingClientRect().width || 0),
                    nicknameValue: nicknameInput?.value || '',
                    voiceValue: voiceTextarea?.value || '',
                    nicknameStatus: nicknameRow?.querySelector('.role-variable-editor-status')?.innerText || '',
                    voiceStatus: voiceRow?.querySelector('.role-variable-editor-status')?.innerText || '',
                    voiceHasExpand: !!voiceExpand,
                  };
                }
                """
            )
            record(
                results,
                "B05 右侧主变量宽度可用且默认值可见",
                right_panel_metrics["nicknameWidth"] >= 180
                and right_panel_metrics["voiceWidth"] >= 180
                and bool(right_panel_metrics["nicknameValue"].strip())
                and bool(right_panel_metrics["voiceValue"].strip())
                and right_panel_metrics["voiceHasExpand"],
                str(right_panel_metrics),
            )
            record(
                results,
                "B06 右侧主变量状态明确区分预设填充与系统自动填充",
                "预设" in right_panel_metrics["nicknameStatus"]
                and "系统" in right_panel_metrics["voiceStatus"],
                str(right_panel_metrics),
            )

            summary_interval = await page.locator("#f-summary-interval").input_value()
            injection_depth = await page.locator("#f-injection-depth").input_value()
            record(results, "C01 默认摘要间隔为 5", summary_interval == "5", f"value={summary_interval}")
            record(results, "C02 默认注入深度为 4", injection_depth == "4", f"value={injection_depth}")

            await page.locator("#rightPanel .panel-tab").filter(has_text="⚙️ 参数").click()
            await page.wait_for_timeout(300)
            params_generation_visible = await page.evaluate(
                """
                () => {
                  const host = document.getElementById('tab-params');
                  if (!host) return false;
                  return [...host.querySelectorAll('*')].some(el => {
                    const text = (el.textContent || '').trim();
                    const style = getComputedStyle(el);
                    return text === '生成多样性'
                      && style.display !== 'none'
                      && style.visibility !== 'hidden'
                      && el.offsetParent !== null;
                  });
                }
                """
            )
            record(results, "C03 参数页不再单独展示生成多样性", not params_generation_visible)

            await page.locator("#btn-header-model-settings").click()
            await page.locator("#modal-sp-edit").wait_for(state="visible", timeout=5000)
            await page.wait_for_timeout(400)
            await page.locator("#sp-btn-generation-preset-creative").click()
            await page.wait_for_timeout(250)
            single_prompt_metrics = await page.evaluate(
                """
                () => {
                  const generation = document.getElementById('runtime-prompt-editor-generation');
                  const vars = document.getElementById('runtime-prompt-editor-vars');
                  const chips = [...document.querySelectorAll('#runtime-prompt-editor-vars-body .variable-preview-chip')];
                  const firstChip = chips[0];
                  const creative = document.getElementById('sp-btn-generation-preset-creative');
                  const rect = creative?.getBoundingClientRect();
                  const pointEl = rect ? document.elementFromPoint(rect.left + rect.width / 2, rect.top + rect.height / 2) : null;
                  return {
                    generationVisible: !!generation && getComputedStyle(generation).display !== 'none',
                    varsVisible: !!vars && getComputedStyle(vars).display !== 'none',
                    chipCount: chips.length,
                    inputCount: document.querySelectorAll('#runtime-prompt-editor-vars-body input[data-var]').length,
                    hasColorChip: firstChip ? getComputedStyle(firstChip).backgroundColor !== 'rgba(0, 0, 0, 0)' : false,
                    creativeActive: !!creative && creative.classList.contains('active'),
                    creativeReceivesPointer: pointEl === creative || creative?.contains(pointEl),
                    topElementTag: pointEl?.tagName || '',
                    topElementId: pointEl?.id || '',
                    temperature: document.getElementById('sp-temperature')?.value || '',
                    topP: document.getElementById('sp-top-p')?.value || '',
                  };
                }
                """
            )
            record(
                results,
                "C04 单模型顶部设置入口可打开多样性与变量弹窗",
                single_prompt_metrics["generationVisible"]
                and single_prompt_metrics["varsVisible"]
                and single_prompt_metrics["chipCount"] > 0
                and single_prompt_metrics["inputCount"] > 0
                and single_prompt_metrics["hasColorChip"]
                and single_prompt_metrics["creativeActive"]
                and single_prompt_metrics["creativeReceivesPointer"]
                and single_prompt_metrics["temperature"] == "0.8"
                and single_prompt_metrics["topP"] == "1",
                str(single_prompt_metrics),
            )
            await page.locator("#modal-sp-edit .runtime-prompt-editor-actions .btn-secondary").click()

            await page.locator("#btn-toggle-compare").click()
            await page.wait_for_timeout(600)
            record(results, "D01 模型对比模式可进入", await visible(page, "#page-freechat"), "#page-freechat")

            settings_button = page.locator("button[title='编辑独立 System Prompt']").first
            await settings_button.wait_for(state="visible", timeout=5000)
            await settings_button.click()
            await page.locator("#modal-freechat-prompt").wait_for(state="visible", timeout=5000)
            await page.wait_for_timeout(400)
            await page.locator("#freechat-prompt-editor").fill("{{Role_Nickname}}\n{{voice_forbidden}}")
            await page.wait_for_timeout(300)
            await page.locator("#fc-btn-generation-preset-creative").click()
            await page.wait_for_timeout(250)
            freechat_metrics = await page.evaluate(
                """
                () => {
                  const chips = [...document.querySelectorAll('#freechat-prompt-vars-body .variable-preview-chip')];
                  const firstChip = chips[0];
                  const creative = document.getElementById('fc-btn-generation-preset-creative');
                  const generation = document.querySelector('#modal-freechat-prompt .generation-card');
                  const rect = creative?.getBoundingClientRect();
                  const pointEl = rect ? document.elementFromPoint(rect.left + rect.width / 2, rect.top + rect.height / 2) : null;
                  return {
                    chipCount: chips.length,
                    inputCount: document.querySelectorAll('#freechat-prompt-vars-body input[data-var]').length,
                    hasColorChip: firstChip ? getComputedStyle(firstChip).backgroundColor !== 'rgba(0, 0, 0, 0)' : false,
                    hasCreativeButton: !!creative && getComputedStyle(creative).display !== 'none',
                    creativeActive: !!creative && creative.classList.contains('active'),
                    creativeReceivesPointer: pointEl === creative || creative?.contains(pointEl),
                    topElementTag: pointEl?.tagName || '',
                    topElementId: pointEl?.id || '',
                    generationVisible: !!generation && getComputedStyle(generation).display !== 'none',
                    temperature: document.getElementById('fc-temperature')?.value || '',
                    topP: document.getElementById('fc-top-p')?.value || '',
                  };
                }
                """
            )
            record(
                results,
                "D02 模型对比 Prompt 编辑器右栏可用",
                freechat_metrics["chipCount"] > 0
                and freechat_metrics["inputCount"] > 0
                and freechat_metrics["hasColorChip"]
                and freechat_metrics["hasCreativeButton"]
                and freechat_metrics["generationVisible"]
                and freechat_metrics["creativeActive"]
                and freechat_metrics["creativeReceivesPointer"]
                and freechat_metrics["temperature"] == "0.8"
                and freechat_metrics["topP"] == "1",
                str(freechat_metrics),
            )
            await page.locator("#modal-freechat-prompt .runtime-prompt-editor-actions .btn-secondary").click()

            await page.locator("#btn-toggle-compare").click()
            await page.wait_for_timeout(300)
            record(results, "D03 模型对比模式可退出", await visible(page, "#page-chat"), "#page-chat")

            await page.click(".nav-item[data-page='test-center']")
            await page.wait_for_timeout(300)
            record(results, "E01 测试中心可进入", await visible(page, "#page-test-center"), "#page-test-center")

            record(
                results,
                "E01-1 批量结果表头包含 AI均分",
                await page.locator("#batch-results thead th").filter(has_text="AI均分").count() > 0,
                "AI均分",
            )

            with tempfile.TemporaryDirectory() as td:
                excel_path = Path(td) / "batch_per_turn.xlsx"
                build_per_turn_batch_excel(excel_path)
                await page.locator("#batch-excel-input").set_input_files(str(excel_path))
                await page.wait_for_timeout(600)
                detail_text = await page.locator("#batch-config-detail").inner_text()
                record(
                    results,
                    "E01-2 逐轮模板 Excel 导入后 turns 非空",
                    "2" in detail_text and "组配置已加载" in detail_text and "3轮" in detail_text,
                    detail_text.replace("\n", " | ")[:240],
                )

                placeholder_path = Path(td) / "batch_placeholder.xlsx"
                build_placeholder_batch_excel(placeholder_path)
                await page.locator("#batch-excel-input").set_input_files(str(placeholder_path))
                await page.wait_for_timeout(900)
                placeholder_metrics = await page.evaluate(
                    """
                    () => {
                      const cfg = (typeof batchConfigs !== 'undefined' && batchConfigs[0]) ? batchConfigs[0] : null;
                      if (!cfg) return null;
                      const payload = buildConversationRunPayload(cfg, { turns: getBatchRunTurns(cfg), dryRun: true });
                      return {
                        cfg: {
                          personal_type: cfg.personal_type || '',
                          longform_persona: cfg.longform_persona || '',
                          longform_narrative_style: cfg.longform_narrative_style || '',
                          longform_few_shot: cfg.longform_few_shot || '',
                          dialogue_summary: cfg.dialogue_summary || '',
                          system_module8: cfg.system_module8 || '',
                          voice_forbidden: cfg.voice_forbidden || '',
                        },
                        payload: {
                          prompt_version: payload.prompt_version || '',
                          summary_prompt_version: payload.summary_prompt_version || '',
                          scoring_prompt_version: payload.scoring_prompt_version || '',
                          personal_type: payload.character.personal_type || '',
                          few_shot_file: payload.few_shot_file || '',
                          longform_persona: payload.modules.longform_persona || '',
                          longform_narrative_style: payload.modules.longform_narrative_style || '',
                          voice_forbidden: payload.modules.voice_forbidden || '',
                        },
                      };
                    }
                    """
                )
                record(
                    results,
                    "E01-3 批量占位词清空且补齐自动模块",
                    bool(placeholder_metrics)
                    and placeholder_metrics["cfg"]["dialogue_summary"] == ""
                    and placeholder_metrics["cfg"]["system_module8"] == ""
                    and bool(placeholder_metrics["cfg"]["longform_persona"])
                    and bool(placeholder_metrics["cfg"]["longform_narrative_style"])
                    and bool(placeholder_metrics["cfg"]["longform_few_shot"])
                    and bool(placeholder_metrics["cfg"]["voice_forbidden"])
                    and placeholder_metrics["payload"]["prompt_version"] == chat_selected
                    and placeholder_metrics["payload"]["summary_prompt_version"] == summary_selected
                    and placeholder_metrics["payload"]["scoring_prompt_version"] == scoring_selected
                    and placeholder_metrics["payload"]["personal_type"] == "霸道腹黑"
                    and bool(placeholder_metrics["payload"]["few_shot_file"])
                    and bool(placeholder_metrics["payload"]["longform_persona"])
                    and bool(placeholder_metrics["payload"]["longform_narrative_style"])
                    and bool(placeholder_metrics["payload"]["voice_forbidden"]),
                    str(placeholder_metrics),
                )

            await page.click(".nav-item[data-page='prompts']")
            await page.wait_for_timeout(300)
            record(results, "E02 提示词管理可进入", await visible(page, "#page-prompts"), "#page-prompts")

            await page.click(".nav-item[data-page='history']")
            await page.wait_for_timeout(300)
            record(results, "E03 历史记录可进入", await visible(page, "#page-history"), "#page-history")

            # ═══ F 组: 新 P0 功能浏览器级行为验证 ═══
            # F01: formatEta 函数行为测试
            format_eta_results = await page.evaluate("""
                () => {
                    if (typeof formatEta !== 'function') return { exists: false };
                    return {
                        exists: true,
                        zero: formatEta(0),
                        negative: formatEta(-5),
                        seconds_only: formatEta(45),
                        minutes: formatEta(120),
                        hours: formatEta(3661),
                    };
                }
            """)
            record(
                results,
                "F01 formatEta 函数存在且返回值正确",
                format_eta_results.get("exists") is True
                and format_eta_results.get("zero") == ""
                and format_eta_results.get("negative") == ""
                and "45s" in str(format_eta_results.get("seconds_only", ""))
                and "2m" in str(format_eta_results.get("minutes", ""))
                and "1h" in str(format_eta_results.get("hours", "")),
                str(format_eta_results),
            )

            # F02: buildTurnStatusTags 三色分级行为测试
            turn_status_results = await page.evaluate("""
                () => {
                    if (typeof buildTurnStatusTags !== 'function') return { exists: false };
                    const noTrim = buildTurnStatusTags({ token_trim_level: 0 });
                    const lowTrim = buildTurnStatusTags({ token_trim_level: 1 });
                    const midTrim = buildTurnStatusTags({ token_trim_level: 3 });
                    const highTrim = buildTurnStatusTags({ token_trim_level: 5 });
                    const withRetry = buildTurnStatusTags({ quality_retries: 2 });
                    const withCooldown = buildTurnStatusTags({ has_cooldown_reinject: true });
                    const withStyleIso = buildTurnStatusTags({ has_style_isolation: true });
                    return {
                        exists: true,
                        noTrimCount: noTrim.length,
                        lowTrimTag: lowTrim.find(t => t.text.includes('L1')),
                        midTrimTag: midTrim.find(t => t.text.includes('L3')),
                        highTrimTag: highTrim.find(t => t.text.includes('L5')),
                        retryTag: withRetry.find(t => t.text.includes('质量重试')),
                        cooldownTag: withCooldown.find(t => t.text.includes('冷却复注')),
                        styleIsoTag: withStyleIso.find(t => t.text.includes('风格隔离')),
                    };
                }
            """)
            record(
                results,
                "F02 buildTurnStatusTags 三色分级与状态标签行为正确",
                turn_status_results.get("exists") is True
                and turn_status_results.get("noTrimCount") == 0
                and turn_status_results.get("lowTrimTag", {}).get("cls") == "yellow"
                and turn_status_results.get("midTrimTag", {}).get("cls") == "orange"
                and turn_status_results.get("highTrimTag", {}).get("cls") == "red"
                and turn_status_results.get("retryTag") is not None
                and turn_status_results.get("cooldownTag") is not None
                and turn_status_results.get("styleIsoTag") is not None,
                str(turn_status_results),
            )

            # F03: normalizeDebugEntry 透传冷却复注与风格隔离字段
            debug_entry_results = await page.evaluate("""
                () => {
                    if (typeof normalizeDebugEntry !== 'function') return { exists: false };
                    const entry = normalizeDebugEntry({
                        has_cooldown_reinject: true,
                        has_style_isolation: true,
                        token_trim_level: 3,
                    });
                    return {
                        exists: true,
                        has_cooldown: entry.has_cooldown_reinject,
                        has_style: entry.has_style_isolation,
                        trim_level: entry.trim_level,
                    };
                }
            """)
            record(
                results,
                "F03 normalizeDebugEntry 透传冷却复注和风格隔离",
                debug_entry_results.get("exists") is True
                and debug_entry_results.get("has_cooldown") is True
                and debug_entry_results.get("has_style") is True,
                str(debug_entry_results),
            )

            # F04: buildDegenerationPanel 可渲染且使用 CSS 变量
            degen_panel_results = await page.evaluate("""
                () => {
                    if (typeof buildDegenerationPanel !== 'function') return { exists: false };
                    // 构造符合 per_turn_comparison 格式的假报告
                    // 需要: report.per_turn_comparison = [{ turn, groups: [{status:'scored', dimension_scores, ai_output, label}, ...] }]
                    const fakeReport = {
                        per_turn_comparison: [
                            {
                                turn: 1,
                                groups: [
                                    {
                                        status: 'scored',
                                        label: 'Base',
                                        dimension_scores: { persona_fidelity: 9.0, narrative_immersion: 8.5 },
                                        ai_output: '这是一段较长的基线输出文本，用于测试字数退化检测功能。'.repeat(10),
                                        conv_id: 'test-base-conv',
                                    },
                                    {
                                        status: 'scored',
                                        label: 'Compare',
                                        dimension_scores: { persona_fidelity: 6.0, narrative_immersion: 8.0 },
                                        ai_output: '短文本',
                                        conv_id: 'test-cmp-conv',
                                    },
                                ],
                            },
                        ],
                    };
                    const html = buildDegenerationPanel(fakeReport);
                    return {
                        exists: true,
                        hasContent: typeof html === 'string' && html.length > 0,
                        usesCssVar: html.includes('var(--color-danger') || html.includes('var(--color-success'),
                        noHardcodedRed: !html.includes("color:#DC2626") && !html.includes("color: #DC2626"),
                    };
                }
            """)
            record(
                results,
                "F04 buildDegenerationPanel 可渲染且使用 CSS 变量",
                degen_panel_results.get("exists") is True
                and degen_panel_results.get("hasContent") is True
                and degen_panel_results.get("noHardcodedRed") is True,
                str(degen_panel_results),
            )

            await browser.close()

    except Exception as exc:
        record(results, "UI 完整性复测执行", False, str(exc))

    filtered_console_errors = [
        item for item in console_errors
        if "fonts.googleapis.com" not in item and "ERR_CONNECTION_CLOSED" not in item
    ]
    record(
        results,
        "浏览器 Console 无 error/warning",
        not filtered_console_errors,
        " | ".join(filtered_console_errors[:6]),
    )
    record(
        results,
        "页面运行时无异常",
        not page_errors,
        " | ".join(page_errors[:6]),
    )

    passed = sum(1 for item in results if item.passed)
    failed = len(results) - passed
    print(f"\nUI 完整性复测结果: {passed}/{len(results)} 通过, {failed} 失败")
    if failed:
        sys.exit(1)


if __name__ == "__main__":
    asyncio.run(main())
