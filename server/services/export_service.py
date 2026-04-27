"""
ExportService: Excel 导入/导出工具。
"""
from __future__ import annotations

from pathlib import Path
import json
import re


class ExportService:
    """Excel 导入/导出。"""

    INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
    DETAIL_HEADERS = [
        "测试对应提示词",
        "轮次",
        "Role_Nickname",
        "gender",
        "age",
        "occupation",
        "personality",
        "speaking_style",
        "personal_type",
        "hobby",
        "background",
        "user_Nickname",
        "user_gender",
        "user_identity",
        "relationship",
        "relation_info",
        "intimacy_boundary",
        "relation_calling",
        "currentTime",
        "weekDay",
        "timeperiod",
        "season",
        "current_scene",
        "last_cst_type",
        "完整时间信息",
        "longform_narrative_style",
        "longform_persona",
        "longform_dialogue_guideline",
        "dialogueStartPrompt",
        "moments",
        "weekly_schedule",
        "monthly_schedule",
        "dialogue_summary",
        "用户输入",
        "AI输出",
        "输入tokens",
        "输出tokens",
        "延迟(秒)",
        "模型ID",
        "AI打分状态",
        "人设一致性",
        "叙事沉浸度",
        "情感张力",
        "关系边界与记忆",
        "格式合规",
        "上下文衔接度",
        "加权总分/10",
        "打分理由",
        "人工星级评分",
        "人工评语",
        "人工评价",
    ]
    DIMENSION_LABELS = [
        ("人设一致性", "score_persona_fidelity"),
        ("叙事沉浸度", "score_narrative_immersion"),
        ("情感张力", "score_emotional_tension"),
        ("关系边界与记忆", "score_boundary_memory"),
        ("格式合规", "score_format_compliance"),
        ("上下文衔接度", "score_context_coherence"),
        ("加权总分/10", "score_total"),
    ]

    @classmethod
    def safe_filename_part(cls, value: str, fallback: str = "unknown") -> str:
        cleaned = cls.INVALID_FILENAME_CHARS.sub("_", str(value or "").strip())
        cleaned = re.sub(r"\s+", "_", cleaned).strip(" ._")
        return cleaned or fallback

    @staticmethod
    def _normalize_cell(value):
        if value is None:
            return ""
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return value

    @staticmethod
    def _safe_float(value):
        try:
            if value in ("", None):
                return None
            return float(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _manual_label(score) -> str:
        value = ExportService._safe_float(score)
        if value is None:
            return ""
        if value >= 9:
            return "强烈推荐"
        if value >= 7:
            return "表现良好"
        if value >= 5:
            return "一般"
        return "需复核"

    @classmethod
    def _style_header(cls, ws, headers):
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border
        return thin_border

    @classmethod
    def _apply_column_widths(cls, ws, headers):
        import openpyxl

        custom_widths = {
            "测试对应提示词": 24,
            "轮次": 8,
            "用户输入": 32,
            "AI输出": 64,
            "dialogue_summary": 40,
            "longform_narrative_style": 38,
            "longform_persona": 38,
            "longform_dialogue_guideline": 38,
            "moments": 40,
            "weekly_schedule": 32,
            "monthly_schedule": 32,
            "完整时间信息": 28,
            "personality": 28,
            "dialogueStartPrompt": 30,
            "打分理由": 40,
            "人工评语": 28,
        }
        for col_idx, header in enumerate(headers, start=1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = custom_widths.get(header, 14)

    @classmethod
    def _build_detail_row(cls, result: dict, config: dict, turn_index: int) -> list:
        character = config.get("character", {})
        context = config.get("context", {})
        modules = config.get("modules", {})
        prompt_name = config.get("prompt_file", "unknown")
        manual_score = result.get("manual_star_score", "")

        return [
            prompt_name,
            result.get("turn", turn_index + 1),
            character.get("Role_Nickname", ""),
            character.get("gender", ""),
            character.get("age", ""),
            character.get("occupation", ""),
            character.get("personality", ""),
            character.get("speaking_style", ""),
            character.get("personal_type", ""),
            character.get("hobby", ""),
            character.get("background", ""),
            modules.get("user_Nickname", ""),
            modules.get("user_gender", ""),
            modules.get("user_identity", ""),
            context.get("relationship", ""),
            context.get("relation_info", ""),
            context.get("intimacy_boundary", ""),
            context.get("relation_calling", ""),
            context.get("currentTime", ""),
            context.get("weekDay", ""),
            context.get("timeperiod", ""),
            context.get("season", ""),
            context.get("current_scene", ""),
            context.get("last_cst_type", ""),
            context.get("完整时间信息", ""),
            modules.get("longform_narrative_style", ""),
            modules.get("longform_persona", ""),
            modules.get("longform_dialogue_guideline", ""),
            modules.get("dialogueStartPrompt", ""),
            modules.get("moments", ""),
            modules.get("weekly_schedule", ""),
            modules.get("monthly_schedule", ""),
            result.get("dialogue_summary", ""),
            result.get("user_input", ""),
            result.get("ai_output", ""),
            result.get("input_tokens", 0),
            result.get("output_tokens", 0),
            result.get("latency_s", 0),
            result.get("model_id", config.get("runtime", {}).get("model_ids", [""])[0]),
            result.get("score_status", "unscored"),
            result.get("score_persona_fidelity", 0),
            result.get("score_narrative_immersion", 0),
            result.get("score_emotional_tension", 0),
            result.get("score_boundary_memory", 0),
            result.get("score_format_compliance", 0),
            result.get("score_context_coherence", 0),
            result.get("score_total", 0),
            result.get("score_reasoning", ""),
            manual_score,
            result.get("manual_comment", ""),
            cls._manual_label(manual_score),
        ]

    @classmethod
    def export_to_excel(
        cls, results: list, config: dict, output_path: str, summary: bool = False
    ):
        """导出对话结果。"""
        import openpyxl
        from openpyxl.styles import Alignment

        wb = openpyxl.Workbook()
        ws = wb.active

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        if summary:
            ws.title = "评分摘要"
            headers = ["字段", "值"]
            cls._style_header(ws, headers)

            character = config.get("character", {})
            prompt_name = config.get("prompt_file", "unknown")
            scored_results = [row for row in results if row.get("score_status") == "scored"]
            failed_count = len([row for row in results if row.get("score_status") == "failed"])
            manual_scores = [
                cls._safe_float(row.get("manual_star_score"))
                for row in results
                if cls._safe_float(row.get("manual_star_score")) is not None
            ]

            # 基础统计
            total_scores = [row.get("score_total", 0) for row in scored_results]
            avg_total = round(sum(total_scores) / len(scored_results), 2) if scored_results else 0
            max_total = max(total_scores) if total_scores else 0
            min_total = min(total_scores) if total_scores else 0
            pass_count = len([s for s in total_scores if s >= 7])
            pass_rate = round(pass_count / len(total_scores) * 100, 1) if total_scores else 0

            rows = [
                ("测试对应提示词", prompt_name),
                ("角色", character.get("Role_Nickname", "")),
                ("总轮数", len(results)),
                ("AI已评分轮数", len(scored_results)),
                ("AI失败轮数", failed_count),
                ("AI平均总分", avg_total),
                ("AI最高总分", max_total),
                ("AI最低总分", min_total),
                ("通过率(≥7分)", f"{pass_rate}%  ({pass_count}/{len(total_scores)})"),
                (
                    "人工平均分",
                    round(sum(manual_scores) / len(manual_scores), 2)
                    if manual_scores
                    else "",
                ),
            ]
            # 各维度：均分 + 最高 + 最低
            for label, key in cls.DIMENSION_LABELS:
                if not scored_results:
                    rows.append((f"{label}（均分）", 0))
                    rows.append((f"{label}（最高）", 0))
                    rows.append((f"{label}（最低）", 0))
                else:
                    vals = [row.get(key, 0) for row in scored_results]
                    rows.append((f"{label}（均分）", round(sum(vals) / len(vals), 2)))
                    rows.append((f"{label}（最高）", max(vals)))
                    rows.append((f"{label}（最低）", min(vals)))

            # Top3 最优/最劣轮次
            if scored_results:
                sorted_desc = sorted(scored_results, key=lambda r: r.get("score_total", 0), reverse=True)
                sorted_asc = sorted(scored_results, key=lambda r: r.get("score_total", 0))
                rows.append(("", ""))  # 分隔行
                rows.append(("── Top3 优秀轮次 ──", ""))
                for i, r in enumerate(sorted_desc[:3], 1):
                    rows.append((
                        f"Top{i} 优秀（轮次{r.get('turn', '?')}）",
                        f"总分{r.get('score_total', 0)} | {r.get('score_reasoning', '')[:120]}",
                    ))
                rows.append(("── Top3 待改进轮次 ──", ""))
                for i, r in enumerate(sorted_asc[:3], 1):
                    rows.append((
                        f"Bottom{i} 待改进（轮次{r.get('turn', '?')}）",
                        f"总分{r.get('score_total', 0)} | {r.get('score_reasoning', '')[:120]}",
                    ))

            for row_idx, (label, value) in enumerate(rows, start=2):
                ws.cell(row=row_idx, column=1, value=label)
                ws.cell(row=row_idx, column=2, value=cls._normalize_cell(value))

            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 28
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        else:
            ws.title = "长文模式多轮对话"
            headers = list(cls.DETAIL_HEADERS)
            thin_border = cls._style_header(ws, headers)
            wrap_columns = {
                "personality",
                "longform_narrative_style",
                "dialogueStartPrompt",
                "dialogue_summary",
                "用户输入",
                "AI输出",
                "打分理由",
                "人工评语",
            }
            wrap_indexes = {
                idx
                for idx, header in enumerate(headers, start=1)
                if header in wrap_columns
            }
            for turn_index, result in enumerate(results):
                row_idx = turn_index + 2
                values = cls._build_detail_row(result, config, turn_index)
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=cls._normalize_cell(value),
                    )
                    cell.border = thin_border
                    if col_idx in wrap_indexes:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
            cls._apply_column_widths(ws, headers)

        wb.save(output_path)
        wb.close()
        return output_path

    @classmethod
    def export_rows_to_excel(
        cls, rows: list[dict], output_path: str, sheet_name: str = "评分结果"
    ):
        """导出任意行数据，保持键顺序并补全后续新增列。"""
        import openpyxl
        from openpyxl.styles import Alignment

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        headers: list[str] = []
        for row in rows:
            for key in row.keys():
                key_text = str(key)
                if key_text not in headers:
                    headers.append(key_text)

        cls._style_header(ws, headers)
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=cls._normalize_cell(row.get(header, "")),
                )
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        cls._apply_column_widths(ws, headers)
        wb.save(output_path)
        wb.close()
        return output_path

    @staticmethod
    def import_from_excel(file_path: str) -> list[dict]:
        """从 Excel 导入数据，返回按表头映射的行列表。"""
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            wb.close()
            return []

        headers = [str(header or "").strip() for header in rows[0]]
        results = []
        for row in rows[1:]:
            entry = {}
            for index, value in enumerate(row):
                if index < len(headers) and headers[index]:
                    entry[headers[index]] = value if value is not None else ""
            if any(value not in ("", None) for value in entry.values()):
                results.append(entry)
        wb.close()
        return results
