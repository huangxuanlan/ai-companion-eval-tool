"""
A/B/C 多组对比报告路由: /api/reports
"""
from __future__ import annotations

from datetime import datetime

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field

import database as db
from config import PROJECT_DIR
from services.scoring_service import ScoringService

router = APIRouter(prefix="/api/reports", tags=["reports"])

DIMENSIONS = [
    "persona_fidelity",
    "narrative_immersion",
    "emotional_tension",
    "boundary_memory",
    "format_compliance",
    "context_coherence",
]
DIMENSION_NAMES = {
    "persona_fidelity": "人设一致性",
    "narrative_immersion": "叙事沉浸度",
    "emotional_tension": "情感张力",
    "boundary_memory": "关系边界与记忆",
    "format_compliance": "格式合规",
    "context_coherence": "上下文衔接度",
    "total": "总分",
}


class CompareReportCreate(BaseModel):
    """创建对比报告。"""

    groups: list[dict] = Field(
        ..., min_length=2, max_length=3, description="2-3 组，每组包含 conv_id 和 label"
    )


class HistorySelectionReportCreate(BaseModel):
    """创建历史多选评分摘要报告。"""

    conversation_ids: list[str] = Field(
        ..., min_length=2, max_length=24, description="待汇总的历史对话 ID 列表"
    )


_scoring_service: ScoringService | None = None


def _get_scoring_service() -> ScoringService:
    global _scoring_service
    if _scoring_service is None:
        _scoring_service = ScoringService()
    return _scoring_service


def _build_turn_dimension_scores(row: dict) -> dict:
    return {
        dimension: row.get(f"score_{dimension}", 0)
        for dimension in DIMENSIONS
    }


def _build_group_result(group: dict) -> dict:
    conversation = db.get_conversation(group["conv_id"])
    if not conversation:
        raise HTTPException(status_code=404, detail=f"对话不存在: {group['conv_id']}")

    config = conversation.get("config", {})
    model_id = group.get("model_id") or conversation.get("model_id", "")
    prompt_version = (
        group.get("prompt_version")
        or conversation.get("prompt_version")
        or config.get("prompt_file", "")
    )
    label = group.get("label") or prompt_version or model_id or group["conv_id"]
    role_name = config.get("character", {}).get("Role_Nickname", "") or ""

    results = conversation.get("results", [])
    scored_rows = [row for row in results if row.get("score_status") == "scored"]
    failed_rows = [row for row in results if row.get("score_status") == "failed"]
    pending_rows = [
        row
        for row in results
        if row.get("score_status") not in {"scored", "failed"}
    ]
    if not scored_rows:
        raise HTTPException(
            status_code=400, detail=f"对话 {group['conv_id']} 无打分数据"
        )

    avg_scores = {}
    for dimension in DIMENSIONS:
        avg_scores[dimension] = round(
            sum(row.get(f"score_{dimension}", 0) for row in scored_rows) / len(scored_rows),
            2,
        )
    avg_scores["total"] = round(
        sum(row.get("score_total", 0) for row in scored_rows) / len(scored_rows), 2
    )

    manual_scores = [
        float(row["manual_star_score"])
        for row in results
        if row.get("manual_star_score") not in (None, "")
    ]
    total_input_tokens = sum(int(row.get("input_tokens", 0) or 0) for row in results)
    total_output_tokens = sum(int(row.get("output_tokens", 0) or 0) for row in results)
    latency_values = [
        float(row.get("latency_s", 0) or 0)
        for row in results
        if row.get("latency_s") not in (None, "")
    ]
    pass_count = len([row for row in scored_rows if float(row.get("score_total", 0) or 0) >= 8])

    return {
        "conv_id": group["conv_id"],
        "label": label,
        "status": conversation.get("status", ""),
        "created_at": conversation.get("created_at", ""),
        "role_name": role_name,
        "model_id": model_id,
        "prompt_version": prompt_version,
        "avg_scores": avg_scores,
        "per_turn": [
            {
                "turn": row.get("turn", 0),
                "total": row.get("score_total", 0),
                "status": row.get("score_status", "unscored"),
                "manual_star_score": row.get("manual_star_score"),
                "input_tokens": row.get("input_tokens", 0),
                "output_tokens": row.get("output_tokens", 0),
                "latency_s": row.get("latency_s", 0),
                "reasoning": row.get("score_reasoning", ""),
                "ai_output": row.get("ai_output", ""),
                "dimension_scores": _build_turn_dimension_scores(row),
            }
            for row in results
        ],
        "turn_count": len(results),
        "scored_count": len(scored_rows),
        "failed_count": len(failed_rows),
        "pending_count": len(pending_rows),
        "pass_count": pass_count,
        "manual_avg": round(sum(manual_scores) / len(manual_scores), 2)
        if manual_scores
        else None,
        "total_input_tokens": total_input_tokens,
        "total_output_tokens": total_output_tokens,
        "avg_latency_s": round(sum(latency_values) / len(latency_values), 2)
        if latency_values
        else 0,
    }


def _validate_group_results(group_results: list[dict]) -> None:
    statuses = {str(group.get("status", "")).strip() for group in group_results}
    if statuses - {"completed"}:
        raise HTTPException(status_code=400, detail="仅支持已完成记录生成对比报告")

    turn_counts = {int(group.get("turn_count", 0) or 0) for group in group_results}
    if len(turn_counts) > 1:
        raise HTTPException(status_code=400, detail="仅支持轮数一致的记录进行对比")

    compare_mode = _detect_compare_mode(group_results)
    if compare_mode == "mixed":
        raise HTTPException(
            status_code=400,
            detail="仅支持同类型（提示词对比或模型对比）的记录进行对比",
        )


def _compute_winners(group_results: list[dict]) -> dict:
    winners = {}
    for dimension in DIMENSIONS + ["total"]:
        best_score = max(group["avg_scores"].get(dimension, 0) for group in group_results)
        labels = [
            group["label"]
            for group in group_results
            if group["avg_scores"].get(dimension, 0) == best_score
        ]
        winners[dimension] = labels if len(labels) > 1 else labels[0]
    return winners


def _build_per_dim_comparison(group_results: list[dict], winners: dict) -> dict:
    return {
        dimension: {
            "scores": {
                group["label"]: group["avg_scores"].get(dimension, 0)
                for group in group_results
            },
            "winner": winners.get(dimension, ""),
        }
        for dimension in DIMENSIONS + ["total"]
    }


def _detect_compare_mode(group_results: list[dict]) -> str:
    model_ids = {group.get("model_id", "") for group in group_results if group.get("model_id")}
    prompt_versions = {
        group.get("prompt_version", "")
        for group in group_results
        if group.get("prompt_version")
    }
    if len(prompt_versions) > 1 and len(model_ids) <= 1:
        return "prompt"
    if len(model_ids) > 1 and len(prompt_versions) <= 1:
        return "model"
    return "mixed"


def _normalize_compare_axis_value(value: str, fallback: str = "-") -> str:
    normalized = str(value or "").strip()
    return normalized or fallback


def _build_history_selection_meta(group_results: list[dict]) -> dict:
    roles = {
        _normalize_compare_axis_value(group.get("role_name", ""), group.get("conv_id", "未命名角色"))
        for group in group_results
    }
    models = {
        _normalize_compare_axis_value(group.get("model_id", ""), group.get("conv_id", "未命名模型"))
        for group in group_results
    }
    prompts = {
        _normalize_compare_axis_value(group.get("prompt_version", ""), group.get("conv_id", "未命名提示词"))
        for group in group_results
    }
    selection_count = len(group_results)
    role_count = len(roles)
    model_count = len(models)
    prompt_count = len(prompts)

    if selection_count <= 1:
        summary_type = "single_combination"
        report_title = "单组合评分摘要"
    elif role_count == 1 and model_count > 1 and prompt_count == 1:
        summary_type = "model_summary"
        report_title = "模型评分摘要"
    elif role_count == 1 and model_count == 1 and prompt_count > 1:
        summary_type = "prompt_summary"
        report_title = "提示词评分摘要"
    elif role_count > 1 and model_count == 1 and prompt_count == 1:
        summary_type = "role_summary"
        report_title = "角色评分摘要"
    elif prompt_count == 1:
        summary_type = "mode_summary"
        report_title = "模式评分摘要"
    elif role_count == 1 or model_count == 1 or prompt_count == 1:
        summary_type = "matrix_summary"
        report_title = "实验矩阵评分摘要"
    else:
        summary_type = "mixed_summary"
        report_title = "混合样本评分盘点"

    return {
        "summary_type": summary_type,
        "report_title": report_title,
        "selection_count": selection_count,
        "role_count": role_count,
        "model_count": model_count,
        "prompt_count": prompt_count,
    }


def _build_history_selection_group_label(group: dict, report_meta: dict) -> str:
    summary_type = str(report_meta.get("summary_type", "") or "").strip()
    if summary_type == "model_summary":
        return _normalize_compare_axis_value(group.get("model_id", ""), group.get("conv_id", "未命名模型"))
    if summary_type == "prompt_summary":
        return _normalize_compare_axis_value(group.get("prompt_version", ""), group.get("conv_id", "未命名提示词"))
    if summary_type == "role_summary":
        return _normalize_compare_axis_value(group.get("role_name", ""), group.get("conv_id", "未命名角色"))
    if summary_type == "single_combination":
        created_at = str(group.get("created_at", "") or "").strip()
        return created_at or str(group.get("conv_id", "") or "未命名记录")
    role_name = _normalize_compare_axis_value(group.get("role_name", ""), "")
    model_id = _normalize_compare_axis_value(group.get("model_id", ""), "")
    prompt_version = _normalize_compare_axis_value(group.get("prompt_version", ""), "")
    parts = [part for part in [role_name, model_id, prompt_version] if part]
    return " · ".join(parts) or str(group.get("conv_id", "") or "未命名记录")


def _build_history_selection_groups(group_results: list[dict], report_meta: dict) -> list[dict]:
    groups = []
    for group in group_results:
        label = _build_history_selection_group_label(group, report_meta)
        group["label"] = label
        groups.append(
            {
                "conv_id": group.get("conv_id", ""),
                "label": label,
                "model_id": group.get("model_id", ""),
                "prompt_version": group.get("prompt_version", ""),
                "summary_type": report_meta.get("summary_type", ""),
                "report_title": report_meta.get("report_title", ""),
                "selection_count": report_meta.get("selection_count", 0),
                "role_count": report_meta.get("role_count", 0),
                "model_count": report_meta.get("model_count", 0),
                "prompt_count": report_meta.get("prompt_count", 0),
            }
        )
    return groups


def _extract_report_meta(report: dict) -> dict:
    groups = list(report.get("groups", []) or [])
    if not groups:
        return {}
    first_group = dict(groups[0] or {})
    summary_type = str(first_group.get("summary_type", "") or "").strip()
    report_title = str(first_group.get("report_title", "") or "").strip()
    if not summary_type and not report_title:
        return {}
    return {
        "summary_type": summary_type,
        "report_title": report_title,
        "selection_count": int(first_group.get("selection_count", len(groups)) or len(groups)),
        "role_count": int(first_group.get("role_count", 0) or 0),
        "model_count": int(first_group.get("model_count", 0) or 0),
        "prompt_count": int(first_group.get("prompt_count", 0) or 0),
    }


def _resolve_report_title(report_meta: dict, compare_mode: str) -> str:
    title = str(report_meta.get("report_title", "") or "").strip()
    if title:
        return title
    compare_mode_map = {
        "model": "模型对比报告",
        "prompt": "提示词对比报告",
        "mixed": "混合对比报告",
    }
    return compare_mode_map.get(compare_mode, "历史评分报告")


def _build_per_turn_comparison(group_results: list[dict]) -> list[dict]:
    max_turn = max((len(group.get("per_turn", [])) for group in group_results), default=0)
    rows = []
    for turn in range(1, max_turn + 1):
        turn_groups = []
        scored_entries = []
        for group in group_results:
            turn_data = next(
                (item for item in group.get("per_turn", []) if item.get("turn") == turn),
                None,
            )
            entry = {
                "label": group["label"],
                "model_id": group.get("model_id", ""),
                "prompt_version": group.get("prompt_version", ""),
                "turn": turn,
                "total": turn_data.get("total", 0) if turn_data else None,
                "status": turn_data.get("status", "missing") if turn_data else "missing",
                "manual_star_score": (
                    turn_data.get("manual_star_score") if turn_data else None
                ),
                "dimension_scores": (
                    dict(turn_data.get("dimension_scores", {}) or {})
                    if turn_data
                    else {}
                ),
                "reasoning": turn_data.get("reasoning", "") if turn_data else "",
                "ai_output": turn_data.get("ai_output", "") if turn_data else "",
            }
            turn_groups.append(entry)
            if entry["status"] == "scored" and entry["total"] is not None:
                scored_entries.append(entry)

        best_labels = []
        if scored_entries:
            best_score = max(entry["total"] for entry in scored_entries)
            best_labels = [
                entry["label"] for entry in scored_entries if entry["total"] == best_score
            ]

        rows.append(
            {
                "turn": turn,
                "groups": turn_groups,
                "winners": best_labels,
            }
        )
    return rows


def _augment_report(report: dict) -> dict:
    group_results = report.get("group_results", [])
    winners = report.get("winners", {})
    report_meta = _extract_report_meta(report)
    compare_mode = str(report_meta.get("summary_type", "") or "").strip() or _detect_compare_mode(group_results)
    return {
        **report,
        "compare_mode": compare_mode,
        "summary_type": compare_mode,
        "report_title": _resolve_report_title(report_meta, compare_mode),
        "report_meta": report_meta,
        "per_dim_comparison": _build_per_dim_comparison(group_results, winners),
        "per_turn_comparison": _build_per_turn_comparison(group_results),
    }


@router.post("/compare")
async def create_compare_report(data: CompareReportCreate):
    """创建 A/B/C 对比报告。"""
    group_results = [_build_group_result(group) for group in data.groups]
    _validate_group_results(group_results)
    winners = _compute_winners(group_results)
    report_id = db.create_compare_report(data.groups, group_results, winners)

    report = {
        "id": report_id,
        "groups": data.groups,
        "group_results": group_results,
        "winners": winners,
    }
    return _augment_report(report)


@router.post("/history-selection")
async def create_history_selection_report(data: HistorySelectionReportCreate):
    """根据历史多选结果创建评分摘要报告。"""
    conv_ids = [str(item or "").strip() for item in data.conversation_ids if str(item or "").strip()]
    if len(conv_ids) < 2:
        raise HTTPException(status_code=400, detail="至少选择 2 条历史记录")

    unique_ids = list(dict.fromkeys(conv_ids))
    group_results = [_build_group_result({"conv_id": conv_id}) for conv_id in unique_ids]
    if any(group.get("scored_count", 0) <= 0 for group in group_results):
        raise HTTPException(status_code=400, detail="存在无已评分轮次的记录，无法生成摘要报告")

    report_meta = _build_history_selection_meta(group_results)
    groups = _build_history_selection_groups(group_results, report_meta)
    winners = _compute_winners(group_results)
    report_id = db.create_compare_report(groups, group_results, winners)
    return _augment_report(
        {
            "id": report_id,
            "groups": groups,
            "group_results": group_results,
            "winners": winners,
        }
    )


@router.get("/compare/{report_id}")
async def get_compare_report(report_id: str):
    """获取对比报告详情。"""
    report = db.get_compare_report(report_id)
    if not report:
        raise HTTPException(status_code=404, detail="报告不存在")
    return _augment_report(report)


@router.post("/compare/{report_id}/ai-summary")
async def generate_compare_ai_summary(
    report_id: str,
    model_id: str = Query(default=""),
    prompt_version: str = Query(default=""),
):
    """生成历史对比 AI 摘要。"""
    report = db.get_compare_report(report_id)
    if not report:
        raise HTTPException(status_code=404, detail="报告不存在")
    augmented = _augment_report(report)
    service = _get_scoring_service()
    summary = await service.generate_compare_report(
        augmented,
        model_id=model_id.strip() or None,
        prompt_version=prompt_version.strip() or None,
    )
    return {"report_id": report_id, "summary": summary}


@router.get("/compare/{report_id}/export")
async def export_compare_report(report_id: str, summary: bool = False):
    """导出对比报告 Excel。"""
    import openpyxl

    report = db.get_compare_report(report_id)
    if not report:
        raise HTTPException(status_code=404, detail="报告不存在")

    report = _augment_report(report)
    group_results = report.get("group_results", [])
    per_turn_comparison = report.get("per_turn_comparison", [])
    winners = report.get("winners", {})

    output_dir = PROJECT_DIR / "output"
    output_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = "summary" if summary else "full"
    filename = f"compare_report_{timestamp}_{suffix}.xlsx"
    output_path = output_dir / filename

    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "摘要" if summary else "维度对比"

    headers = ["维度"] + [group["label"] for group in group_results] + ["最佳"]
    for column, header in enumerate(headers, start=1):
        summary_sheet.cell(row=1, column=column, value=header)

    for row_index, dimension in enumerate(DIMENSIONS + ["total"], start=2):
        summary_sheet.cell(
            row=row_index, column=1, value=DIMENSION_NAMES.get(dimension, dimension)
        )
        for column, group in enumerate(group_results, start=2):
            summary_sheet.cell(
                row=row_index,
                column=column,
                value=group["avg_scores"].get(dimension, 0),
            )
        winner_value = winners.get(dimension, "")
        if isinstance(winner_value, list):
            winner_value = " / ".join(winner_value)
        summary_sheet.cell(row=row_index, column=len(headers), value=winner_value)

    if not summary:
        turn_sheet = workbook.create_sheet("逐轮对比")
        turn_headers = ["轮次"]
        for group in group_results:
            turn_headers.extend(
                [f"{group['label']}_总分", f"{group['label']}_状态", f"{group['label']}_人工分"]
            )
        turn_headers.append("本轮最佳")

        for column, header in enumerate(turn_headers, start=1):
            turn_sheet.cell(row=1, column=column, value=header)

        for row_index, row in enumerate(per_turn_comparison, start=2):
            turn_sheet.cell(row=row_index, column=1, value=row["turn"])
            column = 2
            for group_entry in row.get("groups", []):
                turn_sheet.cell(row=row_index, column=column, value=group_entry["total"])
                turn_sheet.cell(
                    row=row_index, column=column + 1, value=group_entry["status"]
                )
                turn_sheet.cell(
                    row=row_index,
                    column=column + 2,
                    value=group_entry.get("manual_star_score"),
                )
                column += 3
            turn_sheet.cell(
                row=row_index,
                column=column,
                value=" / ".join(row.get("winners", [])),
            )

    workbook.save(str(output_path))
    workbook.close()
    return FileResponse(
        path=str(output_path),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
