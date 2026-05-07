"""
评分路由: /api/scoring
"""
from __future__ import annotations

import asyncio
from copy import deepcopy
import inspect
import json
import os
import uuid
from datetime import datetime

from fastapi import APIRouter, Body, File, Form, HTTPException, Query, UploadFile, WebSocket
from fastapi.responses import FileResponse

import database as db
from config import PROJECT_DIR
from models import (
    ManualDialogueRequest,
    ManualScoreRequest,
    RescoreAllRequest,
    ScoringConfigUpdate,
    TaskControlRequest,
    TriggerScoringRequest,
)
from services.export_service import ExportService
from services.live_scoring_dispatcher import LiveScoringDispatcher
from services.model_adapter import ModelAdapter
from services.public_demo import ensure_visible_conversation
from services.scoring_service import ScoringService, invoke_score_turn_compat
from services import task_control

router = APIRouter(prefix="/api/scoring", tags=["scoring"])

_scoring = None
_live_scoring_dispatcher = None
_export_service = ExportService()
_score_ws: dict[str, list[WebSocket]] = {}
OUTPUT_DIR = PROJECT_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)
DIMENSIONS = [
    "persona_fidelity",
    "narrative_immersion",
    "emotional_tension",
    "boundary_memory",
    "format_compliance",
    "context_coherence",
]
DIMENSION_DISPLAY = {
    "persona_fidelity": "人设忠实度",
    "narrative_immersion": "叙事沉浸感",
    "emotional_tension": "情感张力",
    "boundary_memory": "边界与记忆",
    "format_compliance": "格式合规",
    "context_coherence": "上下文衔接度",
}
LIVE_SCORING_TIMEOUT_ENV = "SCORING_LIVE_REQUEST_TIMEOUT_S"
LIVE_SCORING_DEFAULT_TIMEOUT_S = 25.0
LIVE_SCORING_MAX_WORKERS_ENV = "SCORING_LIVE_MAX_WORKERS"
LIVE_SCORING_DEFAULT_MAX_WORKERS = 6


def _get_scoring():
    global _scoring
    if _scoring is None:
        _scoring = ScoringService()
    return _scoring


def get_live_scoring_dispatcher() -> LiveScoringDispatcher:
    global _live_scoring_dispatcher
    if _live_scoring_dispatcher is None:
        _live_scoring_dispatcher = LiveScoringDispatcher(
            worker=lambda conv_id, turn, config=None: _run_live_scoring_turn(
                conv_id,
                turn,
                config=config,
            ),
            get_max_workers=lambda: min(
                max(
                    1,
                    int(
                        (
                            getattr(_get_scoring(), "get_max_workers", lambda: 1)()
                            or 1
                        )
                    ),
                ),
                _resolve_live_scoring_max_workers(),
            ),
        )
    return _live_scoring_dispatcher


def _resolve_live_scoring_max_workers() -> int:
    raw = os.environ.get(LIVE_SCORING_MAX_WORKERS_ENV, str(LIVE_SCORING_DEFAULT_MAX_WORKERS))
    try:
        return max(1, min(int(raw), 24))
    except (TypeError, ValueError):
        return LIVE_SCORING_DEFAULT_MAX_WORKERS


def _resolve_scoring_prompt_version(prompt_version: str | None = None) -> str:
    service = _get_scoring()
    service.prompt_store.ensure_initialized()
    requested = str(prompt_version or "").strip() or "latest"
    return service.prompt_store.resolve_filename(requested)


def _get_visible_conversation_or_404(conv_id: str) -> dict:
    return ensure_visible_conversation(db.get_conversation(conv_id), conv_id)


def _record_scoring_event(
    conv_id: str,
    event_type: str,
    *,
    level: str = "info",
    detail: dict | None = None,
) -> None:
    try:
        db.log_conversation_event(
            conv_id,
            scope="scoring",
            level=level,
            event_type=event_type,
            detail=detail or {},
        )
    except Exception as exc:
        print(f"[WARN] scoring event log failed: {conv_id} {event_type}: {exc}")


async def _push_scoring_status(
    conv_id: str,
    status: str,
    *,
    event_type: str | None = None,
    message: str = "",
    **extra,
):
    payload = {
        "type": "task_status",
        "scope": "scoring",
        "conversation_id": conv_id,
        "status": status,
        **extra,
    }
    if message:
        payload["message"] = message
    await _push_score_message(conv_id, payload)
    if event_type:
        event_payload = {
            "type": event_type,
            "conversation_id": conv_id,
            **extra,
        }
        if message:
            event_payload["message"] = message
        await _push_score_message(conv_id, event_payload)


def _invalidate_conversation_scoring_summary(conv_id: str) -> int:
    cleared = db.clear_ai_report_summaries(
        target_type="conversation_scoring",
        target_id=conv_id,
        report_kind="scoring_report",
    )
    _record_scoring_event(
        conv_id,
        "summary_invalidated",
        detail={"cleared_count": int(cleared or 0)},
    )
    return cleared


def _build_ai_summary_items(conversation: dict) -> list[dict]:
    items = []
    for result in conversation.get("results", []) or []:
        status = str(result.get("score_status", "unscored") or "unscored").strip()
        items.append(
            {
                "turn": result.get("turn"),
                "status": status,
                "success": status == "scored",
                "mapped_total": result.get("score_total", 0),
                "reasoning": result.get("score_reasoning", ""),
                "scores": {
                    dimension: result.get(f"score_{dimension}", 0)
                    for dimension in DIMENSIONS
                },
            }
        )
    return items


def _flatten_scored_rows(scored_rows: list[dict]) -> list[dict]:
    flattened = []
    for row in scored_rows:
        item = dict(row)
        scores = item.pop("scores", {}) or {}
        for dimension in DIMENSIONS:
            item[f"score_{dimension}"] = scores.get(dimension, 0)
        item["score_total"] = item.get("mapped_total", 0)
        item["score_reasoning"] = item.get("reasoning", "")
        item["score_status"] = "scored" if item.get("success") else "failed"
        flattened.append(item)
    return flattened


def _build_turn_payload(turn_result: dict, config: dict, *, history_context: str = "") -> dict:
    character = config.get("character", {})
    context = config.get("context", {})
    return {
        "turn": turn_result.get("turn", 0),
        "user_input": turn_result.get("user_input", ""),
        "ai_output": turn_result.get("ai_output", ""),
        "role_name": character.get("Role_Nickname", ""),
        "personality": character.get("personality", ""),
        "relationship": context.get("relationship", ""),
        "prompt_name": config.get("prompt_file", ""),
        "dialogueStartPrompt": dict(config.get("modules", {}) or {}).get("dialogueStartPrompt", ""),
        "moments": dict(config.get("modules", {}) or {}).get("moments", ""),
        "dialogue_summary": turn_result.get("dialogue_summary", dict(config.get("runtime", {}) or {}).get("latest_dialogue_summary", "")),
        "history_context": history_context,
    }


def _resolve_scoring_thinking_effort(
    service,
    *,
    model_id: str,
    thinking_enabled,
    thinking_effort: str,
):
    resolver = getattr(service, "resolve_scoring_thinking_effort", None)
    if callable(resolver):
        return resolver(
            model_id,
            thinking_enabled,
            thinking_effort,
            None,
        )
    return ModelAdapter.resolve_thinking_effort(
        model_id,
        thinking_enabled,
        thinking_effort,
    )


def _find_turn_result(conversation: dict, turn: int) -> dict | None:
    return next(
        (item for item in conversation.get("results", []) if item.get("turn") == turn),
        None,
    )


def _normalize_turn_score_status(result: dict) -> str:
    status = str(result.get("score_status", "unscored") or "unscored").strip().lower()
    if status in {"scored", "failed", "skipped", "unscored"}:
        return status
    return "scored" if float(result.get("score_total", 0) or 0) > 0 else "unscored"


def _select_pending_scoring_results(results: list[dict]) -> list[dict]:
    pending_results = []
    for result in results or []:
        if _normalize_turn_score_status(result) in {"unscored", "failed"}:
            pending_results.append(result)
    return pending_results


def _should_force_full_rescore(
    conversation: dict,
    data: TriggerScoringRequest | None,
) -> bool:
    if not data or data.scoring_model_id is None:
        return False

    runtime = ((conversation or {}).get("config", {}) or {}).get("runtime", {}) or {}
    requested_model = (
        ModelAdapter.normalize_model_id(data.scoring_model_id)
        or str(data.scoring_model_id or "").strip()
    )
    stored_model = str(runtime.get("scoring_model_id", "") or "").strip()
    if requested_model and not stored_model:
        return True
    current_model = (
        ModelAdapter.normalize_model_id(stored_model)
        or stored_model
    )
    return bool(requested_model and requested_model != current_model)


def _merge_scoring_runtime(
    config: dict,
    data: TriggerScoringRequest | None,
) -> dict:
    merged = deepcopy(config or {})
    runtime = merged.setdefault("runtime", {})
    if not data:
        return merged

    if data.scoring_model_id is not None:
        runtime["scoring_model_id"] = (
            ModelAdapter.normalize_model_id(data.scoring_model_id)
            or str(data.scoring_model_id or "").strip()
        )
    if data.scoring_prompt_version is not None:
        runtime["scoring_prompt_version"] = _resolve_scoring_prompt_version(
            data.scoring_prompt_version
        )
    if data.scoring_thinking_enabled is not None:
        runtime["scoring_thinking_enabled"] = bool(data.scoring_thinking_enabled)
    if data.scoring_thinking_effort is not None:
        runtime["scoring_thinking_effort"] = str(data.scoring_thinking_effort or "").strip()
    if data.max_workers is not None:
        runtime["scoring_max_workers"] = max(1, min(int(data.max_workers), 24))
    if data.scoring_retry_count is not None:
        runtime["scoring_retry_count"] = max(0, min(int(data.scoring_retry_count), 10))
    return merged


def _persist_scoring_runtime(conv_id: str, config: dict) -> None:
    db.update_conversation_config(conv_id, config)


def _build_scoring_summary(conversation: dict) -> dict:
    results = conversation.get("results", [])
    turns = [
        {
            "turn": result.get("turn", 0),
            "scores": {
                dimension: result.get(f"score_{dimension}", 0)
                for dimension in DIMENSIONS
            },
            "total": result.get("score_total", 0),
            "status": result.get("score_status", "unscored"),
            "manual_star_score": result.get("manual_star_score"),
        }
        for result in results
    ]
    scored_turns = [item for item in turns if item["status"] == "scored"]
    failed_turns = [item for item in turns if item["status"] == "failed"]
    skipped_turns = [item for item in turns if item["status"] == "skipped"]
    manual_scores = [
        float(item["manual_star_score"])
        for item in turns
        if item.get("manual_star_score") not in (None, "")
    ]
    avg_scores = {}
    dim_details = {}
    for dimension in DIMENSIONS:
        values = [item["scores"][dimension] for item in scored_turns]
        avg_scores[dimension] = round(sum(values) / len(values), 2) if values else 0
        dim_details[dimension] = {
            "display_name": DIMENSION_DISPLAY.get(dimension, dimension),
            "avg": round(sum(values) / len(values), 2) if values else 0,
            "max": max(values) if values else 0,
            "min": min(values) if values else 0,
        }
    totals = [float(item["total"] or 0) for item in scored_turns]
    stats = db.recalculate_conversation_avg(conversation.get("id", ""))
    return {
        "avg_scores": avg_scores,
        "avg_total": stats["avg_total"],
        "manual_avg": round(sum(manual_scores) / len(manual_scores), 2)
        if manual_scores
        else None,
        "manual_count": len(manual_scores),
        "scored_count": stats["scored_count"],
        "failed_count": stats["failed_count"],
        "skipped_count": stats["skipped_count"],
        "total_count": len(turns),
        "max_total": max(totals) if totals else 0,
        "min_total": min(totals) if totals else 0,
        "dim_avgs": avg_scores,
        "dim_details": dim_details,
    }


def _build_ai_report_meta(conversation: dict, summary: dict | None = None) -> dict:
    resolved_summary = summary or _build_scoring_summary(conversation)
    return db.get_conversation_ai_report_meta(
        str(conversation.get("id", "")).strip(),
        conversation_status=conversation.get("status", ""),
        total_turns=len(conversation.get("results", []) or []),
        scored_turns=resolved_summary.get("scored_count", 0),
        failed_turns=resolved_summary.get("failed_count", 0),
        skipped_turns=resolved_summary.get("skipped_count", 0),
    )


def _get_scoring_runtime_status(conv_id: str) -> tuple[bool, str]:
    ctrl = task_control.get(f"score_{conv_id}")
    ctrl_status = str(getattr(ctrl, "status", "") or "").strip().lower()
    if ctrl_status in {"running", "paused"}:
        return True, ctrl_status
    if get_live_scoring_dispatcher().has_activity(conv_id):
        return True, ctrl_status or "running"
    return False, ctrl_status


def _build_scoring_action_state(
    conv_id: str,
    conversation: dict,
    summary: dict | None = None,
    ai_report_meta: dict | None = None,
) -> dict:
    resolved_summary = summary or _build_scoring_summary(conversation)
    resolved_report_meta = ai_report_meta or _build_ai_report_meta(
        conversation, resolved_summary
    )
    total_count = int(resolved_summary.get("total_count", 0) or 0)
    scored_count = int(resolved_summary.get("scored_count", 0) or 0)
    failed_count = int(resolved_summary.get("failed_count", 0) or 0)
    skipped_count = int(resolved_summary.get("skipped_count", 0) or 0)
    settled_count = scored_count + failed_count + skipped_count
    pending_count = max(0, total_count - settled_count)
    scoring_active, scoring_task_status = _get_scoring_runtime_status(conv_id)
    report_status = str(
        resolved_report_meta.get("ai_report_status", "") or ""
    ).strip().lower()
    repair_summary_needed = (
        total_count > 0
        and pending_count == 0
        and scored_count > 0
        and report_status != "ready"
    )

    if scoring_active:
        recommended_action = "resume_sync"
        recommended_action_label = "继续同步"
        recommended_action_detail = "后台处理中，继续同步结果"
    elif repair_summary_needed:
        recommended_action = "repair_summary"
        recommended_action_label = "汇总评分"
        recommended_action_detail = "重算均分并补生成 AI 报告"
    elif pending_count > 0 and scored_count > 0:
        recommended_action = "retry_failed_turns"
        recommended_action_label = "重试失败项"
        recommended_action_detail = "仅补失败或未完成轮次"
    elif pending_count > 0 and failed_count > 0:
        recommended_action = "rescore_all"
        recommended_action_label = "重新全部打分"
        recommended_action_detail = "当前没有可用评分结果，整段重打"
    elif pending_count > 0:
        recommended_action = "start_scoring"
        recommended_action_label = "开始打分"
        recommended_action_detail = "首次发起整段评分"
    elif scored_count > 0:
        recommended_action = "view_results"
        recommended_action_label = "查看结果"
        recommended_action_detail = "当前评分结果已可查看"
    elif failed_count > 0:
        recommended_action = "rescore_all"
        recommended_action_label = "重新全部打分"
        recommended_action_detail = "当前没有可用评分结果，整段重打"
    else:
        recommended_action = "start_scoring"
        recommended_action_label = "开始打分"
        recommended_action_detail = "首次发起整段评分"

    return {
        "scoring_active": scoring_active,
        "scoring_task_status": scoring_task_status,
        "has_scored_turns": scored_count > 0,
        "has_pending_turns": pending_count > 0,
        "pending_count": pending_count,
        "all_turns_settled": total_count > 0 and pending_count == 0,
        "repair_summary_needed": repair_summary_needed,
        "recommended_action": recommended_action,
        "recommended_action_label": recommended_action_label,
        "recommended_action_detail": recommended_action_detail,
    }


def _build_scoring_state_view(
    conv_id: str,
    conversation: dict,
    summary: dict | None = None,
    ai_report_meta: dict | None = None,
) -> dict:
    resolved_summary = summary or _build_scoring_summary(conversation)
    resolved_report_meta = ai_report_meta or _build_ai_report_meta(
        conversation, resolved_summary
    )
    action_state = _build_scoring_action_state(
        conv_id,
        conversation,
        resolved_summary,
        resolved_report_meta,
    )
    summary_payload = {
        **resolved_summary,
        "report_status": resolved_report_meta.get("ai_report_status", ""),
        "report_label": resolved_report_meta.get("ai_report_label", ""),
        "report_ready": resolved_report_meta.get("ai_report_ready", False),
        "report_updated_at": resolved_report_meta.get("ai_report_updated_at", ""),
        **action_state,
    }
    meta_payload = {
        "model_id": conversation.get("model_id", ""),
        "summary_prompt_version": conversation.get("summary_prompt_version", ""),
        "scoring_prompt_version": conversation.get("scoring_prompt_version", ""),
        "scoring_model_id": conversation.get("scoring_model_id", ""),
        "prompt_version": conversation.get("prompt_version", ""),
        "ai_report_status": resolved_report_meta.get("ai_report_status", ""),
        "ai_report_label": resolved_report_meta.get("ai_report_label", ""),
        "ai_report_ready": resolved_report_meta.get("ai_report_ready", False),
        "ai_report_updated_at": resolved_report_meta.get("ai_report_updated_at", ""),
        "dialogue_summary": next(
            (
                str(item.get("dialogue_summary", "")).strip()
                for item in reversed(conversation.get("results", []))
                if str(item.get("dialogue_summary", "")).strip()
            ),
            "",
        ),
        **action_state,
    }
    return {
        "summary": summary_payload,
        "meta": meta_payload,
        "action": action_state,
    }


async def _push_score_refresh(conv_id: str, turn: int | None = None):
    refreshed = _get_visible_conversation_or_404(conv_id)
    summary = _build_scoring_summary(refreshed)
    ai_report_meta = _build_ai_report_meta(refreshed, summary)
    await _push_score_message(
        conv_id,
        {
            "type": "score_updated",
            "conversation_id": conv_id,
            "turn": turn,
            "avg_total": summary.get("avg_total", 0),
            "scored_count": summary.get("scored_count", 0),
            "failed_count": summary.get("failed_count", 0),
            "skipped_count": summary.get("skipped_count", 0),
            "total_count": summary.get("total_count", 0),
            "summary": summary,
            "report_status": ai_report_meta.get("ai_report_status", ""),
            "report_label": ai_report_meta.get("ai_report_label", ""),
            "report_ready": ai_report_meta.get("ai_report_ready", False),
            "report_updated_at": ai_report_meta.get("ai_report_updated_at", ""),
            "report_event": ai_report_meta.get("ai_report_event", ""),
            "report_meta": ai_report_meta,
        },
    )


async def _generate_conversation_ai_summary(
    conv_id: str,
    conversation: dict,
    config: dict,
    service: ScoringService,
    summary: dict,
) -> dict | None:
    if int(summary.get("scored_count", 0) or 0) <= 0:
        return None
    _record_scoring_event(
        conv_id,
        "summary_generation_started",
        detail={"scored_count": summary.get("scored_count", 0)},
    )
    await _push_score_message(
        conv_id,
        {
            "type": "report_status",
            "conversation_id": conv_id,
            "report_status": "generating",
            "report_label": "报告生成中",
        },
    )
    try:
        report = await service.generate_ai_summary(
            _build_ai_summary_items(conversation),
            config,
            conversation_id=conv_id,
        )
    except Exception as exc:
        report = {"error": str(exc)}

    if report and report.get("error"):
        _record_scoring_event(
            conv_id,
            "summary_preheat_failed",
            level="warn",
            detail={"error": str(report.get("error") or "")},
        )
        await _push_score_message(
            conv_id,
            {
                "type": "report_status",
                "conversation_id": conv_id,
                "report_status": "failed",
                "report_label": "报告生成失败",
                "message": str(report.get("error") or ""),
            },
        )
    else:
        await _push_score_message(
            conv_id,
            {
                "type": "report_status",
                "conversation_id": conv_id,
                "report_status": "ready",
                "report_label": "报告就绪",
                "report_ready": True,
            },
        )
    await _push_score_refresh(conv_id)
    return report


async def _score_single_turn(
    conv_id: str,
    turn_result: dict,
    config: dict,
    require_available: bool,
):
    history_context = db.get_history_context(
        conv_id,
        turn_result.get("turn", 0),
    )
    service = _get_scoring()
    runtime = config.get("runtime", {})
    scoring_model_id = runtime.get("scoring_model_id", "")
    if not service.is_available(scoring_model_id):
        if require_available:
            raise HTTPException(
                status_code=503,
                detail=service.get_last_error() or "评分服务不可用",
            )
        return None

    scoring_thinking_effort = _resolve_scoring_thinking_effort(
        service,
        model_id=scoring_model_id,
        thinking_enabled=runtime.get("scoring_thinking_enabled", None),
        thinking_effort=runtime.get("scoring_thinking_effort", ""),
    )
    score_result = await invoke_score_turn_compat(
        service,
        _build_turn_payload(turn_result, config, history_context=history_context),
        prompt_version=runtime.get("scoring_prompt_version", ""),
        model_id=scoring_model_id,
        thinking_effort=scoring_thinking_effort,
    )
    db.update_turn_scores(
        conv_id,
        turn_result.get("turn", 0),
        {
            **score_result.get("scores", {}),
            "mapped_total": score_result.get("mapped_total", 0),
            "reasoning": score_result.get("reasoning", ""),
            "success": score_result.get("success", False),
        },
    )

    refreshed = _get_visible_conversation_or_404(conv_id)
    if not refreshed:
        return None
    return _find_turn_result(refreshed, turn_result.get("turn", 0))


async def score_turn_if_available(conv_id: str, turn: int, config: dict | None = None):
    """供对话链自动评分复用；评分服务不可用时静默跳过。"""
    conversation = _get_visible_conversation_or_404(conv_id)
    turn_result = _find_turn_result(conversation, turn)
    if not turn_result:
        return None
    return await _score_single_turn(
        conv_id=conv_id,
        turn_result=turn_result,
        config=config or conversation.get("config", {}),
        require_available=False,
    )


def _resolve_retry_schedule(service: ScoringService) -> tuple[float, ...]:
    schedule = tuple(getattr(service, "_default_retry_delays", ()) or ())
    return schedule or (5.0, 15.0, 30.0)


def _resolve_retry_count(config: dict) -> int:
    runtime = dict(config.get("runtime", {}) or {})
    try:
        return max(1, int(runtime.get("scoring_retry_count", 3) or 3))
    except (TypeError, ValueError):
        return 3


def _sanitize_scoring_error(service: ScoringService, message: str) -> str:
    sanitizer = getattr(service, "_sanitize_error_message", None)
    if callable(sanitizer):
        return sanitizer(message)
    return str(message or "").strip()


def _resolve_live_scoring_timeout(config: dict) -> float:
    runtime = dict((config or {}).get("runtime", {}) or {})
    raw = runtime.get(
        "live_scoring_timeout_s",
        os.environ.get(LIVE_SCORING_TIMEOUT_ENV, str(LIVE_SCORING_DEFAULT_TIMEOUT_S)),
    )
    try:
        return max(1.0, float(raw))
    except (TypeError, ValueError):
        return LIVE_SCORING_DEFAULT_TIMEOUT_S


async def _push_live_score_event(
    conv_id: str,
    event_type: str,
    *,
    turn: int,
    **extra,
) -> None:
    await _push_score_message(
        conv_id,
        {
            "type": event_type,
            "conversation_id": conv_id,
            "turn": int(turn or 0),
            **extra,
        },
    )


async def _sleep_with_control(ctrl, delay_s: float) -> None:
    if delay_s <= 0:
        return
    if ctrl is None:
        await asyncio.sleep(delay_s)
        return
    loop = asyncio.get_running_loop()
    deadline = loop.time() + delay_s
    while True:
        await ctrl.checkpoint()
        remaining = deadline - loop.time()
        if remaining <= 0:
            return
        await asyncio.sleep(min(0.25, remaining))


async def _push_live_score_progress(
    conv_id: str,
    turn: int,
    *,
    score_result: dict,
    summary: dict,
) -> None:
    settled = (
        int(summary.get("scored_count", 0) or 0)
        + int(summary.get("failed_count", 0) or 0)
        + int(summary.get("skipped_count", 0) or 0)
    )
    await _push_score_message(
        conv_id,
        {
            "type": "score_progress",
            "conversation_id": conv_id,
            "turn": turn,
            "total": int(summary.get("total_count", 0) or 0),
            "current": settled,
            "score": score_result.get("mapped_total", 0),
            "success": bool(score_result.get("success", False)),
            "failed_count": int(summary.get("failed_count", 0) or 0),
            "skipped_count": int(summary.get("skipped_count", 0) or 0),
        },
    )


async def _run_live_scoring_turn(
    conv_id: str,
    turn: int,
    config: dict | None = None,
) -> dict | None:
    try:
        conversation = _get_visible_conversation_or_404(conv_id)
    except HTTPException:
        return None

    turn_result = _find_turn_result(conversation, turn)
    if not turn_result:
        return None
    if _normalize_turn_score_status(turn_result) in {"scored", "failed", "skipped"}:
        return None

    config = deepcopy(config or conversation.get("config", {}) or {})
    service = _get_scoring()
    runtime = dict(config.get("runtime", {}) or {})
    scoring_model_id = str(runtime.get("scoring_model_id", "") or "").strip()
    retry_schedule = _resolve_retry_schedule(service)
    max_attempts = _resolve_retry_count(config)
    live_timeout_s = _resolve_live_scoring_timeout(config)
    task_id = f"live_score_{conv_id}"
    ctrl = task_control.get(task_id) or task_control.get_or_create(task_id)
    scoring_thinking_effort = _resolve_scoring_thinking_effort(
        service,
        model_id=scoring_model_id,
        thinking_enabled=runtime.get("scoring_thinking_enabled", None),
        thinking_effort=runtime.get("scoring_thinking_effort", ""),
    )
    service_config = getattr(service, "_config", None) or {}
    dims = list(service_config.get("dimensions", []) or DIMENSIONS)
    _record_scoring_event(
        conv_id,
        "score_started",
        detail={"turn": turn, "timeout_s": live_timeout_s},
    )
    await _push_live_score_event(
        conv_id,
        "score_started",
        turn=turn,
        timeout_s=live_timeout_s,
    )

    if not str(turn_result.get("ai_output", "") or "").strip():
        score_result = {
            "success": False,
            "scores": {dimension: 0 for dimension in dims},
            "weighted_total": 0,
            "mapped_total": 0,
            "reasoning": "[跳过] ai_output 为空，未发起打分",
            "reasoning_content": "",
            "error": "",
            "model_id": scoring_model_id,
            "score_status": "skipped",
        }
    else:
        score_result = None
        for attempt in range(1, max_attempts + 1):
            await ctrl.checkpoint()
            _record_scoring_event(
                conv_id,
                "score_attempt",
                detail={"turn": turn, "attempt": attempt, "max_attempts": max_attempts},
            )
            await _push_live_score_event(
                conv_id,
                "score_attempt",
                turn=turn,
                attempt=attempt,
                max_attempts=max_attempts,
            )
            try:
                if not service.is_available(scoring_model_id):
                    raise RuntimeError(service.get_last_error() or "评分服务不可用")
                history_context = db.get_history_context(conv_id, turn_result.get("turn", 0))
                score_result = await invoke_score_turn_compat(
                    service,
                    _build_turn_payload(turn_result, config, history_context=history_context),
                    timeout_s=live_timeout_s,
                    retry_delays=(),
                    provider_retry_delays=(),
                    prompt_version=runtime.get("scoring_prompt_version", ""),
                    model_id=scoring_model_id,
                    thinking_effort=scoring_thinking_effort,
                )
            except asyncio.CancelledError:
                raise
            except Exception as exc:
                sanitized_err = _sanitize_scoring_error(service, str(exc))
                score_result = {
                    "success": False,
                    "scores": {dimension: 0 for dimension in dims},
                    "weighted_total": 0,
                    "mapped_total": 0,
                    "reasoning": f"[打分异常] {sanitized_err}",
                    "reasoning_content": "",
                    "error": sanitized_err,
                    "model_id": scoring_model_id,
                    "score_status": "failed",
                }
            if score_result.get("success") or score_result.get("score_status") == "skipped":
                break
            if attempt >= max_attempts:
                break
            delay_s = retry_schedule[min(attempt - 1, len(retry_schedule) - 1)]
            _record_scoring_event(
                conv_id,
                "score_waiting_retry",
                level="warn",
                detail={
                    "turn": turn,
                    "attempt": attempt,
                    "max_retries": max_attempts,
                    "next_delay_s": delay_s,
                },
            )
            await _push_live_score_event(
                conv_id,
                "score_waiting_retry",
                turn=turn,
                attempt=attempt,
                max_retries=max_attempts,
                next_delay_s=delay_s,
            )
            await _push_score_message(
                conv_id,
                {
                    "type": "retry",
                    "conversation_id": conv_id,
                    "turn": turn,
                    "attempt": attempt,
                    "max_retries": max_attempts,
                    "next_delay_s": delay_s,
                },
            )
            await _sleep_with_control(ctrl, delay_s)

    _invalidate_conversation_scoring_summary(conv_id)
    db.update_turn_scores(
        conv_id,
        turn,
        {
            **score_result.get("scores", {}),
            "mapped_total": score_result.get("mapped_total", 0),
            "reasoning": score_result.get("reasoning", ""),
            "success": bool(score_result.get("success", False)),
            "score_status": score_result.get("score_status", ""),
        },
    )
    refreshed = _get_visible_conversation_or_404(conv_id)
    summary = _build_scoring_summary(refreshed)
    if score_result.get("score_status") == "failed":
        _record_scoring_event(
            conv_id,
            "score_final_failed",
            level="warn",
            detail={
                "turn": turn,
                "error": score_result.get("error", "") or score_result.get("reasoning", ""),
            },
        )
        await _push_live_score_event(
            conv_id,
            "score_final_failed",
            turn=turn,
            error=score_result.get("error", "") or score_result.get("reasoning", ""),
        )
        _record_scoring_event(
            conv_id,
            "turn_failed",
            level="warn",
            detail={
                "turn": turn,
                "error": score_result.get("error", "") or score_result.get("reasoning", ""),
            },
        )
    else:
        _record_scoring_event(
            conv_id,
            "turn_scored",
            detail={
                "turn": turn,
                "score": score_result.get("mapped_total", 0),
                "success": bool(score_result.get("success", False)),
                "status": score_result.get("score_status", ""),
            },
        )
    await _push_live_score_progress(
        conv_id,
        turn,
        score_result=score_result,
        summary=summary,
    )
    await _push_score_refresh(conv_id, turn)
    return score_result


async def enqueue_live_score_turn(
    conv_id: str,
    turn: int,
    *,
    config: dict | None = None,
) -> bool:
    enqueued = await get_live_scoring_dispatcher().enqueue(conv_id, turn, config=config)
    if enqueued:
        _record_scoring_event(conv_id, "score_enqueued", detail={"turn": turn})
        await _push_live_score_event(conv_id, "score_enqueued", turn=turn)
    return enqueued


async def enqueue_pending_live_scores(conv_id: str, *, config: dict | None = None) -> int:
    conversation = _get_visible_conversation_or_404(conv_id)
    resolved_config = config or conversation.get("config", {})
    pending_turns = [
        int(result.get("turn", 0) or 0)
        for result in conversation.get("results", []) or []
        if _normalize_turn_score_status(result) == "unscored"
    ]
    return await get_live_scoring_dispatcher().enqueue_pending(
        conv_id,
        pending_turns,
        config=resolved_config,
    )


def _store_failed_scores(conv_id: str, results: list[dict], reason: str):
    for result in results:
        if result.get("score_status") == "scored":
            continue
        db.update_turn_scores(
            conv_id,
            result.get("turn", 0),
            {
                **{dimension: 0 for dimension in DIMENSIONS},
                "mapped_total": 0,
                "reasoning": f"[评分失败] {reason}",
                "success": False,
            },
        )


async def _push_score_message(conv_id: str, data: dict):
    connections = list(_score_ws.get(conv_id) or [])
    if not connections:
        return

    message = json.dumps(data, ensure_ascii=False, default=str)
    dead_connections = []
    for websocket in connections:
        try:
            await websocket.send_text(message)
        except Exception:
            dead_connections.append(websocket)

    current_connections = _score_ws.get(conv_id)
    if not current_connections:
        return
    for websocket in dead_connections:
        if websocket in current_connections:
            current_connections.remove(websocket)
    if conv_id in _score_ws and not current_connections:
        del _score_ws[conv_id]


async def _invoke_score_rows_compat(
    service,
    rows: list[dict],
    *,
    on_progress=None,
    model_id: str | None = None,
    prompt_version: str = "",
    thinking_effort: str = "",
):
    score_rows = service.score_rows
    forwarded = {
        "on_progress": on_progress,
        "model_id": model_id,
        "prompt_version": prompt_version,
        "thinking_effort": thinking_effort,
    }
    try:
        signature = inspect.signature(score_rows)
        parameters = signature.parameters
        accepts_var_kwargs = any(
            param.kind == inspect.Parameter.VAR_KEYWORD
            for param in parameters.values()
        )
    except (TypeError, ValueError):
        parameters = {}
        accepts_var_kwargs = False

    if accepts_var_kwargs:
        filtered_kwargs = forwarded
    else:
        filtered_kwargs = {
            key: value for key, value in forwarded.items() if key in parameters
        }

    result = score_rows(rows, **filtered_kwargs)
    if inspect.isawaitable(result):
        return await result
    return result


@router.get("/status")
async def scoring_status(model_id: str = Query(default="")):
    """检查评分服务可用性。"""
    service = _get_scoring()
    available = service.is_available(model_id or None)
    prompt_meta = service.get_prompt_meta() if available else {}
    return {
        "available": available,
        "prompts": service.get_scoring_prompts(),
        "active_prompt": prompt_meta.get("active_filename", ""),
        "max_workers": service.get_max_workers(),
        "error": service.get_last_error() if not available else "",
    }


@router.post("/config")
async def update_scoring_config(data: ScoringConfigUpdate):
    """更新打分配置（当前仅支持并发上限）。"""
    service = _get_scoring()
    if data.max_workers is not None:
        service.set_max_workers(data.max_workers)
        await get_live_scoring_dispatcher().notify_capacity_changed()
    return {"max_workers": service.get_max_workers()}


@router.get("/config")
async def get_scoring_config():
    """获取当前评分配置。"""
    service = _get_scoring()
    return {"max_workers": service.get_max_workers()}


@router.get("/dimensions")
async def scoring_dimensions(model_id: str = Query(default="")):
    """获取评分维度配置。"""
    service = _get_scoring()
    if not service.is_available(model_id or None):
        raise HTTPException(
            status_code=503,
            detail=service.get_last_error() or "评分服务不可用",
        )
    return service.get_dimensions(model_id or None)


@router.post("/upload")
async def upload_scoring_excel(
    file: UploadFile = File(...),
    scoring_model_id: str = Form(default=""),
    scoring_prompt_version: str = Form(default=""),
    scoring_thinking_enabled: bool | None = Form(default=None),
    scoring_thinking_effort: str = Form(default=""),
):
    """上传 Excel 并直接批量评分。"""
    filename = file.filename or ""
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix not in {"xlsx", "xls"}:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx/.xls 文件")

    service = _get_scoring()
    if not service.is_available(scoring_model_id or None):
        raise HTTPException(
            status_code=503,
            detail=service.get_last_error() or "评分服务不可用",
        )

    temp_name = f"upload_{uuid.uuid4().hex}.{suffix}"
    temp_path = OUTPUT_DIR / temp_name
    with temp_path.open("wb") as handle:
        handle.write(await file.read())

    try:
        rows = _export_service.import_from_excel(str(temp_path))
        if not rows:
            raise HTTPException(status_code=400, detail="Excel 中没有可评分数据")

        resolved_scoring_thinking_effort = _resolve_scoring_thinking_effort(
            service,
            model_id=scoring_model_id or "",
            thinking_enabled=scoring_thinking_enabled,
            thinking_effort=scoring_thinking_effort,
        )
        scored_rows = await _invoke_score_rows_compat(
            service,
            rows,
            model_id=scoring_model_id or None,
            prompt_version=scoring_prompt_version or "",
            thinking_effort=resolved_scoring_thinking_effort,
        )
        flattened_rows = _flatten_scored_rows(scored_rows)

        original_stem = filename.rsplit(".", 1)[0] if "." in filename else "batch"
        safe_stem = _export_service.safe_filename_part(original_stem, fallback="batch")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_name = f"{safe_stem}_{timestamp}_scored.xlsx"
        output_path = OUTPUT_DIR / output_name
        _export_service.export_rows_to_excel(
            flattened_rows, str(output_path), sheet_name="批量评分结果"
        )

        return {
            "status": "completed",
            "rows_scored": len(flattened_rows),
            "output_file": str(output_path),
            "download_url": f"/api/scoring/output/{output_name}",
        }
    finally:
        if temp_path.exists():
            temp_path.unlink()


@router.post("/test-file/import")
async def import_scoring_test_file(file: UploadFile = File(...)):
    """导入测试 Excel，仅做结构预览。"""
    filename = file.filename or ""
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix not in {"xlsx", "xls"}:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx/.xls 文件")

    temp_name = f"preview_{uuid.uuid4().hex}.{suffix}"
    temp_path = OUTPUT_DIR / temp_name
    with temp_path.open("wb") as handle:
        handle.write(await file.read())

    try:
        rows = _export_service.import_from_excel(str(temp_path))
        headers = list(rows[0].keys()) if rows else []
        return {
            "rows": len(rows),
            "headers": headers,
            "preview": rows[:3],
        }
    finally:
        if temp_path.exists():
            temp_path.unlink()


@router.get("/template/download")
async def download_scoring_template():
    """下载评分模板。"""
    import openpyxl

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "打分模板"
    headers = [
        "session_id",
        "turn_order",
        "测试对应提示词",
        "Role_Nickname",
        "gender",
        "age",
        "occupation",
        "Role_info_works",
        "personality",
        "speaking_style",
        "personal_type",
        "background",
        "hobby",
        "system_module8",
        "user_Nickname",
        "user_gender",
        "user_identity",
        "relationship",
        "relation_calling",
        "intimacy_boundary",
        "current_scene",
        "currentTime",
        "weekDay",
        "timeperiod",
        "season",
        "last_cst_type",
        "完整时间信息",
        "longform_persona",
        "longform_narrative_style",
        "longform_dialogue_guideline",
        "longform_few_shot",
        "moments",
        "weekly_schedule",
        "monthly_schedule",
        "conversation_history",
        "dialogueStartPrompt",
        "dialogue_summary",
        "用户输入",
        "AI输出",
    ]
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column, value=header)

    output_path = OUTPUT_DIR / "长文模式_打分模板.xlsx"
    workbook.save(str(output_path))
    workbook.close()
    return FileResponse(
        path=str(output_path),
        filename="长文模式_打分模板.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/output/{filename}")
async def download_scored_output(filename: str):
    """下载批量评分输出文件。"""
    safe_name = _export_service.safe_filename_part(filename, fallback="")
    if not safe_name or safe_name != filename:
        raise HTTPException(status_code=400, detail="非法文件名")

    output_path = OUTPUT_DIR / filename
    if not output_path.exists():
        raise HTTPException(status_code=404, detail="文件不存在")

    return FileResponse(
        path=str(output_path),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.post("/manual-dialogue")
async def score_manual_dialogue(data: ManualDialogueRequest):
    """对手动输入的多轮对话逐轮评分。"""
    if not data.turns:
        raise HTTPException(status_code=400, detail="turns 不能为空")

    service = _get_scoring()
    config = data.config or {}
    runtime = config.get("runtime", {})
    scoring_model_id = runtime.get("scoring_model_id", "")
    if not service.is_available(scoring_model_id):
        raise HTTPException(
            status_code=503,
            detail=service.get_last_error() or "评分服务不可用",
        )

    scoring_thinking_effort = _resolve_scoring_thinking_effort(
        service,
        model_id=scoring_model_id,
        thinking_enabled=runtime.get("scoring_thinking_enabled", None),
        thinking_effort=runtime.get("scoring_thinking_effort", ""),
    )
    character = config.get("character", {})
    context = config.get("context", {})
    prompt_name = config.get("prompt_file", "")

    rows = []
    for index, turn in enumerate(data.turns, start=1):
        if not turn.ai_output:
            raise HTTPException(
                status_code=400,
                detail=f"第 {index} 轮缺少 ai_output，无法直接评分",
            )
        rows.append(
            {
                "turn": turn.turn or index,
                "用户输入": turn.user_input,
                "AI输出": turn.ai_output,
                "测试对应提示词": prompt_name,
                "Role_Nickname": character.get("Role_Nickname", ""),
                "personality": character.get("personality", ""),
                "relationship": context.get("relationship", ""),
                "dialogueStartPrompt": dict(config.get("modules", {}) or {}).get("dialogueStartPrompt", ""),
                "moments": dict(config.get("modules", {}) or {}).get("moments", ""),
            }
        )

    scored_rows = []
    for row in rows:
        scored_rows.append(
            {
                **row,
                **await invoke_score_turn_compat(
                    service,
                    {
                        "turn": row.get("turn", 0),
                        "user_input": row.get("用户输入", ""),
                        "ai_output": row.get("AI输出", ""),
                        "role_name": row.get("Role_Nickname", ""),
                        "personality": row.get("personality", ""),
                        "relationship": row.get("relationship", ""),
                        "prompt_name": row.get("测试对应提示词", ""),
                        "dialogueStartPrompt": row.get("dialogueStartPrompt", ""),
                        "moments": row.get("moments", ""),
                        "history_context": row.get("history_context", "") or row.get("近期对话历史", ""),
                    },
                    prompt_version=runtime.get("scoring_prompt_version", ""),
                    model_id=scoring_model_id,
                    thinking_effort=scoring_thinking_effort,
                ),
            }
        )
    flattened_rows = _flatten_scored_rows(scored_rows)
    turns = []
    for row in flattened_rows:
        turns.append(
            {
                "turn": row.get("turn", 0),
                "scores": {
                    dimension: row.get(f"score_{dimension}", 0)
                    for dimension in DIMENSIONS
                },
                "total": row.get("score_total", 0),
                "reasoning": row.get("score_reasoning", ""),
                "status": row.get("score_status", "unscored"),
            }
        )
    scored_turns = [item for item in turns if item["status"] == "scored"]
    return {
        "turns": turns,
        "summary": {
            "avg_total": round(
                sum(item["total"] for item in scored_turns) / len(scored_turns),
                2,
            ) if scored_turns else 0,
            "total_count": len(turns),
            "scored_count": len(scored_turns),
        },
    }


@router.get("/multi-model/export")
async def export_multi_model_scoring(
    conv_ids: str = Query(..., description="逗号分隔的对话 ID"),
    summary: bool = False,
):
    """导出多模型评分结果。"""
    import openpyxl
    from openpyxl.styles import Alignment

    conv_id_list = [item.strip() for item in conv_ids.split(",") if item.strip()]
    if not conv_id_list:
        raise HTTPException(status_code=400, detail="conv_ids 不能为空")

    conversations = []
    for conv_id in conv_id_list:
        conversation = _get_visible_conversation_or_404(conv_id)
        conversations.append(conversation)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = "summary" if summary else "full"
    filename = f"multi_model_scoring_{timestamp}_{suffix}.xlsx"
    output_path = OUTPUT_DIR / filename

    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "摘要"
    summary_headers = ["conv_id", "角色", "模型", "总轮数", "已评分轮数", "平均总分"]
    _export_service._style_header(summary_sheet, summary_headers)

    for row_index, conversation in enumerate(conversations, start=2):
        results = conversation.get("results", [])
        scored_rows = [row for row in results if row.get("score_status") == "scored"]
        config = conversation.get("config", {})
        summary_values = [
            conversation.get("id"),
            config.get("character", {}).get("Role_Nickname", ""),
            conversation.get("model_id", ""),
            len(results),
            len(scored_rows),
            round(
                sum(row.get("score_total", 0) for row in scored_rows) / len(scored_rows),
                2,
            ) if scored_rows else 0,
        ]
        for column, value in enumerate(summary_values, start=1):
            summary_sheet.cell(row=row_index, column=column, value=value)

    if not summary:
        for conversation in conversations:
            config = conversation.get("config", {})
            role_name = config.get("character", {}).get("Role_Nickname", "conversation")
            sheet = workbook.create_sheet(
                _export_service.safe_filename_part(
                    f"{role_name}_{conversation.get('id', '')}",
                    fallback="sheet",
                )[:28]
            )
            headers = list(_export_service.DETAIL_HEADERS)
            border = _export_service._style_header(sheet, headers)
            for turn_index, result in enumerate(conversation.get("results", []), start=2):
                values = _export_service._build_detail_row(
                    result, config, turn_index - 2
                )
                for col_idx, value in enumerate(values, start=1):
                    cell = sheet.cell(row=turn_index, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            _export_service._apply_column_widths(sheet, headers)

    workbook.save(str(output_path))
    workbook.close()
    return FileResponse(
        path=str(output_path),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/dialogue/export")
async def export_dialogue_scoring(session_id: str, summary: bool = False):
    """兼容 PRD 路径：按 session_id 导出多轮对话评分结果。"""
    return await export_scoring(session_id, summary=summary)


@router.post("/{conv_id}")
async def trigger_scoring(
    conv_id: str,
    data: TriggerScoringRequest | None = Body(default=None),
):
    """触发对话评分。"""
    conversation = _get_visible_conversation_or_404(conv_id)

    results = conversation.get("results", [])
    if not results:
        raise HTTPException(status_code=400, detail="对话无结果数据")

    pending_results = _select_pending_scoring_results(results)
    force_full_rescore = not pending_results and _should_force_full_rescore(conversation, data)
    if not pending_results and not force_full_rescore:
        scoring_view = _build_scoring_state_view(conv_id, conversation)
        return {
            "status": "already_scored",
            "conversation_id": conv_id,
            "turns_to_score": 0,
            "summary": scoring_view["summary"],
            "meta": scoring_view["meta"],
            "action": scoring_view["action"],
        }
    target_results = list(results if force_full_rescore else pending_results)

    service = _get_scoring()
    config = _merge_scoring_runtime(conversation.get("config", {}), data)
    runtime = config.get("runtime", {})
    scoring_model_id = runtime.get("scoring_model_id", "")
    if data and data.max_workers is not None:
        service.set_max_workers(data.max_workers)
    if not service.is_available(scoring_model_id):
        raise HTTPException(
            status_code=503,
            detail=service.get_last_error() or "评分服务不可用",
        )
    _persist_scoring_runtime(conv_id, config)

    existing_ctrl = task_control.get(f"score_{conv_id}")
    if existing_ctrl and existing_ctrl.status in {"running", "paused"}:
        return {
            "status": existing_ctrl.status,
            "conversation_id": conv_id,
            "turns_to_score": len(target_results),
        }

    await get_live_scoring_dispatcher().cancel_conversation(conv_id)
    if force_full_rescore:
        db.reset_conversation_scores(conv_id)
    _invalidate_conversation_scoring_summary(conv_id)
    task_control.remove(f"score_{conv_id}")
    task_control.get_or_create(f"score_{conv_id}")
    _record_scoring_event(
        conv_id,
        "rescore_started" if force_full_rescore else "started",
        detail={"turns_to_score": len(target_results)},
    )
    asyncio.create_task(
        _run_scoring(conv_id, target_results, config, service)
    )
    return {
        "status": "rescore_started" if force_full_rescore else "scoring_started",
        "conversation_id": conv_id,
        "turns_to_score": len(target_results),
    }


@router.post("/{conv_id}/resume-sync")
async def resume_scoring_sync(conv_id: str):
    """继续同步评分状态，不重复触发新评分任务。"""
    conversation = _get_visible_conversation_or_404(conv_id)
    scoring_view = _build_scoring_state_view(conv_id, conversation)
    await _push_score_refresh(conv_id)
    return {
        "status": "sync_resumed" if scoring_view["action"]["scoring_active"] else "sync_refreshed",
        "conversation_id": conv_id,
        "summary": scoring_view["summary"],
        "meta": scoring_view["meta"],
        "action": scoring_view["action"],
    }


@router.post("/{conv_id}/repair-summary")
async def repair_scoring_summary(
    conv_id: str,
    data: TriggerScoringRequest | None = Body(default=None),
):
    """重算会话均分并补生成评分摘要报告。"""
    conversation = _get_visible_conversation_or_404(conv_id)
    results = conversation.get("results", [])
    if not results:
        raise HTTPException(status_code=400, detail="对话无结果数据")

    config = _merge_scoring_runtime(conversation.get("config", {}), data)
    _persist_scoring_runtime(conv_id, config)

    summary = _build_scoring_summary(conversation)
    ai_report_meta = _build_ai_report_meta(conversation, summary)
    report = None
    if (
        int(summary.get("scored_count", 0) or 0) > 0
        and str(ai_report_meta.get("ai_report_status", "") or "").strip().lower() != "ready"
    ):
        report = await _generate_conversation_ai_summary(
            conv_id,
            conversation,
            config,
            _get_scoring(),
            summary,
        )
    else:
        await _push_score_refresh(conv_id)

    refreshed = _get_visible_conversation_or_404(conv_id)
    scoring_view = _build_scoring_state_view(conv_id, refreshed)
    return {
        "status": "summary_repaired" if report is not None else "summary_refreshed",
        "conversation_id": conv_id,
        "summary": scoring_view["summary"],
        "meta": scoring_view["meta"],
        "action": scoring_view["action"],
        "report": report,
    }


@router.post("/{conv_id}/retry-failed-turns")
async def retry_failed_turns(
    conv_id: str,
    data: TriggerScoringRequest | None = Body(default=None),
):
    """仅补失败或未完成轮次。"""
    result = await trigger_scoring(conv_id, data)
    if result.get("status") == "already_scored":
        return result
    return {
        **result,
        "action": {
            **(result.get("action") or {}),
            "recommended_action": "retry_failed_turns",
            "recommended_action_label": "重试失败项",
        },
    }


@router.post("/{conv_id}/turn/{turn}")
async def trigger_single_turn_scoring(
    conv_id: str,
    turn: int,
    data: TriggerScoringRequest | None = Body(default=None),
):
    """触发单轮 AI 打分。"""
    conversation = _get_visible_conversation_or_404(conv_id)

    turn_result = _find_turn_result(conversation, turn)
    if not turn_result:
        raise HTTPException(status_code=404, detail=f"轮次不存在: {turn}")

    config = _merge_scoring_runtime(conversation.get("config", {}), data)
    _invalidate_conversation_scoring_summary(conv_id)
    _persist_scoring_runtime(conv_id, config)
    updated = await _score_single_turn(
        conv_id=conv_id,
        turn_result=turn_result,
        config=config,
        require_available=True,
    )
    _record_scoring_event(conv_id, "turn_scored", detail={"turn": turn})
    await _push_score_refresh(conv_id, turn)
    return {
        "message": "AI 打分完成",
        "conversation_id": conv_id,
        "turn": turn,
        "result": updated,
    }


@router.post("/{conv_id}/rescore-all")
async def rescore_all(conv_id: str, data: RescoreAllRequest):
    """切换评分模型或参数后，全量清空旧分并重新打分。"""
    conversation = _get_visible_conversation_or_404(conv_id)
    results = conversation.get("results", [])
    if not results:
        raise HTTPException(status_code=400, detail="对话无结果数据")

    service = _get_scoring()
    config = _merge_scoring_runtime(conversation.get("config", {}), data)
    runtime = config.get("runtime", {})
    scoring_model_id = runtime.get("scoring_model_id", "")
    if data.max_workers is not None:
        service.set_max_workers(data.max_workers)
    if not service.is_available(scoring_model_id):
        raise HTTPException(
            status_code=503,
            detail=service.get_last_error() or "评分服务不可用",
        )
    _persist_scoring_runtime(conv_id, config)

    await get_live_scoring_dispatcher().cancel_conversation(conv_id)
    db.reset_conversation_scores(conv_id)
    _invalidate_conversation_scoring_summary(conv_id)
    await _push_score_refresh(conv_id)
    task_control.remove(f"score_{conv_id}")
    task_control.get_or_create(f"score_{conv_id}")
    _record_scoring_event(conv_id, "rescore_started", detail={"turns_to_score": len(results)})
    asyncio.create_task(_run_scoring(conv_id, results, config, service))
    return {
        "status": "rescore_started",
        "conversation_id": conv_id,
        "turns_to_score": len(results),
    }


@router.post("/{conv_id}/control")
async def control_scoring(conv_id: str, data: TaskControlRequest = Body(...)):
    conversation = _get_visible_conversation_or_404(conv_id)
    ctrl = task_control.get(f"score_{conv_id}")
    live_dispatcher = get_live_scoring_dispatcher()
    has_live_activity = live_dispatcher.has_activity(conv_id)
    if ctrl is None and not has_live_activity:
        raise HTTPException(status_code=400, detail="当前会话没有可控制的整段评分任务")

    action = data.action
    runtime = ((conversation.get("config", {}) or {}).get("runtime", {}) or {})
    total_turns = int(runtime.get("total_turns", 0) or 0)
    next_turn_index = int(runtime.get("next_turn_index", len(conversation.get("results", []) or [])) or 0)
    generation_complete = total_turns > 0 and next_turn_index >= total_turns
    allow_status_mutation = ctrl is not None or generation_complete or conversation.get("status") in {"completed", "paused", "interrupted"}
    if action == "pause":
        if ctrl is not None:
            ctrl.pause()
        if has_live_activity:
            await live_dispatcher.pause_conversation(conv_id)
        if allow_status_mutation:
            db.update_conversation_status(conv_id, "paused")
        _record_scoring_event(conv_id, "paused", detail={"reason": "control_request", "source": "scoring_control"})
        await _push_scoring_status(
            conv_id,
            "paused",
            event_type="paused",
            message="整段评分已暂停",
            reason="control_request",
            source="scoring_control",
        )
        return {"conversation_id": conv_id, "status": "paused"}

    if action == "resume":
        if ctrl is not None:
            ctrl.resume()
        if has_live_activity:
            await live_dispatcher.resume_conversation(conv_id)
        if allow_status_mutation:
            db.update_conversation_status(conv_id, "completed" if generation_complete else "running")
        _record_scoring_event(conv_id, "resumed", detail={"reason": "control_request", "source": "scoring_control"})
        await _push_scoring_status(
            conv_id,
            "completed" if generation_complete else "running",
            event_type="resumed",
            message="整段评分已恢复",
            reason="control_request",
            source="scoring_control",
        )
        return {"conversation_id": conv_id, "status": "running"}

    if ctrl is not None:
        ctrl.cancel()
    if has_live_activity:
        await live_dispatcher.cancel_conversation(conv_id)
    if allow_status_mutation:
        db.update_conversation_status(conv_id, "cancelled")
    _record_scoring_event(conv_id, "cancelled", detail={"reason": "control_request", "source": "scoring_control"})
    return {"conversation_id": conv_id, "status": "cancelling"}


async def _run_scoring(
    conv_id: str, results: list[dict], config: dict, service: ScoringService
):
    ctrl = task_control.get_or_create(f"score_{conv_id}")

    async def _on_progress(data: dict):
        payload = dict(data or {})
        await _push_score_message(conv_id, payload)
        payload_type = str(payload.get("type", "") or "").strip()
        if payload_type == "retry":
            _record_scoring_event(
                conv_id,
                "retry",
                level="warn",
                detail={
                    "turn": payload.get("turn"),
                    "attempt": payload.get("attempt"),
                    "max_retries": payload.get("max_retries"),
                    "next_delay_s": payload.get("next_delay_s"),
                },
            )
        elif payload_type == "score_progress":
            _record_scoring_event(
                conv_id,
                "turn_scored",
                detail={
                    "turn": payload.get("turn"),
                    "current": payload.get("current"),
                    "total": payload.get("total"),
                    "score": payload.get("score"),
                    "success": payload.get("success"),
                },
            )

    try:
        db.update_conversation_status(conv_id, "running")
        await _push_scoring_status(
            conv_id,
            "running",
            event_type="started",
            message="整段评分开始执行",
        )
        runtime = config.get("runtime", {})
        scored = await service.score_conversation(
            conv_id,
            results,
            config,
            on_progress=_on_progress,
            max_workers=runtime.get("scoring_max_workers", None),
        )

        for item in scored:
            db.update_turn_scores(
                conv_id,
                item["turn"],
                {
                    **item.get("scores", {}),
                    "mapped_total": item.get("mapped_total", 0),
                    "reasoning": item.get("reasoning", ""),
                    "success": item.get("success", False),
                    "score_status": item.get("score_status", ""),
                },
            )
            await _push_score_refresh(conv_id, item["turn"])

        refreshed = _get_visible_conversation_or_404(conv_id)
        summary = _build_scoring_summary(refreshed)
        ctrl_state = task_control.get(f"score_{conv_id}")
        if ctrl_state and ctrl_state.is_cancelled:
            db.update_conversation_status(conv_id, "cancelled")
            _record_scoring_event(
                conv_id,
                "cancelled",
                detail={
                    "scored_count": summary.get("scored_count", 0),
                    "failed_count": summary.get("failed_count", 0),
                },
            )
            await _push_scoring_status(
                conv_id,
                "cancelled",
                event_type="cancelled",
                message="整段评分已取消",
                summary=summary,
            )
            return

        ctrl.complete()
        db.update_conversation_status(conv_id, "completed")
        refreshed = _get_visible_conversation_or_404(conv_id)
        summary = _build_scoring_summary(refreshed)
        _record_scoring_event(
            conv_id,
            "completed",
            detail={
                "total_scored": summary.get("scored_count", 0),
                "total_failed": summary.get("failed_count", 0),
                "avg_score": summary.get("avg_total", 0),
            },
        )
        await _push_scoring_status(
            conv_id,
            "completed",
            event_type="completed",
            total_scored=summary.get("scored_count", 0),
            total_failed=summary.get("failed_count", 0),
            avg_score=summary.get("avg_total", 0),
            summary=summary,
        )
        if summary.get("scored_count", 0) > 0:
            await _generate_conversation_ai_summary(
                conv_id,
                refreshed,
                config,
                service,
                summary,
            )
    except asyncio.CancelledError:
        refreshed = _get_visible_conversation_or_404(conv_id)
        summary = _build_scoring_summary(refreshed)
        db.update_conversation_status(conv_id, "cancelled")
        _record_scoring_event(
            conv_id,
            "cancelled",
            detail={
                "scored_count": summary.get("scored_count", 0),
                "failed_count": summary.get("failed_count", 0),
            },
        )
        await _push_scoring_status(
            conv_id,
            "cancelled",
            event_type="cancelled",
            message="整段评分已取消",
            summary=summary,
        )
    except Exception as exc:
        sanitized = ScoringService._sanitize_error_message(str(exc))
        db.update_conversation_status(conv_id, "failed")
        _store_failed_scores(conv_id, results, sanitized)
        _record_scoring_event(
            conv_id,
            "failed",
            level="error",
            detail={"error": sanitized},
        )
        await _push_scoring_status(
            conv_id,
            "failed",
            event_type="error",
            message=sanitized,
        )
    finally:
        task_control.remove(f"score_{conv_id}")


@router.get("/{conv_id}/results")
async def get_scoring_results(conv_id: str):
    """获取评分结果。"""
    conversation = _get_visible_conversation_or_404(conv_id)
    scoring_view = _build_scoring_state_view(conv_id, conversation)
    summary = scoring_view["summary"]
    meta = scoring_view["meta"]

    turns = []
    for result in conversation.get("results", []):
        turns.append(
            {
                "turn": result.get("turn", 0),
                "scores": {
                    dimension: result.get(f"score_{dimension}", 0)
                    for dimension in DIMENSIONS
                },
                "total": result.get("score_total", 0),
                "reasoning": result.get("score_reasoning", ""),
                "status": result.get("score_status", "unscored"),
                "manual_star_score": result.get("manual_star_score"),
                "manual_comment": result.get("manual_comment", ""),
                "manual_label": _export_service._manual_label(
                    result.get("manual_star_score")
                ),
            }
        )

    return {
        "conversation_id": conv_id,
        "meta": meta,
        "turns": turns,
        "summary": summary,
        "action": scoring_view["action"],
    }


@router.get("/{conv_id}")
async def get_scoring_results_alias(conv_id: str):
    """兼容 PRD 路径：获取评分结果。"""
    return await get_scoring_results(conv_id)


@router.put("/{conv_id}/turn/{turn}/manual")
async def manual_score(conv_id: str, turn: int, data: ManualScoreRequest):
    """人工评分指定轮次。"""
    conversation = _get_visible_conversation_or_404(conv_id)

    if not any(result.get("turn") == turn for result in conversation.get("results", [])):
        raise HTTPException(status_code=404, detail=f"轮次不存在: {turn}")

    db.update_manual_score(conv_id, turn, data.star_score, data.comment)
    return {
        "message": "人工打分已保存",
        "conversation_id": conv_id,
        "turn": turn,
        "star_score": data.star_score,
    }


@router.get("/{conv_id}/export")
async def export_scoring(conv_id: str, summary: bool = False):
    """导出评分结果 Excel。"""
    conversation = _get_visible_conversation_or_404(conv_id)

    results = conversation.get("results", [])
    if not results:
        raise HTTPException(status_code=400, detail="对话无结果数据")

    config = conversation.get("config", {})
    role_name = config.get("character", {}).get("Role_Nickname", "unknown")
    safe_name = _export_service.safe_filename_part(role_name, fallback="scoring")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = "评分摘要" if summary else "评分结果"
    filename = f"{safe_name}_{timestamp}_{suffix}.xlsx"
    output_path = OUTPUT_DIR / filename
    _export_service.export_to_excel(
        results, config, str(output_path), summary=summary
    )

    return FileResponse(
        path=str(output_path),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.post("/{conv_id}/ai-summary")
async def generate_ai_summary(
    conv_id: str,
    model_id: str = Query(default=""),
    prompt_version: str = Query(default=""),
):
    """生成单会话评分摘要报告。"""
    conversation = _get_visible_conversation_or_404(conv_id)
    results = conversation.get("results", [])
    if not results:
        raise HTTPException(status_code=400, detail="对话无结果数据")

    service = _get_scoring()
    config = conversation.get("config", {})
    summary_data = await service.generate_ai_summary(
        _build_ai_summary_items(conversation),
        config,
        model_id=model_id.strip() or None,
        prompt_version=prompt_version.strip() or None,
        conversation_id=conv_id,
    )
    return {"conversation_id": conv_id, "summary": summary_data}


@router.websocket("/ws/{conv_id}")
async def scoring_websocket(websocket: WebSocket, conv_id: str):
    """实时推送评分进度。"""
    await websocket.accept()
    _score_ws.setdefault(conv_id, []).append(websocket)

    try:
        while True:
            payload = await websocket.receive_text()
            try:
                command = json.loads(payload)
            except json.JSONDecodeError:
                continue

            if command.get("action") == "ping":
                await websocket.send_text(json.dumps({"type": "pong"}))
    except Exception:
        connections = _score_ws.get(conv_id)
        if connections and websocket in connections:
            connections.remove(websocket)
            if not connections:
                del _score_ws[conv_id]
