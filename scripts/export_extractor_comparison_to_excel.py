#!/usr/bin/env python3
"""将互动要点抽取模型对比结果导出为Excel"""
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

results_file = Path("output/mode_switching_switch_state/extractor_comparison_real_v2_20260513/comparison_results.jsonl")
output_file = Path("output/mode_switching_switch_state/extractor_comparison_real_v2_20260513/对比结果.xlsx")

# 读取结果
data = []
with open(results_file, encoding='utf-8') as f:
    for line in f:
        data.append(json.loads(line))

# 创建Excel
wb = openpyxl.Workbook()

# 样式定义
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Sheet 1: 整体对比
ws1 = wb.active
ws1.title = "整体对比"

headers1 = ["模型", "总测试次数", "格式通过", "通过率", "平均延迟(s)", "中位数延迟(s)", "p95延迟(s)", "最小延迟(s)", "最大延迟(s)"]
ws1.append(headers1)

# 应用表头样式
for col_num, header in enumerate(headers1, 1):
    cell = ws1.cell(1, col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# 按模型统计
from collections import defaultdict
import statistics

model_stats = defaultdict(lambda: {'latencies': [], 'format_passes': 0})
for r in data:
    model = r['model_id']
    model_stats[model]['latencies'].append(r['latency'])
    if r['format_ok']:
        model_stats[model]['format_passes'] += 1

for model, stats in sorted(model_stats.items()):
    latencies = stats['latencies']
    total = len(latencies)
    passes = stats['format_passes']
    
    row = [
        model,
        total,
        passes,
        f"{passes/total*100:.1f}%",
        f"{statistics.mean(latencies):.3f}",
        f"{statistics.median(latencies):.3f}",
        f"{statistics.quantiles(latencies, n=20)[18] if len(latencies)>=20 else max(latencies):.3f}",
        f"{min(latencies):.3f}",
        f"{max(latencies):.3f}"
    ]
    ws1.append(row)
    
    # 应用样式
    row_num = ws1.max_row
    for col_num in range(1, len(headers1) + 1):
        cell = ws1.cell(row_num, col_num)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 通过率单元格着色
        if col_num == 4:
            if passes/total >= 0.9:
                cell.fill = pass_fill
            elif passes/total < 0.5:
                cell.fill = fail_fill

# 调整列宽
for col_num in range(1, len(headers1) + 1):
    ws1.column_dimensions[get_column_letter(col_num)].width = 15

# Sheet 2: 详细结果
ws2 = wb.create_sheet("详细结果")

headers2 = ["模型", "测试用例", "对话类型", "角色类型", "角色", "关系", "运行次数", "延迟(s)", "格式通过", "失败原因", "互动要点数", "字符数"]
ws2.append(headers2)

# 应用表头样式
for col_num, header in enumerate(headers2, 1):
    cell = ws2.cell(1, col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# 填充数据
for r in sorted(data, key=lambda x: (x['model_id'], x.get('dialogue_type', ''), x['case_id'], x['run'])):
    row = [
        r['model_id'],
        r['case_id'],
        r.get('dialogue_type', ''),
        r.get('role_type', ''),
        r['role_name'],
        r['relationship'],
        r['run'],
        f"{r['latency']:.3f}",
        "通过" if r['format_ok'] else "失败",
        "; ".join(r['issues'][:2]) if not r['format_ok'] else "",
        r['metrics'].get('point_count', 0),
        r['metrics'].get('char_count', 0)
    ]
    ws2.append(row)
    
    # 应用样式
    row_num = ws2.max_row
    for col_num in range(1, len(headers2) + 1):
        cell = ws2.cell(row_num, col_num)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 格式通过单元格着色
        if col_num == 9:
            if r['format_ok']:
                cell.fill = pass_fill
            else:
                cell.fill = fail_fill

# 调整列宽
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 15
ws2.column_dimensions['F'].width = 10
ws2.column_dimensions['G'].width = 10
ws2.column_dimensions['H'].width = 12
ws2.column_dimensions['I'].width = 12
ws2.column_dimensions['J'].width = 40
ws2.column_dimensions['K'].width = 12
ws2.column_dimensions['L'].width = 12

# Sheet 3: 输出样本对比
ws3 = wb.create_sheet("输出样本对比")

headers3 = ["模型", "测试用例", "运行次数", "延迟(s)", "格式通过", "输出内容（前500字）"]
ws3.append(headers3)

# 应用表头样式
for col_num, header in enumerate(headers3, 1):
    cell = ws3.cell(1, col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# 只取每个case的第一次运行
seen_cases = set()
for r in sorted(data, key=lambda x: (x['case_id'], x['model_id'], x['run'])):
    case_key = (r['case_id'], r['model_id'])
    if case_key in seen_cases:
        continue
    seen_cases.add(case_key)
    
    row = [
        r['model_id'],
        r['case_id'],
        r['run'],
        f"{r['latency']:.3f}",
        "通过" if r['format_ok'] else "失败",
        r['output'][:500] + "..." if len(r['output']) > 500 else r['output']
    ]
    ws3.append(row)
    
    # 应用样式
    row_num = ws3.max_row
    for col_num in range(1, len(headers3) + 1):
        cell = ws3.cell(row_num, col_num)
        cell.border = border
        
        if col_num == 6:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 格式通过单元格着色
        if col_num == 5:
            if r['format_ok']:
                cell.fill = pass_fill
            else:
                cell.fill = fail_fill

# 调整列宽和行高
ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 25
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 12
ws3.column_dimensions['E'].width = 12
ws3.column_dimensions['F'].width = 80

for row in ws3.iter_rows(min_row=2):
    ws3.row_dimensions[row[0].row].height = 60

# 保存
wb.save(output_file)
print(f"Excel已生成: {output_file}")
print(f"包含3个sheet:")
print(f"  1. 整体对比 - 两个模型的整体性能对比")
print(f"  2. 详细结果 - 每次测试的详细数据")
print(f"  3. 输出样本对比 - 实际输出内容对比")
