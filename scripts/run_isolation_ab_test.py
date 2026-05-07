from __future__ import annotations

import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import json
import multiprocessing as mp
import os
import re
import sys
from copy import deepcopy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from statistics import mean
from time import perf_counter

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_DIR = Path(__file__).resolve().parents[1]
SERVER_DIR = PROJECT_DIR / "server"
for path in (PROJECT_DIR, SERVER_DIR):
    text = str(path)
    if text not in sys.path:
        sys.path.insert(0, text)

from config import DEFAULT_PRIMARY_MODEL, DEFAULT_SUMMARY_MODEL, get_latest_prompt_file  # noqa: E402
from services.conversation_service import ConversationService  # noqa: E402
from services.message_assembler import (  # noqa: E402
    LONGFORM_HISTORY_PREFIX,
    LONGFORM_HISTORY_SUFFIX,
    SHORTFORM_HISTORY_PREFIX,
    SHORTFORM_HISTORY_SUFFIX,
)


DEFAULT_INPUT = PROJECT_DIR / "tests" / "data" / "0403测试用例_温暖霸道.xlsx"
DEFAULT_OUTPUT_DIR = PROJECT_DIR / "output" / "isolation_ab"
DEFAULT_LONG_SUMMARY_PROMPT = Path(
    r"E:\工作资料\产品资料\提示词资料\长文模式\摘要提示词\长文模式摘要提示词_v2.7_20260425.md"
)
DEFAULT_SHORT_SUMMARY_PROMPT = Path(
    r"E:\工作资料\产品资料\提示词资料\短期记忆\短文模式摘要提示词.md"
)
DEFAULT_SHORT_SYSTEM_SOURCE = Path(
    r"E:\工作资料\产品资料\提示词资料\问题排查\缺少动作描述括号.md"
)
DEFAULT_TURNS = 22
EXCEL_CELL_LIMIT = 32000


@dataclass
class TestCase:
    turn: int
    user_input: str
    row: dict[str, str]


@dataclass
class ScenarioCase:
    name: str
    target_mode: str
    model_id: str
    current_input: str
    history: list[dict]
    turn: int = 21


def _text(value) -> str:
    return str(value or "").strip()


def excel_chunks(value: str, limit: int = EXCEL_CELL_LIMIT) -> list[str]:
    text = _text(value)
    if not text:
        return [""]
    return [text[i:i + limit] for i in range(0, len(text), limit)]


def safe_excel_text(value: str) -> str:
    text = _text(value)
    if len(text) <= EXCEL_CELL_LIMIT:
        return text
    return text[: EXCEL_CELL_LIMIT - 40] + f"\n...[已截断，完整内容见完整拼接日志，原长{len(text)}字]"


def read_text_file(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def extract_short_system_prompt(path: Path) -> str:
    """从问题排查样本中只提取真实请求里的短文system prompt，忽略反馈正文。"""
    raw = read_text_file(path)
    payload = json.loads(raw)
    prompt_value = (
        payload.get("prompt")
        or payload.get("systemPrompt")
        or (payload.get("modelRequest") or {}).get("prompt")
    )
    if not prompt_value:
        raise ValueError(f"未在问题排查文件中找到 prompt/systemPrompt: {path}")
    messages = json.loads(prompt_value)
    # 部分样本外层 system.content 又嵌了一层 messages JSON。
    if len(messages) == 1 and messages[0].get("role") == "system":
        content = _text(messages[0].get("content", ""))
        if content.startswith("["):
            try:
                nested = json.loads(content)
                messages = nested
            except json.JSONDecodeError:
                pass
    system_msg = next((msg for msg in messages if msg.get("role") == "system"), None)
    if not system_msg:
        raise ValueError(f"问题排查样本 prompt 中没有 system 消息: {path}")
    return _text(system_msg.get("content", ""))


def format_conversation_log(history: list[dict]) -> str:
    lines: list[str] = []
    for index, msg in enumerate(history, start=1):
        role = "用户" if msg.get("role") == "user" else "角色"
        content = _text(msg.get("content", ""))
        source_mode = _text(msg.get("source_mode", ""))
        mode_suffix = f"({source_mode})" if source_mode else ""
        lines.append(f"[04-28 12:{index % 60:02d}][{role}{mode_suffix}] {content}")
    return "\n".join(lines)


def render_summary_prompt(
    template: str,
    *,
    conversation_log: str,
    current_mode: str,
    existing_summary: str,
    role_name: str,
    personal_type: str,
    relationship: str,
) -> str:
    replacements = {
        "{conversation_log}": conversation_log,
        "{conversation_text}": conversation_log,
        "{current_mode}": current_mode,
        "{existing_summary}": existing_summary or "(首次生成，无旧摘要)",
        "{role_name}": role_name,
        "{personal_type}": personal_type,
        "{relationship}": relationship,
    }
    prompt = template
    for placeholder, value in replacements.items():
        prompt = prompt.replace(placeholder, value)
    return prompt


def load_cases(path: Path, turns: int) -> tuple[list[TestCase], dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, str]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: _text(values[i]) for i in range(min(len(headers), len(values))) if headers[i]}
        if row.get("用户输入"):
            rows.append(row)
    if len(rows) < turns:
        raise ValueError(f"有效用户输入不足 {turns} 轮，当前只有 {len(rows)} 行")
    cases = [
        TestCase(turn=i + 1, user_input=rows[i]["用户输入"], row=rows[i])
        for i in range(turns)
    ]
    return cases, rows[0]


def build_config(seed: dict[str, str]) -> dict:
    character_keys = {
        "Role_Nickname",
        "gender",
        "personal_type",
        "age",
        "occupation",
        "personality",
        "speaking_style",
        "background",
        "hobby",
        "Role_info_works",
        "system_module8",
    }
    context_keys = {
        "relationship",
        "user_Nickname",
        "user_gender",
        "user_identity",
        "relation_calling",
        "last_cst_type",
    }
    module_keys = {
        "monthly_schedule",
        "weekly_schedule",
        "voice_forbidden",
        "moments",
        "dialogueStartPrompt",
        "dialogue_summary",
        "longform_dialogue_guideline",
        "longform_persona",
        "longform_narrative_style",
        "longform_few_shot",
        "intimacy_boundary",
    }
    normalized = {k.strip(): v for k, v in seed.items()}
    return {
        "prompt_file": normalized.get("测试对应提示词", "") or get_latest_prompt_file(),
        "character": {k: normalized.get(k, "") for k in character_keys},
        "context": {k: normalized.get(k, "") for k in context_keys},
        "modules": {k: normalized.get(k, "") for k in module_keys},
        "custom_variables": {
            "完整时间信息": normalized.get("完整时间信息", ""),
        },
        "runtime": {
            "summary_interval": 10,
            "conversation_mode": "isolation_ab",
        },
    }


def shortform_contaminant(turn: int) -> str:
    samples = [
        "（低头蹭了蹭你的额头）别闹，先让我看看你是不是又在逞强。",
        "嗯？这么看着我干什么，是不是又想让我哄你？",
        "（把外套披到你肩上）冷就直说，别每次都靠眼神让我猜。",
        "乖一点，今天先不逗你，等你缓过来再说。",
    ]
    return samples[(turn - 1) % len(samples)]


def longform_contaminant(turn: int) -> str:
    samples = [
        "窗外的风把窗帘掀起一角，他站在灯影边缘，视线落在你攥紧的指尖上。**\"过来。\"**",
        "走廊尽头的声控灯亮了一下，他没有立刻说话，只把外套搭到你肩上。**\"别硬撑。\"**",
        "热茶的雾气在杯沿散开，他低头看着你，语气放得很慢。**\"这件事先听我说。\"**",
        "夜色压在玻璃窗上，他伸手替你挡开冷风，掌心停在半空。**\"我在这儿。\"**",
    ]
    return samples[(turn - 1) % len(samples)]


def build_extreme_switch_cases(cases: list[TestCase], *, long_model: str, short_model: str) -> list[ScenarioCase]:
    if len(cases) < 21:
        raise ValueError("极端切换测试至少需要 21 条用户输入：20轮历史 + 1轮当前输入")

    def make_history(mode_sequence: list[str]) -> list[dict]:
        history: list[dict] = []
        for idx, source_mode in enumerate(mode_sequence, start=1):
            user_input = cases[idx - 1].user_input
            assistant_content = (
                shortform_contaminant(idx)
                if source_mode == "short"
                else longform_contaminant(idx)
            )
            history.extend([
                {"role": "user", "content": user_input},
                {
                    "role": "assistant",
                    "content": assistant_content,
                    "source_mode": source_mode,
                    "data_type": f"{source_mode}form",
                },
            ])
        return history

    current_input = cases[20].user_input
    return [
        ScenarioCase(
            name="短切长_19短1长",
            target_mode="long",
            model_id=long_model,
            current_input=current_input,
            history=make_history(["short"] * 19 + ["long"]),
        ),
        ScenarioCase(
            name="长切短_19长1短",
            target_mode="short",
            model_id=short_model,
            current_input=current_input,
            history=make_history(["long"] * 19 + ["short"]),
        ),
    ]


def add_contaminated_history(history: list[dict], turn: int, user_input: str, ai_output: str) -> None:
    history.append({"role": "user", "content": user_input})
    if turn % 3 == 0:
        history.append({
            "role": "assistant",
            "content": shortform_contaminant(turn),
            "source_mode": "short",
            "data_type": "shortform_residue",
        })
    history.append({
        "role": "assistant",
        "content": ai_output,
        "source_mode": "long",
        "data_type": "longform",
    })


def strip_wrapped_text(content: str) -> tuple[str, str]:
    text = _text(content)
    if text.startswith(SHORTFORM_HISTORY_PREFIX):
        inner = text[len(SHORTFORM_HISTORY_PREFIX):].strip()
        if inner.endswith(SHORTFORM_HISTORY_SUFFIX):
            inner = inner[: -len(SHORTFORM_HISTORY_SUFFIX)].strip()
        return "short", inner
    if text.startswith(LONGFORM_HISTORY_PREFIX):
        inner = text[len(LONGFORM_HISTORY_PREFIX):].strip()
        if inner.endswith(LONGFORM_HISTORY_SUFFIX):
            inner = inner[: -len(LONGFORM_HISTORY_SUFFIX)].strip()
        return "long", inner
    return "", text


def depth_injection_message(role_name: str, personality: str, relationship: str) -> dict:
    return {
        "role": "system",
        "content": (
            f"切换当轮轻量 Depth Injection：你是{role_name}，性格{personality}，"
            f"当前关系阶段为{relationship}。该消息只做身份/关系近因锚定，不携带剧情事实。"
        ),
    }


def insert_depth_before_penultimate_history_round(
    messages: list[dict],
    role_name: str,
    personality: str,
    relationship: str,
) -> list[dict]:
    """上下文达到4轮后，将Depth插到倒数第二轮历史user前方。"""
    historical_user_indices = [
        idx
        for idx, msg in enumerate(messages[:-1])
        if msg.get("role") == "user"
    ]
    if len(historical_user_indices) < 4:
        return messages
    insert_at = historical_user_indices[-2]
    return [
        *messages[:insert_at],
        depth_injection_message(role_name, personality, relationship),
        *messages[insert_at:],
    ]


def to_system_sandwich(messages: list[dict]) -> list[dict]:
    converted: list[dict] = []
    for index, msg in enumerate(messages):
        if msg.get("role") == "assistant":
            source_mode, inner = strip_wrapped_text(msg.get("content", ""))
            if source_mode == "short":
                converted.extend([
                    {
                        "role": "system",
                        "content": "以下为短文模式回复记录，仅供剧情事实参考，不要模仿其字数、括号动作、语气格式",
                    },
                    {"role": "assistant", "content": inner},
                    {
                        "role": "system",
                        "content": "短文模式记录结束，请继续以长文模式格式回复",
                    },
                ])
                continue
            if source_mode == "long":
                converted.extend([
                    {
                        "role": "system",
                        "content": "以下为长文模式回复记录，仅供剧情事实参考，不要模仿其第三人称旁白、长段落、加粗对白格式",
                    },
                    {"role": "assistant", "content": inner},
                    {
                        "role": "system",
                        "content": "长文模式记录结束，请继续以目标模式格式回复",
                    },
                ])
                continue
        converted.append(msg)
    return converted


def count_pattern(messages: list[dict], pattern: str) -> int:
    return sum(1 for msg in messages if pattern in _text(msg.get("content", "")))


def evaluate_payload(
    messages: list[dict],
    strategy: str,
    expected_cross_mode_count: int,
    target_mode: str = "long",
) -> dict:
    roles = [msg.get("role", "") for msg in messages]
    current_user_last = bool(messages and messages[-1].get("role") == "user")
    current_user = _text(messages[-1].get("content", "")) if messages else ""
    text_marker_count = count_pattern(messages, "模式回复记录")
    assistant_text_marker_count = sum(
        1
        for msg in messages
        if msg.get("role") == "assistant" and "模式回复记录" in _text(msg.get("content", ""))
    )
    system_boundary_count = sum(
        1
        for msg in messages
        if msg.get("role") == "system" and "模式回复记录" in _text(msg.get("content", ""))
    )
    depth_positions = [
        idx for idx, msg in enumerate(messages)
        if msg.get("role") == "system" and "轻量 Depth Injection" in _text(msg.get("content", ""))
    ]
    historical_user_indices = [
        idx for idx, msg in enumerate(messages[:-1])
        if msg.get("role") == "user"
    ]
    expected_depth = len(historical_user_indices) >= 4
    expected_depth_position = historical_user_indices[-2] - 1 if expected_depth else -1
    summary_as_system = any(
        msg.get("role") == "system" and "动态摘要" in _text(msg.get("content", ""))
        for msg in messages[1:]
    )
    if target_mode == "short":
        pass_core = (
            current_user_last
            and not current_user.startswith("<Core_Constraints>")
            and "<user_input>" not in current_user
        )
    else:
        pass_core = current_user_last and current_user.startswith("<Core_Constraints>") and "<user_input>" in current_user
    if strategy == "assistant_text_marker":
        pass_isolation = (
            assistant_text_marker_count > 0 and system_boundary_count == 0
            if expected_cross_mode_count > 0
            else assistant_text_marker_count == 0 and system_boundary_count == 0
        )
        pass_depth = (
            depth_positions == [expected_depth_position]
            if expected_depth
            else not depth_positions
        )
    else:
        pass_isolation = (
            system_boundary_count > 0 and assistant_text_marker_count == 0
            if expected_cross_mode_count > 0
            else system_boundary_count == 0 and assistant_text_marker_count == 0
        )
        pass_depth = (
            depth_positions == [expected_depth_position]
            if expected_depth
            else not depth_positions
        )
    return {
        "role_sequence": " > ".join(roles),
        "message_count": len(messages),
        "system_count": roles.count("system"),
        "assistant_count": roles.count("assistant"),
        "user_count": roles.count("user"),
        "text_marker_count": text_marker_count,
        "assistant_text_marker_count": assistant_text_marker_count,
        "system_boundary_count": system_boundary_count,
        "expected_cross_mode_count": expected_cross_mode_count,
        "depth_injection_count": len(depth_positions),
        "context_round_count": len(historical_user_indices),
        "expected_depth": expected_depth,
        "expected_depth_position": expected_depth_position,
        "actual_depth_positions": ",".join(str(pos) for pos in depth_positions),
        "depth_before_penultimate_context": bool(depth_positions == [expected_depth_position]) if expected_depth else not depth_positions,
        "depth_before_current_user": bool(depth_positions and depth_positions[-1] == len(messages) - 2),
        "summary_as_system": summary_as_system,
        "current_user_core_wrapped": pass_core,
        "payload_pass": pass_isolation and pass_depth and pass_core and not summary_as_system,
    }


def evaluate_output(text: str, target_mode: str = "long") -> dict:
    output = _text(text)
    leak_terms = ["以下为", "记录结束", "动态摘要", "内部认知记录", "短文模式回复记录", "Core_Constraints"]
    has_leak = any(term in output for term in leak_terms)
    bracket_actions = len(re.findall(r"[（(][^）)]{1,40}[）)]", output))
    chinese_bracket_actions = len(re.findall(r"（[^）]{1,40}）", output))
    has_bold_dialogue = "**" in output and '"' in output
    if target_mode == "short":
        format_pass = 30 <= len(output) <= 60 and chinese_bracket_actions > 0 and not has_leak
    else:
        format_pass = 300 <= len(output) <= 500 and not has_leak
    return {
        "word_count": len(output),
        "has_template_leak": has_leak,
        "bracket_action_count": bracket_actions,
        "chinese_bracket_action_count": chinese_bracket_actions,
        "has_bold_dialogue": has_bold_dialogue,
        "format_pass": format_pass,
    }


def build_messages(
    service: ConversationService,
    runtime_bundle,
    config: dict,
    history: list[dict],
    case: TestCase,
    model_id: str,
    strategy: str,
    target_mode: str,
    short_system_prompt: str = "",
) -> list[dict]:
    old_flag = os.environ.get("LONGFORM_V52_MESSAGE_CONTRACT")
    os.environ["LONGFORM_V52_MESSAGE_CONTRACT"] = "1"
    try:
        memory_context, _ = service._build_memory_context_block(
            runtime_bundle.memory_profile,
            runtime_bundle.memory_moments,
            runtime_bundle.seed_dialogue_summary,
        )
        contract_model_id = "doubao-pro" if target_mode == "short" else model_id
        messages = service._build_messages_internal(
            rendered_system=runtime_bundle.rendered_system,
            system_after=runtime_bundle.rendered_after,
            few_shot_messages=runtime_bundle.few_shot_messages,
            conversation_history=list(history)[-40:],
            dialogue_summary=runtime_bundle.seed_dialogue_summary,
            memory_context=memory_context,
            current_input=case.user_input,
            relationship=runtime_bundle.relationship,
            role_name=runtime_bundle.role_name,
            personality=runtime_bundle.personality,
            turn_num=case.turn,
            injection_depth=runtime_bundle.injection_depth,
            model_id=contract_model_id,
        )
    finally:
        if old_flag is None:
            os.environ.pop("LONGFORM_V52_MESSAGE_CONTRACT", None)
        else:
            os.environ["LONGFORM_V52_MESSAGE_CONTRACT"] = old_flag

    if target_mode == "short" and short_system_prompt:
        messages[0]["content"] = short_system_prompt
    if strategy == "system_sandwich":
        messages = to_system_sandwich(messages)
    messages = insert_depth_before_penultimate_history_round(
        messages,
        runtime_bundle.role_name,
        runtime_bundle.personality,
        runtime_bundle.relationship,
    )
    return messages


def _model_call_worker(model_id: str, messages: list[dict], queue) -> None:
    from services.model_adapter import ModelAdapter

    started = perf_counter()
    try:
        result = ModelAdapter().chat(
            model_id,
            messages,
            thinking_effort="disabled",
        )
        latency = perf_counter() - started
        queue.put({
            "success": bool(result.success),
            "content": _text(result.content),
            "input_tokens": int(result.input_tokens or 0),
            "output_tokens": int(result.output_tokens or 0),
            "latency_s": round(float(result.latency_s or latency), 2),
            "error": _text(result.error),
        })
    except Exception as exc:
        queue.put({
            "success": False,
            "content": "",
            "input_tokens": 0,
            "output_tokens": 0,
            "latency_s": round(perf_counter() - started, 2),
            "error": f"{type(exc).__name__}: {exc}",
        })


def call_model(
    service: ConversationService,
    model_id: str,
    messages: list[dict],
    dry_run: bool,
    turn: int,
    strategy: str,
    request_timeout_s: float,
) -> dict:
    if dry_run:
        return {
            "success": True,
            "content": f"[dry-run:{strategy}] 第{turn}轮模拟长文输出。系统仅验证 payload 结构，不验证真实模型文风效果。",
            "input_tokens": 0,
            "output_tokens": 0,
            "latency_s": 0.0,
            "error": "",
        }
    timeout_s = max(1.0, float(request_timeout_s or 180))
    ctx = mp.get_context("spawn")
    queue = ctx.Queue(maxsize=1)
    process = ctx.Process(target=_model_call_worker, args=(model_id, messages, queue))
    started = perf_counter()
    process.start()
    process.join(timeout_s)
    if process.is_alive():
        process.terminate()
        process.join(5)
        return {
            "success": False,
            "content": "",
            "input_tokens": 0,
            "output_tokens": 0,
            "latency_s": round(perf_counter() - started, 2),
            "error": f"REQUEST_TIMEOUT after {timeout_s:.1f}s",
        }
    if not queue.empty():
        return queue.get()
    return {
        "success": False,
        "content": "",
        "input_tokens": 0,
        "output_tokens": 0,
        "latency_s": round(perf_counter() - started, 2),
        "error": f"worker exited with code {process.exitcode}",
    }


def generate_summary_event(
    service: ConversationService,
    *,
    strategy: str,
    turn: int,
    summary_type: str,
    prompt_path: Path,
    summary_model: str,
    history: list[dict],
    existing_summary: str,
    runtime_bundle,
    dry_run: bool,
    request_timeout_s: float,
) -> dict:
    conversation_log = format_conversation_log(history)
    prompt_template = read_text_file(prompt_path)
    current_mode = "longform" if summary_type == "long" else "shortform"
    prompt_text = render_summary_prompt(
        prompt_template,
        conversation_log=conversation_log,
        current_mode=current_mode,
        existing_summary=existing_summary,
        role_name=runtime_bundle.role_name,
        personal_type=runtime_bundle.personal_type,
        relationship=runtime_bundle.relationship,
    )
    messages = [{"role": "user", "content": prompt_text}]
    result = call_model(
        service,
        summary_model,
        messages,
        dry_run,
        turn,
        f"{strategy}:summary:{summary_type}",
        request_timeout_s,
    )
    return {
        "strategy": strategy,
        "turn": turn,
        "summary_type": summary_type,
        "summary_model": summary_model,
        "summary_prompt_path": str(prompt_path),
        "summary_success": result["success"],
        "summary_error": result["error"],
        "summary_latency_s": result["latency_s"],
        "summary_input_tokens": result["input_tokens"],
        "summary_output_tokens": result["output_tokens"],
        "history_message_count": len(history),
        "conversation_log": conversation_log,
        "summary_prompt": prompt_text,
        "summary_output": result["content"] if result["success"] else f"[ERROR] {result['error']}",
    }


def run_strategy(
    cases: list[TestCase],
    seed_config: dict,
    model_id: str,
    strategy: str,
    target_mode: str,
    dry_run: bool,
    request_timeout_s: float,
    long_summary_prompt: Path,
    short_summary_prompt: Path,
    summary_model: str,
    short_system_prompt: str = "",
) -> list[dict]:
    service = ConversationService()
    runtime_bundle = service._prepare_runtime_bundle(seed_config, web_search=False)
    history: list[dict] = []
    summary_buffer: list[dict] = []
    rows: list[dict] = []
    active_dialogue_summary = _text(runtime_bundle.seed_dialogue_summary)
    for case in cases:
        runtime_bundle.seed_dialogue_summary = active_dialogue_summary
        expected_cross_mode_count = sum(
            1
            for msg in history[-40:]
            if msg.get("role") == "assistant"
            and msg.get("source_mode")
            and msg.get("source_mode") != target_mode
        )
        messages = build_messages(
            service,
            runtime_bundle,
            seed_config,
            history,
            case,
            model_id,
            strategy,
            target_mode,
            short_system_prompt,
        )
        payload_eval = evaluate_payload(messages, strategy, expected_cross_mode_count, target_mode)
        model_result = call_model(
            service,
            model_id,
            messages,
            dry_run,
            case.turn,
            strategy,
            request_timeout_s,
        )
        output_eval = evaluate_output(model_result["content"], target_mode)
        ai_output = model_result["content"] if model_result["success"] else f"[ERROR] {model_result['error']}"
        if target_mode == "short":
            history.append({"role": "user", "content": case.user_input})
            history.append({
                "role": "assistant",
                "content": ai_output,
                "source_mode": "short",
                "data_type": "shortform",
            })
            turn_history_items = history[-2:]
        else:
            add_contaminated_history(history, case.turn, case.user_input, ai_output)
            turn_history_items = history[-3:] if case.turn % 3 == 0 else history[-2:]
        summary_buffer.extend(deepcopy(turn_history_items))
        summary_events: list[dict] = []
        summary_input = deepcopy(summary_buffer)
        if target_mode == "long" and case.turn in {10, 20}:
            long_summary_event = generate_summary_event(
                service,
                strategy=strategy,
                turn=case.turn,
                summary_type="long",
                prompt_path=long_summary_prompt,
                summary_model=summary_model,
                history=summary_input,
                existing_summary=active_dialogue_summary,
                runtime_bundle=runtime_bundle,
                dry_run=dry_run,
                request_timeout_s=request_timeout_s,
            )
            summary_events.append(long_summary_event)
            if long_summary_event["summary_success"]:
                active_dialogue_summary = _text(long_summary_event["summary_output"])
                runtime_bundle.seed_dialogue_summary = active_dialogue_summary
                history.clear()
        if target_mode == "short" and case.turn == 20:
            short_summary_event = generate_summary_event(
                service,
                strategy=strategy,
                turn=case.turn,
                summary_type="short",
                prompt_path=short_summary_prompt,
                summary_model=summary_model,
                history=summary_input,
                existing_summary="",
                runtime_bundle=runtime_bundle,
                dry_run=dry_run,
                request_timeout_s=request_timeout_s,
            )
            summary_events.append(short_summary_event)
            if short_summary_event["summary_success"]:
                active_dialogue_summary = _text(short_summary_event["summary_output"])
                runtime_bundle.seed_dialogue_summary = active_dialogue_summary
                history.clear()
        if summary_events and all(event.get("summary_success") for event in summary_events):
            summary_buffer.clear()
        rows.append({
            "scenario": f"rolling_{target_mode}",
            "target_mode": target_mode,
            "strategy": strategy,
            "turn": case.turn,
            "user_input": case.user_input,
            "ai_output": ai_output,
            "model_id": model_id,
            "model_success": model_result["success"],
            "model_error": model_result["error"],
            "input_tokens": model_result["input_tokens"],
            "output_tokens": model_result["output_tokens"],
            "latency_s": model_result["latency_s"],
            **payload_eval,
            **output_eval,
            "messages_json": json.dumps(messages, ensure_ascii=False),
            "history_json": json.dumps(history[-40:], ensure_ascii=False),
            "current_user_payload": messages[-1].get("content", "") if messages else "",
            "dialogue_summary": active_dialogue_summary,
            "summary_events_json": json.dumps(summary_events, ensure_ascii=False),
        })
        print(
            f"{strategy} turn={case.turn:02d} "
            f"payload_pass={payload_eval['payload_pass']} "
            f"format_pass={output_eval['format_pass']} "
            f"success={model_result['success']} "
            f"summary_calls={len(summary_events)}"
        )
    return rows


def run_extreme_switch(
    scenarios: list[ScenarioCase],
    seed_config: dict,
    strategy: str,
    dry_run: bool,
    request_timeout_s: float,
    short_system_prompt: str = "",
) -> list[dict]:
    service = ConversationService()
    runtime_bundle = service._prepare_runtime_bundle(seed_config, web_search=False)
    rows: list[dict] = []
    for scenario in scenarios:
        expected_cross_mode_count = sum(
            1
            for msg in scenario.history[-40:]
            if msg.get("role") == "assistant" and msg.get("source_mode") != scenario.target_mode
        )
        case = TestCase(
            turn=scenario.turn,
            user_input=scenario.current_input,
            row={},
        )
        messages = build_messages(
            service,
            runtime_bundle,
            seed_config,
            scenario.history,
            case,
            scenario.model_id,
            strategy,
            scenario.target_mode,
            short_system_prompt,
        )
        payload_eval = evaluate_payload(
            messages,
            strategy,
            expected_cross_mode_count,
            scenario.target_mode,
        )
        model_result = call_model(
            service,
            scenario.model_id,
            messages,
            dry_run,
            scenario.turn,
            f"{strategy}:{scenario.name}",
            request_timeout_s,
        )
        output_eval = evaluate_output(model_result["content"], scenario.target_mode)
        rows.append({
            "scenario": scenario.name,
            "target_mode": scenario.target_mode,
            "strategy": strategy,
            "turn": scenario.turn,
            "user_input": scenario.current_input,
            "ai_output": model_result["content"] if model_result["success"] else f"[ERROR] {model_result['error']}",
            "model_id": scenario.model_id,
            "model_success": model_result["success"],
            "model_error": model_result["error"],
            "input_tokens": model_result["input_tokens"],
            "output_tokens": model_result["output_tokens"],
            "latency_s": model_result["latency_s"],
            **payload_eval,
            **output_eval,
            "messages_json": json.dumps(messages, ensure_ascii=False),
            "history_json": json.dumps(scenario.history[-40:], ensure_ascii=False),
            "current_user_payload": messages[-1].get("content", "") if messages else "",
            "dialogue_summary": runtime_bundle.seed_dialogue_summary,
            "summary_events_json": "[]",
        })
        print(
            f"{scenario.name} {strategy} "
            f"payload_pass={payload_eval['payload_pass']} "
            f"format_pass={output_eval['format_pass']} "
            f"success={model_result['success']}"
        )
    return rows


def summarize(rows: list[dict], turns: int, dry_run: bool, model_id: str) -> list[list]:
    result = []
    group_keys = sorted({(row.get("scenario", "rolling"), row["strategy"]) for row in rows})
    for scenario, strategy in group_keys:
        group = [
            row for row in rows
            if row.get("scenario", "rolling") == scenario and row["strategy"] == strategy
        ]
        result.append([
            scenario,
            strategy,
            len(group),
            group[0].get("model_id", model_id),
            "dry-run" if dry_run else "live",
            sum(1 for row in group if row["model_success"]),
            sum(1 for row in group if row["payload_pass"]),
            sum(1 for row in group if row["format_pass"]),
            round(mean(row["message_count"] for row in group), 1),
            round(mean(row["system_count"] for row in group), 1),
            sum(row["system_boundary_count"] for row in group),
            sum(row["assistant_text_marker_count"] for row in group),
            sum(1 for row in group if row["has_template_leak"]),
            sum(row["bracket_action_count"] for row in group),
            round(mean(row["word_count"] for row in group), 1),
        ])
    return result


def write_report(output_path: Path, rows: list[dict], turns: int, dry_run: bool, model_id: str, input_path: Path) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "摘要"
    summary_headers = [
        "隔离策略", "样本数", "模型", "运行模式", "模型成功样本数", "Payload通过样本数", "输出格式通过样本数",
        "平均消息数", "平均system数", "system边界总数", "assistant文本标记总数", "模板泄漏轮数",
        "括号动作总数", "平均输出字数",
    ]
    summary_headers.insert(0, "场景")
    ws_summary.append(["长短文异质上下文隔离 A/B 测试报告"])
    ws_summary.append([f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws_summary.append([f"输入文件：{input_path}"])
    ws_summary.append([])
    ws_summary.append(summary_headers)
    for row in summarize(rows, turns, dry_run, model_id):
        ws_summary.append(row)

    ws_detail = wb.create_sheet("逐轮结果")
    detail_headers = [
        "scenario", "target_mode", "strategy", "turn", "user_input", "ai_output", "model_id", "model_success", "model_error",
        "payload_pass", "format_pass", "message_count", "system_count", "system_boundary_count",
        "assistant_text_marker_count", "expected_cross_mode_count", "depth_injection_count", "depth_before_current_user",
        "context_round_count", "expected_depth", "expected_depth_position", "actual_depth_positions", "depth_before_penultimate_context",
        "current_user_core_wrapped", "summary_as_system", "word_count", "has_template_leak",
        "bracket_action_count", "chinese_bracket_action_count", "has_bold_dialogue", "input_tokens", "output_tokens", "latency_s",
    ]
    ws_detail.append(detail_headers)
    for row in rows:
        ws_detail.append([safe_excel_text(row.get(header, "")) for header in detail_headers])

    ws_prompt_log = wb.create_sheet("完整拼接日志")
    ws_prompt_log.append([
        "scenario",
        "target_mode",
        "strategy",
        "turn",
        "model_id",
        "message_index",
        "role",
        "chunk_index",
        "chunk_count",
        "content",
    ])
    for row in rows:
        try:
            messages = json.loads(row.get("messages_json", "[]") or "[]")
        except json.JSONDecodeError:
            messages = []
        for msg_index, message in enumerate(messages):
            chunks = excel_chunks(message.get("content", ""))
            for chunk_index, chunk in enumerate(chunks, start=1):
                ws_prompt_log.append([
                    row.get("scenario", "rolling"),
                    row.get("target_mode", "long"),
                    row.get("strategy", ""),
                    row.get("turn", ""),
                    row.get("model_id", model_id),
                    msg_index,
                    message.get("role", ""),
                    chunk_index,
                    len(chunks),
                    chunk,
                ])

    ws_context = wb.create_sheet("上下文")
    ws_context.append([
        "scenario",
        "target_mode",
        "strategy",
        "turn",
        "history_index",
        "role",
        "source_mode",
        "data_type",
        "chunk_index",
        "chunk_count",
        "content",
    ])
    for row in rows:
        try:
            history_items = json.loads(row.get("history_json", "[]") or "[]")
        except json.JSONDecodeError:
            history_items = []
        for history_index, item in enumerate(history_items):
            chunks = excel_chunks(item.get("content", ""))
            for chunk_index, chunk in enumerate(chunks, start=1):
                ws_context.append([
                    row.get("scenario", "rolling"),
                    row.get("target_mode", "long"),
                    row.get("strategy", ""),
                    row.get("turn", ""),
                    history_index,
                    item.get("role", ""),
                    item.get("source_mode", ""),
                    item.get("data_type", ""),
                    chunk_index,
                    len(chunks),
                    chunk,
                ])

    ws_user_input = wb.create_sheet("用户输入")
    ws_user_input.append([
        "scenario",
        "target_mode",
        "strategy",
        "turn",
        "raw_user_input",
        "final_user_payload",
        "dialogue_summary",
    ])
    for row in rows:
        ws_user_input.append([
            row.get("scenario", "rolling"),
            row.get("target_mode", "long"),
            row.get("strategy", ""),
            row.get("turn", ""),
            safe_excel_text(row.get("user_input", "")),
            safe_excel_text(row.get("current_user_payload", "")),
            safe_excel_text(row.get("dialogue_summary", "")),
        ])

    ws_summary_log = wb.create_sheet("摘要日志")
    ws_summary_log.append([
        "scenario",
        "strategy",
        "turn",
        "summary_type",
        "summary_model",
        "summary_prompt_path",
        "summary_success",
        "summary_error",
        "summary_latency_s",
        "summary_input_tokens",
        "summary_output_tokens",
        "history_message_count",
        "conversation_log",
        "summary_prompt",
        "summary_output",
    ])
    for row in rows:
        try:
            events = json.loads(row.get("summary_events_json", "[]") or "[]")
        except json.JSONDecodeError:
            events = []
        for event in events:
            ws_summary_log.append([
                row.get("scenario", "rolling"),
                event.get("strategy", row.get("strategy", "")),
                event.get("turn", row.get("turn", "")),
                event.get("summary_type", ""),
                event.get("summary_model", ""),
                event.get("summary_prompt_path", ""),
                event.get("summary_success", ""),
                safe_excel_text(event.get("summary_error", "")),
                event.get("summary_latency_s", ""),
                event.get("summary_input_tokens", ""),
                event.get("summary_output_tokens", ""),
                event.get("history_message_count", ""),
                safe_excel_text(event.get("conversation_log", "")),
                safe_excel_text(event.get("summary_prompt", "")),
                safe_excel_text(event.get("summary_output", "")),
            ])

    ws_messages = wb.create_sheet("Payload抽样")
    ws_messages.append(["scenario", "strategy", "turn", "role_sequence", "messages_json"])
    for row in rows:
        if row["turn"] in {1, 2, 3, 10, 20, 21, turns}:
            ws_messages.append([
                row.get("scenario", "rolling"),
                row["strategy"],
                row["turn"],
                row["role_sequence"],
                safe_excel_text(row["messages_json"]),
            ])

    ws_criteria = wb.create_sheet("验收口径")
    criteria = [
        ["项目", "验收标准"],
        ["轮数", "rolling 场景每种隔离策略至少生成 20 轮；extreme-switch 场景固定构造 20 轮满窗口历史 + 1 条当前输入"],
        ["极端场景", "短切长=19轮短文上下文+1轮长文上下文；长切短=19轮长文上下文+1轮短文上下文"],
        ["A策略", "assistant 文本标记：异质短文 assistant 保持 assistant role，内容内含短文模式回复记录标记"],
        ["B策略", "system 三明治：system(开始) -> assistant(原文) -> system(结束)"],
        ["Depth", "当拼接上下文达到4轮及以上时，以system注入到倒数第二轮历史user前方；摘要清空历史后重新计数"],
        ["当前 user", "长文目标请求最后一条 user 必须以 <Core_Constraints> 开头并包含 <user_input>；短文目标请求不得新增 Core 或 user_input 包裹"],
        ["短文输出", "短文目标请求按 30-60 字验收，动作旁白必须由中文括号（）包裹"],
        ["摘要", "真实 rolling 链路调用 doubao-lite 生成摘要；长文每10轮摘要后清空已摘要上下文，短文第20轮摘要后清空已摘要上下文；下一轮为system+摘要assistant+当前user"],
        ["Provider重试", "保留 provider 默认重试策略；单请求外层超时只负责避免整批卡死"],
        ["完整日志", "完整拼接提示词按 message_index/chunk_index 写入“完整拼接日志”，超长内容不只依赖 Payload 抽样"],
        ["输出污染", "不得泄漏“以下为/记录结束/动态摘要/Core_Constraints”等模板词"],
    ]
    for row in criteria:
        ws_criteria.append(row)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for col_idx, column in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in column:
                value = _text(cell.value)
                max_len = max(max_len, min(len(value), 80))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 55))
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="长短文异质上下文隔离 A/B 20+轮测试")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--turns", type=int, default=DEFAULT_TURNS)
    parser.add_argument("--model", default=DEFAULT_PRIMARY_MODEL)
    parser.add_argument("--long-model", default="qwen3.6-plus")
    parser.add_argument("--short-model", default="doubao-pro")
    parser.add_argument("--summary-model", default=DEFAULT_SUMMARY_MODEL)
    parser.add_argument("--long-summary-prompt", type=Path, default=DEFAULT_LONG_SUMMARY_PROMPT)
    parser.add_argument("--short-summary-prompt", type=Path, default=DEFAULT_SHORT_SUMMARY_PROMPT)
    parser.add_argument("--short-system-source", type=Path, default=DEFAULT_SHORT_SYSTEM_SOURCE)
    parser.add_argument(
        "--target-mode",
        choices=("long", "short"),
        default="long",
        help="rolling 场景主生成目标模式：long=长文10轮摘要清历史；short=短文20轮摘要清历史",
    )
    parser.add_argument(
        "--scenario",
        choices=("rolling", "extreme-switch"),
        default="rolling",
        help="rolling=连续多轮滚动；extreme-switch=20轮满窗口极端切换",
    )
    parser.add_argument(
        "--request-timeout-s",
        type=float,
        default=180.0,
        help="单次模型请求超时秒数；超时样本写入 Excel，不阻塞整批",
    )
    parser.add_argument(
        "--no-parallel-strategies",
        action="store_true",
        help="关闭 A/B 策略级并发，改为串行执行",
    )
    parser.add_argument("--dry-run", action="store_true")
    return parser.parse_args()


def run_parallel_or_serial(jobs: list[tuple[str, callable]]) -> list[dict]:
    rows: list[dict] = []
    with ThreadPoolExecutor(max_workers=len(jobs)) as executor:
        future_map = {executor.submit(fn): name for name, fn in jobs}
        for future in as_completed(future_map):
            name = future_map[future]
            try:
                rows.extend(future.result())
            except Exception as exc:
                raise RuntimeError(f"策略任务失败: {name}: {exc}") from exc
    return rows


def main() -> int:
    args = parse_args()
    if args.turns < 20:
        raise ValueError("turns 必须 >= 20，避免单轮/短链路误判隔离效果")
    if args.scenario == "rolling":
        if not args.long_summary_prompt.exists():
            raise FileNotFoundError(f"长文摘要提示词不存在: {args.long_summary_prompt}")
        if not args.short_summary_prompt.exists():
            raise FileNotFoundError(f"短文摘要提示词不存在: {args.short_summary_prompt}")
        if args.target_mode == "short" and not args.short_system_source.exists():
            raise FileNotFoundError(f"短文 system 样本不存在: {args.short_system_source}")
    short_system_prompt = (
        extract_short_system_prompt(args.short_system_source)
        if args.short_system_source.exists()
        else ""
    )
    cases, seed = load_cases(args.input, args.turns)
    config = build_config(seed)
    all_rows: list[dict] = []
    if args.scenario == "extreme-switch":
        scenarios = build_extreme_switch_cases(
            cases,
            long_model=args.long_model,
            short_model=args.short_model,
        )
        jobs = [
            (
                strategy,
                lambda s=strategy: run_extreme_switch(
                    scenarios,
                    deepcopy(config),
                    s,
                    args.dry_run,
                    args.request_timeout_s,
                    short_system_prompt,
                ),
            )
            for strategy in ("assistant_text_marker", "system_sandwich")
        ]
    else:
        generation_model = args.model if args.target_mode == "long" else args.short_model
        jobs = [
            (
                strategy,
                lambda s=strategy: run_strategy(
                    cases,
                    deepcopy(config),
                    generation_model,
                    s,
                    args.target_mode,
                    args.dry_run,
                    args.request_timeout_s,
                    args.long_summary_prompt,
                    args.short_summary_prompt,
                    args.summary_model,
                    short_system_prompt,
                ),
            )
            for strategy in ("assistant_text_marker", "system_sandwich")
        ]
    if args.no_parallel_strategies:
        for _, job in jobs:
            all_rows.extend(job())
    else:
        all_rows.extend(run_parallel_or_serial(jobs))
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = "dryrun" if args.dry_run else "live"
    output_path = args.output_dir / f"isolation_ab_{args.scenario}_{args.target_mode}_{suffix}_{args.turns}turns_{timestamp}.xlsx"
    report_model = args.model if args.target_mode == "long" else args.short_model
    write_report(output_path, all_rows, args.turns, args.dry_run, report_model, args.input)
    print(f"REPORT={output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
