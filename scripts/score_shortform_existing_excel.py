#!/usr/bin/env python3
"""给已生成的短文模式批量测试 Excel 补充 v1.1 单条评分。"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import re
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))

from shortform_model_switch_batch_test import (
    DEFAULT_SCORER,
    OUTPUT_DIR,
    CallResult,
    ModelSpec,
    render_template,
)

SCORE_COLUMNS = [
    "score_v11_ai_flavor_signals",
    "score_v11_rhythm_verdict",
    "score_v11_format_verdict",
    "score_v11_context_verdict",
    "score_v11_input_relevance",
    "score_v11_persona_fidelity",
    "score_v11_continuity",
    "score_v11_shortform_rhythm",
    "score_v11_memory_discipline",
    "score_v11_safety_boundary",
    "score_v11_weighted_total",
    "score_v11_mapped_total",
    "score_v11_qualification_status",
    "score_v11_reasoning",
    "score_v11_input_tokens",
    "score_v11_output_tokens",
    "score_v11_latency_s",
    "score_v11_retry_count",
    "score_v11_parse_error",
    "score_v11_error",
]


def load_records(worksheet) -> tuple[list[str], list[dict[str, Any]]]:
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    records: list[dict[str, Any]] = []
    for excel_row, values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        record = dict(zip(headers, values))
        record["_excel_row"] = excel_row
        records.append(record)
    return [str(header) for header in headers], records


def ensure_columns(worksheet, headers: list[str]) -> dict[str, int]:
    current = list(headers)
    for column in SCORE_COLUMNS:
        if column not in current:
            current.append(column)
            worksheet.cell(row=1, column=len(current), value=column)
    return {column: index + 1 for index, column in enumerate(current)}


def value(record: dict[str, Any], key: str) -> str:
    raw = record.get(key)
    return "" if raw is None else str(raw)


def variables_from_row(record: dict[str, Any], history_context: str) -> dict[str, str]:
    variables = {
        key.removeprefix("var_"): value(record, key)
        for key in record
        if str(key).startswith("var_")
    }
    role_name = variables.get("Role_Nickname") or value(record, "role_name")
    relationship = variables.get("relationship") or value(record, "relationship")
    variables.update(
        {
            "Role_Nickname": role_name,
            "relationship": relationship,
            "user_message": value(record, "user_input"),
            "output": value(record, "assistant_output"),
            "history_context": history_context,
            "call_name": variables.get("call_name")
            or variables.get("Tacall_name")
            or value(record, "var_Tacall_name"),
            "intimacy_boundary": variables.get("intimacy_boundary", ""),
        }
    )
    return variables


def build_history_context(
    records: list[dict[str, Any]],
    current: dict[str, Any],
) -> str:
    group_id = current.get("group_id")
    turn = int(current.get("turn") or 0)
    previous = [
        row
        for row in records
        if row.get("group_id") == group_id and int(row.get("turn") or 0) < turn
    ]
    previous.sort(key=lambda row: int(row.get("turn") or 0))
    snippets = []
    for row in previous[-5:]:
        snippets.append(
            f"第{row.get('turn')}轮 用户：{value(row, 'user_input')}\n"
            f"第{row.get('turn')}轮 AI：{value(row, 'assistant_output')}"
        )
    return "\n".join(snippets)


def parse_score(content: str) -> dict[str, Any]:
    match = re.search(r"\{.*\}", content, flags=re.S)
    if not match:
        return {"score_v11_parse_error": content[:500]}
    try:
        data = json.loads(match.group(0))
    except json.JSONDecodeError:
        return {"score_v11_parse_error": content[:500]}

    defect = (
        data.get("defect_prescan")
        if isinstance(data.get("defect_prescan"), dict)
        else {}
    )
    scores = data.get("scores") if isinstance(data.get("scores"), dict) else {}
    return {
        "score_v11_ai_flavor_signals": defect.get("ai_flavor_signals", ""),
        "score_v11_rhythm_verdict": defect.get("rhythm_verdict", ""),
        "score_v11_format_verdict": defect.get("format_verdict", ""),
        "score_v11_context_verdict": defect.get("context_verdict", ""),
        "score_v11_input_relevance": scores.get("input_relevance", ""),
        "score_v11_persona_fidelity": scores.get("persona_fidelity", ""),
        "score_v11_continuity": scores.get("continuity", ""),
        "score_v11_shortform_rhythm": scores.get("shortform_rhythm", ""),
        "score_v11_memory_discipline": scores.get("memory_discipline", ""),
        "score_v11_safety_boundary": scores.get("safety_boundary", ""),
        "score_v11_weighted_total": data.get("weighted_total", ""),
        "score_v11_mapped_total": data.get("mapped_total", ""),
        "score_v11_qualification_status": data.get("qualification_status", ""),
        "score_v11_reasoning": data.get("reasoning", ""),
    }


def score_call_once(
    model: ModelSpec,
    messages: list[dict[str, str]],
    timeout_s: float,
) -> CallResult:
    from openai import OpenAI

    api_key = os.environ.get(model.api_key_env, "").strip()
    if not api_key:
        raise RuntimeError(f"缺少环境变量 {model.api_key_env}")
    client = OpenAI(
        api_key=api_key,
        base_url=model.base_url,
        timeout=timeout_s,
    )
    started = time.time()
    response = client.chat.completions.create(
        model=model.model,
        messages=messages,
        **model.parameters,
    )
    content = response.choices[0].message.content or ""
    content = re.sub(
        r"(?is)<(?:think|thought)>.*?</(?:think|thought)>",
        "",
        content,
    ).strip()
    usage = getattr(response, "usage", None)
    return CallResult(
        content=content,
        input_tokens=int(getattr(usage, "prompt_tokens", 0) or 0),
        output_tokens=int(getattr(usage, "completion_tokens", 0) or 0),
        latency_s=round(time.time() - started, 3),
    )


async def score_call_with_retry(
    model: ModelSpec,
    messages: list[dict[str, str]],
    retries: int,
    retry_delay: float,
    timeout_s: float,
) -> CallResult:
    last_error = ""
    for attempt in range(retries + 1):
        try:
            result = await asyncio.to_thread(
                score_call_once,
                model,
                messages,
                timeout_s,
            )
            result.retry_count = attempt
            return result
        except Exception as exc:  # noqa: BLE001 - 记录外部 API 失败
            last_error = str(exc)
            if attempt < retries:
                await asyncio.sleep(retry_delay * (attempt + 1))
    return CallResult(error=last_error, retry_count=retries)


async def score_one(
    record: dict[str, Any],
    all_records: list[dict[str, Any]],
    prompt_template: str,
    scorer: ModelSpec,
    semaphore: asyncio.Semaphore,
    retries: int,
    retry_delay: float,
    timeout_s: float,
) -> dict[str, Any]:
    if value(record, "error"):
        return {"score_v11_error": "skip_generation_error"}
    history_context = build_history_context(all_records, record)
    prompt = render_template(
        prompt_template,
        variables_from_row(record, history_context),
    )
    async with semaphore:
        result = await score_call_with_retry(
            scorer,
            [{"role": "user", "content": prompt}],
            retries=retries,
            retry_delay=retry_delay,
            timeout_s=timeout_s,
        )
    parsed = {
        "score_v11_input_tokens": result.input_tokens,
        "score_v11_output_tokens": result.output_tokens,
        "score_v11_latency_s": result.latency_s,
        "score_v11_retry_count": result.retry_count,
    }
    if result.error:
        parsed["score_v11_error"] = result.error
        return parsed
    parsed.update(parse_score(result.content))
    return parsed


async def score_one_with_record(
    record: dict[str, Any],
    all_records: list[dict[str, Any]],
    prompt_template: str,
    scorer: ModelSpec,
    semaphore: asyncio.Semaphore,
    retries: int,
    retry_delay: float,
    timeout_s: float,
) -> tuple[dict[str, Any], dict[str, Any]]:
    score = await score_one(
        record,
        all_records,
        prompt_template,
        scorer,
        semaphore,
        retries,
        retry_delay,
        timeout_s,
    )
    return record, score


def is_scored(record: dict[str, Any]) -> bool:
    return any(
        record.get(column) not in (None, "")
        for column in (
            "score_v11_mapped_total",
            "score_v11_parse_error",
            "score_v11_error",
        )
    )


def write_score_to_sheet(
    worksheet,
    column_map: dict[str, int],
    record: dict[str, Any],
    score: dict[str, Any],
) -> None:
    record.update(score)
    excel_row = int(record["_excel_row"])
    for column in SCORE_COLUMNS:
        worksheet.cell(
            row=excel_row,
            column=column_map[column],
            value=record.get(column, ""),
        )


def truthy(raw: Any) -> bool:
    return str(raw).strip().lower() in {"true", "1", "yes", "y"}


def write_summary(
    workbook,
    records: list[dict[str, Any]],
    include_baseline: bool,
) -> None:
    if "score_v11_summary" in workbook.sheetnames:
        del workbook["score_v11_summary"]
    sheet = workbook.create_sheet("score_v11_summary")
    sheet.append(
        [
            "model_name",
            "rows",
            "score_errors",
            "parse_errors",
            "avg_mapped_total",
            "pass_count",
            "fail_count",
            "avg_input_relevance",
            "avg_persona_fidelity",
            "avg_continuity",
            "avg_shortform_rhythm",
            "avg_memory_discipline",
            "avg_safety_boundary",
        ]
    )
    summary_records = [
        row for row in records if include_baseline or not truthy(row.get("is_baseline"))
    ]
    for model_name in sorted({value(row, "model_name") for row in summary_records}):
        rows = [
            row
            for row in summary_records
            if value(row, "model_name") == model_name
        ]

        def avg(column: str) -> float | str:
            values = []
            for row in rows:
                try:
                    values.append(float(row.get(column)))
                except (TypeError, ValueError):
                    pass
            return round(sum(values) / len(values), 3) if values else ""

        sheet.append(
            [
                model_name,
                len(rows),
                sum(1 for row in rows if row.get("score_v11_error")),
                sum(1 for row in rows if row.get("score_v11_parse_error")),
                avg("score_v11_mapped_total"),
                sum(
                    1
                    for row in rows
                    if row.get("score_v11_qualification_status") == "PASS"
                ),
                sum(
                    1
                    for row in rows
                    if row.get("score_v11_qualification_status") == "FAIL"
                ),
                avg("score_v11_input_relevance"),
                avg("score_v11_persona_fidelity"),
                avg("score_v11_continuity"),
                avg("score_v11_shortform_rhythm"),
                avg("score_v11_memory_discipline"),
                avg("score_v11_safety_boundary"),
            ]
        )


async def main_async(args: argparse.Namespace) -> Path:
    input_path = Path(args.input)
    prompt_template = Path(args.prompt).read_text(encoding="utf-8")
    output_path = (
        Path(args.output)
        if args.output
        else OUTPUT_DIR
        / (
            f"{input_path.stem}_scored_v11_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    )
    if output_path.exists():
        workbook_path = output_path
    else:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(input_path, output_path)
        workbook_path = output_path

    workbook = load_workbook(workbook_path)
    worksheet = (
        workbook["detail"] if "detail" in workbook.sheetnames else workbook.active
    )
    headers, records = load_records(worksheet)
    column_map = ensure_columns(worksheet, headers)

    scorer_config = {
        **DEFAULT_SCORER,
        "parameters": {
            "temperature": args.temperature,
            "max_tokens": args.max_tokens,
        },
    }
    scorer = ModelSpec.from_dict(scorer_config)
    semaphore = asyncio.Semaphore(args.workers)
    pending_records = [
        record
        for record in records
        if not is_scored(record)
        and (not args.skip_baseline or not truthy(record.get("is_baseline")))
    ]
    total_pending = len(pending_records)
    if args.limit:
        pending_records = pending_records[: args.limit]
        total_pending = len(pending_records)
    if total_pending == 0:
        write_summary(workbook, records, include_baseline=not args.skip_baseline)
        workbook.save(output_path)
        return output_path

    tasks = [
        asyncio.create_task(
            score_one_with_record(
                record,
                records,
                prompt_template,
                scorer,
                semaphore,
                args.retries,
                args.retry_delay,
                args.timeout,
            )
        )
        for record in pending_records
    ]

    completed = 0
    for task in asyncio.as_completed(tasks):
        record, score = await task
        write_score_to_sheet(worksheet, column_map, record, score)
        completed += 1
        if completed % args.save_every == 0:
            workbook.save(output_path)
            print(f"saved {completed}/{total_pending}", flush=True)

    workbook.save(output_path)
    write_summary(workbook, records, include_baseline=not args.skip_baseline)
    workbook.save(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="给短文模式既有结果表补充 v1.1 评分")
    parser.add_argument("--input", required=True, help="既有生成结果 Excel")
    parser.add_argument("--prompt", required=True, help="短文模式打分提示词 Markdown")
    parser.add_argument("--output", help="输出 Excel 路径")
    parser.add_argument("--workers", type=int, default=8, help="评分并发数")
    parser.add_argument("--retries", type=int, default=3, help="失败重试次数")
    parser.add_argument("--save-every", type=int, default=20, help="每 N 条保存一次")
    parser.add_argument("--limit", type=int, default=0, help="只处理前 N 条待评分记录")
    parser.add_argument("--temperature", type=float, default=0.0, help="评分温度")
    parser.add_argument(
        "--max-tokens",
        type=int,
        default=1200,
        help="评分最大输出 token",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=90.0,
        help="单次评分请求超时秒数",
    )
    parser.add_argument(
        "--skip-baseline",
        action="store_true",
        help="只给候选模型评分，不补 baseline 行",
    )
    parser.add_argument(
        "--retry-delay",
        type=float,
        default=1.5,
        help="重试基础等待秒数",
    )
    return parser.parse_args()


def main() -> None:
    output_path = asyncio.run(main_async(parse_args()))
    print(output_path)


if __name__ == "__main__":
    main()
