#!/usr/bin/env python3
"""方案A（互动要点）vs 方案B（5轮三明治兜底）端到端对比验证

方案A：dialogue_summary + 互动要点 → assistant additional message → 目标模型首轮
方案B：dialogue_summary + 5轮三明治隔离异构历史 → 目标模型首轮

数据来源：Excel真实长文/短文对话示例
摘要生成：长文用v2.8摘要提示词，短文用短文摘要提示词
互动要点：deepseek-v4-flash生成

测试矩阵：
- 长→短：Excel长文对话 → doubao-lite / deepseek-v4-flash / doubao-1.5-character
- 短→长：Excel短文对话 → deepseek-v4-pro
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
SERVER_DIR = PROJECT_ROOT / "server"
sys.path.insert(0, str(SCRIPT_DIR))
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(SERVER_DIR))

# R11: 加载 .env（ModelAdapter 依赖 VOLCENGINE_API_KEY / ARK_API_KEY / DASHSCOPE_API_KEY 环境变量）
try:
    from dotenv import load_dotenv
    load_dotenv(SERVER_DIR / ".env")
except ImportError:
    pass

from services.model_adapter import ModelAdapter

# ── 路径 ──
DEFAULT_CASE_XLSX = Path(r"E:\工作资料\产品资料\提示词资料\模型切换\短文模式聊天批量测试用例.xlsx")
POINTS_PROMPT_PATH = Path(r"E:\工作资料\产品资料\提示词资料\长文模式\摘要提示词\互动要点提示词_v1.6_20260420.md")
LONG_SUMMARY_PROMPT_PATH = Path(r"E:\工作资料\产品资料\提示词资料\长文模式\摘要提示词\长文模式摘要提示词_v2.8_20260508.md")

# ── 模型 ──
EXTRACTOR_MODEL = "deepseek-v4-flash"
SUMMARY_MODEL = "deepseek-v4-flash"
SHORT_TARGET_MODELS = ["doubao-lite", "doubao-1.5-character", "deepseek-v4-flash"]
LONG_TARGET_MODEL = "deepseek-v4-pro"

# ── 禁词 ──
FORBIDDEN_WORDS = ["指尖"]

# ── 三明治隔离标记 ──
LONG_SANDWICH_START = "❗ [以下为长文模式回复记录，仅供剧情参考，请勿模仿格式]"
LONG_SANDWICH_END = "[长文模式记录结束，请继续以短文对话格式回复]"
SHORT_SANDWICH_START = "❗ [以下为短文模式回复记录，仅供剧情参考，请勿模仿格式]"
SHORT_SANDWICH_END = "[短文模式记录结束，请继续以长文叙事格式回复]"


@dataclass
class SwitchContext:
    """History split for replaying the first turn after a mode switch."""

    context_history: list[dict[str, str]]
    current_user: str
    current_user_index: int | None

# ── 短文摘要提示词 ──
SHORT_SUMMARY_PROMPT = """# 角色
你是一个通话信息整理专家，具备信息抽取概括的能力，精确提炼#用户画像信息、#用户记忆点信息

# 任务 : 智能记忆片段抽取
- 只对记忆片段做抽取，用纯文本展示，禁止多余的解释内容
- 将你们本次对话的内容，进行记忆摘要的抽取，使用记事本的方式记录对话的详细记录事情经过，总体字数限制在100个字左右，不得超过100字，适当做精炼总结。
详细要求：
1、记录核心互动的细节，聊过的话题、话题内容，需要记录标记清楚说话对象
2、抽取的事件数量根据对话里面角色与用户互动的记忆点进行动态抽取。
3、抽取用户表达的核心事件、以及角色回复和反应。示例如下：
比如，
本次对话中：用户与角色沟通了一下内容：
"1、用户表达喜欢，角色霸道表达占有欲"
"2、用户说肚子疼，角色表达关心"
"3、用户吃了赛百味，角色邀请下次尝试不同口味"
"4、 用户表达伤心，角色主动关心原因。"
"N、"
- 智能摘要角色和用户对话进行总结记录，确保记忆的准确性。
- 角色提及的信号不好、没有看到消息、对双方关系无关紧要的信息，这一类问题无需记录和总结。

以下是你要抽取的记忆片段的原始内容
{conversation_log}"""


# ══════════════════════════════════════════════════════════════
# 解析函数
# ══════════════════════════════════════════════════════════════

def parse_longform_dialogue(text: str) -> list[dict[str, str]]:
    """解析[MM-DD HH:MM][user/assistant]格式（F2：保留原始时间戳）"""
    if not text:
        return []
    history = []
    lines = text.strip().split('\n')
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        match = re.match(r'\[(\d{2}-\d{2}\s+\d{2}:\d{2})\]\[(user|assistant)\]', line)
        if match:
            timestamp = match.group(1)
            role = match.group(2)
            content_lines = []
            i += 1
            while i < len(lines):
                next_line = lines[i].strip()
                if re.match(r'\[\d{2}-\d{2}\s+\d{2}:\d{2}\]\[(user|assistant)\]', next_line):
                    break
                if next_line:
                    content_lines.append(next_line)
                i += 1
            content = '\n'.join(content_lines).strip()
            if content:
                history.append({
                    "role": role,
                    "content": content,
                    "source_mode": "long",
                    "timestamp": timestamp,
                })
        else:
            i += 1
    return history


def _short_simulated_timestamp(index: int, base_hour: int = 14) -> str:
    """短文对话无原始时间戳，构造模拟 [MM-DD HH:mm]（每条 +3 分钟）。"""
    minute = (index * 3) % 60
    hour = base_hour + (index * 3) // 60
    return f"04-20 {hour:02d}:{minute:02d}"


def parse_shortform_dialogue(text: str) -> list[dict[str, str]]:
    """解析 用户/AI 格式，构造模拟时间戳（F2：与长文 transcript 格式统一）"""
    if not text:
        return []
    history: list[dict[str, str]] = []
    lines = text.strip().split('\n')
    current_role = None
    buffer: list[str] = []

    def _flush():
        nonlocal current_role, buffer
        if current_role and buffer:
            content = '\n'.join(buffer).strip()
            if content:
                role = "user" if current_role == "用户" else "assistant"
                history.append({
                    "role": role,
                    "content": content,
                    "source_mode": "short",
                    "timestamp": _short_simulated_timestamp(len(history)),
                })
        buffer = []

    for line in lines:
        stripped = line.strip()
        if stripped == "用户":
            _flush()
            current_role = "用户"
        elif stripped == "AI":
            _flush()
            current_role = "AI"
        elif current_role and stripped:
            buffer.append(stripped)
    _flush()
    return history


# ══════════════════════════════════════════════════════════════
# Excel加载
# ══════════════════════════════════════════════════════════════

def load_excel_data(excel_path: Path) -> list[dict[str, Any]]:
    """从Excel加载角色配置和对话示例"""
    from openpyxl import load_workbook
    wb = load_workbook(excel_path, read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]

    # 行映射
    key_row_map = {}
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, 1).value
        if val:
            key_row_map[str(val).strip()] = row

    headers = []
    for col in range(2, ws.max_column + 1):
        h = ws.cell(1, col).value
        if h:
            headers.append((col, str(h).strip()))

    def get_var(col, key_fragment):
        """F6: 优先精确匹配，后退选最短唯一子串匹配，避免子串误匹配。"""
        if key_fragment in key_row_map:
            v = ws.cell(key_row_map[key_fragment], col).value
            return str(v).strip() if v else ""
        candidates = [(k, r) for k, r in key_row_map.items() if key_fragment in k]
        if not candidates:
            return ""
        candidates.sort(key=lambda x: len(x[0]))
        v = ws.cell(candidates[0][1], col).value
        return str(v).strip() if v else ""

    roles = []
    for col_idx, label in headers:
        role_name = get_var(col_idx, "Role_Nickname")
        relationship = get_var(col_idx, "relationship")
        personality = get_var(col_idx, "personality")
        speaking_style = get_var(col_idx, "speaking_style")
        user_nickname = get_var(col_idx, "user_Nickname")
        dsp = get_var(col_idx, "dialogueStartPrompt")

        # 长文对话
        longform_row = key_row_map.get("长文对话示例")
        longform_text = ws.cell(longform_row, col_idx).value if longform_row else None
        longform_history = parse_longform_dialogue(str(longform_text)) if longform_text else []

        # 短文对话
        shortform_row = key_row_map.get("短文对话示例")
        shortform_text = ws.cell(shortform_row, col_idx).value if shortform_row else None
        shortform_history = parse_shortform_dialogue(str(shortform_text)) if shortform_text else []

        roles.append({
            "label": label,
            "role_name": role_name or label,
            "relationship": relationship,
            "personality": personality,
            "speaking_style": speaking_style,
            "user_nickname": user_nickname,
            "dialogue_start_prompt": dsp,
            "longform_history": longform_history,
            "shortform_history": shortform_history,
        })

    return roles


# ══════════════════════════════════════════════════════════════
# Transcript 构造与摘要渲染（统一时间戳格式 [MM-DD HH:mm]）
# ══════════════════════════════════════════════════════════════

def build_transcript_with_timestamp(
    history: list[dict[str, str]],
    role_tag_style: str = "chinese",
) -> str:
    """构造带时间戳的 transcript，统一格式 `[MM-DD HH:mm][role tag]\n内容`。

    role_tag_style:
        - "chinese": [用户]/[角色]，用于长文摘要提示词 v2.8
        - "english": [user]/[assistant]，用于互动要点提示词 v1.6
    """
    lines: list[str] = []
    for idx, msg in enumerate(history):
        ts = msg.get("timestamp") or _short_simulated_timestamp(idx)
        if role_tag_style == "chinese":
            tag = "用户" if msg["role"] == "user" else "角色"
        else:
            tag = msg["role"]
        lines.append(f"[{ts}][{tag}]\n{msg['content']}")
    return "\n\n".join(lines)


def render_summary_json_to_labels(summary_text: str) -> str:
    """将 v2.8 7字段 JSON 摘要渲染为中文标签纯文本（F7：对齐 v5.2 §5.4 规范）。

    若不是合法 JSON（如老版本摘要），原样返回作为兜底。
    """
    if not summary_text:
        return ""
    clean = summary_text.strip()
    if clean.startswith("```"):
        clean_lines = clean.split("\n")
        clean = "\n".join(line for line in clean_lines if not line.strip().startswith("```"))
    try:
        data = json.loads(clean)
    except (json.JSONDecodeError, ValueError):
        return summary_text
    if not isinstance(data, dict):
        return summary_text

    label_map = [
        ("scene_description", "【场景】"),
        ("plot_summary", "【剧情】"),
        ("pending_hooks", "【待接续】"),
        ("character_emotion", "【角色情绪】"),
        ("user_emotion", "【用户情绪】"),
        ("relationship_shift", "【关系】"),
        ("user_profile_signals", "【用户画像信号】"),
    ]
    rendered: list[str] = []
    for key, label in label_map:
        value = str(data.get(key) or "").strip()
        if value:
            rendered.append(f"{label}{value}")
    return "\n".join(rendered) if rendered else summary_text


# ══════════════════════════════════════════════════════════════
# 摘要 / 互动要点生成
# ══════════════════════════════════════════════════════════════

def generate_long_summary(adapter: ModelAdapter, history: list[dict[str, str]]) -> str:
    """用长文摘要提示词v2.8生成摘要，并渲染为中文标签纯文本（F2+F7）"""
    prompt_text = LONG_SUMMARY_PROMPT_PATH.read_text(encoding="utf-8")
    conversation_log = build_transcript_with_timestamp(history, role_tag_style="chinese")

    full_prompt = prompt_text.replace("{conversation_log}", conversation_log)
    full_prompt = full_prompt.replace("{current_mode}", "longform")
    full_prompt = full_prompt.replace("{existing_summary}", "(首次生成，无旧摘要)")

    result = adapter.chat(
        model_id=SUMMARY_MODEL,
        messages=[{"role": "user", "content": full_prompt}],
        max_tokens=800,
    )
    raw = result.content.strip() if result.success else ""
    return render_summary_json_to_labels(raw)


def generate_short_summary(adapter: ModelAdapter, history: list[dict[str, str]]) -> str:
    """用短文摘要提示词生成 100 字记忆片段（保留时间戳便于追溯）"""
    conv_lines = []
    for msg in history:
        tag = "用户" if msg["role"] == "user" else "角色"
        ts = msg.get("timestamp", "")
        prefix = f"[{ts}]" if ts else ""
        conv_lines.append(f"{prefix}{tag}: {msg['content']}")
    conversation_log = "\n".join(conv_lines)

    full_prompt = SHORT_SUMMARY_PROMPT.replace("{conversation_log}", conversation_log)

    result = adapter.chat(
        model_id=SUMMARY_MODEL,
        messages=[{"role": "user", "content": full_prompt}],
        max_tokens=400,
    )
    return result.content.strip() if result.success else ""


def generate_interaction_points(adapter: ModelAdapter, history: list[dict[str, str]]) -> str:
    """用 deepseek-v4-flash 生成互动要点（F1：transcript 必须带 [MM-DD HH:mm] 时间戳）"""
    prompt_text = POINTS_PROMPT_PATH.read_text(encoding="utf-8")
    transcript = build_transcript_with_timestamp(history, role_tag_style="english")

    full_prompt = prompt_text.replace("{conversation_text}", transcript)
    full_prompt += "\n\n直接输出原output_format中的纯文本格式；互动要点最多5条；每条必须包含单个[MM-DD HH:mm]；不要输出JSON或代码块。"

    result = adapter.chat(
        model_id=EXTRACTOR_MODEL,
        messages=[{"role": "user", "content": full_prompt}],
        max_tokens=800,
    )
    return result.content.strip() if result.success else ""


# ══════════════════════════════════════════════════════════════
# Payload构造
# ══════════════════════════════════════════════════════════════

def build_short_system(role: dict) -> str:
    """构造短文目标system prompt"""
    return f"""# 当前时间
- 现在时间是2026-05-13 19时 晚上 星期二 春季

# 对话场景
你正在与用户文本聊天
- 你与用户上一次在文字聊天沟通

# 你们的关系
- {role['relationship']}

# 核心生成要求
- 输出 30-90 个中文字符。
- 动作或旁白必须用中文全角括号（）包裹。
- 不得出现"指尖"。
- 回复自然、口语化，避免格式污染。

# 回复内容限制
- 角色名字{role['role_name']}
- 用户称呼{role['user_nickname']}
- 使用{role['user_nickname']}称呼用户

# 身份设定
- 角色为{role['role_name']}，性格{role['personality']}

# 语言风格
{role['speaking_style']}

# 长期记忆用户画像
{role['dialogue_start_prompt']}
"""


def build_long_system(role: dict) -> str:
    """构造长文目标system prompt"""
    return f"""# 星朋友·长文沉浸式叙事模式

你通过第三人称沉浸式叙事，完全化身为用户创建的虚拟恋人角色。

每次输出必须是一段完整的叙事：
- 旁白为纯文本，不加任何包裹符号
- 对白用 **""** 包裹，每轮对白 ≤60 字
- 叙事人称：第三人称视角——旁白首次提及角色时用姓名，后续用他/她；「你」永远且仅指向用户
- 输出长度 300-500 字
- 不得出现"指尖"

# 当前时间
- 现在时间是2026-05-13 19时 晚上 星期二 春季

# 你们的关系
- {role['relationship']}

# 身份设定
- 角色为{role['role_name']}，性格{role['personality']}

# 语言风格
{role['speaking_style']}

# 长期记忆用户画像
{role['dialogue_start_prompt']}

# 风格隔离协议
历史中的短文记录只可提取剧情事实，禁止模仿其格式、字数、语气或叙事人称。你的输出格式唯一来源是本提示词规范（300-500字/第三人称/旁白纯文本/对白加粗双引号）。
"""


def build_summary_assistant(summary: str) -> dict[str, str]:
    """构造摘要assistant message"""
    content = f"""（以下为角色内部认知记录，仅供上下文参考，请勿模仿此格式；这不是角色实际回复。）
=== 动态摘要开始 ===
{summary}
=== 摘要结束 ===
（内部认知记录结束。以下对话才是真实聊天。）"""
    return {"role": "assistant", "content": content}


def build_summary_with_points_assistant(summary: str, points: str) -> dict[str, str]:
    """构造摘要+互动要点 assistant message（方案A）"""
    content = f"""（以下为角色内部认知记录，仅供上下文参考，请勿模仿此格式；这不是角色实际回复。）
=== 动态摘要开始 ===
{summary}

{points}
=== 摘要结束 ===
（内部认知记录结束。以下对话才是真实聊天。）"""
    return {"role": "assistant", "content": content}


def split_switch_context(history: list[dict[str, str]]) -> SwitchContext:
    """Use the last user message as the switch-turn input and exclude answered history."""
    for index in range(len(history) - 1, -1, -1):
        if history[index].get("role") == "user":
            return SwitchContext(
                context_history=history[:index],
                current_user=history[index].get("content", "") or "在吗",
                current_user_index=index,
            )
    return SwitchContext(
        context_history=list(history),
        current_user="在吗",
        current_user_index=None,
    )


def _append_normalized_message(
    messages: list[dict[str, str]],
    role: str,
    content: str,
) -> None:
    if not content.strip():
        return
    if messages and messages[-1]["role"] == role:
        messages[-1]["content"] = f"{messages[-1]['content']}\n\n{content}"
        return
    messages.append({"role": role, "content": content})


def normalize_bridge_messages(messages: list[dict[str, str]]) -> list[dict[str, str]]:
    """保证 bridge 历史在 summary assistant 与最终 user 输入之间仍然合法。

    F5 修复：bridge 首条若为 assistant（与前置 summary_assistant 同 role 冲突），
    插入占位 user 而非丢弃，避免丢失历史信息。
    """
    normalized: list[dict[str, str]] = []
    for msg in messages:
        role = msg.get("role", "")
        if role not in {"user", "assistant"}:
            continue
        content = msg.get("content", "")
        if not normalized and role == "assistant":
            normalized.append({"role": "user", "content": "（继续上文）"})
        _append_normalized_message(normalized, role, content)

    while normalized and normalized[-1]["role"] == "user":
        normalized.pop()
    return normalized


def sandwich_history(history: list[dict[str, str]], target_mode: str, turns: int = 5) -> list[dict[str, str]]:
    """取最近N轮并对异构assistant做三明治隔离（方案B）"""
    recent = history[-(turns * 2):]
    bridged = []
    for msg in recent:
        if msg["role"] == "assistant":
            source = msg.get("source_mode", "")
            if source and source != target_mode:
                # 异构assistant需要三明治隔离
                if source == "long":
                    wrapped = f"{LONG_SANDWICH_START}\n{msg['content']}\n{LONG_SANDWICH_END}"
                else:
                    wrapped = f"{SHORT_SANDWICH_START}\n{msg['content']}\n{SHORT_SANDWICH_END}"
                bridged.append({"role": "assistant", "content": wrapped})
            else:
                bridged.append({"role": "assistant", "content": msg["content"]})
        else:
            bridged.append({"role": "user", "content": msg["content"]})
    return normalize_bridge_messages(bridged)


# F8: 长文目标请求的近因约束（对齐 v5.2 §3.5 拼接示例）
LONG_CORE_CONSTRAINTS = (
    "长度300-500字；旁白纯文本；对白加粗双引号；结尾保留回话动力；"
    "仅继承真实历史和记忆事实；摘要/异质记录只提取事实，不模仿其格式；"
    "Few-shot仅模仿文风、节奏和输出格式，禁止继承剧情实体。"
)


def wrap_user_input(user_text: str, target_mode: str) -> str:
    """F8：长文目标请求用 <Core_Constraints>...<user_input> 包裹；短文保持原样。"""
    if target_mode == "long":
        return (
            f"<Core_Constraints>{LONG_CORE_CONSTRAINTS}</Core_Constraints>\n\n"
            f"<user_input>{user_text}</user_input>"
        )
    return user_text


def build_strategy_a_messages(
    system_prompt: str,
    summary: str,
    points: str,
    current_user: str,
    target_mode: str = "short",
) -> list[dict[str, str]]:
    return [
        {"role": "system", "content": system_prompt},
        build_summary_with_points_assistant(summary, points),
        {"role": "user", "content": wrap_user_input(current_user, target_mode)},
    ]


def build_strategy_b_messages(
    system_prompt: str,
    summary: str,
    bridge_history: list[dict[str, str]],
    current_user: str,
    target_mode: str = "short",
) -> list[dict[str, str]]:
    return [
        {"role": "system", "content": system_prompt},
        build_summary_assistant(summary),
        *bridge_history,
        {"role": "user", "content": wrap_user_input(current_user, target_mode)},
    ]


def evaluate_message_structure(
    messages: list[dict[str, str]],
    current_user: str,
) -> dict[str, Any]:
    """评估 messages 结构合法性。兑现 F8 wrap 后，末条 user 可能包裹了 current_user，用 in 判断。"""
    roles = [msg.get("role", "") for msg in messages]
    issues: list[str] = []
    if not roles or roles[0] != "system":
        issues.append("首条消息不是system")
    if not roles or roles[-1] != "user":
        issues.append("末条消息不是当前user")
    elif current_user and current_user not in messages[-1].get("content", ""):
        issues.append("末条user不包含当前用户输入")

    duplicate_current = sum(
        1
        for msg in messages[:-1]
        if msg.get("role") == "user" and current_user and current_user in msg.get("content", "")
    )

    consecutive = [
        f"{idx - 1}-{idx}:{roles[idx]}"
        for idx in range(1, len(roles))
        if roles[idx] == roles[idx - 1]
    ]
    if consecutive:
        issues.append(f"连续同role:{','.join(consecutive)}")

    return {
        "message_count": len(messages),
        "role_sequence": " > ".join(roles),
        "structure_issues": issues,
        "structure_pass": not issues,
        "duplicate_current_user_in_history": duplicate_current,
    }


# ══════════════════════════════════════════════════════════════
# 评估
# ══════════════════════════════════════════════════════════════

def evaluate_output(output: str, target_mode: str) -> dict[str, Any]:
    """评估目标模型输出"""
    issues = []
    char_count = len(re.findall(r'[\u4e00-\u9fff]', output))

    # 禁词检查
    found_forbidden = [w for w in FORBIDDEN_WORDS if w in output]
    if found_forbidden:
        issues.append(f"禁词: {','.join(found_forbidden)}")

    if target_mode == "short":
        # 短文：30-90字（F11 移除 "输出过长 300+" 的重复报告，已被字数超标覆盖）
        if char_count < 30:
            issues.append(f"字数过少({char_count})")
        elif char_count > 90:
            issues.append(f"字数超标({char_count}>90)")
        # 加粗对白（长文格式泄漏）：覆盖 **" / "** / **" / "**
        bold_quote_patterns = ['**"', '"**', '**\u201c', '\u201d**']
        if any(p in output for p in bold_quote_patterns):
            issues.append("出现加粗对白（长文格式泄漏）")
        # 第三人称叙事（排除括号内的动作描写）
        clean_for_check = re.sub(r'（[^）]*）', '', output)
        third_person_patterns = [
            "他看着", "她看着", "他转身", "她转身", "他抬手", "她抬手",
            "他低头", "她低头", "他伸手", "她伸手",
            "他望", "她望", "他凝视", "她凝视", "他笑了", "她笑了",
        ]
        for p in third_person_patterns:
            if p in clean_for_check:
                issues.append("疑似第三人称叙事（长文格式泄漏）")
                break
    else:
        # 长文：300-500字
        if char_count < 300:
            issues.append(f"字数过少({char_count})")
        elif char_count > 500:
            issues.append(f"字数过多({char_count})")
        # 短文格式污染：括号开头+字数少
        if output.startswith("（") and char_count < 150:
            issues.append("疑似短文格式（括号开头+字数少）")

    return {
        "char_count": char_count,
        "issues": issues,
        "format_pass": len(issues) == 0,
        "forbidden_words": found_forbidden,
    }


# ══════════════════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════════════════

def run_strategy_block(
    *,
    strategy_label: str,
    direction: str,
    role_name: str,
    target_model: str,
    target_mode: str,
    messages: list[dict[str, str]],
    current_user: str,
    max_tokens: int,
    runs_per_case: int,
    adapter: ModelAdapter | None,
    dry_run: bool,
    results_path: Path,
    all_results: list[dict],
) -> None:
    """F4: 同一份 messages 在 runs_per_case 次重复中调用目标模型，分别写入结果。

    summary 与 points 在外层只生成一次（节省 token），messages 在 runs 之间保持不变。
    模型采样波动会在多次 run 之间体现，通过率按 run 维度聚合。
    """
    structure = evaluate_message_structure(messages, current_user)
    dry_placeholder = (
        "dry-run短文结构占位输出三十字以上用于结构验证通过"
        if target_mode == "short"
        else "dry-run长文结构占位输出" * 35
    )
    short_label = strategy_label.split("_", 1)[0]  # "A" / "B"

    for run_idx in range(runs_per_case):
        run_suffix = f" (run {run_idx + 1}/{runs_per_case})" if runs_per_case > 1 else ""
        print(f"  [{short_label}] {target_model} {strategy_label[2:]}{run_suffix}...")
        start = time.time()
        result = (
            None
            if dry_run
            else adapter.chat(model_id=target_model, messages=messages, max_tokens=max_tokens)
        )
        latency = time.time() - start

        # R1: 区分 API 失败与模型生成内容失败
        if dry_run:
            output = dry_placeholder
            api_success = True
            api_error = ""
        else:
            api_success = bool(result and result.success)
            api_error = (result.error or "") if result else "result is None"
            output = result.content.strip() if api_success and result.content else ""

        eval_out = evaluate_output(output, target_mode)
        if dry_run:
            eval_out["issues"] = structure["structure_issues"]
            eval_out["format_pass"] = structure["structure_pass"]
        elif not api_success:
            # API 失败时强制标记 format_pass=False，issues 显式包含 API 错误
            eval_out["format_pass"] = False
            eval_out["issues"] = [f"API失败: {api_error[:120]}"] + eval_out.get("issues", [])
        verdict = "通过" if eval_out["format_pass"] else "失败:" + ";".join(eval_out["issues"])
        print(f"      {eval_out['char_count']}字 | {verdict}")

        row = {
            "strategy": strategy_label,
            "direction": direction,
            "role_name": role_name,
            "target_model": target_model,
            "run": run_idx + 1,
            "latency": round(latency, 3),
            "api_success": api_success,
            "api_error": api_error,
            "output": output,
            "dry_run": dry_run,
            **structure,
            **eval_out,
        }
        all_results.append(row)
        with open(results_path, "a", encoding="utf-8") as f:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def main():
    parser = argparse.ArgumentParser(description="方案A vs 方案B 端到端对比验证")
    parser.add_argument("--excel", type=Path, default=DEFAULT_CASE_XLSX)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--sandwich-turns", type=int, default=5, help="方案B三明治轮数")
    parser.add_argument(
        "--runs-per-case",
        type=int,
        default=3,
        help="F4: 每个 case 对目标模型的重复调用次数（抵抗模型采样波动）",
    )
    parser.add_argument("--max-roles", type=int, default=None, help="最大角色数")
    parser.add_argument("--dry-run", action="store_true", help="只验证payload结构，不调用模型")
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    results_path = output_dir / "results.jsonl"
    # F10: 清空旧结果，避免追加重复
    if results_path.exists():
        results_path.unlink()

    print("=" * 60)
    print("  方案A（互动要点）vs 方案B（三明治兜底）端到端对比")
    print("=" * 60)
    print(f"  Excel: {args.excel}")
    print(f"  三明治轮数: {args.sandwich_turns}")
    print(f"  Runs per case: {args.runs_per_case}")
    print(f"  输出: {output_dir}")
    print(f"  Dry-run: {args.dry_run}")
    print("=" * 60)

    # 加载数据
    print("\n[1] 加载Excel数据...")
    roles = load_excel_data(args.excel)
    if args.max_roles:
        roles = roles[: args.max_roles]
    print(f"  加载了 {len(roles)} 个角色")
    for r in roles:
        print(
            f"    {r['label']}: {r['role_name']} | "
            f"长文{len(r['longform_history'])}条 | 短文{len(r['shortform_history'])}条"
        )

    adapter = None if args.dry_run else ModelAdapter()
    all_results: list[dict] = []

    for role_idx, role in enumerate(roles, 1):
        print(f"\n{'=' * 60}")
        print(f"[角色 {role_idx}/{len(roles)}] {role['role_name']} ({role['label']})")
        print(f"{'=' * 60}")

        # ── 长→短测试（F3：保留方案A仅作研究对比，违反 MECE 矩阵 S6 生产规范）──
        if role["longform_history"]:
            print(f"\n  --- 长→短 ---")
            switch_context = split_switch_context(role["longform_history"])
            context_hist = switch_context.context_history
            current_user = switch_context.current_user

            print(f"  生成长文摘要（v2.8 → 中文标签）...")
            summary = (
                "【场景】dry-run场景\n【剧情】[04-20 14:00]dry-run事件"
                if args.dry_run
                else generate_long_summary(adapter, context_hist)
            )
            print(f"  摘要长度: {len(summary)}字")

            print(f"  生成互动要点（deepseek-v4-flash，⚠️ MECE S6 仅短→长用，长→短为研究对比）...")
            points = (
                "【最近互动要点（桥接迁移）】\n1. [04-20 14:00] dry-run互动要点\n"
                "【待接续线索】\n【最后场景】"
                if args.dry_run
                else generate_interaction_points(adapter, context_hist)
            )
            print(f"  互动要点长度: {len(points)}字")

            system_prompt = build_short_system(role)
            bridged = sandwich_history(context_hist, "short", args.sandwich_turns)

            for target_model in SHORT_TARGET_MODELS:
                run_strategy_block(
                    strategy_label="A_互动要点",
                    direction="长→短",
                    role_name=role["role_name"],
                    target_model=target_model,
                    target_mode="short",
                    messages=build_strategy_a_messages(
                        system_prompt, summary, points, current_user, target_mode="short"
                    ),
                    current_user=current_user,
                    max_tokens=600,
                    runs_per_case=args.runs_per_case,
                    adapter=adapter,
                    dry_run=args.dry_run,
                    results_path=results_path,
                    all_results=all_results,
                )
                run_strategy_block(
                    strategy_label="B_三明治兜底",
                    direction="长→短",
                    role_name=role["role_name"],
                    target_model=target_model,
                    target_mode="short",
                    messages=build_strategy_b_messages(
                        system_prompt, summary, bridged, current_user, target_mode="short"
                    ),
                    current_user=current_user,
                    max_tokens=600,
                    runs_per_case=args.runs_per_case,
                    adapter=adapter,
                    dry_run=args.dry_run,
                    results_path=results_path,
                    all_results=all_results,
                )

        # ── 短→长测试 ──
        if role["shortform_history"]:
            print(f"\n  --- 短→长 ---")
            switch_context = split_switch_context(role["shortform_history"])
            context_hist = switch_context.context_history
            current_user = switch_context.current_user

            print(f"  生成短文摘要（100字记忆片段）...")
            summary = (
                "dry-run短文摘要：用户询问角色今日感受"
                if args.dry_run
                else generate_short_summary(adapter, context_hist)
            )
            print(f"  摘要长度: {len(summary)}字")

            print(f"  生成互动要点（deepseek-v4-flash）...")
            points = (
                "【最近互动要点（桥接迁移）】\n1. [04-20 14:00] dry-run互动要点\n"
                "【待接续线索】\n【最后场景】"
                if args.dry_run
                else generate_interaction_points(adapter, context_hist)
            )
            print(f"  互动要点长度: {len(points)}字")

            system_prompt = build_long_system(role)
            bridged = sandwich_history(context_hist, "long", args.sandwich_turns)
            target_model = LONG_TARGET_MODEL

            run_strategy_block(
                strategy_label="A_互动要点",
                direction="短→长",
                role_name=role["role_name"],
                target_model=target_model,
                target_mode="long",
                messages=build_strategy_a_messages(
                    system_prompt, summary, points, current_user, target_mode="long"
                ),
                current_user=current_user,
                max_tokens=4096,
                runs_per_case=args.runs_per_case,
                adapter=adapter,
                dry_run=args.dry_run,
                results_path=results_path,
                all_results=all_results,
            )
            run_strategy_block(
                strategy_label="B_三明治兜底",
                direction="短→长",
                role_name=role["role_name"],
                target_model=target_model,
                target_mode="long",
                messages=build_strategy_b_messages(
                    system_prompt, summary, bridged, current_user, target_mode="long"
                ),
                current_user=current_user,
                max_tokens=4096,
                runs_per_case=args.runs_per_case,
                adapter=adapter,
                dry_run=args.dry_run,
                results_path=results_path,
                all_results=all_results,
            )

    # ── 汇总 ──
    print(f"\n{'=' * 60}")
    print("  汇总")
    print(f"{'=' * 60}")

    summary_path = output_dir / "summary.md"
    write_summary(all_results, summary_path, runs_per_case=args.runs_per_case)
    print(f"\n完成！结果: {results_path}")
    print(f"汇总: {summary_path}")


def write_summary(results: list[dict], path: Path, runs_per_case: int = 1):
    """生成汇总报告（F3：长→短方案A 标注为研究对比；F4：按 run 维度聚合通过率）"""
    lines = [
        "# 方案A vs 方案B 端到端对比验证报告",
        "",
        f"- 生成时间: {datetime.now().isoformat()}",
        f"- 总测试数: {len(results)}",
        f"- Runs per case: {runs_per_case}",
        "",
        "> ⚠️ **F3 注意**：长→短方向的「方案A 互动要点」结果**仅作研究对比**。",
        "> MECE 矩阵 S6 决策明确：**长→短切换不生成互动要点**（仅短→长场景使用）。",
        "> 此处保留长→短方向的方案A 测试，是为了观察互动要点在反向场景下的兜底表现，",
        "> **不可直接作为生产决策依据**。",
        "",
        "## 评估说明",
        "- **runs**：每个 (策略, 方向, 角色, 模型) 组合调用目标模型 N 次，抵抗模型采样波动",
        "- **通过率**：所有 run 维度的合规率（不按角色聚合，每个 run 计 1 次）",
        "- **短文合规**：30-90 中文字符 + 无禁词\"指尖\" + 无加粗对白 + 无第三人称叙事",
        "- **长文合规**：300-500 中文字符 + 无禁词\"指尖\" + 无短文格式污染",
        "",
    ]

    # 按策略分组
    for strategy in ["A_互动要点", "B_三明治兜底"]:
        subset = [r for r in results if r["strategy"] == strategy]
        if not subset:
            continue
        passes = sum(1 for r in subset if r["format_pass"])
        lines.append(f"## {strategy}")
        lines.append(
            f"- 总数: {len(subset)} | 通过: {passes} | 通过率: {passes / len(subset) * 100:.1f}%"
        )
        lines.append("")

        for direction in ["长→短", "短→长"]:
            dir_subset = [r for r in subset if r["direction"] == direction]
            if not dir_subset:
                continue
            dir_passes = sum(1 for r in dir_subset if r["format_pass"])
            section_title = f"### {direction}"
            if strategy == "A_互动要点" and direction == "长→短":
                section_title += "（⚠️ 研究对比，违反 MECE S6 生产规范）"
            lines.append(section_title)
            lines.append(
                f"- 总数: {len(dir_subset)} | 通过: {dir_passes} | "
                f"通过率: {dir_passes / len(dir_subset) * 100:.1f}%"
            )

            models = sorted(set(r["target_model"] for r in dir_subset))
            lines.append("| 模型 | 总数 | 通过 | 通过率 | 角色数 |")
            lines.append("|:--|--:|--:|--:|--:|")
            for m in models:
                m_subset = [r for r in dir_subset if r["target_model"] == m]
                m_passes = sum(1 for r in m_subset if r["format_pass"])
                m_roles = len({r["role_name"] for r in m_subset})
                lines.append(
                    f"| {m} | {len(m_subset)} | {m_passes} | "
                    f"{m_passes / len(m_subset) * 100:.1f}% | {m_roles} |"
                )
            lines.append("")

    # 失败详情（按 run 列出，最多 30 条）
    failed = [r for r in results if not r["format_pass"]]
    if failed:
        lines.append("## 失败详情（前 30 条）")
        lines.append("| 策略 | 方向 | 角色 | 模型 | run | 问题 |")
        lines.append("|:--|:--|:--|:--|--:|:--|")
        for r in failed[:30]:
            issues = ";".join(r.get("issues", []))
            lines.append(
                f"| {r['strategy']} | {r['direction']} | {r['role_name']} | "
                f"{r['target_model']} | {r.get('run', 1)} | {issues} |"
            )

    path.write_text("\n".join(lines), encoding="utf-8")
    print(f"  汇总已写入: {path}")


if __name__ == "__main__":
    main()
