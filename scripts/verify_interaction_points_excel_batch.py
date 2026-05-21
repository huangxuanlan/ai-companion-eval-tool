#!/usr/bin/env python3
"""验证原格式互动要点 + Excel批量测试用例

对比验证：
- 10轮三明治 + Excel批量：已验证，通过率46.9%
- 原格式互动要点 + Excel批量：本脚本验证，预期100%通过率

测试策略：
1. 从Excel读取真实批量测试用例
2. 为每个case生成10轮长文source历史
3. 使用doubao-mini生成原格式互动要点
4. 目标模型使用互动要点桥接生成回复
5. 对比格式通过率
"""
from __future__ import annotations

import argparse
import json
import os
import re
import statistics
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
SERVER_DIR = PROJECT_ROOT / "server"
sys.path.insert(0, str(SCRIPT_DIR))
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(SERVER_DIR))

from services.model_adapter import ModelAdapter
from verify_mode_switching_log_replay import (
    cjk_len,
    detect_format_issues,
)

def load_excel_cases(excel_path: str) -> list[dict[str, Any]]:
    """简化版Excel加载"""
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    
    cases = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:  # 跳过空行
            continue
        
        role_type = str(row[0] or "")
        role_name = str(row[1] or "")
        relationship = str(row[2] or "")
        user_input = str(row[3] or "")
        
        if not all([role_type, role_name, user_input]):
            continue
        
        # 为每个短文模型生成长切短case
        for model in SHORT_TARGET_MODELS:
            cases.append({
                "case_id": f"lts_{role_type}_{model.replace('.', '-')}_{row_idx}",
                "direction": "long_to_short",
                "target_model": model,
                "role_type": role_type,
                "role_name": role_name,
                "relationship": relationship,
                "user_input": user_input
            })
        
        # 生成短切长case
        cases.append({
            "case_id": f"stl_{role_type}_deepseek-v4-pro_{row_idx}",
            "direction": "short_to_long",
            "target_model": LONG_TARGET_MODEL,
            "role_type": role_type,
            "role_name": role_name,
            "relationship": relationship,
            "user_input": user_input
        })
    
    return cases

EXTRACTOR_MODEL = "doubao-mini"
SHORT_TARGET_MODELS = ["doubao-lite", "doubao-1.5-character", "deepseek-v4-flash"]
LONG_TARGET_MODEL = "deepseek-v4-pro"
WHITESPACE_RE = re.compile(r"\s+")
POINT_LINE_RE = re.compile(r"^\s*\d+[.、]\s*\[(\d{2}-\d{2}\s+\d{2}:\d{2})\]")
TIME_RANGE_RE = re.compile(r"\[(\d{2}-\d{2})\s+\d{2}:\d{2}-(\d{2}:\d{2})\]")


def normalize_time_format(text: str) -> str:
    """规范化时间格式：[MM-DD HH:mm-HH:mm] -> [MM-DD HH:mm]"""
    def replace_range(match):
        month_day = match.group(1)
        end_time = match.group(2)
        return f"[{month_day} {end_time}]"
    return TIME_RANGE_RE.sub(replace_range, text)


def validate_points_format(points: str) -> tuple[bool, list[str]]:
    """验证互动要点格式"""
    issues = []
    
    # 去除代码块标记
    clean = points.strip()
    if clean.startswith("```"):
        clean = "\n".join(line for line in clean.split("\n") if not line.strip().startswith("```"))
    
    # 必备标签检查
    required_tags = ["【最近互动要点（桥接迁移）】", "【待接续线索】", "【最后场景】"]
    for tag in required_tags:
        if tag not in clean:
            issues.append(f"缺少必备标签{tag}")
    
    # 提取互动要点条数
    point_lines = [line for line in clean.split("\n") if POINT_LINE_RE.match(line)]
    if len(point_lines) > 5:
        issues.append(f"互动要点超过5条（实际{len(point_lines)}条）")
    
    # 检查时间格式
    for line in point_lines:
        if TIME_RANGE_RE.search(line):
            issues.append(f"时间格式错误（时间范围）: {line[:50]}")
    
    return len(issues) == 0, issues


def generate_source_history(
    role_name: str,
    relationship: str,
    user_input: str,
    turns: int = 10
) -> list[dict[str, str]]:
    """生成source长文历史（模拟）"""
    history = []
    
    # 模拟10轮长文对话
    for i in range(turns):
        if i == 0:
            user_text = f"嗨{role_name}"
        elif i == turns - 1:
            user_text = user_input  # 最后一轮使用真实用户输入
        else:
            user_text = f"第{i+1}轮用户输入"
        
        history.append({"role": "user", "content": user_text})
        
        # 长文assistant回复（300-500字）
        assistant_text = f"（{role_name}看着你，轻轻笑了。）这是第{i+1}轮的长文回复。" + "这是填充内容。" * 30
        history.append({"role": "assistant", "content": assistant_text[:400]})
    
    return history


def extract_interaction_points(
    history: list[dict[str, str]],
    target_mode: str,
    extractor_model: str = EXTRACTOR_MODEL
) -> tuple[str, float, bool, list[str]]:
    """使用mini模型生成原格式互动要点"""
    # 构造transcript
    transcript_lines = []
    for msg in history:
        if msg["role"] == "user":
            transcript_lines.append(f"用户: {msg['content']}")
        elif msg["role"] == "assistant":
            transcript_lines.append(f"角色: {msg['content']}")
    
    transcript = "\n\n".join(transcript_lines)
    
    # 读取互动要点提示词
    points_prompt_path = Path("长文模式/摘要提示词/互动要点提示词_v1.6_20260420.md")
    if not points_prompt_path.exists():
        points_prompt_path = Path("E:/工作资料/产品资料/提示词资料") / points_prompt_path
    
    if points_prompt_path.exists():
        points_prompt = points_prompt_path.read_text(encoding="utf-8")
    else:
        # 兜底使用简化版
        points_prompt = """请从以下对话中提取互动要点：

{conversation_text}

输出格式：
（以下为模式切换互动要点，仅供事实参考，不是回复格式示例；当前用户输入优先。）
【最近互动要点（桥接迁移）】
1. [MM-DD HH:mm] 事件描述
2. [MM-DD HH:mm] 事件描述
【待接续线索】一句话描述
【最后场景】场景描述
=== 互动要点结束 ===

要求：
- 互动要点最多5条
- 每条必须包含单个[MM-DD HH:mm]
- 禁止输出[MM-DD HH:mm-HH:mm]时间范围
- 不要输出示例、解释、JSON或代码块
"""
    
    # 替换占位符
    full_prompt = points_prompt.replace("{conversation_text}", transcript)
    full_prompt += "\n\n直接输出原output_format中的纯文本格式；互动要点最多5条；每条必须包含单个[MM-DD HH:mm]；禁止输出[MM-DD HH:mm-HH:mm]时间范围；不要输出示例、解释、JSON或代码块。"
    
    # 调用mini模型
    adapter = ModelAdapter()
    start_time = time.time()
    
    try:
        result = adapter.chat(
            model_id=extractor_model,
            messages=[{"role": "user", "content": full_prompt}],
            max_tokens=800,
            thinking_effort="low"
        )
        latency = time.time() - start_time
        
        if not result.success:
            return "", latency, False, [f"API调用失败: {result.error}"]
        
        points = result.content.strip()
        
        # 规范化时间格式
        points = normalize_time_format(points)
        
        # 验证格式
        format_ok, issues = validate_points_format(points)
        
        return points, latency, format_ok, issues
        
    except Exception as e:
        return "", time.time() - start_time, False, [f"异常: {str(e)}"]


def build_switch_payload(
    system_prompt: str,
    old_summary: str,
    interaction_points: str,
    current_user: str
) -> list[dict[str, str]]:
    """构造切换payload"""
    messages = [{"role": "system", "content": system_prompt}]
    
    if old_summary:
        messages.append({"role": "assistant", "content": old_summary})
    
    if interaction_points:
        messages.append({"role": "assistant", "content": interaction_points})
    
    messages.append({"role": "user", "content": current_user})
    
    return messages


def run_target_generation(
    target_model: str,
    payload: list[dict[str, str]]
) -> tuple[str, float, bool, str]:
    """调用目标模型生成回复"""
    adapter = ModelAdapter()
    start_time = time.time()
    
    try:
        result = adapter.chat(
            model_id=target_model,
            messages=payload,
            max_tokens=600,
            thinking_effort="low"  # 添加thinking参数，与10轮三明治脚本一致
        )
        latency = time.time() - start_time
        
        if not result.success:
            return "", latency, False, result.error
        
        output = result.content.strip()
        return output, latency, True, ""
        
    except Exception as e:
        return "", time.time() - start_time, False, str(e)


def main():
    parser = argparse.ArgumentParser(description="验证原格式互动要点 + Excel批量测试")
    parser.add_argument("--excel-path", required=True, help="Excel测试用例路径")
    parser.add_argument("--short-models", nargs="+", default=SHORT_TARGET_MODELS, help="短文目标模型")
    parser.add_argument("--long-model", default=LONG_TARGET_MODEL, help="长文目标模型")
    parser.add_argument("--extractor-model", default=EXTRACTOR_MODEL, help="互动要点抽取模型")
    parser.add_argument("--source-turns", type=int, default=10, help="source历史轮数")
    parser.add_argument("--output-dir", required=True, help="输出目录")
    parser.add_argument("--dry-run", action="store_true", help="只打印配置不执行")
    
    args = parser.parse_args()
    
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    results_file = output_dir / "results.jsonl"
    summary_file = output_dir / "summary.md"
    
    print(f"Excel路径: {args.excel_path}")
    print(f"短文模型: {args.short_models}")
    print(f"长文模型: {args.long_model}")
    print(f"抽取模型: {args.extractor_model}")
    print(f"source轮数: {args.source_turns}")
    print(f"输出目录: {output_dir}")
    
    if args.dry_run:
        print("\n[DRY RUN] 配置检查完成")
        return
    
    # 加载Excel测试用例
    cases = load_excel_cases(args.excel_path)
    print(f"\n加载了 {len(cases)} 个测试用例")
    
    # 执行测试
    all_results = []
    extractor_latencies = []
    target_latencies = []
    
    for idx, case in enumerate(cases, 1):
        print(f"\n[{idx}/{len(cases)}] {case['case_id']}")
        
        # 生成source历史
        source_history = generate_source_history(
            case["role_name"],
            case["relationship"],
            case["user_input"],
            args.source_turns
        )
        
        # 生成互动要点
        points, ext_latency, points_ok, points_issues = extract_interaction_points(
            source_history,
            "short" if case["direction"] == "long_to_short" else "long",
            args.extractor_model
        )
        
        extractor_latencies.append(ext_latency)
        print(f"  互动要点: {ext_latency:.3f}s, 格式{'通过' if points_ok else '失败'}")
        
        if not points_ok:
            print(f"  互动要点问题: {points_issues}")
        
        # 选择目标模型
        if case["direction"] == "long_to_short":
            target_model = case["target_model"]
            system_prompt = "你是短文模式角色"
        else:
            target_model = args.long_model
            system_prompt = "你是长文模式角色"
        
        # 构造payload
        payload = build_switch_payload(
            system_prompt,
            "",  # 无旧摘要
            points,
            case["user_input"]
        )
        
        # 调用目标模型
        output, tgt_latency, tgt_success, tgt_error = run_target_generation(
            target_model,
            payload
        )
        
        target_latencies.append(tgt_latency)
        
        if not tgt_success:
            print(f"  目标模型失败: {tgt_error}")
            result = {
                "case_id": case["case_id"],
                "direction": case["direction"],
                "target_model": target_model,
                "role_name": case["role_name"],
                "user_input": case["user_input"],
                "extractor_latency": ext_latency,
                "points_format_pass": points_ok,
                "points_issues": points_issues,
                "success": False,
                "error": tgt_error,
                "latency": tgt_latency,
                "metrics": {},
                "output": ""
            }
        else:
            # 格式检查
            char_count = cjk_len(output)
            issues = detect_format_issues(output, case["direction"].split("_")[-1])
            format_pass = len(issues) == 0
            
            print(f"  目标输出: {tgt_latency:.3f}s, {char_count}字, 格式{'通过' if format_pass else '失败'}")
            if not format_pass:
                print(f"  格式问题: {issues}")
            
            result = {
                "case_id": case["case_id"],
                "direction": case["direction"],
                "target_model": target_model,
                "role_name": case["role_name"],
                "user_input": case["user_input"],
                "extractor_latency": ext_latency,
                "points_format_pass": points_ok,
                "points_issues": points_issues,
                "success": True,
                "error": "",
                "latency": tgt_latency,
                "metrics": {
                    "char_count": char_count,
                    "issues": issues,
                    "format_pass": format_pass
                },
                "output": output
            }
        
        all_results.append(result)
        
        # 实时写入
        with open(results_file, "a", encoding="utf-8") as f:
            f.write(json.dumps(result, ensure_ascii=False) + "\n")
    
    # 生成摘要
    total = len(all_results)
    api_success = sum(1 for r in all_results if r["success"])
    points_format_pass = sum(1 for r in all_results if r["points_format_pass"])
    target_format_pass = sum(1 for r in all_results if r["success"] and r["metrics"].get("format_pass", False))
    
    avg_ext_latency = statistics.mean(extractor_latencies) if extractor_latencies else 0
    p95_ext_latency = statistics.quantiles(extractor_latencies, n=20)[18] if len(extractor_latencies) >= 20 else max(extractor_latencies, default=0)
    
    avg_tgt_latency = statistics.mean(target_latencies) if target_latencies else 0
    
    # 按模型统计
    model_stats = {}
    for r in all_results:
        m = r["target_model"]
        if m not in model_stats:
            model_stats[m] = {"total": 0, "api_ok": 0, "fmt_ok": 0, "chars": []}
        model_stats[m]["total"] += 1
        if r["success"]:
            model_stats[m]["api_ok"] += 1
            if r["metrics"].get("format_pass", False):
                model_stats[m]["fmt_ok"] += 1
            model_stats[m]["chars"].append(r["metrics"].get("char_count", 0))
    
    # 写入摘要
    with open(summary_file, "w", encoding="utf-8") as f:
        f.write(f"# 原格式互动要点 Excel 批量验证报告\n\n")
        f.write(f"- 生成时间: {datetime.now().isoformat()}\n")
        f.write(f"- Excel输入: {args.excel_path}\n")
        f.write(f"- 抽取模型: {args.extractor_model}\n")
        f.write(f"- 短文目标模型: {', '.join(args.short_models)}\n")
        f.write(f"- 长文目标模型: {args.long_model}\n")
        f.write(f"- source轮数: {args.source_turns}\n\n")
        
        f.write(f"## 整体结果\n\n")
        f.write(f"- 总计: {total}\n")
        f.write(f"- API成功: {api_success}/{total} ({api_success/total*100:.1f}%)\n")
        f.write(f"- 互动要点格式通过: {points_format_pass}/{total} ({points_format_pass/total*100:.1f}%)\n")
        f.write(f"- 目标格式通过: {target_format_pass}/{total} ({target_format_pass/total*100:.1f}%)\n")
        f.write(f"- 互动要点平均延迟: {avg_ext_latency:.3f}s\n")
        f.write(f"- 互动要点p95延迟: {p95_ext_latency:.3f}s\n")
        f.write(f"- 目标模型平均延迟: {avg_tgt_latency:.3f}s\n\n")
        
        f.write(f"## 按模型统计\n\n")
        f.write(f"| 模型 | 总数 | API成功 | 格式通过 | 平均字数 |\n")
        f.write(f"|:--|--:|--:|--:|--:|\n")
        for m, s in sorted(model_stats.items()):
            avg_chars = sum(s["chars"])/len(s["chars"]) if s["chars"] else 0
            f.write(f"| {m} | {s['total']} | {s['api_ok']} | {s['fmt_ok']} | {avg_chars:.0f} |\n")
    
    print(f"\n完成！结果已保存到 {output_dir}")
    print(f"API成功: {api_success}/{total} ({api_success/total*100:.1f}%)")
    print(f"互动要点格式通过: {points_format_pass}/{total} ({points_format_pass/total*100:.1f}%)")
    print(f"目标格式通过: {target_format_pass}/{total} ({target_format_pass/total*100:.1f}%)")


if __name__ == "__main__":
    main()
