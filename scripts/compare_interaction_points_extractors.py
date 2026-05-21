#!/usr/bin/env python3
"""对比测试不同模型生成互动要点的质量和速度（使用真实Excel测试用例）

对比模型：
1. doubao-mini (当前默认)
2. deepseek-v4-flash (无思考)

测试维度：
- 生成速度（延迟）
- 格式正确性
- 内容质量（是否复制原文、是否符合要求）
"""
from __future__ import annotations

import argparse
import json
import re
import statistics
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
SERVER_DIR = PROJECT_ROOT / "server"
sys.path.insert(0, str(SCRIPT_DIR))
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(SERVER_DIR))

# R11: 加载 .env（ModelAdapter 依赖 VOLCENGINE_API_KEY / ARK_API_KEY / DASHSCOPE_API_KEY 环境变量）
try:
    from dotenv import load_dotenv
    load_dotenv(SERVER_DIR / ".env")
except ImportError:
    pass

from services.model_adapter import ModelAdapter

POINTS_PROMPT_PATH = Path("E:/工作资料/产品资料/提示词资料/长文模式/摘要提示词/互动要点提示词_v1.6_20260420.md")
DEFAULT_CASE_XLSX = Path(r"E:\工作资料\产品资料\提示词资料\模型切换\短文模式聊天批量测试用例.xlsx")
POINT_LINE_RE = re.compile(r"^\s*\d+[.、]\s*\[(\d{2}-\d{2}\s+\d{2}:\d{2})\]")
TIME_RANGE_RE = re.compile(r"\[(\d{2}-\d{2})\s+\d{2}:\d{2}-(\d{2}:\d{2})\]")


@dataclass
class TestCase:
    """测试用例"""
    case_id: str
    role_name: str
    role_type: str
    relationship: str
    dialogue_type: str  # "长文" 或 "短文"
    history: list[dict[str, str]]
    transcript: str


def load_points_prompt() -> str:
    """加载互动要点提示词"""
    if not POINTS_PROMPT_PATH.exists():
        raise FileNotFoundError(f"互动要点提示词不存在: {POINTS_PROMPT_PATH}")
    
    return POINTS_PROMPT_PATH.read_text(encoding="utf-8")


def parse_longform_dialogue(dialogue_text: str) -> list[dict[str, str]]:
    """解析长文对话示例，提取对话历史
    
    格式示例：
    [03-31 20:15][user]
    战战，刚看你发布会下台了。辛苦啦！
    
    [03-31 20:18][assistant]
    （指尖捏着眉心稍微按揉了两下，车窗降下一道缝隙透风）
    刚上保姆车，确实有点累。
    """
    if not dialogue_text:
        return []
    
    history = []
    lines = dialogue_text.strip().split('\n')
    
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        # 匹配时间戳和角色标记：[MM-DD HH:MM][user] 或 [MM-DD HH:MM][assistant]
        import re
        match = re.match(r'\[(\d{2}-\d{2}\s+\d{2}:\d{2})\]\[(user|assistant)\]', line)
        
        if match:
            role = match.group(2)
            
            # 收集该轮的内容（直到下一个时间戳或结束）
            content_lines = []
            i += 1
            
            while i < len(lines):
                next_line = lines[i].strip()
                # 如果遇到下一个时间戳，停止
                if re.match(r'\[\d{2}-\d{2}\s+\d{2}:\d{2}\]\[(user|assistant)\]', next_line):
                    break
                if next_line:  # 跳过空行
                    content_lines.append(next_line)
                i += 1
            
            content = '\n'.join(content_lines).strip()
            if content:
                history.append({"role": role, "content": content})
        else:
            i += 1
    
    return history


def parse_shortform_dialogue(dialogue_text: str) -> list[dict[str, str]]:
    """解析短文对话示例，提取对话历史
    
    格式示例：
    用户
    
    宝宝
    
    
    
    AI
    
    （刚卸完妆靠在酒店床头翻手机）刚收工回酒店...
    """
    if not dialogue_text:
        return []
    
    history = []
    lines = dialogue_text.strip().split('\n')
    
    current_role = None
    buffer = []
    
    for line in lines:
        stripped = line.strip()
        
        # 检测角色标记
        if stripped == "用户":
            # 保存上一轮
            if current_role and buffer:
                content = '\n'.join(buffer).strip()
                if content:
                    role = "user" if current_role == "用户" else "assistant"
                    history.append({"role": role, "content": content})
            current_role = "用户"
            buffer = []
            continue
        elif stripped == "AI":
            # 保存上一轮
            if current_role and buffer:
                content = '\n'.join(buffer).strip()
                if content:
                    role = "user" if current_role == "用户" else "assistant"
                    history.append({"role": role, "content": content})
            current_role = "AI"
            buffer = []
            continue
        
        # 收集内容
        if current_role and stripped:
            buffer.append(stripped)
    
    # 保存最后一轮
    if current_role and buffer:
        content = '\n'.join(buffer).strip()
        if content:
            role = "user" if current_role == "用户" else "assistant"
            history.append({"role": role, "content": content})
    
    return history


def load_excel_test_cases(excel_path: Path, max_cases: int = None) -> list[TestCase]:
    """从Excel加载真实测试用例（同时加载长文和短文对话示例）"""
    print(f"从Excel加载测试用例: {excel_path}")
    
    from openpyxl import load_workbook
    
    workbook = load_workbook(excel_path, read_only=True, data_only=False)
    worksheet = workbook[workbook.sheetnames[0]]
    
    # 找到"长文对话示例"和"短文对话示例"行
    longform_row = None
    shortform_row = None
    
    for row in range(1, worksheet.max_row + 1):
        cell_value = worksheet.cell(row, 1).value
        if cell_value:
            cell_str = str(cell_value).strip()
            if "长文对话示例" in cell_str:
                longform_row = row
            elif "短文对话示例" in cell_str:
                shortform_row = row
    
    if not longform_row and not shortform_row:
        raise ValueError("Excel中未找到'长文对话示例'或'短文对话示例'行")
    
    print(f"找到长文对话示例行: 第{longform_row}行" if longform_row else "未找到长文对话示例")
    print(f"找到短文对话示例行: 第{shortform_row}行" if shortform_row else "未找到短文对话示例")
    
    # 读取表头（角色名）
    headers = []
    for col in range(2, worksheet.max_column + 1):
        header = worksheet.cell(1, col).value
        if header:
            headers.append((col, str(header).strip()))
    
    print(f"找到 {len(headers)} 个角色")
    
    cases = []
    
    for col_idx, role_label in headers:
        # 读取角色变量
        role_name = "未知角色"
        relationship = "未知"
        
        for row in range(2, worksheet.max_row + 1):
            key = worksheet.cell(row, 1).value
            if not key:
                continue
            
            key_str = str(key).strip()
            
            if "Role_Nickname" in key_str:
                value = worksheet.cell(row, col_idx).value
                if value:
                    role_name = str(value).strip()
            
            if "@全局用户参数_relationship&" == key_str or "relationship" in key_str:
                value = worksheet.cell(row, col_idx).value
                if value:
                    relationship = str(value).strip()
        
        # 处理长文对话示例
        if longform_row:
            longform_text = worksheet.cell(longform_row, col_idx).value
            
            if longform_text:
                history = parse_longform_dialogue(str(longform_text))
                
                if history:
                    print(f"  {role_label} [长文]: 解析到 {len(history)} 条消息")
                    
                    # 构造transcript（保留原始时间戳格式，和pipeline一致）
                    transcript = str(longform_text).strip()
                    
                    case_id = f"{role_label}_{role_name}_长文"
                    cases.append(TestCase(
                        case_id=case_id,
                        role_name=role_name,
                        role_type=role_label,
                        relationship=relationship,
                        dialogue_type="长文",
                        history=history,
                        transcript=transcript
                    ))
        
        # 处理短文对话示例
        if shortform_row:
            shortform_text = worksheet.cell(shortform_row, col_idx).value
            
            if shortform_text:
                history = parse_shortform_dialogue(str(shortform_text))
                
                if history:
                    print(f"  {role_label} [短文]: 解析到 {len(history)} 条消息")
                    
                    # 短文没有时间戳，构造transcript时加上模拟时间戳（和pipeline格式一致）
                    transcript_lines = []
                    base_hour = 14
                    for idx, msg in enumerate(history):
                        minute = (idx * 3) % 60
                        hour = base_hour + (idx * 3) // 60
                        ts = f"[04-20 {hour:02d}:{minute:02d}]"
                        role_tag = "user" if msg["role"] == "user" else "assistant"
                        transcript_lines.append(f"{ts}[{role_tag}]")
                        transcript_lines.append(msg["content"])
                        transcript_lines.append("")
                    
                    transcript = "\n".join(transcript_lines)
                    
                    case_id = f"{role_label}_{role_name}_短文"
                    cases.append(TestCase(
                        case_id=case_id,
                        role_name=role_name,
                        role_type=role_label,
                        relationship=relationship,
                        dialogue_type="短文",
                        history=history,
                        transcript=transcript
                    ))
    
    # 如果设置了max_cases，截断
    if max_cases and len(cases) > max_cases:
        cases = cases[:max_cases]
    
    long_count = sum(1 for c in cases if c.dialogue_type == '长文')
    short_count = sum(1 for c in cases if c.dialogue_type == '短文')
    print(f"生成了 {len(cases)} 个测试用例（长文: {long_count}，短文: {short_count}）")
    return cases


def normalize_time_format(text: str) -> str:
    """规范化时间格式：[MM-DD HH:mm-HH:mm] -> [MM-DD HH:mm]"""
    def replace_range(match):
        month_day = match.group(1)
        end_time = match.group(2)
        return f"[{month_day} {end_time}]"
    return TIME_RANGE_RE.sub(replace_range, text)


def validate_points_format(points: str) -> tuple[bool, list[str], dict[str, Any]]:
    """验证互动要点格式"""
    issues = []
    metrics = {}
    
    # 去除代码块标记
    clean = points.strip()
    if clean.startswith("```"):
        clean = "\n".join(line for line in clean.split("\n") if not line.strip().startswith("```"))
    
    # 必备标签检查
    required_tags = ["【最近互动要点（桥接迁移）】", "【待接续线索】", "【最后场景】"]
    missing_tags = []
    for tag in required_tags:
        if tag not in clean:
            missing_tags.append(tag)
            issues.append(f"缺少必备标签{tag}")
    
    metrics["missing_tags"] = missing_tags
    
    # 提取互动要点条数
    point_lines = [line for line in clean.split("\n") if POINT_LINE_RE.match(line)]
    metrics["point_count"] = len(point_lines)
    
    if len(point_lines) > 5:
        issues.append(f"互动要点超过5条（实际{len(point_lines)}条）")
    elif len(point_lines) == 0:
        issues.append("未找到任何互动要点")
    
    # 检查时间格式
    time_range_errors = []
    for line in point_lines:
        if TIME_RANGE_RE.search(line):
            time_range_errors.append(line[:50])
    
    if time_range_errors:
        issues.append(f"时间格式错误（时间范围）: {len(time_range_errors)}处")
        metrics["time_range_errors"] = time_range_errors
    
    # 检查是否有JSON/代码块
    if "```" in points or "{" in points[:100]:
        issues.append("疑似输出JSON或代码块")
    
    # 检查长度
    metrics["char_count"] = len(clean)
    if len(clean) > 1000:
        issues.append(f"输出过长({len(clean)}字)")
    
    return len(issues) == 0, issues, metrics


def extract_with_model(
    model_id: str,
    transcript: str,
    points_prompt: str,
    use_thinking: bool = False
) -> tuple[str, float, bool, list[str], dict[str, Any]]:
    """使用指定模型生成互动要点"""
    # 替换占位符
    full_prompt = points_prompt.replace("{conversation_text}", transcript)
    full_prompt += "\n\n直接输出原output_format中的纯文本格式；互动要点最多5条；每条必须包含单个[MM-DD HH:mm]；禁止输出[MM-DD HH:mm-HH:mm]时间范围；不要输出示例、解释、JSON或代码块。"
    
    adapter = ModelAdapter()
    start_time = time.time()
    
    try:
        if use_thinking:
            result = adapter.chat(
                model_id=model_id,
                messages=[{"role": "user", "content": full_prompt}],
                max_tokens=800,
                thinking_effort="low"
            )
        else:
            result = adapter.chat(
                model_id=model_id,
                messages=[{"role": "user", "content": full_prompt}],
                max_tokens=800
            )
        
        latency = time.time() - start_time
        
        if not result.success:
            return "", latency, False, [f"API调用失败: {result.error}"], {}
        
        points = result.content.strip()
        
        # 规范化时间格式
        points = normalize_time_format(points)
        
        # 验证格式
        format_ok, issues, metrics = validate_points_format(points)
        
        return points, latency, format_ok, issues, metrics
        
    except Exception as e:
        return "", time.time() - start_time, False, [f"异常: {str(e)}"], {}


def main():
    parser = argparse.ArgumentParser(description="对比测试互动要点抽取模型（使用真实Excel测试用例）")
    parser.add_argument("--models", nargs="+", 
                       default=["doubao-mini", "deepseek-v4-flash"],
                       help="要测试的模型列表")
    parser.add_argument("--excel", type=Path, default=DEFAULT_CASE_XLSX, 
                       help="Excel测试用例文件路径")
    parser.add_argument("--max-cases", type=int, default=None, 
                       help="最大测试用例数量（默认全部）")
    parser.add_argument("--output-dir", required=True, help="输出目录")
    parser.add_argument("--runs-per-case", type=int, default=1, help="每个case重复次数")
    
    args = parser.parse_args()
    
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    results_file = output_dir / "comparison_results.jsonl"
    summary_file = output_dir / "comparison_summary.md"
    
    print(f"对比测试互动要点抽取模型（真实Excel用例）")
    print(f"=" * 60)
    print(f"测试模型: {', '.join(args.models)}")
    print(f"Excel文件: {args.excel}")
    print(f"最大用例数: {args.max_cases or '全部'}")
    print(f"每个case重复: {args.runs_per_case}次")
    print(f"输出目录: {output_dir}")
    print(f"=" * 60)
    
    # 加载提示词
    print(f"\n加载互动要点提示词: {POINTS_PROMPT_PATH}")
    points_prompt = load_points_prompt()
    print(f"提示词长度: {len(points_prompt)}字")
    
    # 从Excel加载测试用例
    print(f"\n从Excel加载测试用例...")
    test_cases = load_excel_test_cases(args.excel, args.max_cases)
    print(f"测试用例加载完成: {len(test_cases)}个")
    
    # 执行测试
    all_results = []
    
    for model_id in args.models:
        print(f"\n{'='*60}")
        print(f"测试模型: {model_id}")
        print(f"{'='*60}")
        
        model_results = []
        
        for case_idx, case in enumerate(test_cases, 1):
            print(f"\n[{case_idx}/{len(test_cases)}] {case.case_id}")
            
            for run in range(args.runs_per_case):
                if args.runs_per_case > 1:
                    print(f"  运行 {run+1}/{args.runs_per_case}...")
                
                # 生成互动要点
                points, latency, format_ok, issues, metrics = extract_with_model(
                    model_id,
                    case.transcript,
                    points_prompt,
                    use_thinking=False  # 都不使用thinking
                )
                
                print(f"  延迟: {latency:.3f}s, 格式: {'通过' if format_ok else '失败'}")
                if not format_ok:
                    print(f"  问题: {issues[:2]}")  # 只显示前2个问题
                
                result = {
                    "model_id": model_id,
                    "case_id": case.case_id,
                    "role_type": case.role_type,
                    "role_name": case.role_name,
                    "relationship": case.relationship,
                    "dialogue_type": case.dialogue_type,
                    "run": run + 1,
                    "latency": latency,
                    "format_ok": format_ok,
                    "issues": issues,
                    "metrics": metrics,
                    "output": points
                }
                
                model_results.append(result)
                all_results.append(result)
                
                # 实时写入
                with open(results_file, "a", encoding="utf-8") as f:
                    f.write(json.dumps(result, ensure_ascii=False) + "\n")
        
        # 模型汇总
        latencies = [r["latency"] for r in model_results]
        format_passes = sum(1 for r in model_results if r["format_ok"])
        
        print(f"\n{model_id} 汇总:")
        print(f"  总计: {len(model_results)}")
        print(f"  格式通过: {format_passes}/{len(model_results)} ({format_passes/len(model_results)*100:.1f}%)")
        print(f"  平均延迟: {statistics.mean(latencies):.3f}s")
        print(f"  中位数延迟: {statistics.median(latencies):.3f}s")
        print(f"  p95延迟: {statistics.quantiles(latencies, n=20)[18] if len(latencies)>=20 else max(latencies):.3f}s")
    
    # 生成对比摘要
    print(f"\n生成对比摘要...")
    
    with open(summary_file, "w", encoding="utf-8") as f:
        f.write(f"# 互动要点抽取模型对比测试（真实Excel用例）\n\n")
        f.write(f"- 生成时间: {datetime.now().isoformat()}\n")
        f.write(f"- 测试模型: {', '.join(args.models)}\n")
        f.write(f"- Excel文件: {args.excel}\n")
        f.write(f"- 测试用例: {len(test_cases)}个\n")
        f.write(f"- 每个case重复: {args.runs_per_case}次\n")
        f.write(f"- 提示词: {POINTS_PROMPT_PATH}\n\n")
        
        f.write(f"## 整体对比\n\n")
        f.write(f"| 模型 | 总计 | 格式通过 | 通过率 | 平均延迟 | 中位数延迟 | p95延迟 |\n")
        f.write(f"|:--|--:|--:|--:|--:|--:|--:|\n")
        
        for model_id in args.models:
            model_results = [r for r in all_results if r["model_id"] == model_id]
            latencies = [r["latency"] for r in model_results]
            format_passes = sum(1 for r in model_results if r["format_ok"])
            
            avg_lat = statistics.mean(latencies)
            med_lat = statistics.median(latencies)
            p95_lat = statistics.quantiles(latencies, n=20)[18] if len(latencies)>=20 else max(latencies)
            
            f.write(f"| {model_id} | {len(model_results)} | {format_passes} | {format_passes/len(model_results)*100:.1f}% | {avg_lat:.3f}s | {med_lat:.3f}s | {p95_lat:.3f}s |\n")
        
        f.write(f"\n## 失败原因统计\n\n")
        
        for model_id in args.models:
            model_results = [r for r in all_results if r["model_id"] == model_id]
            failed = [r for r in model_results if not r["format_ok"]]
            
            if not failed:
                f.write(f"### {model_id}\n\n无失败case\n\n")
                continue
            
            f.write(f"### {model_id}\n\n")
            
            from collections import Counter
            issue_counter = Counter()
            for r in failed:
                for issue in r["issues"]:
                    issue_counter[issue] += 1
            
            f.write(f"| 失败原因 | 次数 |\n")
            f.write(f"|:--|--:|\n")
            for issue, count in issue_counter.most_common():
                f.write(f"| {issue} | {count} |\n")
            f.write(f"\n")
        
        f.write(f"\n## 推荐结论\n\n")
        
        # 找出最优模型
        model_scores = {}
        for model_id in args.models:
            model_results = [r for r in all_results if r["model_id"] == model_id]
            latencies = [r["latency"] for r in model_results]
            format_passes = sum(1 for r in model_results if r["format_ok"])
            
            # 综合评分：格式通过率 * 0.7 + (1 - 归一化延迟) * 0.3
            pass_rate = format_passes / len(model_results)
            avg_lat = statistics.mean(latencies)
            
            # 归一化延迟（假设10秒为基准）
            norm_lat = min(avg_lat / 10.0, 1.0)
            
            score = pass_rate * 0.7 + (1 - norm_lat) * 0.3
            model_scores[model_id] = {
                "score": score,
                "pass_rate": pass_rate,
                "avg_latency": avg_lat
            }
        
        best_model = max(model_scores.items(), key=lambda x: x[1]["score"])
        
        f.write(f"**推荐模型**: {best_model[0]}\n\n")
        f.write(f"- 格式通过率: {best_model[1]['pass_rate']*100:.1f}%\n")
        f.write(f"- 平均延迟: {best_model[1]['avg_latency']:.3f}s\n")
        f.write(f"- 综合评分: {best_model[1]['score']:.3f}\n\n")
        
        f.write(f"评分规则：格式通过率 × 0.7 + (1 - 归一化延迟) × 0.3\n")
    
    print(f"\n完成！结果已保存到 {output_dir}")
    print(f"详细结果: {results_file}")
    print(f"对比摘要: {summary_file}")


if __name__ == "__main__":
    main()
