#!/usr/bin/env python3
"""
短文模式模型切换批量测试脚本。

核心口径：
- 候选主生成模型按 2 角色 x 3 关系阶段 x 20 轮执行。
- 默认包含基线组用于给候选评分提供 baseline 输出，也支持候选-only 生成。
- 组内串行，组间用 asyncio.Semaphore 控制并发。
- API Key 仅从环境变量读取，不写入输出。
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import re
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
try:
    from dotenv import load_dotenv

    load_dotenv(PROJECT_ROOT / "server" / ".env")
except Exception:
    pass

DASHSCOPE_BASE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"
ARK_BASE_URL = "https://ark.cn-beijing.volces.com/api/v3"
OUTPUT_DIR = PROJECT_ROOT / "output" / "shortform_model_switch"

RELATIONSHIPS = ("熟人", "暧昧", "恋人")
PERSONAL_TYPE = "温暖陪伴"
CONTINUATION_MARKER = "用户侧暂无新交互"

WEIGHTS = {
    "role_consistency": 0.25,
    "naturalness": 0.25,
    "stage_boundary": 0.20,
    "context_continuity": 0.20,
    "safety": 0.10,
}

VARIABLE_NAMES = [
    "完整时间信息",
    "voice_forbidden",
    "last_cst_type",
    "relationship",
    "relation_info",
    "weekly_schedule",
    "monthly_schedule",
    "Role_Nickname",
    "age",
    "occupation",
    "background",
    "Role_info_works",
    "user_Nickname",
    "call_name",
    "Tacall_name",
    "system_module3",
    "system_module7",
    "personality",
    "hobby",
    "system_module9",
    "speaking_style",
    "system_module11",
    "moments",
    "dialogueStartPrompt",
]

DEFAULT_SYSTEM_TEMPLATE = """# 当前时间
- 现在时间是{{完整时间信息}}
- 记住当前的时间，并遵循这个季节的气温感知输出回复。
{{voice_forbidden}}

# 对话场景
你正在与用户文本聊天
- 你与用户{{last_cst_type}}

# 你们的关系
- {{relationship}}，{{relation_info}}

# 你正在做的事情与聊天话题
- 你正在做的事情：{{weekly_schedule}}

# 核心生成要求
- 输出 30-90 个中文字符。
- 动作或旁白必须用中文全角括号（）包裹。
- 不得出现“指尖”。
- 回复自然、口语化，避免格式污染。

# 回复内容限制
- 角色名字{{Role_Nickname}}
- 用户大名{{user_Nickname}}，称呼{{Tacall_name}}
- 使用{{Tacall_name}}称呼用户

# 对话表达风格
{{system_module3}}

# 身份设定
- 角色为{{Role_Nickname}}，年龄{{age}}，职业{{occupation}}
- 近期行动：{{monthly_schedule}}
{{background}}
- 已拍摄作品：{{Role_info_works}}

# 性格特征
{{system_module7}}
{{personality}}

# 次要偏好
{{hobby}}

# 语言风格
{{system_module9}}
{{speaking_style}}

# 当前关系阶段
{{system_module11}}

# 用户朋友圈记忆模块
{{moments}}

{{dialogueStartPrompt}}
"""

DEFAULT_ASSISTANT_SEED = (
    "角色内部认知记录：当前为文字聊天场景，保持温暖陪伴型表达，"
    "承接上下文，不跳出角色，不解释规则。"
)
DRY_RUN_RESPONSE = (
    "（我把手里的杯子放低了一点，语气放软）嗯，我在听，你慢慢说，"
    "别急着自己扛着。"
)

EXCEL_KEY_PREFIX = "@全局用户参数_"
EXCEL_KEY_SUFFIX = "&"
EXCEL_KEY_ALIASES = {
    "weekly_schedul": "weekly_schedule",
    "monthly_schedul": "monthly_schedule",
}
OPTIONAL_VARIABLE_NAMES = {
    "voice_forbidden",
    "weekly_schedule",
    "monthly_schedule",
    "Role_info_works",
    "call_name",
    "Tacall_name",
    "system_module7",
    "hobby",
    "system_module9",
    "moments",
}

DEFAULT_RELATION_INFO = {
    "熟人": "刚认识不久的朋友，保持自然礼貌和轻微熟悉感",
    "暧昧": "互有好感但尚未确认关系，允许轻微试探与含蓄关心",
    "恋人": "已经确认恋爱关系，可以表达稳定亲近和日常陪伴",
}

DEFAULT_USER_MESSAGES = {
    "熟人": [
        "在忙吗？",
        "今天过得怎么样？",
        "这周有什么安排？",
        "新剧拍得顺利吗？",
        "我刚吃完晚饭，有点撑。",
        "你平时放松会做什么？",
        "有机会一起看电影吧。",
        "你相信缘分吗？",
        "今天遇到一件有趣的事。",
        "最近天气变冷了。",
        "你对未来有什么规划？",
        "下次见面想去海边。",
        "你会不会偶尔想起朋友？",
        "我今天工作有点累。",
        "你觉得我们算熟了吗？",
        "今晚好像很安静。",
        "如果明天休息你想做什么？",
        "我刚看到一个很好笑的视频。",
        "你现在方便聊两句吗？",
        "那你早点休息。",
    ],
    "暧昧": [
        "在忙吗？",
        "今天有没有想起我？",
        "这周有什么安排？",
        "新剧拍得顺利吗？",
        "我刚吃完晚饭，有点想找你说话。",
        "你平时会怎么哄人开心？",
        "有机会一起看电影吧。",
        "你相信缘分吗？",
        "今天遇到一件有趣的事，想讲给你听。",
        "最近天气变冷了，你注意点。",
        "你对未来有什么规划？",
        "下次见面想去海边。",
        "你觉得我们是什么关系？",
        "我今天工作有点累。",
        "如果我说想见你，会不会太突然？",
        "今晚好像很安静。",
        "如果明天休息你想做什么？",
        "我刚看到一个很好笑的视频。",
        "你现在方便陪我聊两句吗？",
        "那你早点休息，别太累。",
    ],
    "恋人": [
        "在忙吗？",
        "今天有没有想我？",
        "这周有什么安排？",
        "新剧拍得顺利吗？",
        "我刚吃完晚饭，有点想你。",
        "你平时累了会不会想抱一下？",
        "有机会一起看电影吧。",
        "你相信我们会一直好吗？",
        "今天遇到一件有趣的事，想第一个讲给你听。",
        "最近天气变冷了，你多穿点。",
        "你对我们的未来有什么规划？",
        "下次见面想去海边。",
        "你会不会突然很想见我？",
        "我今天工作有点累。",
        "如果我现在说想抱你呢？",
        "今晚好像很安静。",
        "如果明天休息你想和我做什么？",
        "我刚看到一个很好笑的视频。",
        "你现在方便陪我聊两句吗？",
        "那你早点休息，我会想你的。",
    ],
}

DEFAULT_CANDIDATES = [
    {
        "plan": "Plan A",
        "name": "deepseek-v4-flash",
        "model": "deepseek-v4-flash",
        "base_url": DASHSCOPE_BASE_URL,
        "api_key_env": "DASHSCOPE_API_KEY",
        "input_price": 1.0,
        "output_price": 2.0,
        "parameters": {},
    },
    {
        "plan": "Plan B",
        "name": "deepseek-v3.2",
        "model": "deepseek-v3.2",
        "base_url": DASHSCOPE_BASE_URL,
        "api_key_env": "DASHSCOPE_API_KEY",
        "input_price": 2.0,
        "output_price": 3.0,
        "parameters": {},
    },
    {
        "plan": "Plan C",
        "name": "doubao-seed-character",
        "model": "doubao-seed-character-251128",
        "base_url": ARK_BASE_URL,
        "api_key_env": "VOLCENGINE_API_KEY",
        "input_price": 0.8,
        "output_price": 2.0,
        "parameters": {},
    },
    {
        "plan": "Plan D",
        "name": "character-250715",
        "model": "doubao-1-5-pro-32k-character-250715",
        "base_url": ARK_BASE_URL,
        "api_key_env": "VOLCENGINE_API_KEY",
        "input_price": 0.6,
        "output_price": 1.5,
        "parameters": {},
    },
]

DEFAULT_BASELINE = {
    "plan": "Baseline",
    "name": "deepseek-v3-1-terminus",
    "model": "deepseek-v3-1-terminus",
    "base_url": ARK_BASE_URL,
    "api_key_env": "VOLCENGINE_API_KEY",
    "input_price": 4.0,
    "output_price": 12.0,
    "parameters": {},
}

DEFAULT_SCORER = {
    "name": "qwen3.6-plus",
    "model": "qwen3.6-plus",
    "base_url": DASHSCOPE_BASE_URL,
    "api_key_env": "DASHSCOPE_API_KEY",
    "parameters": {},
}


@dataclass(frozen=True)
class ModelSpec:
    plan: str
    name: str
    model: str
    base_url: str
    api_key_env: str
    input_price: float = 0.0
    output_price: float = 0.0
    parameters: dict[str, Any] = field(default_factory=dict)

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "ModelSpec":
        return cls(
            plan=str(data.get("plan", "")),
            name=str(data.get("name", data.get("model", ""))),
            model=str(data.get("model", data.get("model_name", ""))),
            base_url=str(data.get("base_url", DASHSCOPE_BASE_URL)),
            api_key_env=str(data.get("api_key_env", "DASHSCOPE_API_KEY")),
            input_price=float(data.get("input_price", 0.0) or 0.0),
            output_price=float(data.get("output_price", 0.0) or 0.0),
            parameters=dict(data.get("parameters", {}) or {}),
        )


@dataclass(frozen=True)
class RoleCase:
    role_type: str
    variables: dict[str, Any]


@dataclass(frozen=True)
class GroupSpec:
    group_id: str
    model: ModelSpec
    role: RoleCase
    relationship: str
    turns: list[str]
    is_baseline: bool = False


@dataclass
class CallResult:
    content: str = ""
    input_tokens: int = 0
    output_tokens: int = 0
    latency_s: float = 0.0
    error: str = ""
    retry_count: int = 0


def render_template(template: str, variables: dict[str, Any]) -> str:
    result = template
    for key, value in variables.items():
        text = "" if value is None else str(value)
        result = result.replace("{{" + key + "}}", text)
        result = result.replace("{{ " + key + " }}", text)
        result = result.replace("{" + key + "}", text)
    return result


def chinese_time_info(now: datetime | None = None) -> str:
    current = now or datetime.now()
    weekday = "一二三四五六日"[current.weekday()]
    hour = current.hour
    if 5 <= hour < 11:
        period = "上午"
    elif 11 <= hour < 14:
        period = "中午"
    elif 14 <= hour < 18:
        period = "下午"
    else:
        period = "晚上"
    return current.strftime(f"%Y-%m-%d %H时%M分 星期{weekday} {period}")


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def clean_cell(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    return "" if text == "/" else text


def normalize_excel_key(raw_key: Any) -> str:
    key = clean_cell(raw_key)
    if key.startswith(EXCEL_KEY_PREFIX) and key.endswith(EXCEL_KEY_SUFFIX):
        key = key[len(EXCEL_KEY_PREFIX) : -len(EXCEL_KEY_SUFFIX)]
    return EXCEL_KEY_ALIASES.get(key, key)


def clean_excel_variable(name: str, value: Any) -> str:
    return clean_cell(value)


def extract_user_turns(example_text: str) -> list[str]:
    turns: list[str] = []
    current_role: str | None = None
    buffer: list[str] = []
    for raw_line in str(example_text or "").splitlines():
        line = raw_line.strip()
        if line in {"用户", "AI"}:
            if current_role == "用户":
                content = "\n".join(buffer).strip()
                if content:
                    turns.append(content)
            current_role = line
            buffer = []
            continue
        if current_role and line:
            buffer.append(line)
    if current_role == "用户":
        content = "\n".join(buffer).strip()
        if content:
            turns.append(content)
    return turns


def load_excel_cases(path: Path) -> tuple[list[RoleCase], dict[str, list[str]]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [
        clean_cell(worksheet.cell(1, column).value)
        for column in range(1, worksheet.max_column + 1)
    ]
    key_to_row = {
        clean_cell(worksheet.cell(row, 1).value): row
        for row in range(2, worksheet.max_row + 1)
    }
    roles: list[RoleCase] = []
    user_messages: dict[str, list[str]] = {
        relationship: [] for relationship in RELATIONSHIPS
    }

    for column in range(2, worksheet.max_column + 1):
        role_label = headers[column - 1]
        if not role_label:
            continue
        variables: dict[str, Any] = {}
        for row in range(2, worksheet.max_row + 1):
            key = normalize_excel_key(worksheet.cell(row, 1).value)
            if not key or key in {"短文对话示例", "长文对话示例"}:
                continue
            variables[key] = clean_excel_variable(
                key,
                worksheet.cell(row, column).value,
            )
        variables.setdefault("Tacall_name", variables.get("call_name", ""))
        roles.append(RoleCase(role_type=role_label, variables=variables))

        short_row = key_to_row.get("短文对话示例")
        example_value = worksheet.cell(short_row, column).value if short_row else ""
        turns = extract_user_turns(example_value)
        relationship = str(variables.get("relationship") or "").strip()
        if relationship in user_messages and turns:
            user_messages[relationship].extend(turns)

    for relationship in RELATIONSHIPS:
        if not user_messages[relationship]:
            user_messages[relationship] = list(DEFAULT_USER_MESSAGES[relationship])
    return roles, user_messages


def default_role(role_type: str, name: str) -> RoleCase:
    return RoleCase(
        role_type=role_type,
        variables={
            "Role_Nickname": name,
            "age": "28",
            "occupation": "演员",
            "background": "长期在剧组工作，行程紧但会认真回应重要的人。",
            "Role_info_works": "",
            "user_Nickname": "用户",
            "call_name": "你",
            "Tacall_name": "你",
            "weekly_schedule": "这周主要在剧组拍摄，晚上有少量休息时间",
            "monthly_schedule": "本月以拍摄和宣传为主，中间有两天休整",
            "system_module3": "表达克制但有温度，少用夸张词，避免油腻。",
            "system_module7": "",
            "personality": "温暖陪伴型，可靠、细心，习惯用具体行动表达关心。",
            "hobby": "",
            "system_module9": "",
            "speaking_style": "短句为主，自然口语化，偶尔带轻微玩笑。",
            "moments": "",
            "dialogueStartPrompt": "<dialogue_history>暂无历史对话</dialogue_history>",
        },
    )


def load_cases(
    config: dict[str, Any],
) -> tuple[list[RoleCase], dict[str, list[str]], str, str]:
    if config.get("case_xlsx"):
        roles, user_messages = load_excel_cases(Path(config["case_xlsx"]))
        template = str(config.get("system_template") or DEFAULT_SYSTEM_TEMPLATE)
        assistant_seed = str(config.get("assistant_seed") or DEFAULT_ASSISTANT_SEED)
        return roles, user_messages, template, assistant_seed

    role_items = config.get("roles") or [
        {
            "role_type": "官方角色",
            "variables": default_role("官方角色", "肖战").variables,
        },
        {
            "role_type": "自定义角色",
            "variables": default_role("自定义角色", "顾风").variables,
        },
    ]
    roles = [
        RoleCase(str(item.get("role_type", "")), dict(item.get("variables", {}) or {}))
        for item in role_items
    ]
    user_messages = {
        rel: list(
            (config.get("user_messages", {}) or {}).get(rel)
            or DEFAULT_USER_MESSAGES[rel]
        )
        for rel in RELATIONSHIPS
    }
    template = str(config.get("system_template") or DEFAULT_SYSTEM_TEMPLATE)
    assistant_seed = str(config.get("assistant_seed") or DEFAULT_ASSISTANT_SEED)
    return roles, user_messages, template, assistant_seed


def normalize_turns(raw_turns: list[str], turn_count: int) -> list[str]:
    if not raw_turns:
        raise ValueError("用户消息库为空")
    turns = [str(item) for item in raw_turns]
    while len(turns) < turn_count:
        turns.extend(str(item) for item in raw_turns)
    turns = turns[:turn_count]
    if CONTINUATION_MARKER not in turns:
        turns[min(9, turn_count - 1)] = CONTINUATION_MARKER
        if turn_count >= 16:
            turns[15] = CONTINUATION_MARKER
    return turns


def build_variables(role: RoleCase, relationship: str) -> dict[str, Any]:
    variables = dict(role.variables)
    variables.setdefault("完整时间信息", chinese_time_info())
    variables.setdefault("voice_forbidden", "")
    variables.setdefault("last_cst_type", "上一次在文字聊天沟通")
    variables["relationship"] = relationship
    variables.setdefault("relation_info", DEFAULT_RELATION_INFO[relationship])
    variables.setdefault(
        "system_module11",
        f"personal_type={PERSONAL_TYPE}；当前关系阶段={relationship}；行为表达必须匹配关系边界。",
    )
    variables.setdefault("call_name", variables.get("Tacall_name", ""))
    variables.setdefault("Tacall_name", variables.get("call_name", ""))
    for name in VARIABLE_NAMES:
        variables.setdefault(name, "")
    return variables


def validate_variables(variables: dict[str, Any], relationship: str) -> list[str]:
    errors: list[str] = []
    for name in VARIABLE_NAMES:
        if name not in variables:
            errors.append(f"{name} 字段缺失")

    required = [
        "完整时间信息",
        "last_cst_type",
        "relationship",
        "relation_info",
        "Role_Nickname",
        "age",
        "occupation",
        "background",
        "user_Nickname",
        "Tacall_name",
        "system_module3",
        "personality",
        "speaking_style",
        "system_module11",
        "dialogueStartPrompt",
    ]
    required = [name for name in required if name not in OPTIONAL_VARIABLE_NAMES]
    for name in required:
        if not str(variables.get(name, "")).strip():
            errors.append(f"{name} 不能为空")

    last_cst_type = str(variables.get("last_cst_type", ""))
    if "电话聊天沟通" not in last_cst_type and "文字聊天沟通" not in last_cst_type:
        errors.append("last_cst_type 值非法")
    if variables.get("relationship") not in set(RELATIONSHIPS):
        errors.append("relationship 值非法")
    if not str(variables.get("age", "")).isdigit():
        errors.append("age 必须为数字")
    module11 = str(variables.get("system_module11", ""))
    if relationship not in module11:
        errors.append("system_module11 未包含关系阶段")
    return errors


def build_messages(
    system_template: str,
    variables: dict[str, Any],
    assistant_seed: str,
    history: list[dict[str, str]],
    user_message: str,
) -> list[dict[str, str]]:
    system_prompt = render_template(system_template, variables)
    return [
        {"role": "system", "content": system_prompt},
        {"role": "assistant", "content": render_template(assistant_seed, variables)},
        *history,
        {"role": "user", "content": user_message},
    ]


def response_checks(text: str) -> dict[str, Any]:
    chinese_chars = re.findall(r"[\u4e00-\u9fff]", text)
    char_count = len(chinese_chars)
    has_fullwidth_action = bool(re.search(r"（[^（）]+）", text))
    balanced_fullwidth = text.count("（") == text.count("）")
    has_ascii_paren = "(" in text or ")" in text
    banned_word = "指尖" in text
    narrative_chars = sum(len(match) for match in re.findall(r"（([^（）]+)）", text))
    ratio = round(narrative_chars / max(len(text), 1), 4)
    bracket_violation = (
        (not has_fullwidth_action) or (not balanced_fullwidth) or has_ascii_paren
    )
    return {
        "char_count": char_count,
        "word_count_violation": char_count < 30 or char_count > 90,
        "bracket_violation": bracket_violation,
        "banned_word_violation": banned_word,
        "narrative_ratio": ratio,
        "ratio_warning": abs(ratio - 0.5) > 0.2,
        "format_violation": (char_count < 30 or char_count > 90)
        or (not has_fullwidth_action)
        or (not balanced_fullwidth)
        or has_ascii_paren
        or banned_word,
    }


def cost(input_tokens: int, output_tokens: int, model: ModelSpec) -> float:
    return (
        input_tokens * model.input_price + output_tokens * model.output_price
    ) / 1_000_000


def create_client(model: ModelSpec):
    from openai import OpenAI

    api_key = os.environ.get(model.api_key_env, "").strip()
    if not api_key:
        raise RuntimeError(f"缺少环境变量 {model.api_key_env}")
    return OpenAI(api_key=api_key, base_url=model.base_url)


def call_once(model: ModelSpec, messages: list[dict[str, str]]) -> CallResult:
    client = create_client(model)
    started = time.time()
    response = client.chat.completions.create(
        model=model.model,
        messages=messages,
        **model.parameters,
    )
    latency = round(time.time() - started, 3)
    content = response.choices[0].message.content or ""
    content = re.sub(
        r"(?is)<(?:think|thought)>.*?</(?:think|thought)>",
        "",
        content,
    ).strip()
    usage = getattr(response, "usage", None)
    return CallResult(
        content=content,
        input_tokens=int(getattr(usage, "prompt_tokens", 0) or 0),
        output_tokens=int(getattr(usage, "completion_tokens", 0) or 0),
        latency_s=latency,
    )


async def call_with_retry(
    model: ModelSpec,
    messages: list[dict[str, str]],
    retries: int,
    retry_delay: float,
) -> CallResult:
    last_error = ""
    for attempt in range(retries + 1):
        try:
            result = await asyncio.to_thread(call_once, model, messages)
            result.retry_count = attempt
            return result
        except Exception as exc:  # noqa: BLE001 - 需要记录外部 API 失败
            last_error = str(exc)
            if attempt < retries:
                await asyncio.sleep(retry_delay * (attempt + 1))
    return CallResult(error=last_error, retry_count=retries)


async def run_group(
    group: GroupSpec,
    system_template: str,
    assistant_seed: str,
    semaphore: asyncio.Semaphore,
    retries: int,
    retry_delay: float,
    dry_run: bool,
) -> list[dict[str, Any]]:
    async with semaphore:
        variables = build_variables(group.role, group.relationship)
        validation_errors = validate_variables(variables, group.relationship)
        history: list[dict[str, str]] = []
        records: list[dict[str, Any]] = []

        if validation_errors:
            raise ValueError(
                f"{group.group_id} dry-run 变量校验失败: {validation_errors}"
            )

        for turn_index, user_message in enumerate(group.turns, start=1):
            messages = build_messages(
                system_template=system_template,
                variables=variables,
                assistant_seed=assistant_seed,
                history=history,
                user_message=user_message,
            )
            if dry_run:
                result = CallResult(
                    content=DRY_RUN_RESPONSE,
                    input_tokens=sum(len(item["content"]) for item in messages) // 2,
                    output_tokens=28,
                    latency_s=0.0,
                )
            else:
                result = await call_with_retry(
                    group.model,
                    messages,
                    retries,
                    retry_delay,
                )

            checks = response_checks(result.content)
            record = {
                "group_id": group.group_id,
                "is_baseline": group.is_baseline,
                "plan": group.model.plan,
                "model_name": group.model.name,
                "model": group.model.model,
                "role_type": group.role.role_type,
                "role_name": variables.get("Role_Nickname", ""),
                "relationship": group.relationship,
                "turn": turn_index,
                "user_input": user_message,
                "assistant_output": result.content,
                "input_tokens": result.input_tokens,
                "output_tokens": result.output_tokens,
                "cost": cost(result.input_tokens, result.output_tokens, group.model),
                "latency_s": result.latency_s,
                "error": result.error,
                "retry_count": result.retry_count,
                **{f"var_{name}": variables.get(name, "") for name in VARIABLE_NAMES},
                **checks,
            }
            records.append(record)

            if result.error:
                break

            history.append({"role": "user", "content": user_message})
            history.append({"role": "assistant", "content": result.content})

        return records


def build_groups(
    roles: list[RoleCase],
    messages_by_relationship: dict[str, list[str]],
    baseline: ModelSpec | None,
    candidates: list[ModelSpec],
    turn_count: int,
) -> list[GroupSpec]:
    groups: list[GroupSpec] = []
    for role in roles:
        role_relationship = str(role.variables.get("relationship", "")).strip()
        relationships = (
            (role_relationship,)
            if role_relationship in RELATIONSHIPS
            else RELATIONSHIPS
        )
        for relationship in relationships:
            turns = normalize_turns(messages_by_relationship[relationship], turn_count)
            if baseline is not None:
                groups.append(
                    GroupSpec(
                        group_id=f"baseline::{role.role_type}::{relationship}",
                        model=baseline,
                        role=role,
                        relationship=relationship,
                        turns=turns,
                        is_baseline=True,
                    )
                )
            for candidate in candidates:
                groups.append(
                    GroupSpec(
                        group_id=f"{candidate.plan}::{role.role_type}::{relationship}",
                        model=candidate,
                        role=role,
                        relationship=relationship,
                        turns=turns,
                    )
                )
    return groups


def score_prompt(candidate: dict[str, Any], baseline: dict[str, Any]) -> str:
    return f"""你是短文模式模型切换评估员。请分别给基线回复和候选回复打分。

只输出 JSON，不要输出 Markdown。
输出结构：
{{
  "baseline": {{
    "role_consistency": 0-100,
    "naturalness": 0-100,
    "stage_boundary": 0-100,
    "context_continuity": 0-100,
    "safety": 0-100,
    "total": 0-100,
    "reason": "简短中文理由"
  }},
  "candidate": {{
    "role_consistency": 0-100,
    "naturalness": 0-100,
    "stage_boundary": 0-100,
    "context_continuity": 0-100,
    "safety": 0-100,
    "total": 0-100,
    "reason": "简短中文理由"
  }},
  "comparison_reason": "候选相对基线的简短结论"
}}

关系阶段：{candidate["relationship"]}
角色：{candidate["role_name"]}
用户输入：{candidate["user_input"]}

基线回复：
{baseline["assistant_output"]}

候选回复：
{candidate["assistant_output"]}
"""


def _parse_score_object(data: dict[str, Any], prefix: str) -> dict[str, Any]:
    parsed: dict[str, Any] = {}
    for key in WEIGHTS:
        parsed[f"{prefix}_{key}"] = float(data.get(key, 0) or 0)
    weighted = sum(
        parsed[f"{prefix}_{key}"] * weight for key, weight in WEIGHTS.items()
    )
    parsed[f"{prefix}_total"] = float(data.get("total", weighted) or weighted)
    parsed[f"{prefix}_weighted_total"] = round(weighted, 3)
    parsed[f"{prefix}_reason"] = str(data.get("reason", ""))
    return parsed


def parse_score(content: str) -> dict[str, Any]:
    match = re.search(r"\{.*\}", content, flags=re.S)
    if not match:
        return {"score_parse_error": content[:200]}
    try:
        data = json.loads(match.group(0))
    except json.JSONDecodeError:
        return {"score_parse_error": content[:200]}

    baseline_data = (
        data.get("baseline") if isinstance(data.get("baseline"), dict) else {}
    )
    candidate_data = (
        data.get("candidate") if isinstance(data.get("candidate"), dict) else data
    )
    parsed = _parse_score_object(candidate_data, "score")
    if baseline_data:
        parsed.update(_parse_score_object(baseline_data, "baseline_score"))
    parsed["comparison_reason"] = str(data.get("comparison_reason", ""))
    return parsed


async def score_records(
    records: list[dict[str, Any]],
    scorer: ModelSpec,
    retries: int,
    retry_delay: float,
    dry_run: bool,
) -> None:
    baseline_by_key = {
        (row["role_type"], row["relationship"], row["turn"]): row
        for row in records
        if row["is_baseline"] and not row["error"]
    }
    for row in records:
        if row["is_baseline"] or row["error"]:
            continue
        baseline = baseline_by_key.get(
            (row["role_type"], row["relationship"], row["turn"])
        )
        if not baseline:
            row["score_error"] = "missing_baseline"
            continue
        if dry_run:
            row.update(
                {
                    "score_role_consistency": 90,
                    "score_naturalness": 90,
                    "score_stage_boundary": 90,
                    "score_context_continuity": 90,
                    "score_safety": 95,
                    "score_total": 91,
                    "score_weighted_total": 90.5,
                    "score_reason": "dry-run mock score",
                    "baseline_score_role_consistency": 90,
                    "baseline_score_naturalness": 90,
                    "baseline_score_stage_boundary": 90,
                    "baseline_score_context_continuity": 90,
                    "baseline_score_safety": 95,
                    "baseline_score_total": 91,
                    "baseline_score_weighted_total": 90.5,
                    "baseline_score_reason": "dry-run mock baseline score",
                    "comparison_reason": "dry-run mock comparison",
                }
            )
            continue
        messages = [{"role": "user", "content": score_prompt(row, baseline)}]
        result = await call_with_retry(scorer, messages, retries, retry_delay)
        if result.error:
            row["score_error"] = result.error
            row["score_retry_count"] = result.retry_count
            continue
        row.update(parse_score(result.content))
        row["score_retry_count"] = result.retry_count


def write_excel(records: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    detail = workbook.active
    detail.title = "detail"

    columns = [
        "group_id",
        "is_baseline",
        "plan",
        "model_name",
        "model",
        "role_type",
        "role_name",
        "relationship",
        "turn",
        "user_input",
        "assistant_output",
        "input_tokens",
        "output_tokens",
        "cost",
        "latency_s",
        "error",
        "retry_count",
        "char_count",
        "word_count_violation",
        "bracket_violation",
        "banned_word_violation",
        "narrative_ratio",
        "ratio_warning",
        "format_violation",
        "score_role_consistency",
        "score_naturalness",
        "score_stage_boundary",
        "score_context_continuity",
        "score_safety",
        "score_total",
        "score_weighted_total",
        "score_reason",
        "baseline_score_role_consistency",
        "baseline_score_naturalness",
        "baseline_score_stage_boundary",
        "baseline_score_context_continuity",
        "baseline_score_safety",
        "baseline_score_total",
        "baseline_score_weighted_total",
        "baseline_score_reason",
        "comparison_reason",
        "score_parse_error",
        "score_error",
        *[f"var_{name}" for name in VARIABLE_NAMES],
    ]
    detail.append(columns)
    for row in records:
        detail.append([row.get(column, "") for column in columns])

    summary = workbook.create_sheet("summary")
    summary.append(
        ["model_name", "rows", "errors", "format_violations", "total_cost", "avg_score"]
    )
    model_names = sorted({str(row["model_name"]) for row in records})
    for model_name in model_names:
        rows = [row for row in records if row["model_name"] == model_name]
        score_values = [
            float(row["score_weighted_total"])
            for row in rows
            if row.get("score_weighted_total") not in (None, "")
        ]
        summary.append(
            [
                model_name,
                len(rows),
                sum(1 for row in rows if row.get("error")),
                sum(1 for row in rows if row.get("format_violation")),
                round(sum(float(row.get("cost") or 0) for row in rows), 6),
                round(sum(score_values) / len(score_values), 3) if score_values else "",
            ]
        )

    workbook.save(output_path)


def load_specs(config: dict[str, Any]) -> tuple[ModelSpec, list[ModelSpec], ModelSpec]:
    baseline = ModelSpec.from_dict(config.get("baseline") or DEFAULT_BASELINE)
    candidates = [
        ModelSpec.from_dict(item)
        for item in (config.get("candidates") or DEFAULT_CANDIDATES)
    ]
    scorer = ModelSpec.from_dict(config.get("scorer") or DEFAULT_SCORER)
    return baseline, candidates, scorer


async def async_main(args: argparse.Namespace) -> Path:
    config = load_json(Path(args.config)) if args.config else {}
    case_xlsx = getattr(args, "case_xlsx", "")
    if case_xlsx:
        config["case_xlsx"] = case_xlsx
    if not args.config and not case_xlsx and not args.dry_run:
        raise SystemExit("正式执行必须通过 --config 或 --case-xlsx 提供测试配置")

    (
        roles,
        messages_by_relationship,
        system_template,
        assistant_seed,
    ) = load_cases(config)
    baseline, candidates, scorer = load_specs(config)
    if getattr(args, "skip_baseline", False):
        baseline = None
    groups = build_groups(
        roles=roles,
        messages_by_relationship=messages_by_relationship,
        baseline=baseline,
        candidates=candidates,
        turn_count=args.turns,
    )
    semaphore = asyncio.Semaphore(args.workers)
    records_nested = await asyncio.gather(
        *[
            run_group(
                group=group,
                system_template=system_template,
                assistant_seed=assistant_seed,
                semaphore=semaphore,
                retries=args.retries,
                retry_delay=args.retry_delay,
                dry_run=args.dry_run,
            )
            for group in groups
        ]
    )
    records = [row for group_rows in records_nested for row in group_rows]
    if not args.no_score:
        await score_records(
            records=records,
            scorer=scorer,
            retries=args.retries,
            retry_delay=args.retry_delay,
            dry_run=args.dry_run,
        )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = (
        Path(args.output)
        if args.output
        else OUTPUT_DIR / f"shortform_model_switch_{timestamp}.xlsx"
    )
    write_excel(records, output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="短文模式模型切换批量测试")
    parser.add_argument("--config", help="测试配置 JSON 路径")
    parser.add_argument("--case-xlsx", help="变量转置表 Excel 路径")
    parser.add_argument("--output", help="Excel 输出路径")
    parser.add_argument("--workers", type=int, default=30, help="并发对话组数")
    parser.add_argument("--turns", type=int, default=20, help="每组轮数")
    parser.add_argument("--retries", type=int, default=3, help="失败后的重试次数")
    parser.add_argument(
        "--retry-delay",
        type=float,
        default=1.0,
        help="首次重试延迟秒数",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="只校验拼接和输出，不调用模型",
    )
    parser.add_argument(
        "--skip-baseline",
        action="store_true",
        help="只生成候选模型结果；需配合 --no-score 使用",
    )
    parser.add_argument("--no-score", action="store_true", help="跳过评分")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.skip_baseline and not args.no_score:
        raise SystemExit("--skip-baseline 需要配合 --no-score 使用")
    output_path = asyncio.run(async_main(args))
    print(f"[OK] Excel 已输出: {output_path}")


if __name__ == "__main__":
    main()
