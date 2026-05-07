from __future__ import annotations

import argparse
import importlib.util
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = PROJECT_ROOT / "scripts" / "shortform_model_switch_batch_test.py"

spec = importlib.util.spec_from_file_location(
    "shortform_model_switch_batch_test",
    SCRIPT_PATH,
)
shortform = importlib.util.module_from_spec(spec)
assert spec and spec.loader
sys.modules[spec.name] = shortform
spec.loader.exec_module(shortform)


def test_validate_variables_accepts_complete_warm_companion_case():
    role = shortform.default_role("官方角色", "肖战")
    variables = shortform.build_variables(role, "暧昧")

    assert shortform.validate_variables(variables, "暧昧") == []
    assert variables["relationship"] == "暧昧"
    assert "温暖陪伴" in variables["system_module11"]


def test_normalize_turns_inserts_continuation_marker():
    turns = shortform.normalize_turns(["你好", "今天忙吗"], 20)

    assert len(turns) == 20
    assert turns[9] == "用户侧暂无新交互"
    assert turns[15] == "用户侧暂无新交互"


def test_dry_run_writes_excel_with_one_row_per_turn(tmp_path):
    output = tmp_path / "shortform.xlsx"
    args = argparse.Namespace(
        config="",
        case_xlsx="",
        output=str(output),
        workers=30,
        turns=2,
        retries=3,
        retry_delay=0.0,
        dry_run=True,
        no_score=False,
    )

    result_path = shortform.asyncio.run(shortform.async_main(args))

    assert result_path == output
    workbook = load_workbook(output)
    assert set(workbook.sheetnames) == {"detail", "summary"}
    detail = workbook["detail"]
    headers = [cell.value for cell in detail[1]]
    rows = list(detail.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 60
    assert "assistant_output" in headers
    assert "score_weighted_total" in headers
    format_violation_idx = headers.index("format_violation")
    assert sum(1 for row in rows if row[format_violation_idx]) == 0


def test_load_excel_cases_maps_transposed_variable_table(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    rows = [
        ["变量代码", "角色1"],
        ["@全局用户参数_完整时间信息&", "现在时间是2026-05-07 10时 上午"],
        ["@全局用户参数_voice_forbidden&", "/"],
        ["@全局用户参数_last_cst_type&", "你与用户上一次在文字聊天沟通"],
        ["@全局用户参数_relationship&", "暧昧"],
        ["@全局用户参数_relation_info&", "朋友，并且在暧昧期"],
        ["@全局用户参数_Role_Nickname&", "肖战"],
        ["@全局用户参数_age&", "28"],
        ["@全局用户参数_occupation&", "歌手"],
        ["@全局用户参数_weekly_schedul&", "/"],
        ["@全局用户参数_monthly_schedul&", "本月拍摄和宣传"],
        ["@全局用户参数_background&", "演员背景"],
        ["@全局用户参数_personality&", "温暖陪伴"],
        ["@全局用户参数_speaking_style&", "自然口语"],
        ["@全局用户参数_user_Nickname&", "琴琴"],
        ["@全局用户参数_call_name&", "/"],
        [
            "@全局用户参数_dialogueStartPrompt&",
            "<dialogue_history>无</dialogue_history>",
        ],
        ["@全局用户参数_system_module3&", "表达风格"],
        ["@全局用户参数_system_module11&", "暧昧阶段规则"],
        ["短文对话示例", "用户\n在忙吗\n\nAI\n（看着你）在\n\n用户\n想你"],
    ]
    for row in rows:
        sheet.append(row)
    path = tmp_path / "cases.xlsx"
    workbook.save(path)

    roles, user_messages = shortform.load_excel_cases(path)

    assert len(roles) == 1
    variables = roles[0].variables
    assert variables["完整时间信息"] == "现在时间是2026-05-07 10时 上午"
    assert variables["last_cst_type"] == "你与用户上一次在文字聊天沟通"
    assert variables["weekly_schedule"] == ""
    assert variables["monthly_schedule"] == "本月拍摄和宣传"
    assert variables["call_name"] == ""
    assert user_messages["暧昧"] == ["在忙吗", "想你"]

    baseline = shortform.ModelSpec.from_dict(shortform.DEFAULT_BASELINE)
    candidates = [shortform.ModelSpec.from_dict(shortform.DEFAULT_CANDIDATES[0])]
    groups = shortform.build_groups(roles, user_messages, baseline, candidates, 2)
    assert {group.relationship for group in groups} == {"暧昧"}
    assert len(groups) == 2
