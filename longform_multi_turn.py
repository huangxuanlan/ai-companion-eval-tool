"""
长文模式多轮对话批量测试工具 v2.0

核心能力：
  1. 读取 v2.0 提示词模板，自动注入 {{variable}} 变量
  2. 按白皮书 v1.6 消息架构组装 messages 数组
  3. 每轮的 AI 回复自动拼接为下一轮的 conversation_history
  4. 每 5 轮暂停 → 调用 mini 模型生成 7 字段 dialogue_summary → 注入模板 → 继续生成

用法：
  python longform_multi_turn.py test_conversation_萧璟言.json
  python longform_multi_turn.py test_conversation_萧璟言.json --dry-run
  python longform_multi_turn.py test_conversation_萧璟言.json --turns 3
"""

import argparse
import builtins
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent
SERVER_DIR = PROJECT_ROOT / "server"
for _path in (str(PROJECT_ROOT), str(SERVER_DIR)):
    if _path not in sys.path:
        sys.path.insert(0, _path)

try:
    from server.services.quality_guard import QualityGuard
except ImportError:
    QualityGuard = None

try:
    from server.services.prompt_service import PromptService
except ImportError:
    PromptService = None

try:
    from server.services.runtime_config import normalize_longform_config_contract
except ImportError:
    normalize_longform_config_contract = None


def init_console_io():
    """尽量统一 stdout/stderr 编码，避免 Windows 控制台直接崩溃。"""
    for stream in (sys.stdout, sys.stderr):
        reconfigure = getattr(stream, "reconfigure", None)
        if callable(reconfigure):
            try:
                reconfigure(encoding="utf-8", errors="replace")
            except (LookupError, OSError, ValueError):
                continue


def _console_safe(text, stream=None) -> str:
    """Normalize console text for legacy Windows terminals."""
    normalized = (
        str(text)
        .replace("✓", "[OK]")
        .replace("✅", "[OK]")
        .replace("✗", "[FAIL]")
        .replace("❌", "[FAIL]")
        .replace("⚠", "[WARN]")
        .replace("🔵", "[INFO]")
    )
    target_stream = stream or sys.stdout
    encoding = getattr(target_stream, "encoding", None) or "utf-8"
    return normalized.encode(encoding, errors="replace").decode(
        encoding, errors="replace"
    )


def console_print(*args, **kwargs):
    """ASCII-safe print wrapper used across the CLI."""
    stream = kwargs.get("file", sys.stdout)
    safe_args = tuple(_console_safe(arg, stream=stream) for arg in args)
    if "sep" in kwargs:
        kwargs["sep"] = _console_safe(kwargs["sep"], stream=stream)
    if "end" in kwargs:
        kwargs["end"] = _console_safe(kwargs["end"], stream=stream)
    builtins.print(*safe_args, **kwargs)


def safe_preview(text: str, limit: int = 80) -> str:
    """控制台预览统一使用 ASCII 安全换行符。"""
    return text[:limit].replace("\n", "\\n")


def normalize_personal_type(personal_type: str) -> str:
    """统一性格类型命名，兼容“霸道腹黑”/“霸道腹黑型”两种写法。"""
    clean = re.sub(r"\s+", "", personal_type or "")
    return clean.replace("型", "")


init_console_io()
print = console_print

# ── 模型配置 ──────────────────────────────────────────────────
BASE_URL = os.environ.get("DOUBAO_BASE_URL", "https://ark.cn-beijing.volces.com/api/v3")
API_KEY = os.environ.get("DOUBAO_API_KEY", "")
MODEL_PRO = os.environ.get("DOUBAO_MODEL", "doubao-seed-2-0-pro-260215")
MODEL_MINI = os.environ.get("DOUBAO_MODEL_MINI", "doubao-seed-2-0-mini-260215")
TEMPERATURE = float(os.environ.get("DOUBAO_TEMPERATURE", "1.0"))
MAX_TOKENS = int(os.environ.get("DOUBAO_MAX_TOKENS", "4096"))
TOP_P = float(os.environ.get("DOUBAO_TOP_P", "0.95"))

# 摘要生成间隔（每 N 轮生成一次 dialogue_summary）
SUMMARY_INTERVAL = 5

SCRIPT_DIR = Path(__file__).resolve().parent
try:
    from server.config import (  # type: ignore
        PROMPT_DIR as SHARED_PROMPT_DIR,
        TEST_PROMPT_DIR as SHARED_TEST_PROMPT_DIR,
        VARIABLE_DIR as SHARED_VARIABLE_DIR,
    )
except Exception:
    SHARED_PROMPT_DIR = SCRIPT_DIR / "prompt"
    SHARED_TEST_PROMPT_DIR = SCRIPT_DIR / "prompt"
    SHARED_VARIABLE_DIR = SCRIPT_DIR / "few_shot"

DEFAULT_PROMPT_FILE = SCRIPT_DIR / "prompt" / "星朋友长文模式_提示词_v2.0.md"
DEFAULT_FEWSHOT_FILE = SCRIPT_DIR / "few_shot" / "长文模式_Few-shot示例库.md"
QUALITY_GUARD = QualityGuard() if QualityGuard else None
PROMPT_SERVICE = PromptService() if PromptService else None


def normalize_cli_config(config: dict) -> dict:
    """Keep CLI JSON/Excel configs aligned with the Web runtime contract."""
    if normalize_longform_config_contract is None:
        return config
    normalized = normalize_longform_config_contract(config)
    for key, value in config.items():
        if key not in normalized:
            normalized[key] = value
    if config.get("_mode"):
        normalized["_mode"] = config.get("_mode")
    if config.get("_session_id"):
        normalized["_session_id"] = config.get("_session_id")
    return normalized

# ── 固定消息模板（白皮书 v1.6 §3.6-3.10 + S1/S3/S5 首轮隔离修复）──────────
FEW_SHOT_PREFIX_MSG = (
    "【写作风格示例开始】以下一段对话仅用于展示文学写作风格和输出格式。"
    "示例中的场景、地点、道具、事件均为虚构，与当前角色、用户及对话没有任何关联。"
    "你从未经历过这段示例对话，请勿将其中任何细节引入本次叙事。"
    "【风格示例正文如下，请仅参考格式与句式】"
)

SEPARATOR_MSG = (
    "---风格示例结束---\n"
    "【你的任务】从现在开始，完全基于下方的真实角色设定和用户输入来生成回复。"
    "你的回复场景、人物、事件必须来自当前角色设定，而非上方示例。\n"
    "---以下才是真实的对话历史---"
)

# S3: 首轮哨兵消息——利用 Recency 位置优势明确告知模型这是首次对话
FIRST_TURN_SENTINEL = (
    "这是你与用户的【第一次对话】。"
    "你们之前从未有过任何交流。"
    "请根据你的角色设定和当前场景，自然地回应用户的第一句话。"
)

STYLE_ISOLATION_MSG = (
    "=== 历史对话记录 ===\n"
    "以下对话仅供参考剧情上下文。你的写作风格和角色表现"
    "必须严格遵循System Prompt定义，而非继承历史回复的风格。"
)

# v2.6 Core_Constraints（含去重+记忆，正面引导）
CORE_CONSTRAINTS_TEMPLATE = """<Core_Constraints>
- 长度：300-500字完整叙事
- 格式：旁白为纯文本不加包裹符号，对白用 **""** 包裹
- 结尾：以带情感张力的引导性钩子收束（动作悬念/情感暗示/反问诱导/开放留白）
- 人设：使用锚点词，保持角色风格
- 去重：每轮切换感官焦点、身体语言区域和核心意象，与上一轮形成变化
- 记忆：仅基于用户画像和当前输入生成内容，不虚构共同回忆，不引用 Few-shot 示例细节
- 关系阶段：{relationship}——肢体接触和情感表达匹配当前阶段
</Core_Constraints>"""

# ── 摘要生成 Prompt（给 mini 模型用）──────────────
SUMMARY_GENERATION_PROMPT = """你是一个对话分析助手。请根据以下多轮对话历史，生成结构化的剧情摘要。

要求：
1. 严格按以下 7 字段 JSON 格式输出
2. 每个字段尽量精简（总量控制在 300 tokens 以内）
3. 仅基于实际对话内容提取，不虚构

输出格式（仅输出 JSON，不要其他文字）:
{{
    "scene_description": "当前场景位置 + 1-2个感官锚点",
    "plot_summary": "关键事件因果链（≤3句）",
    "pending_hooks": "未兑现的承诺/未完成动作/悬念线索",
    "character_emotion": "角色当前情绪 + 触发原因",
    "user_emotion": "用户当前情绪 + 触发原因",
    "relationship_shift": "关系阶段 + 本轮微变化",
    "user_profile_signals": "可沉淀的用户行为模式"
}}

角色信息：
- 角色名：{role_name}
- 性格类型：{personal_type}
- 当前关系阶段：{relationship}

以下是最近的对话历史：
{conversation_text}

请输出 7 字段 JSON 摘要："""

# 摘要注入模板（白皮书 §9.5.3）
SUMMARY_INJECT_TEMPLATE = """=== 之前剧情摘要 ===
- 场景：{scene_description}
- 剧情：{plot_summary}
- 悬念：{pending_hooks}
- 角色情绪：{character_emotion}
- 用户情绪：{user_emotion}
- 关系动态：{relationship_shift}
- 用户画像信号：{user_profile_signals}
=== 摘要结束 ==="""


# ── 变量注入 ──────────────────────────────────────────────────

def render_template(template: str, variables: dict, clean_residual: bool = False) -> str:
    """
    将模板中的 {{variable_name}} 替换为 variables 字典中的值。
    clean_residual=True 时清除未匹配的 {{}} 残留（CLI独立运行推荐）。
    clean_residual=False 时保留原样（后端二次注入场景）。
    """
    def replacer(match):
        key = match.group(1).strip()
        return variables.get(key, match.group(0))  # 未找到则保留原样

    result = re.sub(r"\{\{(\s*\w+\s*)\}\}", replacer, template)
    if clean_residual:
        result = re.sub(r"\{\{[^}]+\}\}", "", result)
    return result


# ── 提示词加载与拆分 ──────────────────────────────────────────

def load_prompt_template(prompt_path: str) -> str:
    """加载提示词模板文件。"""
    p = Path(prompt_path)
    used_external_fallback = False
    if not p.is_absolute():
        candidates = [
            SCRIPT_DIR / prompt_path,
            SCRIPT_DIR / "prompt" / prompt_path,
            SHARED_PROMPT_DIR / prompt_path,
            SHARED_TEST_PROMPT_DIR / prompt_path,
        ]
        for idx, c in enumerate(candidates):
            if c.exists():
                p = c
                used_external_fallback = idx > 1
                break
    if not p.exists():
        print(f"[错误] 找不到提示词文件: {prompt_path}")
        sys.exit(1)
    content = p.read_text(encoding="utf-8")
    if used_external_fallback:
        print(f"  [WARN] 提示词命中外部回退路径: {p}")
    print(f"  [OK] 加载提示词模板: {p.name} ({len(content)} 字符)")
    return content


def extract_system_prompt(template_content: str) -> str:
    """
    从模板内容中提取 messages[0] system prompt 部分。
    截止到 <!-- ======================== 以上为 messages[0] --> 注释处。
    Few-shot 区域 ({{longform_few_shot}}) 在 system prompt 内留空由后端处理。
    """
    # 截取到 messages[0] 结束标记
    marker = "<!-- ======================== 以上为 messages[0]"
    idx = template_content.find(marker)
    if idx > 0:
        return template_content[:idx].strip()
    # 没找到标记，取到第一个 <!-- ======= 消息架构拼接 处
    marker2 = "<!-- ======================== 消息架构拼接说明"
    idx2 = template_content.find(marker2)
    if idx2 > 0:
        return template_content[:idx2].strip()
    # 兜底：全部作为 system
    return template_content.strip()


def split_fewshot_from_system(system_prompt: str):
    """
    将 system prompt 中的 {{longform_few_shot}} 变量区域拆分出来。
    返回: (system_before_fewshot, system_after_fewshot)

    v2.0 模板在 L5 Few-shot 区域放了 {{longform_few_shot}} 占位符，
    后端需要将其替换为实际的 user/assistant 消息对。
    我们在此处将其拆分，system prompt 中不包含 few-shot 内容。
    """
    # 找到 ---L5 Few-shot 示例注入区--- 到下一个 --- 之间的内容
    pattern = r"(---L5 Few-shot 示例注入区---.*?\n)(.*?)((?:\n---\n|\n---\s*\n))"
    match = re.search(pattern, system_prompt, re.DOTALL)

    if match:
        before = system_prompt[:match.start()].rstrip()
        after = system_prompt[match.end():].strip()
        return before, after
    else:
        # 没有 L5 标记，保持原样
        return system_prompt, ""


# ── Few-shot 示例解析 ──────────────────────────────────────────

def parse_few_shot_library(content: str) -> dict[str, list[list[dict]]]:
    """
    解析 Few-shot 示例库，按 personal_type 分组。

    兼容：
      - 标题：### 示例 N / ### 风格示例 N
      - 角色标签：**[User]** / **[Assistant]** / **用户说** / **你的回复**
    """
    library: dict[str, list[list[dict]]] = {}
    parts = re.split(r"^##\s+(.+?)\s*$", content, flags=re.MULTILINE)

    for idx in range(1, len(parts), 2):
        type_name = normalize_personal_type(parts[idx].strip())
        body = parts[idx + 1]
        if not type_name:
            continue

        examples = []
        example_heads = list(re.finditer(r"^###\s+.*$", body, re.MULTILINE))
        for head_idx, head in enumerate(example_heads):
            block_start = head.end()
            block_end = (
                example_heads[head_idx + 1].start()
                if head_idx + 1 < len(example_heads)
                else len(body)
            )
            block = body[block_start:block_end].strip()
            if not block:
                continue

            user_match = re.search(
                r"\*\*(?:\[User\]|用户说)\*\*[：:]\s*(.*?)(?:\n\s*\n\*\*(?:\[Assistant\]|你的回复)\*\*[：:]|\Z)",
                block,
                re.DOTALL,
            )
            assistant_match = re.search(
                r"\*\*(?:\[Assistant\]|你的回复)\*\*[：:]\s*(.*)",
                block,
                re.DOTALL,
            )
            if not user_match or not assistant_match:
                continue

            assistant_content = re.sub(
                r"\n\s*---\s*$", "", assistant_match.group(1).strip()
            ).strip()
            examples.append(
                [
                    {"role": "user", "content": user_match.group(1).strip()},
                    {"role": "assistant", "content": assistant_content},
                ]
            )

        if examples:
            library[type_name] = examples

    return library


def resolve_few_shot_path(few_shot_path: str | None) -> Path:
    """优先显式路径，其次使用仓内默认 Few-shot 文件。"""
    if not few_shot_path:
        return DEFAULT_FEWSHOT_FILE

    p = Path(few_shot_path)
    if p.is_absolute():
        return p

    candidates = [
        SCRIPT_DIR / few_shot_path,
        SCRIPT_DIR / "few_shot" / few_shot_path,
        SHARED_VARIABLE_DIR / few_shot_path,
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def load_few_shot_examples(
    few_shot_path: str | None,
    personal_type: str = "",
    relationship: str = "",
    gender: str = "",
    current_scene: str = "",
    max_examples: int = 2,
) -> list:
    """
    从 Few-shot 文件加载示例，按 personal_type 路由并返回扁平 messages。
    """
    if PROMPT_SERVICE is not None:
        messages = PROMPT_SERVICE.load_few_shot_examples(
            few_shot_path or "",
            relationship=relationship,
            personal_type=personal_type,
            gender=gender,
            current_scene=current_scene,
        )
        if messages:
            selected_groups = max(1, min(max_examples, len(messages) // 2))
            normalized_type = normalize_personal_type(personal_type)
            print(f"  [OK] Few-shot: {normalized_type} -> {selected_groups} 组示例")
            return messages[: max_examples * 2]

    p = resolve_few_shot_path(few_shot_path)
    if not p.exists():
        print(f"  [WARN] Few-shot 文件不存在: {p}，跳过注入")
        return []

    content = p.read_text(encoding="utf-8")
    library = parse_few_shot_library(content)
    normalized_type = normalize_personal_type(personal_type)
    if not normalized_type:
        print("  [WARN] personal_type 为空，跳过 Few-shot 注入")
        return []

    selected = library.get(normalized_type, [])[:max_examples]
    if not selected:
        print(f"  [WARN] Few-shot 未找到匹配分组: {personal_type}")
        return []

    messages = []
    for example in selected:
        messages.extend(example)

    print(f"  [OK] Few-shot: {normalized_type} -> {len(selected)} 组示例")
    return messages


# ── API 调用 ──────────────────────────────────────────────────

def call_api(messages: list, model: str = None, max_tokens: int = None) -> dict:
    """
    调用豆包 API（chat.completions 标准接口）。

    Args:
        messages: 完整消息数组
        model: 模型名（默认 PRO）
        max_tokens: 最大输出 tokens
    """
    if not API_KEY:
        raise RuntimeError("DOUBAO_API_KEY 未配置，无法调用模型接口")

    from volcenginesdkarkruntime import Ark

    client = Ark(base_url=BASE_URL, api_key=API_KEY)
    use_model = model or MODEL_PRO
    use_max_tokens = max_tokens or MAX_TOKENS

    api_kwargs = {
        "model": use_model,
        "messages": messages,
        "temperature": TEMPERATURE,
        "max_tokens": use_max_tokens,
    }
    if TOP_P < 1.0:
        api_kwargs["top_p"] = TOP_P

    start = time.time()
    response = client.chat.completions.create(**api_kwargs)
    latency = round(time.time() - start, 2)

    content = ""
    if response.choices:
        content = response.choices[0].message.content or ""

    usage = response.usage if hasattr(response, "usage") else None
    input_tokens = getattr(usage, "prompt_tokens", 0) if usage else 0
    output_tokens = getattr(usage, "completion_tokens", 0) if usage else 0

    return {
        "output": content,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "latency_s": latency,
    }


def process_ai_output(text: str, quality_guard=None) -> dict:
    """统一版主链路的质量后处理封装。"""
    guard = quality_guard or QUALITY_GUARD
    if guard is None:
        return {
            "needs_retry": False,
            "retry_reason": "",
            "processed_text": text,
            "fixes_applied": [],
        }
    return guard.check(text)


# ── 摘要生成 ──────────────────────────────────────────────────

def generate_dialogue_summary(
    conversation_history: list,
    role_name: str,
    personal_type: str,
    relationship: str,
    dry_run: bool = False,
) -> str:
    """
    调用 mini 模型生成 7 字段 dialogue_summary。
    阻塞等待返回后，格式化为注入模板的文本。

    Returns:
        格式化后的摘要文本（可直接替换 {{dialogue_summary}}）
    """
    # 将历史对话格式化为可读文本
    conv_lines = []
    for msg in conversation_history:
        role_label = "用户" if msg["role"] == "user" else "角色"
        conv_lines.append(f"[{role_label}]: {msg['content'][:500]}")
    conversation_text = "\n".join(conv_lines)

    prompt = SUMMARY_GENERATION_PROMPT.format(
        role_name=role_name,
        personal_type=personal_type,
        relationship=relationship,
        conversation_text=conversation_text,
    )

    if dry_run:
        print("  [dry-run] 跳过摘要生成 API 调用")
        return (
            "=== 之前剧情摘要 ===\n"
            "- 场景：[dry-run 模拟场景]\n"
            "- 剧情：[dry-run 模拟剧情]\n"
            "- 悬念：[dry-run 模拟悬念]\n"
            "- 角色情绪：[dry-run 模拟角色情绪]\n"
            "- 用户情绪：[dry-run 模拟用户情绪]\n"
            "- 关系动态：[dry-run 模拟关系动态]\n"
            "- 用户画像信号：[dry-run 模拟用户画像]\n"
            "=== 摘要结束 ==="
        )

    print("  [INFO] 调用 mini 模型生成 dialogue_summary ...")
    messages = [
        {"role": "system", "content": "你是一个专业的对话分析助手。请严格按JSON格式输出。"},
        {"role": "user", "content": prompt},
    ]

    response = call_api(messages, model=MODEL_MINI, max_tokens=800)
    raw_output = response["output"].strip()

    print(
        f"  [OK] 摘要生成完成 | "
        f"{response['input_tokens']}+{response['output_tokens']} tokens | "
        f"{response['latency_s']}s"
    )

    # 解析 JSON
    try:
        # 清理可能的 markdown 代码块包裹
        clean = raw_output
        if clean.startswith("```"):
            clean = re.sub(r"^```\w*\n?", "", clean)
            clean = re.sub(r"\n?```$", "", clean)
        summary_data = json.loads(clean)
    except json.JSONDecodeError:
        print("  [WARN] 摘要 JSON 解析失败，使用原始文本")
        print(f"  [INFO] 原始输出: {raw_output[:200]}")
        # 兜底：直接使用原始输出作为摘要
        return f"=== 之前剧情摘要 ===\n{raw_output}\n=== 摘要结束 ==="

    # 格式化为白皮书 §9.5.3 注入格式
    formatted = SUMMARY_INJECT_TEMPLATE.format(
        scene_description=summary_data.get("scene_description", "未知"),
        plot_summary=summary_data.get("plot_summary", "无"),
        pending_hooks=summary_data.get("pending_hooks", "无"),
        character_emotion=summary_data.get("character_emotion", "未知"),
        user_emotion=summary_data.get("user_emotion", "未知"),
        relationship_shift=summary_data.get("relationship_shift", "未知"),
        user_profile_signals=summary_data.get("user_profile_signals", "无"),
    )

    print(f"  [INFO] 摘要预览:\n{formatted[:200]}...")
    return formatted


# ── 消息组装 ──────────────────────────────────────────────────

def build_messages_for_turn(
    rendered_system: str,
    system_after: str,
    few_shot_messages: list,
    conversation_history: list,
    dialogue_summary: str,
    current_input: str,
    relationship: str,
    role_name: str = "",
    personality: str = "",
    turn_num: int = 1,
) -> list:
    """
    按白皮书 v1.6 §4.1 消息架构组装完整 messages 数组。

    消息架构：
      [0]   system  → L0-L4 核心 system prompt（已注入变量，含 dialogue_summary）
      [1-4] user/assistant → Few-shot 示例（2 组）
      [5]   system  → 增强分隔标记
      [6]   system  → 风格隔离声明
      [7]   system  → dialogue_summary（如有，独立 system 消息）
      [7+]  user/assistant → 历史对话轮次
      [N-1] system  → Core_Constraints 重申
      [N]   user    → 当前用户输入
    """
    messages = []

    # messages[0]: System Message (L0-L5 + 系统模块 + 记忆上下文)
    full_system = rendered_system
    if system_after:
        full_system = rendered_system + "\n\n" + system_after
    messages.append({"role": "system", "content": full_system})

    # S1: Few-shot 注入策略
    # - 2026-04-23 切换场景 A/B 测试：首轮注入 RSO 3/3，跳过 0/3
    #   根因：生产 dialogueStartPrompt 含"如果下文有你们的对话历史"句，
    #   导致模型将 Few-shot user/assistant 对误识为历史对话→场景渗透
    # - 首轮（含纯首轮 + 切换首轮）跳过 Few-shot，第 2 轮起注入
    # - Turn 16+: 冷却复注第1组
    is_first_turn = not conversation_history

    # messages[1-4]: Few-shot (§4.7 冷却复注策略)
    injected_few_shot = []
    if not is_first_turn and few_shot_messages:
        if turn_num <= 15:
            injected_few_shot = few_shot_messages
        else:
            injected_few_shot = few_shot_messages[:2]
    if injected_few_shot:
        messages.append({"role": "system", "content": FEW_SHOT_PREFIX_MSG})
        messages.extend(injected_few_shot)
        messages.append({"role": "system", "content": SEPARATOR_MSG})

    # S3: 首轮插入"首次对话"哨兵消息，利用 Recency 位置优势
    if is_first_turn:
        messages.append({"role": "system", "content": FIRST_TURN_SENTINEL})

    # 风格隔离声明（P0，阻断 assistant 历史自强化陷阱）
    # 白皮书 v1.6 §3.10：只要有历史对话或摘要，始终注入
    if conversation_history or dialogue_summary:
        messages.append({"role": "system", "content": STYLE_ISOLATION_MSG})

    # messages[7]: dialogue_summary（如有，独立 system 消息，role=system）
    # 白皮书 v1.6 §4.1：摘要在历史对话区域内、近期对话之前
    if dialogue_summary:
        messages.append({"role": "system", "content": dialogue_summary})

    # messages[7+]: 历史对话轮次
    # CLI 保持保守合同：不额外插入服务端的 Depth Injection system 消息，
    # 避免 dry-run/导出链路的消息数与既有回归基线漂移。
    messages.extend(conversation_history)

    # messages[N-1]: Core_Constraints 重申（Sandwich Pattern 近因端，始终注入）
    # 白皮书 v1.6 §3.8：紧贴用户输入前的 role=system 消息，首轮也需要
    core_constraints = CORE_CONSTRAINTS_TEMPLATE.format(relationship=relationship)
    messages.append({"role": "system", "content": core_constraints})

    # messages[N]: 当前用户输入 (XML 防注入)
    messages.append({
        "role": "user",
        "content": f"<user_input>{current_input}</user_input>",
    })

    return messages


# ── Excel 导出 ────────────────────────────────────────────────





# ── Excel 导出 ────────────────────────────────────────────────

def export_to_excel(results: list, config: dict, output_path: str):
    """导出对话结果为打分用 Excel（含全部打分变量列）。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("[错误] 需要 openpyxl: pip install openpyxl")
        return
    # 打分提示词 default_v2.md 需要的全部变量列
    headers = [
        "测试对应提示词", "轮次",
        # 角色设定（打分维度1: persona_fidelity 需要）
        "Role_Nickname", "gender", "age", "occupation",
        "personality", "speaking_style", "personal_type", "hobby", "background",
        "Role_info_works",
        # 用户信息
        "user_Nickname", "user_gender", "user_identity",
        # 关系与边界（打分维度4: boundary_memory 需要）
        "relationship", "relation_info", "intimacy_boundary", "relation_calling",
        # 时空上下文
        "currentTime", "weekDay", "timeperiod", "season", "current_scene",
        "last_cst_type", "完整时间信息",
        # 叙事策略（打分维度1的锚点词/禁用词标尺）
        "longform_narrative_style", "longform_persona",
        "longform_dialogue_guideline", "system_module8",
        # 记忆上下文（打分维度4: 记忆真实性参考）
        "dialogueStartPrompt", "moments", "weekly_schedule",
        "monthly_schedule", "dialogue_summary",
        "voice_forbidden",
        # 核心输入输出（打分主体）
        "用户输入", "AI输出",
        # 元数据
        "输入tokens", "输出tokens", "延迟(秒)",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # 从 config 提取固定变量（每行相同）
    char = config.get("character", {})
    ctx = config.get("context", {})
    modules = config.get("modules", {})
    prompt_name = config.get("prompt_file", "unknown")

    # 宽文本列集合（输入/输出/策略/摘要/记忆）
    wide_cols = set()
    for i, h in enumerate(headers, 1):
        if h in ("用户输入", "AI输出", "longform_narrative_style",
                 "longform_persona", "dialogue_summary", "dialogueStartPrompt",
                 "moments", "weekly_schedule", "monthly_schedule",
                 "personality", "user_identity"):
            wide_cols.add(i)

    for i, r in enumerate(results):
        row = i + 2
        values = [
            prompt_name, i + 1,
            # 角色设定
            char.get("Role_Nickname", ""), char.get("gender", ""),
            char.get("age", ""), char.get("occupation", ""),
            char.get("personality", ""), char.get("speaking_style", ""),
            char.get("personal_type", ""), char.get("hobby", ""),
            char.get("background", ""),
            char.get("Role_info_works", ""),
            # 用户信息
            modules.get("user_Nickname", ""), modules.get("user_gender", ""),
            modules.get("user_identity", ""),
            # 关系与边界
            ctx.get("relationship", ""), ctx.get("relation_info", ""),
            ctx.get("intimacy_boundary", ""), ctx.get("relation_calling", ""),
            # 时空
            ctx.get("currentTime", ""), ctx.get("weekDay", ""),
            ctx.get("timeperiod", ""), ctx.get("season", ""),
            ctx.get("current_scene", ""), ctx.get("last_cst_type", ""),
            ctx.get("完整时间信息", ""),
            # 叙事策略
            modules.get("longform_narrative_style", ""),
            modules.get("longform_persona", ""),
            modules.get("longform_dialogue_guideline", ""),
            modules.get("system_module8", ""),
            # 记忆
            modules.get("dialogueStartPrompt", ""),
            modules.get("moments", ""),
            modules.get("weekly_schedule", ""),
            modules.get("monthly_schedule", ""),
            r.get("dialogue_summary", ""),
            modules.get("voice_forbidden", ""),
            # 核心输入输出  
            r["user_input"], r["ai_output"],
            # 元数据
            r.get("input_tokens", 0), r.get("output_tokens", 0),
            r.get("latency_s", 0),
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border
            if col in wide_cols:
                cell.alignment = Alignment(wrap_text=True)

    # 列宽设置
    col_widths = {
        "测试对应提示词": 22, "轮次": 5, "用户输入": 30, "AI输出": 60,
        "dialogue_summary": 40, "longform_narrative_style": 35,
        "longform_persona": 35, "personality": 25,
        "dialogueStartPrompt": 30, "moments": 28,
        "weekly_schedule": 26, "monthly_schedule": 24,
        "完整时间信息": 26,
    }
    for col_idx, h in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_widths.get(h, 12)

    wb.save(output_path)
    print(f"  [OK] Excel 导出: {output_path} ({len(headers)} 列)")


# ── Excel 输入 (v1 能力迁移) ───────────────────────────────────

# 角色变量映射（Excel 列名 → config.character 键名）
_CHAR_COLS = [
    "Role_Nickname", "gender", "age", "occupation",
    "personality", "speaking_style", "personal_type", "background", "hobby",
    "Role_info_works",
]
# 上下文变量映射（Excel 列名 → config.context 键名）
_CTX_COLS = [
    "relationship", "relation_info", "intimacy_boundary", "relation_calling",
    "currentTime", "weekDay", "timeperiod", "season", "current_scene",
    "last_cst_type", "完整时间信息",
]
# 模块变量映射（Excel 列名 → config.modules 键名）
_MOD_COLS = [
    "user_Nickname", "user_gender", "user_identity",
    "system_module8", "longform_persona", "longform_narrative_style",
    "longform_dialogue_guideline",
    "system_Role_acting", "weekly_schedule", "monthly_schedule", "moments",
    "dialogueStartPrompt",
    "longform_few_shot", "voice_forbidden", "dialogue_summary",
]


def _safe_str(val) -> str:
    """安全转换 Excel 单元格值为字符串，NaN → ''。"""
    import pandas as pd
    if pd.isna(val):
        return ""
    return str(val).strip()


def load_config_from_excel(
    excel_path: str,
    prompt_path: str = None,
    few_shot_path: str = None,
) -> list[dict]:
    """
    读取 Excel → 返回 config 列表。

    支持两种模式混用（对齐 v1 generate.py）：
      方案A：无 session_id 的行 → 每行独立 config（turns 只有 1 轮）
      方案B：有 session_id 的行 → 同 session 聚合为一个 config（turns 按 turn_order 排序）

    Returns:
        [{"prompt_file", "character", "context", "modules", "turns", "_mode"}, ...]
        _mode: "independent" | "session"
    """
    import pandas as pd
    df = pd.read_excel(excel_path)
    print(f"  [OK] 读取 Excel: {excel_path} ({len(df)} 行, {len(df.columns)} 列)")

    # 兼容 v9 Excel 表头别名
    col_aliases = {"用户输入": "user_message"}
    for old_name, new_name in col_aliases.items():
        if old_name in df.columns and new_name not in df.columns:
            df.rename(columns={old_name: new_name}, inplace=True)
    # dialogue_summary 列名可能有前导空格
    df.columns = [c.strip() for c in df.columns]

    if "user_message" not in df.columns:
        print("[错误] Excel 缺少必填列: user_message 或 用户输入")
        sys.exit(1)

    # 默认提示词路径
    default_prompt = prompt_path or str(DEFAULT_PROMPT_FILE)
    default_fewshot = few_shot_path or str(DEFAULT_FEWSHOT_FILE)

    def _row_to_config_parts(row):
        """将一行 Excel 提取为 character/context/modules 字典。"""
        char = {k: _safe_str(row.get(k, "")) for k in _CHAR_COLS}
        ctx = {k: _safe_str(row.get(k, "")) for k in _CTX_COLS}
        mods = {k: _safe_str(row.get(k, "")) for k in _MOD_COLS}
        return char, ctx, mods

    configs = []
    has_session = "session_id" in df.columns and "turn_order" in df.columns

    if has_session:
        session_mask = df["session_id"].notna()
        # ── 方案B：按 session_id 分组 ──
        session_rows = df[session_mask]
        if not session_rows.empty:
            for sid, group in session_rows.groupby("session_id"):
                sorted_g = group.sort_values("turn_order")
                first_row = sorted_g.iloc[0]
                char, ctx, mods = _row_to_config_parts(first_row)
                turns = [_safe_str(r["user_message"])
                         for _, r in sorted_g.iterrows()]
                configs.append({
                    "prompt_file": default_prompt,
                    "few_shot_file": default_fewshot,
                    "character": char,
                    "context": ctx,
                    "modules": mods,
                    "turns": turns,
                    "_mode": "session",
                    "_session_id": str(sid),
                })
        # ── 方案A：独立行 ──
        indep_rows = df[~session_mask]
    else:
        indep_rows = df

    for idx, row in indep_rows.iterrows():
        char, ctx, mods = _row_to_config_parts(row)
        user_msg = _safe_str(row.get("user_message", ""))
        configs.append({
            "prompt_file": default_prompt,
            "few_shot_file": default_fewshot,
            "character": char,
            "context": ctx,
            "modules": mods,
            "turns": [user_msg],
            "_mode": "independent",
        })

    session_count = sum(1 for c in configs if c["_mode"] == "session")
    indep_count = sum(1 for c in configs if c["_mode"] == "independent")
    print(f"  [OK] 解析完成: {session_count} 个会话(方案B) + "
          f"{indep_count} 个独立行(方案A)")
    return [normalize_cli_config(cfg) for cfg in configs]


# ── 并发批量 (v1 能力迁移) ─────────────────────────────────────

async def _run_single_independent(
    config: dict,
    idx: int,
    semaphore,
    dry_run: bool = False,
) -> tuple[int, list]:
    """并发处理单个独立行（方案A），用 semaphore 限制并发。"""
    import asyncio
    async with semaphore:
        # run_conversation_chain 是同步函数，放到线程池执行
        loop = asyncio.get_event_loop()
        results = await loop.run_in_executor(
            None,
            lambda: run_conversation_chain(config, dry_run=dry_run),
        )
        name = config.get("character", {}).get("Role_Nickname", f"Row{idx}")
        status = "OK" if results else "FAIL"
        print(f"  [{idx}] {name} | {status}")
        return idx, results


async def run_batch_parallel(
    configs: list[dict],
    workers: int = 20,
    dry_run: bool = False,
    output_dir: Path = None,
) -> dict[int, list]:
    """
    批量执行对话配置列表。
    方案A(independent) → 并发；方案B(session) → 串行。

    Returns: {config_index: results_list}
    """
    import asyncio
    output_dir = output_dir or SCRIPT_DIR
    all_results: dict[int, list] = {}

    # ── 先串行执行方案B ──
    session_cfgs = [(i, c) for i, c in enumerate(configs)
                    if c.get("_mode") == "session"]
    if session_cfgs:
        print(f"\n=== 方案B: {len(session_cfgs)} 个会话（串行）===")
        for i, cfg in session_cfgs:
            sid = cfg.get("_session_id", "?")
            turns = cfg["turns"]
            print(f"\n--- Session {sid} ({len(turns)} 轮) ---")
            results = run_conversation_chain(cfg, dry_run=dry_run)
            all_results[i] = results
            if not dry_run:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                role = cfg.get("character", {}).get("Role_Nickname", "unknown")
                excel_path = output_dir / f"{role}_s{sid}_{ts}_待打分.xlsx"
                export_to_excel(results, cfg, str(excel_path))

    # ── 再并发执行方案A ──
    indep_cfgs = [(i, c) for i, c in enumerate(configs)
                  if c.get("_mode") == "independent"]
    if indep_cfgs:
        print(f"\n=== 方案A: {len(indep_cfgs)} 行（并发 {workers}）===")
        sem = asyncio.Semaphore(workers)
        tasks = [
            _run_single_independent(cfg, i, sem, dry_run=dry_run)
            for i, cfg in indep_cfgs
        ]
        done = await asyncio.gather(*tasks)
        for idx, results in done:
            all_results[idx] = results

        # 合并所有独立行结果为一个 Excel
        if not dry_run:
            merged_results = []
            merged_config = indep_cfgs[0][1]  # 用第一行的 config 做模板
            for i, cfg in indep_cfgs:
                for r in all_results.get(i, []):
                    merged_results.append(r)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_path = output_dir / f"batch_{ts}_待打分.xlsx"
            export_to_excel(merged_results, merged_config, str(excel_path))

    return all_results


# ── 主流程 ────────────────────────────────────────────────────

def build_variables(config: dict) -> dict:
    """
    从 JSON 配置中组装所有模板变量，用于 {{variable}} 注入。
    """
    variables = {}

    # 角色变量
    char = config.get("character", {})
    for key in [
        "Role_Nickname", "gender", "age", "occupation",
        "personality", "speaking_style", "personal_type",
        "Role_info_works",
        "background", "hobby",
    ]:
        variables[key] = char.get(key, "")

    # 上下文变量
    ctx = config.get("context", {})
    for key in [
        "relationship", "currentTime", "weekDay", "timeperiod",
        "season", "intimacy_boundary", "relation_calling",
        "relation_info", "relation_rule4", "system_module11",
        "current_scene", "last_cst_type", "完整时间信息",
    ]:
        variables[key] = ctx.get(key, "")
    if not variables.get("完整时间信息"):
        parts = [
            _safe_str(ctx.get("currentTime", "")),
            _safe_str(ctx.get("weekDay", "")),
            _safe_str(ctx.get("timeperiod", "")),
            _safe_str(ctx.get("season", "")),
        ]
        variables["完整时间信息"] = " / ".join(part for part in parts if part)

    # 系统模块变量
    modules = config.get("modules", {})
    for key in [
        "system_module8", "longform_persona", "longform_narrative_style",
        "longform_dialogue_guideline",
        "system_Role_acting", "weekly_schedule", "monthly_schedule",
        "dialogueStartPrompt", "moments",
        "longform_few_shot", "user_Nickname", "user_gender", "user_identity",
        "voice_forbidden",
    ]:
        variables[key] = modules.get(key, ctx.get(key, char.get(key, "")))
    
    # dialogue_summary 初始为空（由脚本动态生成）
    variables.setdefault("dialogue_summary", "")

    # 合并额外的自定义变量（来自前端变量预览的编辑覆盖）
    custom_vars = config.get("custom_variables", {})
    if custom_vars:
        variables.update(custom_vars)
        
    return variables


def run_conversation_chain(config: dict, max_turns: int = None, dry_run: bool = False):
    """
    执行多轮对话链。

    核心流程：
      1. 加载模板 → 注入变量 → 得到 system prompt
      2. 逐轮执行：组装 messages → 调用 pro API → 拼接历史
      3. 每 5 轮暂停：调用 mini → 生成摘要 → 注入变量 → 重新渲染 system prompt
    """
    # ── 1. 加载模板 ──
    prompt_file = config["prompt_file"]
    template_raw = load_prompt_template(prompt_file)
    system_template = extract_system_prompt(template_raw)

    # 拆分 L5 Few-shot 区域
    system_before_fewshot, system_after = split_fewshot_from_system(system_template)

    # 加载 Few-shot 示例（显式路径优先，否则默认仓内资源）
    character = config.get("character", {}) or {}
    context = config.get("context", {}) or {}
    personal_type = character.get("personal_type", "")
    few_shot_file = config.get("few_shot_file") or str(DEFAULT_FEWSHOT_FILE)
    few_shot_messages = load_few_shot_examples(
        few_shot_file,
        personal_type=personal_type,
        relationship=context.get("relationship", ""),
        gender=character.get("gender", ""),
        current_scene=context.get("current_scene", ""),
        max_examples=2,
    )

    # ── 2. 组装变量 ──
    variables = build_variables(config)

    # 首次渲染 system prompt（dialogue_summary 初始为空）
    rendered_system = render_template(system_before_fewshot, variables, clean_residual=True)
    rendered_after = render_template(system_after, variables, clean_residual=True) if system_after else ""

    # ── 3. 对话设置 ──
    turns = config["turns"]
    relationship = config.get("context", {}).get("relationship", "暧昧")
    role_name = config.get("character", {}).get("Role_Nickname", "unknown")
    if max_turns:
        turns = turns[:max_turns]

    print(f"\n{'='*60}")
    print(f"  对话链: {len(turns)} 轮")
    print(f"  角色: {role_name}")
    print(f"  关系: {relationship}")
    print(f"  主模型: {MODEL_PRO}")
    print(f"  摘要模型: {MODEL_MINI}")
    print(f"  摘要间隔: 每 {SUMMARY_INTERVAL} 轮")
    print(f"{'='*60}\n")

    conversation_history = []  # 真实对话历史 [{role, content}, ...]
    dialogue_summary = ""      # 当前摘要文本（格式化后）
    results = []

    for i, user_input in enumerate(turns):
        turn_num = i + 1

        print(f"\n{'-'*50}")
        print(f"  Turn {turn_num}/{len(turns)}: {user_input[:50]}{'...' if len(user_input) > 50 else ''}")
        if dialogue_summary:
            print(f"  [当前 dialogue_summary 已注入]")
        print(f"{'-'*50}")

        # ── 组装消息 ──
        messages = build_messages_for_turn(
            rendered_system=rendered_system,
            system_after=rendered_after,
            few_shot_messages=few_shot_messages,
            conversation_history=conversation_history,
            dialogue_summary=dialogue_summary,
            current_input=user_input,
            relationship=relationship,
            role_name=role_name,
            personality=config.get("character", {}).get("personality", ""),
            turn_num=turn_num,
        )

        if dry_run:
            print(f"  消息数量: {len(messages)}")
            for j, m in enumerate(messages):
                role = m["role"]
                content_preview = safe_preview(m["content"], 80)
                print(f"    [{j}] {role}: {content_preview}...")
            print(f"  历史对话轮次: {len(conversation_history) // 2}")

            ai_output = f"[dry-run] Turn {turn_num} 的模拟AI回复（300-500字叙事内容）"
            result_entry = {
                "turn": turn_num,
                "user_input": user_input,
                "ai_output": ai_output,
                "dialogue_summary": dialogue_summary,
                "msg_count": len(messages),
                "input_tokens": 0,
                "output_tokens": 0,
                "latency_s": 0,
            }
        else:
            # ── 实际调用 pro API ──
            print(f"  [INFO] 调用 {MODEL_PRO} (消息数: {len(messages)})...")
            response = call_api(messages, model=MODEL_PRO)
            quality_result = process_ai_output(response["output"])
            ai_output = quality_result["processed_text"]
            if quality_result["fixes_applied"]:
                fixes = "、".join(quality_result["fixes_applied"])
                print(f"  [INFO] QualityGuard 后处理: {fixes}")
            if quality_result["needs_retry"]:
                print(f"  [WARN] 输出未达质量门槛: {quality_result['retry_reason']}")

            word_count = len(ai_output)
            print(
                f"  [OK] 回复: {word_count} 字 | "
                f"{response['input_tokens']}+{response['output_tokens']} tokens | "
                f"{response['latency_s']}s"
            )
            preview = safe_preview(ai_output, 100)
            print(f"  预览: {preview}...")

            result_entry = {
                "turn": turn_num,
                "user_input": user_input,
                "ai_output": ai_output,
                "dialogue_summary": dialogue_summary,
                "msg_count": len(messages),
                "input_tokens": response["input_tokens"],
                "output_tokens": response["output_tokens"],
                "latency_s": response["latency_s"],
            }

        results.append(result_entry)

        # ── 更新 conversation_history ──
        conversation_history.append({"role": "user", "content": user_input})
        conversation_history.append({"role": "assistant", "content": ai_output})

        # ── 每 SUMMARY_INTERVAL 轮：生成 dialogue_summary ──
        if turn_num % SUMMARY_INTERVAL == 0 and turn_num < len(turns):
            print(f"\n  [INFO] 第 {turn_num} 轮完成，开始生成 dialogue_summary ...")
            dialogue_summary = generate_dialogue_summary(
                conversation_history=conversation_history,
                role_name=role_name,
                personal_type=personal_type,
                relationship=relationship,
                dry_run=dry_run,
            )
            print("  [OK] 摘要已生成，将通过独立 system 消息注入\n")

        # 轮间休息
        if not dry_run and i < len(turns) - 1:
            time.sleep(1)

    return results


def main():
    import asyncio

    parser = argparse.ArgumentParser(
        description="长文模式多轮对话批量测试工具 v3.0 (统一版)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""\
示例:
  # JSON 配置（原 v2 用法）
  python longform_multi_turn.py test_conversation_萧璟言.json
  python longform_multi_turn.py test_conversation_萧璟言.json --dry-run
  python longform_multi_turn.py test_conversation_萧璟言.json --turns 3

  # Excel 输入（原 v1 用法，自动检测方案A/B）
  python longform_multi_turn.py input.xlsx --dry-run
  python longform_multi_turn.py input.xlsx --workers 30
  python longform_multi_turn.py input.xlsx --prompt prompt/v2.md
        """,
    )
    parser.add_argument("input_file",
                        help="输入文件 (.json 或 .xlsx)")
    parser.add_argument("--dry-run", action="store_true",
                        help="仅查看消息结构，不调用 API")
    parser.add_argument("--turns", "-t", type=int, default=None,
                        help="最多执行的轮数 (仅 JSON 模式)")
    parser.add_argument("--output-dir", "-o", default=None,
                        help="输出目录")
    parser.add_argument("--workers", "-w", type=int, default=20,
                        help="并发数 (仅 Excel 方案A, 默认20)")
    parser.add_argument("--prompt", default=None,
                        help="提示词文件路径 (仅 Excel 模式)")
    parser.add_argument("--fewshot", default=None,
                        help="Few-shot 示例文件路径 (仅 Excel 模式)")
    args = parser.parse_args()

    # 解析输入文件
    input_path = Path(args.input_file)
    if not input_path.is_absolute():
        input_path = SCRIPT_DIR / input_path
    if not input_path.exists():
        print(f"[错误] 输入文件不存在: {input_path}")
        sys.exit(1)

    output_dir = Path(args.output_dir) if args.output_dir else SCRIPT_DIR
    if not args.dry_run:
        output_dir.mkdir(parents=True, exist_ok=True)

    print("\n" + "=" * 60)
    print("  长文模式多轮对话批量测试 v3.0 (统一版)")
    print("=" * 60)

    is_excel = input_path.suffix.lower() in ('.xlsx', '.xls')

    if is_excel:
        # ════════════ Excel 输入模式 ════════════
        configs = load_config_from_excel(
            str(input_path),
            prompt_path=args.prompt,
            few_shot_path=args.fewshot,
        )
        asyncio.run(run_batch_parallel(
            configs, workers=args.workers,
            dry_run=args.dry_run, output_dir=output_dir,
        ))
        print(f"\n{'='*60}")
        print(f"  完成! {len(configs)} 个配置已处理")
        print(f"{'='*60}\n")
    else:
        # ════════════ JSON 配置模式（原 v2）════════════
        with open(input_path, "r", encoding="utf-8") as f:
            config = json.load(f)
        config = normalize_cli_config(config)

        if "prompt_file" not in config or "turns" not in config:
            print("[错误] JSON 必须包含 prompt_file 和 turns 字段")
            sys.exit(1)

        results = run_conversation_chain(
            config, max_turns=args.turns, dry_run=args.dry_run,
        )

        log_name = ""
        excel_name = ""
        if not args.dry_run:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            role_name = config.get("character", {}).get(
                "Role_Nickname", "unknown")

            log_name = f"{role_name}_{ts}_对话日志.json"
            log_path = output_dir / log_name
            log_data = {
                "config": config, "timestamp": ts,
                "model_pro": MODEL_PRO, "model_mini": MODEL_MINI,
                "summary_interval": SUMMARY_INTERVAL,
                "results": results,
            }
            with open(log_path, "w", encoding="utf-8") as f:
                json.dump(log_data, f, ensure_ascii=False, indent=2)
            print(f"\n  [OK] 对话日志: {log_path}")

            excel_name = f"{role_name}_{ts}_待打分.xlsx"
            excel_path = output_dir / excel_name
            export_to_excel(results, config, str(excel_path))

        # 汇总
        print(f"\n{'='*60}")
        print(f"  完成! {len(results)} 轮对话")
        if not args.dry_run:
            total_in = sum(r.get("input_tokens", 0) for r in results)
            total_out = sum(r.get("output_tokens", 0) for r in results)
            total_t = sum(r.get("latency_s", 0) for r in results)
            print(f"  总 tokens: {total_in} (输入) + {total_out} (输出)")
            print(f"  总耗时: {total_t:.1f}s")
            print(f"  对话日志: {log_name}")
            print(f"  打分Excel: {excel_name}")
        else:
            print("  [INFO] dry-run 模式未写出日志和 Excel")
        print(f"{'='*60}\n")


if __name__ == "__main__":
    main()

