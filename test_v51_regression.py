"""
v5.1 回归测试。
"""
from __future__ import annotations

import asyncio
import io
import os
import sqlite3
import sys
import tempfile
import threading
import time
import types
from pathlib import Path

import openpyxl
import pytest

PROJECT_DIR = Path(__file__).resolve().parent
SERVER_DIR = PROJECT_DIR / "server"
PIPELINE_SCRIPTS = Path(r"E:\提效工具\promptfoo-pipeline\scripts")
DB_DIR = PROJECT_DIR / "output" / "test_runtime"
DB_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DB_DIR / "v51_regression.db"

os.environ["LONGFORM_DB_PATH"] = str(DB_PATH)

if str(SERVER_DIR) not in sys.path:
    sys.path.insert(0, str(SERVER_DIR))
if str(PIPELINE_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(PIPELINE_SCRIPTS))

import database as db  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402
from config import (  # noqa: E402
    DEFAULT_INJECTION_DEPTH,
    DEFAULT_PROMPT_FILE,
    DEFAULT_SUMMARY_INTERVAL,
    get_latest_prompt_file,
)
# 测试用 prompt_version 优先指向真实存在的最新版本，避免 DEFAULT_PROMPT_FILE
# 配置滞后于实际目录时整组测试因 "提示词版本不存在" 失败。
DEFAULT_PROMPT_FILE = get_latest_prompt_file() or DEFAULT_PROMPT_FILE  # noqa: E402
from main import app  # noqa: E402
from models import (  # noqa: E402
    OrchestrationGroupRequest,
    OrchestrationItemRequest,
    OrchestrationRunCreate,
)
from routers import conversations as conversations_router  # noqa: E402
from routers import ab_sessions as ab_sessions_router  # noqa: E402
from routers import models_router  # noqa: E402
from routers import prompts as prompts_router  # noqa: E402
from routers import scoring as scoring_router  # noqa: E402
from services import model_adapter as model_adapter_module  # noqa: E402
from services import orchestration_service as orchestration_service_module  # noqa: E402
from score_excel import load_scoring_template  # noqa: E402
from services.conversation_service import ConversationService  # noqa: E402
from services.prompt_service import PromptService  # noqa: E402
from services.quality_guard import QualityGuard  # noqa: E402
from services import scoring_service as scoring_service_module  # noqa: E402
from services import task_control as task_control_module  # noqa: E402


def _reset_runtime_singletons():
    conversations_router._conv_service = None
    conversations_router._running_conversations.clear()
    conversations_router._queued_conversations.clear()
    conversations_router._ws_connections.clear()
    scoring_router._scoring = None
    scoring_router._live_scoring_dispatcher = None
    scoring_router._score_ws.clear()
    for task in list(orchestration_service_module._runner_tasks.values()):
        try:
            task.cancel()
        except Exception:
            pass
    orchestration_service_module._runner_tasks.clear()
    for task_id in list(task_control_module.list_active().keys()):
        task_control_module.remove(task_id)


def _prepare_test_database():
    db.init_db()
    db.migrate_add_score_columns()
    db.migrate_add_v51_columns()
    db.migrate_add_compare_reports_table()
    db.migrate_add_ai_report_summaries_table()
    db.migrate_add_conversation_events_table()
    db.migrate_add_orchestration_runs_table()
    db.migrate_add_ab_sessions_table()


def _truncate_test_tables():
    table_names = [
        "turn_results",
        "conversations",
        "presets",
        "saved_configs",
        "ai_report_summaries",
        "conversation_events",
        "compare_reports",
        "orchestration_runs",
        "ab_sessions",
    ]
    for attempt in range(20):
        conn = None
        try:
            conn = db.get_connection()
            tables = {
                row["name"]
                for row in conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'"
                ).fetchall()
            }
            existing_tables = [name for name in table_names if name in tables]
            conn.execute("PRAGMA foreign_keys=OFF")
            for table in existing_tables:
                conn.execute(f"DELETE FROM {table}")
            if "sqlite_sequence" in tables and existing_tables:
                placeholders = ",".join("?" for _ in existing_tables)
                conn.execute(
                    f"DELETE FROM sqlite_sequence WHERE name IN ({placeholders})",
                    existing_tables,
                )
            conn.commit()
            return
        except sqlite3.OperationalError as exc:
            if "locked" not in str(exc).lower() or attempt == 19:
                raise
            time.sleep(0.1)
        finally:
            if conn is not None:
                try:
                    conn.execute("PRAGMA foreign_keys=ON")
                except sqlite3.Error:
                    pass
                conn.close()


def _reset_test_database():
    _reset_runtime_singletons()
    _prepare_test_database()
    _truncate_test_tables()


@pytest.fixture(autouse=True)
def isolate_regression_db():
    _reset_test_database()
    yield
    _reset_runtime_singletons()


@pytest.fixture
def client():
    _reset_runtime_singletons()
    with TestClient(app) as test_client:
        yield test_client
    _reset_runtime_singletons()


def assert_true(condition: bool, message: str):
    if not condition:
        raise AssertionError(message)


def wait_for_turns(conv_id: str, expected_count: int, timeout: float = 6.0):
    deadline = time.time() + timeout
    latest = None
    while time.time() < deadline:
        latest = db.get_conversation(conv_id)
        if latest and len(latest.get("results", [])) >= expected_count:
            return latest
        time.sleep(0.1)
    actual = len(latest.get("results", [])) if latest else 0
    raise AssertionError(
        f"等待对话 {conv_id} 结果超时，期望 {expected_count} 轮，实际 {actual}"
    )


def wait_for_turn_score(
    conv_id: str,
    turn: int,
    expected_status: str = "scored",
    timeout: float = 6.0,
):
    deadline = time.time() + timeout
    latest = None
    while time.time() < deadline:
        latest = db.get_conversation(conv_id)
        if not latest:
            time.sleep(0.1)
            continue
        target = next(
            (item for item in latest.get("results", []) if item.get("turn") == turn),
            None,
        )
        if target and target.get("score_status") == expected_status:
            return latest
        time.sleep(0.1)

    actual = None
    if latest:
        target = next(
            (item for item in latest.get("results", []) if item.get("turn") == turn),
            None,
        )
        if target:
            actual = target.get("score_status")
    raise AssertionError(
        f"等待对话 {conv_id} 第 {turn} 轮评分超时，期望 {expected_status}，实际 {actual}"
    )


def wait_for_conversation_status(
    conv_id: str,
    expected_status: str,
    timeout: float = 8.0,
):
    deadline = time.time() + timeout
    latest = None
    while time.time() < deadline:
        latest = db.get_conversation(conv_id)
        if latest and latest.get("status") == expected_status:
            return latest
        time.sleep(0.1)
    actual = latest.get("status") if latest else None
    raise AssertionError(
        f"等待对话 {conv_id} 状态超时，期望 {expected_status}，实际 {actual}"
    )


def wait_for_condition(predicate, timeout: float = 6.0, message: str = "条件等待超时"):
    deadline = time.time() + timeout
    while time.time() < deadline:
        if predicate():
            return
        time.sleep(0.05)
    raise AssertionError(message)


def create_test_orchestration_run(
    kind: str = "batch",
    *,
    status: str = "running",
    item_status: str = "running",
    title: str = "测试编排任务",
) -> dict:
    manifest = {
        "kind": kind,
        "title": title,
        "concurrency": 1,
        "groups": [
            {
                "key": "group:1",
                "label": "角色A",
                "relationship": "暧昧",
                "planned_turns": 2,
                "items": [
                    {
                        "key": "group:1:item:1",
                        "label": "角色A",
                        "relationship": "暧昧",
                        "model_id": "doubao-pro",
                        "planned_turns": 2,
                        "payload": {
                            "nickname": "角色A",
                            "relationship": "暧昧",
                            "turns": ["第一轮", "第二轮"],
                            "model_id": "doubao-pro",
                        },
                    }
                ],
            }
        ],
    }
    state = {
        "kind": kind,
        "title": title,
        "groups": [
            {
                "key": "group:1",
                "label": "角色A",
                "relationship": "暧昧",
                "planned_turns": 2,
                "status": item_status,
                "items": [
                    {
                        "key": "group:1:item:1",
                        "label": "角色A",
                        "relationship": "暧昧",
                        "model_id": "doubao-pro",
                        "planned_turns": 2,
                        "conversation_id": "",
                        "status": item_status,
                        "turn_count": 1 if item_status != "pending" else 0,
                        "avg_chars": 120 if item_status != "pending" else 0,
                        "avg_score": None,
                        "resume_supported": False,
                        "error": "",
                    }
                ],
            }
        ],
        "summary": {},
    }
    return db.create_orchestration_run(
        kind,
        title=title,
        concurrency=1,
        manifest=manifest,
        state=state,
        status=status,
    )


def count_ai_report_summaries(target_id: str, report_kind: str = "scoring_report") -> int:
    conn = db.get_connection()
    try:
        row = conn.execute(
            """SELECT COUNT(*) AS cnt
               FROM ai_report_summaries
               WHERE target_id=? AND report_kind=?""",
            (target_id, report_kind),
        ).fetchone()
        return int(row["cnt"] or 0)
    finally:
        conn.close()


def create_workbook_bytes(headers: list[str], rows: list[list]):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    buffer.seek(0)
    return buffer


def create_scored_conversation(
    role_name: str,
    totals: list[float],
    manual_score=None,
    *,
    model_id: str = "doubao-pro",
    prompt_file: str = "test_prompt.md",
):
    config = {
        "prompt_file": prompt_file,
        "character": {
            "Role_Nickname": role_name,
            "gender": "男",
            "personality": "冷静",
        },
        "context": {"relationship": "暧昧"},
        "modules": {"user_Nickname": "测试用户"},
        "runtime": {"model_ids": [model_id], "summary_interval": 2},
    }
    conv_id = db.create_conversation(model_id, config)
    for turn, total in enumerate(totals, start=1):
        db.insert_turn_result(
            conv_id,
            {
                "turn": turn,
                "user_input": f"用户输入{turn}",
                "ai_output": f"AI输出{turn}",
                "dialogue_summary": "摘要" if turn > 1 else "",
                "model_id": model_id,
            },
        )
        db.update_turn_scores(
            conv_id,
            turn,
            {
                "persona_fidelity": total,
                "narrative_immersion": total - 0.1,
                "emotional_tension": total - 0.2,
                "boundary_memory": total - 0.3,
                "format_compliance": total - 0.4,
                "context_coherence": total - 0.5,
                "mapped_total": total,
                "reasoning": "测试打分",
                "success": True,
            },
        )
        if manual_score is not None:
            db.update_manual_score(conv_id, turn, manual_score, "人工评语")
    db.update_conversation_status(conv_id, "completed")
    return conv_id


def create_unscored_conversation(
    role_name: str,
    turn_count: int = 1,
    *,
    model_id: str = "doubao-pro",
    prompt_file: str = "test_prompt.md",
    runtime: dict | None = None,
):
    config = {
        "prompt_file": prompt_file,
        "character": {
            "Role_Nickname": role_name,
            "gender": "男",
            "personality": "冷静",
        },
        "context": {"relationship": "暧昧"},
        "modules": {"user_Nickname": "测试用户"},
        "runtime": {
            "model_ids": [model_id],
            "summary_interval": 2,
            **(runtime or {}),
        },
    }
    conv_id = db.create_conversation(model_id, config)
    for turn in range(1, turn_count + 1):
        db.insert_turn_result(
            conv_id,
            {
                "turn": turn,
                "user_input": f"用户输入{turn}",
                "ai_output": f"AI输出{turn}",
                "dialogue_summary": "摘要" if turn > 1 else "",
                "model_id": model_id,
            },
        )
    return conv_id, config


class FakeScoringService:
    def is_available(self, model_id=None):
        return True

    def get_last_error(self):
        return ""

    def get_scoring_prompts(self):
        return ["长文模式"]

    def get_dimensions(self, model_id=None):
        return {}

    def resolve_scoring_thinking_effort(
        self,
        model_id,
        scoring_thinking_enabled,
        scoring_thinking_effort,
        runtime_scoring_thinking_enabled=None,
    ):
        enabled = (
            scoring_thinking_enabled
            if scoring_thinking_enabled is not None
            else runtime_scoring_thinking_enabled
        )
        if enabled is False:
            return "disabled"
        return scoring_thinking_effort or "high"

    async def score_turn(self, turn_data, **kwargs):
        return {
            "success": True,
            "scores": {
                "persona_fidelity": 9.1,
                "narrative_immersion": 8.9,
                "emotional_tension": 8.8,
                "boundary_memory": 9.0,
                "format_compliance": 9.2,
                "context_coherence": 8.7,
            },
            "mapped_total": 9.0,
            "reasoning": f"fake turn scoring: {turn_data.get('turn')}",
        }

    async def score_rows(self, rows, on_progress=None, **kwargs):
        scored = []
        for index, row in enumerate(rows, start=1):
            result = {
                **row,
                "success": True,
                "scores": {
                    "persona_fidelity": 9.1,
                    "narrative_immersion": 8.9,
                    "emotional_tension": 8.8,
                    "boundary_memory": 9.0,
                    "format_compliance": 9.2,
                    "context_coherence": 8.7,
                },
                "mapped_total": 9.0,
                "reasoning": "fake scoring",
            }
            scored.append(result)
            if on_progress:
                await on_progress({"current": index, "total": len(rows)})
        return scored


class FakeConversationService:
    async def run_conversation(
        self,
        conv_id: str,
        config: dict,
        turns: list[str],
        model_id: str = "doubao-pro",
        model_mini: str = "doubao-mini",
        summary_interval: int | None = None,
        dry_run: bool = False,
        on_turn_complete=None,
        on_turn_start=None,
    ):
        db.update_conversation_status(conv_id, "running")
        existing = db.get_conversation(conv_id) or {}
        completed_turns = len(existing.get("results", []))
        for offset, user_input in enumerate(turns, start=1):
            turn_index = completed_turns + offset
            if on_turn_start:
                await on_turn_start(turn_index, completed_turns + len(turns), user_input)
            turn_data = {
                "turn": turn_index,
                "user_input": user_input,
                "ai_output": f"模拟AI输出{turn_index}",
                "dialogue_summary": "",
                "msg_count": 3,
                "word_count": 640,
                "input_tokens": 128,
                "output_tokens": 256,
                "latency_s": 0.1,
                "has_deep_injection": False,
                "has_style_isolation": False,
                "has_cooldown_reinject": False,
                "token_trim_level": 0,
                "quality_retries": 0,
                "messages_snapshot": [],
                "model_id": model_id,
            }
            db.insert_turn_result(conv_id, turn_data)
            if on_turn_complete:
                await on_turn_complete(turn_data)
        db.update_conversation_status(conv_id, "completed")
        return db.get_conversation(conv_id)["results"]


class ControlledConversationService:
    def __init__(self, pause_gate: threading.Event | None = None):
        self.pause_gate = pause_gate or threading.Event()
        self.after_first_turn = threading.Event()

    async def run_conversation(
        self,
        conv_id: str,
        config: dict,
        turns: list[str],
        model_id: str = "doubao-pro",
        model_mini: str = "doubao-mini",
        summary_interval: int | None = None,
        dry_run: bool = False,
        on_turn_complete=None,
        on_turn_start=None,
    ):
        db.update_conversation_status(conv_id, "running")
        existing = db.get_conversation(conv_id) or {}
        completed_turns = len(existing.get("results", []))
        for offset, user_input in enumerate(turns, start=1):
            ctrl = task_control_module.get(conv_id)
            if ctrl:
                await ctrl.checkpoint()
            turn_index = completed_turns + offset
            if on_turn_start:
                await on_turn_start(turn_index, completed_turns + len(turns), user_input)
            turn_data = {
                "turn": turn_index,
                "user_input": user_input,
                "ai_output": f"受控AI输出{turn_index}",
                "dialogue_summary": "",
                "msg_count": 3,
                "word_count": 640,
                "input_tokens": 128,
                "output_tokens": 256,
                "latency_s": 0.1,
                "has_deep_injection": False,
                "has_style_isolation": False,
                "has_cooldown_reinject": False,
                "token_trim_level": 0,
                "quality_retries": 0,
                "messages_snapshot": [],
                "model_id": model_id,
            }
            db.insert_turn_result(conv_id, turn_data)
            if on_turn_complete:
                await on_turn_complete(turn_data)
            if turn_index == 1:
                self.after_first_turn.set()
                while not self.pause_gate.is_set():
                    await asyncio.sleep(0.05)
        db.update_conversation_status(conv_id, "completed")
        return db.get_conversation(conv_id)["results"]


class ControlledScoringService:
    def __init__(self, pause_gate: threading.Event | None = None):
        self.pause_gate = pause_gate or threading.Event()
        self.after_first_turn = threading.Event()
        self.summary_calls = 0

    def set_max_workers(self, _value):
        return 2

    def is_available(self, model_id=None):
        return True

    def get_last_error(self):
        return ""

    def resolve_scoring_thinking_effort(
        self,
        model_id,
        scoring_thinking_enabled,
        scoring_thinking_effort,
        runtime_scoring_thinking_enabled=None,
    ):
        return scoring_thinking_effort or "high"

    async def score_conversation(self, conv_id, results, config, on_progress=None, max_workers=None):
        scored = []
        for index, result in enumerate(results, start=1):
            ctrl = task_control_module.get(f"score_{conv_id}")
            if ctrl:
                await ctrl.checkpoint()
            payload = {
                "turn": result.get("turn", index),
                "scores": {
                    "persona_fidelity": 9,
                    "narrative_immersion": 8,
                    "emotional_tension": 8,
                    "boundary_memory": 8,
                    "format_compliance": 9,
                },
                "mapped_total": 8.8,
                "reasoning": f"受控评分{index}",
                "success": True,
                "score_status": "scored",
            }
            scored.append(payload)
            if on_progress:
                await on_progress(
                    {
                        "type": "score_progress",
                        "turn": payload["turn"],
                        "current": index,
                        "total": len(results),
                        "score": payload["mapped_total"],
                        "success": True,
                    }
                )
            if index == 1:
                self.after_first_turn.set()
                while not self.pause_gate.is_set():
                    await asyncio.sleep(0.05)
        return scored

    async def generate_ai_summary(
        self,
        scored_items,
        config,
        model_id=None,
        prompt_version=None,
        conversation_id="",
    ):
        self.summary_calls += 1
        return {
            "markdown": "# 评分摘要\n\n已评分 1 / 2",
            "model_id": model_id or "qwen-plus",
            "prompt_version": prompt_version or "summary_prompt.md",
            "report_title": "测试摘要",
            "cached": False,
        }


def test_scoring_service_resolves_model_alias_to_api_name():
    service = scoring_service_module.ScoringService()
    assert_true(
        service._resolve_scoring_model_id("doubao-pro") == "doubao-seed-2-0-pro-260215",
        "doubao-pro 未解析到真实评分 endpoint",
    )
    assert_true(
        service._resolve_scoring_model_id("qwen-plus") == "qwen-plus",
        "qwen-plus 在评分链路不应被错误改写为不存在的 endpoint",
    )
    assert_true(
        service._resolve_scoring_model_id("non-existent-model") == "non-existent-model",
        "未知模型不应被错误改写",
    )


def test_scoring_service_openai_client_disables_sdk_retries(monkeypatch: pytest.MonkeyPatch):
    captured = {}

    class FakeOpenAI:
        def __init__(self, **kwargs):
            captured.update(kwargs)

    monkeypatch.setitem(sys.modules, "openai", types.SimpleNamespace(OpenAI=FakeOpenAI))
    service = scoring_service_module.ScoringService()
    service._resolved_scoring_api_keys = ["test-key"]
    service._resolved_scoring_api_key = "test-key"
    service._resolved_scoring_base_url = "https://example.invalid/v1"
    monkeypatch.setattr(service, "_ensure_loaded", lambda require_api_key=True, model_id=None: None)

    service._get_client(timeout_s=17, model_id="qwen-plus")

    assert_true(captured.get("timeout") == 17, f"评分客户端 timeout 未透传: {captured}")
    assert_true(captured.get("max_retries") == 0, f"评分客户端未禁用 SDK 自动重试: {captured}")


def test_model_adapter_provider_loading_is_thread_safe(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    provider_dir = tmp_path / "providers"
    provider_dir.mkdir(parents=True, exist_ok=True)
    (provider_dir / "base.py").write_text(
        """
class ProviderResult:
    def __init__(self, content="", success=True, error="", input_tokens=0, output_tokens=0, latency=0.0):
        self.content = content
        self.success = success
        self.error = error
        self.input_tokens = input_tokens
        self.output_tokens = output_tokens
        self.latency = latency

class BaseProvider:
    def __init__(self, model_config):
        self.config = model_config
        self.parameters = model_config.get("parameters", {})

    def call_with_retry(self, messages, **kwargs):
        return ProviderResult(content="ok")
""".strip(),
        encoding="utf-8",
    )
    (provider_dir / "volcengine.py").write_text(
        """
import time
from .base import BaseProvider, ProviderResult

time.sleep(0.2)

class VolcEngineProvider(BaseProvider):
    def call(self, messages, **kwargs):
        return ProviderResult(content="ok")
""".strip(),
        encoding="utf-8",
    )

    monkeypatch.setattr(model_adapter_module, "_PROVIDER_DIR", provider_dir)
    monkeypatch.setattr(model_adapter_module, "_PROVIDER_BASE_CACHE", None)
    for name in list(sys.modules):
        if name.startswith(f"{model_adapter_module._PROVIDER_PACKAGE_NAME}.") or name == model_adapter_module._PROVIDER_PACKAGE_NAME:
            sys.modules.pop(name, None)

    errors: list[str] = []

    def worker():
        try:
            mod = model_adapter_module._load_provider_module("volcengine")
            base_cls, _ = model_adapter_module._get_provider_base_types()
            provider_cls = getattr(mod, "VolcEngineProvider", None)
            if not isinstance(provider_cls, type) or not issubclass(provider_cls, base_cls):
                raise AssertionError("并发加载后 provider 类不可用")
        except Exception as exc:  # pragma: no cover - 失败分支才会触发
            errors.append(str(exc))

    t1 = threading.Thread(target=worker)
    t2 = threading.Thread(target=worker)
    t1.start()
    t2.start()
    t1.join()
    t2.join()

    assert_true(not errors, f"Provider 并发加载不安全: {errors}")


class BlockingConversationService:
    def __init__(self, release_event: threading.Event):
        self.release_event = release_event

    async def run_conversation(
        self,
        conv_id: str,
        config: dict,
        turns: list[str],
        model_id: str = "doubao-pro",
        model_mini: str = "doubao-mini",
        summary_interval: int | None = None,
        dry_run: bool = False,
        on_turn_complete=None,
        on_turn_start=None,
    ):
        db.update_conversation_status(conv_id, "running")
        while not self.release_event.is_set():
            await asyncio.sleep(0.05)

        if on_turn_start:
            await on_turn_start(1, len(turns), turns[0] if turns else "排队测试")
        turn_data = {
            "turn": 1,
            "user_input": turns[0] if turns else "排队测试",
            "ai_output": "模拟排队输出",
            "dialogue_summary": "",
            "msg_count": 3,
            "word_count": 360,
            "input_tokens": 64,
            "output_tokens": 128,
            "latency_s": 0.1,
            "has_deep_injection": False,
            "has_style_isolation": False,
            "has_cooldown_reinject": False,
            "token_trim_level": 0,
            "quality_retries": 0,
            "messages_snapshot": [],
            "model_id": model_id,
        }
        db.insert_turn_result(conv_id, turn_data)
        if on_turn_complete:
            await on_turn_complete(turn_data)
        db.update_conversation_status(conv_id, "completed")
        return db.get_conversation(conv_id)["results"]


class FakeModelResult:
    def __init__(
        self,
        content: str,
        input_tokens: int = 32,
        output_tokens: int = 64,
        latency_s: float = 0.2,
        success: bool = True,
        error: str = "",
    ):
        self.content = content
        self.input_tokens = input_tokens
        self.output_tokens = output_tokens
        self.latency_s = latency_s
        self.success = success
        self.error = error


class CapturingModelAdapter:
    def __init__(self):
        self.calls: list[dict] = []
        self.response_content = (
            "*图书馆的灯光落在长桌边缘，他合上手里的书，目光从纸页间慢慢抬起，"
            "像是早就知道你会在这个时间出现。窗外的夜色压得很低，玻璃上映着两个人"
            "靠得不远不近的影子，安静里却有一种被刻意放轻的呼吸声。*\n\n"
            "「今晚见。」\n\n"
            "*他把这句话说得很淡，像是随口确认，又像把整晚的耐心都压进了短短三个字里。"
            "指尖在书脊上轻轻点了一下后停住，没有催促，只把身侧的位置让出半步，视线仍稳稳"
            "落在你脸上，像要先看清你到底是在犹豫，还是故意吊着他。空气里有纸张和冷木头混在"
            "一起的味道，他终于低声补了一句：「如果你点头，我就陪你走到闭馆。」语气仍旧克制，"
            "却把那点藏不住的在意暴露得很彻底。你还没回答，他已经替你把椅子轻轻拉开，像是把"
            "催促藏进了体贴里，连等待都维持着体面，只留下那双始终没移开的眼睛，安静却执拗地"
            "守着你的下一句回应。*"
        )

    def chat(self, model_id: str, messages: list[dict], **kwargs):
        self.calls.append(
            {
                "model_id": model_id,
                "messages": list(messages),
                "kwargs": dict(kwargs),
            }
        )
        return FakeModelResult(content=self.response_content)


class SlowCapturingModelAdapter(CapturingModelAdapter):
    def __init__(self, delay_s: float = 0.8):
        super().__init__()
        self.delay_s = delay_s

    def chat(self, model_id: str, messages: list[dict], **kwargs):
        time.sleep(self.delay_s)
        return super().chat(model_id, messages, **kwargs)


def test_batch_create_returns_quickly_and_fourth_request_queues(client: TestClient):
    original_auto_score = scoring_router.score_turn_if_available
    original_enqueue_live_score = scoring_router.enqueue_live_score_turn
    original_max_concurrency = conversations_router._MAX_CONCURRENT_CONVERSATIONS

    async def _skip_auto_score(*args, **kwargs):
        return None

    async def _skip_live_score_enqueue(*args, **kwargs):
        return False

    scoring_router.score_turn_if_available = _skip_auto_score
    scoring_router.enqueue_live_score_turn = _skip_live_score_enqueue
    # 用较小并发上限稳定复现排队逻辑（避免默认并发变更导致用例退化）
    conversations_router._MAX_CONCURRENT_CONVERSATIONS = 3
    conversations_router._conv_service = ConversationService(
        model_adapter=SlowCapturingModelAdapter(delay_s=0.8),
        prompt_service=PromptService(),
    )
    conversations_router._running_conversations.clear()
    conversations_router._queued_conversations.clear()
    conversations_router._ws_connections.clear()

    payload = {
        "preset_id": "xiaoJingYan",
        "model_id": "doubao-pro",
        "dry_run": False,
        "turns": ["后台创建不应被阻塞"],
    }
    timings: list[float] = []
    conv_ids: list[str] = []

    try:
        for index in range(3):
            start = time.perf_counter()
            response = client.post("/api/conversations", json=payload)
            timings.append(time.perf_counter() - start)
            assert_true(response.status_code == 200, response.text)
            body = response.json()
            conv_ids.append(body["id"])
            assert_true(body["status"] == "pending", f"第 {index + 1} 个请求未立即占槽")

        queued_start = time.perf_counter()
        queued_response = client.post("/api/conversations", json=payload)
        timings.append(time.perf_counter() - queued_start)
        assert_true(queued_response.status_code == 200, queued_response.text)
        queued_body = queued_response.json()
        conv_ids.append(queued_body["id"])
        assert_true(queued_body["status"] == "queued", "第 4 个慢请求未进入排队")
        assert_true(queued_body["queue_position"] == 1, "第 4 个慢请求排队位置错误")
        assert_true(
            max(timings) < 0.7,
            f"创建接口被后台生成阻塞: {[round(item, 3) for item in timings]}",
        )

        for conv_id in conv_ids:
            wait_for_conversation_status(conv_id, "completed", timeout=12.0)
    finally:
        deadline = time.time() + 12.0
        while time.time() < deadline:
            if (
                not conversations_router._running_conversations
                and not conversations_router._queued_conversations
            ):
                break
            time.sleep(0.1)
        scoring_router.score_turn_if_available = original_auto_score
        scoring_router.enqueue_live_score_turn = original_enqueue_live_score
        conversations_router._MAX_CONCURRENT_CONVERSATIONS = original_max_concurrency
        conversations_router._conv_service = None
        conversations_router._running_conversations.clear()
        conversations_router._queued_conversations.clear()
        conversations_router._ws_connections.clear()


def test_builtin_preset_and_summary_interval(client: TestClient):
    response = client.post(
        "/api/conversations",
        json={
            "preset_id": "xiaoJingYan",
            "dry_run": True,
            "summary_interval": 2,
            "turns": ["第一轮", "第二轮", "第三轮"],
        },
    )
    assert_true(response.status_code == 200, f"内置预设创建失败: {response.text}")
    conv_id = response.json()["id"]
    conversation = wait_for_turns(conv_id, 3)
    assert_true(
        conversation["config"]["runtime"]["summary_interval"] == 2,
        "summary_interval 未写入运行时配置",
    )
    assert_true(
        conversation["results"][2]["dialogue_summary"] != "",
        "summary_interval=2 未在第 3 轮前注入摘要",
    )


def test_multi_turn_window_summary_and_profile_refresh_dry_run(client: TestClient):
    """验证：10轮滑动窗口 + 第11轮摘要注入 + 第21轮使用更新画像（dry-run 也可验证链路）"""
    turns = [f"U{i}" for i in range(1, 22)]
    response = client.post(
        "/api/conversations",
        json={
            "model_id": "doubao-pro",
            "dry_run": True,
            "summary_interval": 10,
            "injection_depth": 4,
            "turns": turns,
            "character": {
                "Role_Nickname": "多轮测试角色",
                "personality": "稳",
                "gender": "男",
            },
            "context": {
                "relationship": "暧昧",
                "scene": "客厅",
                "time_period": "夜晚",
                "user_nickname": "小鹿",
            },
            "modules": {
                "dialogueStartPrompt": "PROFILE_V0",
            },
        },
    )
    assert_true(response.status_code == 200, response.text)
    conv_id = response.json()["id"]
    conversation = wait_for_turns(conv_id, 21, timeout=30.0)
    results = conversation.get("results") or []
    assert_true(len(results) == 21, "未生成完整 21 轮结果")

    # 第11轮应已注入摘要（第10轮生成，供第11轮使用）
    turn_11 = results[10]
    assert_true(
        "=== 之前剧情摘要 ===" in str(turn_11.get("dialogue_summary", "")),
        "第11轮 dialogue_summary 未包含摘要注入标记",
    )
    messages_11 = (turn_11.get("request_payload_snapshot") or {}).get("messages") or []
    memory_snapshot_11 = (turn_11.get("request_payload_snapshot") or {}).get("memory_context_snapshot") or {}
    assert_true(
        any("【历史对话摘要】" in str(m.get("content", "")) for m in messages_11),
        "第11轮 messages 未包含独立 memory block",
    )
    assert_true(
        "=== 之前剧情摘要 ===" in str(memory_snapshot_11.get("dialogue_summary", "")),
        "第11轮 memory_context_snapshot 未记录摘要",
    )
    assert_true(
        str((turn_11.get("request_payload_snapshot") or {}).get("summary_source", "")) == "completed",
        "第11轮 summary_source 未标记为 completed",
    )

    # 第21轮消息不应再包含第1轮用户输入（滑动窗口：仅保留最近10轮真实历史）
    turn_21 = results[20]
    messages_21 = (turn_21.get("request_payload_snapshot") or {}).get("messages") or []
    assert_true(
        not any(str(m.get("content", "")).strip() == "U1" for m in messages_21),
        "第21轮 messages 仍包含第1轮输入，滑动窗口未生效",
    )

    # 第20轮应触发画像抽取并写回，且第21轮 system_prompt 应使用新画像（运行时 bundle 刷新）
    updated_profile = str(conversation.get("config", {}).get("modules", {}).get("dialogueStartPrompt", ""))
    assert_true("【用户画像信息】" in updated_profile, "画像抽取结果未写回 config.modules.dialogueStartPrompt")
    full_system_1 = str(((results[0].get("request_payload_snapshot") or {}).get("system_prompt", "")))
    full_system_21 = str(((turn_21.get("request_payload_snapshot") or {}).get("system_prompt", "")))
    memory_snapshot_21 = (turn_21.get("request_payload_snapshot") or {}).get("memory_context_snapshot") or {}
    assert_true("PROFILE_V0" not in full_system_1, "第1轮 system_prompt 仍混入初始画像，未剥离 memory block")
    assert_true("【用户画像信息】" not in full_system_21, "第21轮 system_prompt 仍混入画像文本")
    assert_true("PROFILE_V0" not in full_system_21, "第21轮 system_prompt 仍包含旧画像")
    assert_true(
        "【用户画像信息】" in str(memory_snapshot_21.get("dialogueStartPrompt", "")),
        "第21轮 memory_context_snapshot 未包含更新后的画像文本",
    )


def test_custom_preset_reuse(client: TestClient):
    create_response = client.post(
        "/api/presets",
        json={
            "name": "自定义角色",
            "type": "测试",
            "config": {
                "character": {"Role_Nickname": "自定义", "personality": "稳"},
                "context": {"relationship": "朋友"},
                "modules": {},
            },
        },
    )
    assert_true(create_response.status_code == 200, create_response.text)
    preset_id = create_response.json()["id"]

    response = client.post(
        "/api/conversations",
        json={
            "preset_id": preset_id,
            "dry_run": True,
            "turns": ["你好"],
        },
    )
    assert_true(response.status_code == 200, f"自定义预设对话创建失败: {response.text}")
    conv_id = response.json()["id"]
    conversation = wait_for_turns(conv_id, 1)
    assert_true(conversation["preset_id"] == preset_id, "自定义预设未写入 conversation")


def test_preset_flags_and_delete_contract(client: TestClient):
    list_response = client.get("/api/presets")
    assert_true(list_response.status_code == 200, list_response.text)
    presets = list_response.json()["presets"]
    builtin = next((item for item in presets if item.get("id") == "xiaoJingYan"), None)
    assert_true(builtin is not None, "预设列表缺少内置模板")
    assert_true(builtin.get("is_builtin") is True, "内置模板未标记 is_builtin")

    builtin_detail = client.get("/api/presets/xiaoJingYan")
    assert_true(builtin_detail.status_code == 200, builtin_detail.text)
    assert_true(builtin_detail.json().get("is_builtin") is True, "内置模板详情未标记 is_builtin")

    create_response = client.post(
        "/api/presets",
        json={
            "name": "待删除模板",
            "type": "测试",
            "config": {
                "character": {"Role_Nickname": "待删除", "personality": "稳"},
                "context": {"relationship": "朋友"},
                "modules": {},
            },
        },
    )
    assert_true(create_response.status_code == 200, create_response.text)
    preset_id = create_response.json()["id"]

    custom_detail = client.get(f"/api/presets/{preset_id}")
    assert_true(custom_detail.status_code == 200, custom_detail.text)
    assert_true(custom_detail.json().get("is_builtin") is False, "自定义模板详情 is_builtin 错误")

    builtin_delete = client.delete("/api/presets/xiaoJingYan")
    assert_true(builtin_delete.status_code == 403, builtin_delete.text)

    delete_response = client.delete(f"/api/presets/{preset_id}")
    assert_true(delete_response.status_code == 200, delete_response.text)
    missing_response = client.get(f"/api/presets/{preset_id}")
    assert_true(missing_response.status_code == 404, "删除后模板仍可读取")

    linked_create = client.post(
        "/api/presets",
        json={
            "name": "已引用模板",
            "type": "测试",
            "config": {
                "character": {"Role_Nickname": "历史模板", "personality": "稳"},
                "context": {"relationship": "朋友"},
                "modules": {},
            },
        },
    )
    assert_true(linked_create.status_code == 200, linked_create.text)
    linked_preset_id = linked_create.json()["id"]

    conv_response = client.post(
        "/api/conversations",
        json={
            "preset_id": linked_preset_id,
            "dry_run": True,
            "turns": ["删除后历史会话仍应保留"],
        },
    )
    assert_true(conv_response.status_code == 200, conv_response.text)
    linked_conv_id = conv_response.json()["id"]
    linked_conversation = wait_for_turns(linked_conv_id, 1)
    assert_true(
        linked_conversation["preset_id"] == linked_preset_id,
        "测试前置条件失败：历史会话未绑定模板",
    )

    linked_delete = client.delete(f"/api/presets/{linked_preset_id}")
    assert_true(linked_delete.status_code == 200, linked_delete.text)

    linked_detail = client.get(f"/api/conversations/{linked_conv_id}")
    assert_true(linked_detail.status_code == 200, linked_detail.text)
    assert_true(
        linked_detail.json().get("preset_id") in (None, ""),
        "删除模板后历史会话仍保留旧 preset_id",
    )


def test_interactive_runtime_defaults_follow_v55_contract(client: TestClient):
    payload = {
        "model_id": "doubao-pro",
        "character": {"Role_Nickname": "默认值角色", "personality": "理性沉稳"},
        "context": {"relationship": "朋友", "scene": "书房"},
        "modules": {},
    }
    create_response = client.post("/api/conversations/interactive", json=payload)
    assert_true(create_response.status_code == 200, create_response.text)
    conv_id = create_response.json()["id"]

    detail_response = client.get(f"/api/conversations/{conv_id}")
    assert_true(detail_response.status_code == 200, detail_response.text)
    runtime = detail_response.json()["config"]["runtime"]
    assert_true(
        runtime.get("summary_interval") == DEFAULT_SUMMARY_INTERVAL,
        "交互式默认 summary_interval 未对齐 v5.5",
    )
    assert_true(
        runtime.get("injection_depth") == DEFAULT_INJECTION_DEPTH,
        "交互式默认 injection_depth 未对齐 v5.5",
    )
    assert_true(runtime.get("thinking_enabled") is None, "未显式设置时不应写死 thinking_enabled")
    assert_true(runtime.get("thinking_effort", "") in ("", None), "未显式设置时不应写死 thinking_effort")


def test_prompt_version_binding_and_validation(client: TestClient):
    conversations_router._conv_service = FakeConversationService()
    try:
        payload = {
            "model_id": "doubao-pro",
            "prompt_version": DEFAULT_PROMPT_FILE,
            "dry_run": True,
            "turns": ["提示词版本测试"],
            "character": {
                "Role_Nickname": "提示词角色",
                "personality": "冷静",
                "gender": "男",
            },
            "context": {
                "relationship": "暧昧",
                "scene": "客厅",
                "time_period": "晚上",
                "user_nickname": "小鹿",
            },
            "modules": {},
        }
        response = client.post("/api/conversations", json=payload)
        assert_true(response.status_code == 200, response.text)
        conv_id = response.json()["id"]
        conversation = wait_for_turns(conv_id, 1)
        assert_true(
            conversation["config"]["prompt_file"] == DEFAULT_PROMPT_FILE,
            "prompt_version 未映射到真实执行的 prompt_file",
        )
        assert_true(
            conversation["prompt_version"] == DEFAULT_PROMPT_FILE,
            "conversation.prompt_version 未保存真实版本名",
        )

        invalid_payload = dict(payload)
        invalid_payload["prompt_version"] = "does_not_exist.md"
        invalid_response = client.post("/api/conversations", json=invalid_payload)
        assert_true(invalid_response.status_code == 400, invalid_response.text)
    finally:
        conversations_router._conv_service = None


def test_chinese_export_and_manual_score(client: TestClient):
    conv_id = create_scored_conversation("萧钟言", [8.9, 9.3], manual_score=8.5)

    result_response = client.get(f"/api/scoring/{conv_id}/results")
    assert_true(result_response.status_code == 200, result_response.text)
    result_payload = result_response.json()
    assert_true(
        result_payload["turns"][0]["manual_star_score"] == 8.5,
        "评分结果未返回人工分",
    )
    assert_true(
        result_payload["summary"]["manual_avg"] == 8.5,
        "评分结果未统计人工均分",
    )

    export_response = client.get(f"/api/export/{conv_id}")
    assert_true(export_response.status_code == 200, export_response.text)

    scoring_export = client.get(f"/api/scoring/{conv_id}/export")
    assert_true(scoring_export.status_code == 200, scoring_export.text)
    workbook = openpyxl.load_workbook(io.BytesIO(scoring_export.content))
    headers = [cell.value for cell in workbook.active[1]]
    workbook.close()
    assert_true("人工星级评分" in headers, "评分导出缺少人工星级评分列")
    assert_true("人工评语" in headers, "评分导出缺少人工评语列")


def test_scoring_upload_and_test_file_import(client: TestClient):
    scoring_router._scoring = FakeScoringService()
    template_response = client.get("/api/scoring/template/download")
    assert_true(template_response.status_code == 200, template_response.text)
    template_workbook = openpyxl.load_workbook(io.BytesIO(template_response.content))
    template_headers = [cell.value for cell in template_workbook.active[1]]
    template_workbook.close()
    assert_true(
        {"session_id", "turn_order", "用户输入", "AI输出"}.issubset(set(template_headers)),
        f"打分模板表头缺失: {template_headers}",
    )

    workbook_bytes = create_workbook_bytes(
        ["用户输入", "AI输出"],
        [["你好", "你好呀"]],
    )

    import_response = client.post(
        "/api/scoring/test-file/import",
        files={
            "file": (
                "sample.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert_true(import_response.status_code == 200, import_response.text)
    assert_true(import_response.json()["rows"] == 1, "test-file/import 行数错误")

    workbook_bytes.seek(0)
    upload_response = client.post(
        "/api/scoring/upload",
        files={
            "file": (
                "sample.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert_true(upload_response.status_code == 200, upload_response.text)
    payload = upload_response.json()
    assert_true(payload["status"] == "completed", "upload 评分未完成")
    download_response = client.get(payload["download_url"])
    assert_true(download_response.status_code == 200, "upload 结果文件不可下载")
    scoring_router._scoring = None


def test_template_downloads_use_business_filenames_and_v26_headers(client: TestClient):
    variable_response = client.get("/api/configs/variables/template/export")
    assert_true(variable_response.status_code == 200, variable_response.text)
    variable_cd = variable_response.headers.get("content-disposition", "")
    assert_true("variables_template_" not in variable_cd, f"变量模板文件名仍为技术名: {variable_cd}")
    variable_workbook = openpyxl.load_workbook(io.BytesIO(variable_response.content))
    variable_sheet = variable_workbook.active
    variable_names = [
        variable_sheet.cell(row=row_index, column=2).value
        for row_index in range(2, variable_sheet.max_row + 1)
    ]
    variable_workbook.close()
    assert_true(
        {"moments", "monthly_schedule", "last_cst_type", "完整时间信息"}.issubset(set(variable_names)),
        f"变量模板缺少 v2.6 变量: {variable_names}",
    )

    config_template_response = client.get("/api/configs/export-template")
    assert_true(config_template_response.status_code == 200, config_template_response.text)
    config_cd = config_template_response.headers.get("content-disposition", "")
    assert_true("conversation_template_" not in config_cd, f"对话模板文件名仍为技术名: {config_cd}")
    config_workbook = openpyxl.load_workbook(io.BytesIO(config_template_response.content))
    config_headers = [cell.value for cell in config_workbook.active[1]]
    config_workbook.close()
    assert_true(
        {"weekDay", "moments", "monthly_schedule", "last_cst_type", "完整时间信息"}.issubset(set(config_headers)),
        f"对话模板表头缺少 v2.6 字段: {config_headers}",
    )

    scoring_template_response = client.get("/api/scoring/template/download")
    assert_true(scoring_template_response.status_code == 200, scoring_template_response.text)
    scoring_cd = scoring_template_response.headers.get("content-disposition", "")
    assert_true("打分模板.xlsx" not in scoring_cd or "长文模式_打分模板" in scoring_cd, f"打分模板文件名异常: {scoring_cd}")
    scoring_workbook = openpyxl.load_workbook(io.BytesIO(scoring_template_response.content))
    scoring_headers = [cell.value for cell in scoring_workbook.active[1]]
    scoring_workbook.close()
    assert_true(
        {"Role_info_works", "system_module8", "dialogue_summary", "relation_calling", "weekDay", "moments", "monthly_schedule", "last_cst_type", "完整时间信息"}.issubset(set(scoring_headers)),
        f"打分模板表头缺少 v2.6 字段: {scoring_headers}",
    )
    assert_true(
        "relation_info" not in scoring_headers,
        f"打分模板仍包含已废弃的 relation_info 列: {scoring_headers}",
    )


def test_prompt_template_variable_replacement_and_interactive_generate(
    client: TestClient,
    tmp_path: Path,
):
    prompt_path = tmp_path / "custom_prompt.md"
    prompt_path.write_text(
        (
            "你是{{Role_Nickname}}，当前关系是{{relationship}}，所在场景是{{current_scene}}。\n"
            "请用{{personality}}的方式回应。\n"
            "最近朋友圈：{{moments}}\n"
            "月度安排：{{monthly_schedule}}\n"
            "最近互动类型：{{last_cst_type}}\n"
            "完整时间：{{完整时间信息}}\n"
            "<!-- ======================== 以上为 messages[0] role=system 的内容 ======================== -->\n"
        ),
        encoding="utf-8",
    )

    adapter = CapturingModelAdapter()
    conversations_router._conv_service = ConversationService(
        model_adapter=adapter,
        prompt_service=PromptService(),
    )

    try:
        create_response = client.post(
            "/api/conversations/interactive",
            json={
                "model_id": "doubao-pro",
                "model_mini": "doubao-mini",
                "prompt_version": str(prompt_path),
                "summary_interval": 5,
                "injection_depth": 3,
                "temperature": 0.8,
                "top_p": 0.9,
                "character": {
                    "Role_Nickname": "模板角色",
                    "personality": "克制冷静",
                    "gender": "男",
                },
                "context": {
                    "relationship": "暧昧",
                    "scene": "图书馆",
                    "time_period": "夜晚",
                    "season": "春季",
                    "user_nickname": "小鹿",
                },
                "modules": {},
                "custom_variables": {
                    "moments": "她今天发了深夜读书照",
                    "monthly_schedule": "本月最后一周都在准备发布会",
                    "last_cst_type": "晚间文字聊天",
                },
            },
        )
        assert_true(create_response.status_code == 200, create_response.text)
        conv_id = create_response.json()["id"]

        generate_response = client.post(
            f"/api/conversations/{conv_id}/generate",
            json={
                "user_input": "今晚见吗？",
                "model_id": "doubao-pro",
                "temperature": 0.55,
                "top_p": 0.75,
            },
        )
        assert_true(generate_response.status_code == 200, generate_response.text)
        payload = generate_response.json()
        messages = payload["messages_snapshot"]
        system_message = messages[0]["content"]
        request_payload_snapshot = payload["request_payload_snapshot"]
        memory_context_snapshot = request_payload_snapshot.get("memory_context_snapshot", {})

        assert_true("模板角色" in system_message, "提示词中的角色变量未被替换")
        assert_true("暧昧" in system_message and "图书馆" in system_message, "上下文变量未被替换")
        assert_true("她今天发了深夜读书照" not in system_message, "moments 不应再混入 messages[0]")
        assert_true("本月最后一周都在准备发布会" in system_message, "monthly_schedule 自定义变量未注入")
        assert_true("晚间文字聊天" in system_message, "last_cst_type 自定义变量未注入")
        assert_true(" / " in system_message and "完整时间" in system_message, "完整时间信息未注入")
        assert_true("{{" not in system_message, "提示词中仍残留未替换变量")
        assert_true(
            memory_context_snapshot.get("moments") == "她今天发了深夜读书照",
            "moments 自定义变量未进入独立 memory block 快照",
        )
        assert_true(
            any(
                msg.get("role") == "user"
                and "<user_input>今晚见吗？</user_input>" in msg.get("content", "")
                for msg in messages
            ),
            "用户输入未按 XML 包裹进入真实消息栈",
        )
        assert_true(
            payload["ai_output"]
            == QualityGuard().check(adapter.response_content)["processed_text"],
            "单对话生成后的输出未按 v2.6 质量守卫归一化",
        )
        assert_true(
            request_payload_snapshot["custom_variables"]["moments"] == "她今天发了深夜读书照",
            "request_payload_snapshot 未保留自定义变量",
        )
        assert_true(
            request_payload_snapshot["temperature"] == 0.55
            and request_payload_snapshot["top_p"] == 0.75,
            "request_payload_snapshot 未记录运行时采样参数",
        )
        # generate 路径会触发 summary 调度共享 mock adapter，只校验主模型调用；
        # 主模型调用以最后一次 doubao-pro 调用为准。
        primary_calls = [call for call in adapter.calls if call["model_id"] == "doubao-pro"]
        assert_true(
            bool(primary_calls),
            f"未捕获到主模型 doubao-pro 调用: model_ids={[c['model_id'] for c in adapter.calls]}",
        )
        primary_call = primary_calls[-1]
        assert_true(
            primary_call["kwargs"].get("thinking_effort") == "disabled",
            "非 Gemma4 31B 模型的默认 thinking_effort 不应被强制改写",
        )
        assert_true(
            primary_call["kwargs"].get("temperature") == 0.55
            and primary_call["kwargs"].get("top_p") == 0.75,
            "运行时 temperature/top_p 未透传到模型调用",
        )
    finally:
        conversations_router._conv_service = None


def test_interactive_generate_truncates_oversized_single_newline_output(
    client: TestClient,
    tmp_path: Path,
):
    prompt_path = tmp_path / "truncate_prompt.md"
    prompt_path.write_text(
        (
            "你是{{Role_Nickname}}，请严格输出300-500字。\n"
            "<!-- ======================== 以上为 messages[0] role=system 的内容 ======================== -->\n"
        ),
        encoding="utf-8",
    )

    adapter = CapturingModelAdapter()
    adapter.response_content = "\n".join(
        [
            "落地窗外的霓虹在二十九层玻璃上晕开模糊的光斑，萧璟言指尖转着钢笔，刚签完最后一份项目合同。衬衫袖口挽到小臂，腕间那只古董表在台灯下泛着冷光。听到你声音时，笔尖顿了半秒，墨点落在纸角。",
            '**"刚结束。"**',
            "他把文件推到一边，拿起桌角温着的柠檬水喝了一口，喉结动了动。手机屏幕亮着，是下午部门提交的会议纪要，你的名字在参会人列表里排第三，旁边标注了负责整理的待办项。",
            '**"哪个会开这么久？下午的渠道对接会？我让他们压缩到两小时，结果又拖了三个小时？"**',
            "他翻了翻会议记录的最后一页，看到你标注的三个待解决问题，字迹比平时潦草些，末尾的句号都带着点飘。指尖在你名字旁边敲了敲，指腹蹭过纸面上的墨痕。",
            '**"会议室空调开得低，你今天穿的那件针织衫薄，没冻着吧？"**',
            "助理敲门进来拿文件，他抬了抬手示意先放着，目光没离开手机屏幕。听见你说累，眉头微不可察地皱了一下，钢笔在指节间转了个圈停住。",
            '**"我让司机去接你，二十分钟到你楼下。别自己打车，晚高峰不安全。"**',
            "他站起身走到窗边，指尖擦过冰凉的玻璃，下面的车流像一条缓慢流动的光河。停顿了两秒，嗓音比刚才低了些，带着点漫不经心的懒。",
            '**"我让厨房熬了桃胶羹，你上次说想吃的。现在过来，还热着。"**',
        ]
    )
    assert_true(len(adapter.response_content) > 500, "测试前置条件失败：mock 输出未超过500字")
    assert_true("\n\n" not in adapter.response_content, "测试前置条件失败：mock 输出不应包含双换行")

    conversations_router._conv_service = ConversationService(
        model_adapter=adapter,
        prompt_service=PromptService(),
    )

    try:
        create_response = client.post(
            "/api/conversations/interactive",
            json={
                "model_id": "doubao-pro",
                "model_mini": "doubao-mini",
                "prompt_version": str(prompt_path),
                "character": {
                    "Role_Nickname": "萧璟言",
                    "personality": "霸道腹黑",
                    "gender": "男",
                },
                "context": {
                    "relationship": "暧昧",
                    "user_nickname": "小鹿",
                },
                "modules": {},
            },
        )
        assert_true(create_response.status_code == 200, create_response.text)
        conv_id = create_response.json()["id"]

        generate_response = client.post(
            f"/api/conversations/{conv_id}/generate",
            json={
                "user_input": "今天开会累坏了，你还在忙吗？",
                "model_id": "doubao-pro",
            },
        )
        assert_true(generate_response.status_code == 200, generate_response.text)
        payload = generate_response.json()
        expected = QualityGuard().check(adapter.response_content)["processed_text"]

        assert_true(payload["ai_output"] == expected, "交互式生成未使用截断后的质量守卫结果")
        assert_true(300 <= payload["word_count"] <= 500, f"交互式生成字数未被约束到300-500: {payload['word_count']}")

        latest = db.get_conversation(conv_id)
        assert_true(latest is not None, "未读取到会话数据")
        stored_turn = latest["results"][0]
        assert_true(stored_turn["ai_output"] == expected, "落库结果未使用质量守卫后的文本")
        assert_true(300 <= stored_turn["word_count"] <= 500, f"落库字数未被约束到300-500: {stored_turn['word_count']}")
    finally:
        conversations_router._conv_service = None


def test_interactive_generate_closes_dangling_dialogue_line(
    client: TestClient,
    tmp_path: Path,
):
    prompt_path = tmp_path / "dangling_dialogue_prompt.md"
    prompt_path.write_text(
        (
            "你是{{Role_Nickname}}，请严格输出300-500字。\n"
            "<!-- ======================== 以上为 messages[0] role=system 的内容 ======================== -->\n"
        ),
        encoding="utf-8",
    )

    adapter = CapturingModelAdapter()
    adapter.response_content = (
        "办公室的百叶窗半拉着，金色夕阳漏进半格，落在萧璟言面前的红木办公桌上。"
        "他刚结束下午的考察回来，袖口挽到小臂，腕间的古董表在光里转了转冷银色的边。"
        "抬眼扫过你手里皱巴巴的会议记录本，指尖在桌面敲了两下。\n\n"
        '"放我这。"\n\n'
        "他伸手接过你递过来的本子，指节擦过你冰凉的手背，顿了半秒才收回去。"
        "翻到最后一页潦草的字迹处，从笔筒里抽了支钢笔，笔尖划过纸面的声音很轻。"
        "你站在桌前晃了晃，后颈还有点僵。\n\n"
        '"早上给你泡的洋甘菊茶喝了？"\n\n'
        "他头都没抬，钢笔在你写错的数字上圈了个圈，旁边补了两行工整的批注。"
        "办公室里只剩笔尖划纸的声响，墙角的落地钟滴答走了三下。"
        "他把改了小半的纪要推回你面前，指腹按在你没写全的备注那行。\n\n"
        '"下午董事会的决议不用记太细，过两天我让助理把最终版发你。"\n\n'
        "他靠回椅背，抬眼看向你，喉结动了动。桌上刚温好的普洱冒着浅白的热气，"
        "飘过来一点陈香。他伸手把茶杯往你那边推了推，杯底磕在桌面发出轻响。\n\n"
        '**"周五晚上我约了人试新开的火锅店，你跟我一起去。'
    )
    assert_true(len(adapter.response_content) >= 300, "测试前置条件失败：mock 输出应不少于300字")

    conversations_router._conv_service = ConversationService(
        model_adapter=adapter,
        prompt_service=PromptService(),
    )

    try:
        create_response = client.post(
            "/api/conversations/interactive",
            json={
                "model_id": "doubao-pro",
                "model_mini": "doubao-mini",
                "prompt_version": str(prompt_path),
                "character": {
                    "Role_Nickname": "萧璟言",
                    "personality": "霸道腹黑",
                    "gender": "男",
                },
                "context": {
                    "relationship": "暧昧",
                    "user_nickname": "小鹿",
                },
                "modules": {},
            },
        )
        assert_true(create_response.status_code == 200, create_response.text)
        conv_id = create_response.json()["id"]

        generate_response = client.post(
            f"/api/conversations/{conv_id}/generate",
            json={
                "user_input": "今天会议纪要还没整理完，脑子有点乱。",
                "model_id": "doubao-pro",
            },
        )
        assert_true(generate_response.status_code == 200, generate_response.text)
        payload = generate_response.json()
        expected_tail = "周五晚上我约了人试新开的火锅店，你跟我一起去。"

        assert_true(expected_tail in payload["ai_output"], "接口返回未自动闭合残缺对白")
        assert_true(
            '**"' not in payload["ai_output"] and '"**' not in payload["ai_output"],
            "接口返回仍残留旧版对白标记",
        )

        latest = db.get_conversation(conv_id)
        assert_true(latest is not None, "未读取到会话数据")
        stored_turn = latest["results"][0]
        assert_true(expected_tail in stored_turn["ai_output"], "落库结果未自动闭合残缺对白")
        assert_true(
            '**"' not in stored_turn["ai_output"] and '"**' not in stored_turn["ai_output"],
            "数据库中仍残留旧版对白标记",
        )
    finally:
        conversations_router._conv_service = None


def test_last_conversation_type_is_auto_injected_from_previous_session_mode(
    client: TestClient,
    tmp_path: Path,
):
    voice_prompt = tmp_path / "voice_prompt.md"
    voice_prompt.write_text(
        (
            "---L4---\n"
            "- **对话场景**: 你正在与用户1V1语音聊天\n"
            "- **沟通状态**: 你与用户{{last_cst_type}}\n"
            "<!-- ======================== 以上为 messages[0] role=system 的内容 ======================== -->\n"
        ),
        encoding="utf-8",
    )
    text_prompt = tmp_path / "text_prompt.md"
    text_prompt.write_text(
        (
            "---L4---\n"
            "- **对话场景**: 你正在与用户文字聊天\n"
            "- **沟通状态**: 你与用户{{last_cst_type}}\n"
            "<!-- ======================== 以上为 messages[0] role=system 的内容 ======================== -->\n"
        ),
        encoding="utf-8",
    )

    adapter = CapturingModelAdapter()
    conversations_router._conv_service = ConversationService(
        model_adapter=adapter,
        prompt_service=PromptService(),
    )

    try:
        first_session = client.post(
            "/api/conversations/interactive",
            json={
                "model_id": "doubao-pro",
                "model_mini": "doubao-mini",
                "prompt_version": str(voice_prompt),
                "character": {
                    "Role_Nickname": "萧璟言",
                    "personality": "霸道腹黑",
                    "gender": "男",
                },
                "context": {
                    "relationship": "暧昧",
                    "scene": "办公室",
                    "time_period": "夜晚",
                    "season": "春季",
                },
                "modules": {},
                "custom_variables": {},
            },
        )
        assert_true(first_session.status_code == 200, first_session.text)

        second_session = client.post(
            "/api/conversations/interactive",
            json={
                "model_id": "doubao-pro",
                "model_mini": "doubao-mini",
                "prompt_version": str(text_prompt),
                "character": {
                    "Role_Nickname": "萧璟言",
                    "personality": "霸道腹黑",
                    "gender": "男",
                },
                "context": {
                    "relationship": "暧昧",
                    "scene": "书房",
                    "time_period": "深夜",
                    "season": "春季",
                },
                "modules": {},
                "custom_variables": {},
            },
        )
        assert_true(second_session.status_code == 200, second_session.text)
        second_conv_id = second_session.json()["id"]

        generate_response = client.post(
            f"/api/conversations/{second_conv_id}/generate",
            json={"user_input": "还没睡？", "model_id": "doubao-pro"},
        )
        assert_true(generate_response.status_code == 200, generate_response.text)
        payload = generate_response.json()
        system_message = payload["messages_snapshot"][0]["content"]
        stored_conversation = db.get_conversation(second_conv_id)

        assert_true(
            "上一次在电话聊天沟通" in system_message,
            f"上一通类型未注入真实 system prompt: {system_message}",
        )
        assert_true(
            stored_conversation["config"]["context"].get("last_cst_type", "") == "上一次在电话聊天沟通",
            f"会话配置未持久化上一通类型: {stored_conversation['config']['context']}",
        )
        assert_true(
            stored_conversation.get("conversation_channel", "") == "文字聊天沟通",
            f"当前会话通道未按 prompt 正确识别: {stored_conversation.get('conversation_channel', '')}",
        )
    finally:
        conversations_router._conv_service = None


def test_compare_exports(client: TestClient):
    conv_a = create_scored_conversation("角色A", [8.1, 8.4])
    conv_b = create_scored_conversation("角色B", [9.0, 9.2])
    report_response = client.post(
        "/api/reports/compare",
        json={
            "groups": [
                {
                    "conv_id": conv_a,
                    "label": "v2.0",
                    "model_id": "doubao-pro",
                    "prompt_version": "prompt_v2.md",
                },
                {
                    "conv_id": conv_b,
                    "label": "v2.1",
                    "model_id": "doubao-pro",
                    "prompt_version": "prompt_v21.md",
                },
            ]
        },
    )
    assert_true(report_response.status_code == 200, report_response.text)
    payload = report_response.json()
    assert_true(len(payload["per_turn_comparison"]) == 2, "compare 未返回逐轮对比")
    assert_true(payload["compare_mode"] == "prompt", "compare_mode 未识别为 prompt")
    assert_true(
        payload["group_results"][0]["prompt_version"] == "prompt_v2.md",
        "compare 未返回 prompt_version 元数据",
    )
    assert_true("total" in payload["per_dim_comparison"], "compare 未返回 per_dim_comparison")

    report_id = payload["id"]
    full_export = client.get(f"/api/reports/compare/{report_id}/export")
    summary_export = client.get(f"/api/reports/compare/{report_id}/export?summary=true")
    assert_true(full_export.status_code == 200, full_export.text)
    assert_true(summary_export.status_code == 200, summary_export.text)

    full_workbook = openpyxl.load_workbook(io.BytesIO(full_export.content))
    summary_workbook = openpyxl.load_workbook(io.BytesIO(summary_export.content))
    assert_true(
        "逐轮对比" in full_workbook.sheetnames and len(full_workbook.sheetnames) == 2,
        "完整 compare 导出缺少逐轮对比 sheet",
    )
    assert_true(
        summary_workbook.sheetnames == ["摘要"],
        "summary compare 导出未简化为摘要 sheet",
    )
    full_workbook.close()
    summary_workbook.close()


def test_compare_report_rejects_invalid_history_selection(client: TestClient):
    conv_a = create_scored_conversation("角色甲", [8.2, 8.4])
    conv_b = create_scored_conversation("角色乙", [8.9])
    turn_mismatch_response = client.post(
        "/api/reports/compare",
        json={
            "groups": [
                {
                    "conv_id": conv_a,
                    "label": "模型A",
                    "model_id": "doubao-pro",
                    "prompt_version": "prompt_v26.md",
                },
                {
                    "conv_id": conv_b,
                    "label": "模型B",
                    "model_id": "doubao-pro",
                    "prompt_version": "prompt_v26.md",
                },
            ]
        },
    )
    assert_true(turn_mismatch_response.status_code == 400, turn_mismatch_response.text)
    assert_true(
        "轮数一致" in turn_mismatch_response.json().get("detail", ""),
        f"轮数不一致的报错口径异常: {turn_mismatch_response.text}",
    )

    conv_c = create_scored_conversation("角色丙", [8.6, 8.7])
    mixed_mode_response = client.post(
        "/api/reports/compare",
        json={
            "groups": [
                {
                    "conv_id": conv_a,
                    "label": "控制组",
                    "model_id": "doubao-pro",
                    "prompt_version": "prompt_v26.md",
                },
                {
                    "conv_id": conv_c,
                    "label": "实验组",
                    "model_id": "minimax-m27",
                    "prompt_version": "prompt_v27.md",
                },
            ]
        },
    )
    assert_true(mixed_mode_response.status_code == 400, mixed_mode_response.text)
    assert_true(
        "同类型" in mixed_mode_response.json().get("detail", ""),
        f"混合类型报错口径异常: {mixed_mode_response.text}",
    )


def test_history_selection_summary_builds_model_summary_meta(client: TestClient):
    conv_a = create_scored_conversation(
        "同一角色",
        [8.2, 8.4],
        model_id="glm-4.5",
        prompt_file="prompt_same.md",
    )
    conv_b = create_scored_conversation(
        "同一角色",
        [8.8, 8.9],
        model_id="qwen-plus",
        prompt_file="prompt_same.md",
    )

    response = client.post(
        "/api/reports/history-selection",
        json={"conversation_ids": [conv_a, conv_b]},
    )

    assert_true(response.status_code == 200, response.text)
    payload = response.json()
    assert_true(payload["summary_type"] == "model_summary", payload)
    assert_true(payload["report_title"] == "模型评分摘要", payload)
    assert_true(payload["report_meta"]["selection_count"] == 2, payload)
    assert_true(payload["report_meta"]["role_count"] == 1, payload)
    assert_true(payload["report_meta"]["model_count"] == 2, payload)
    labels = [group["label"] for group in payload["group_results"]]
    assert_true("glm-4.5" in labels and "qwen-plus" in labels, labels)


def test_history_list_supports_partial_model_filter(client: TestClient):
    create_scored_conversation(
        "筛选角色",
        [8.3],
        model_id="glm-4.5",
        prompt_file="prompt_filter.md",
    )

    response = client.get("/api/conversations?model_id=glm&include_archived=true")

    assert_true(response.status_code == 200, response.text)
    conversations = response.json().get("conversations", [])
    assert_true(len(conversations) == 1, conversations)
    assert_true(conversations[0]["model_id"] == "glm-4.5", conversations)


def test_root_redirect_uses_latest_app_shell_version(client: TestClient):
    from main import APP_SHELL_VERSION

    response = client.get("/", follow_redirects=False)

    assert_true(response.status_code in {302, 307}, response.text)
    assert_true(
        response.headers.get("location", "")
        == f"/static/index.html?v={APP_SHELL_VERSION}",
        response.headers,
    )


def test_configs_import(client: TestClient):
    workbook_bytes = create_workbook_bytes(
        ["user_message", "Role_Nickname", "relationship"],
        [["测试输入", "导入角色", "暧昧"]],
    )
    response = client.post(
        "/api/configs/import",
        files={
            "file": (
                "configs.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert_true(response.status_code == 200, response.text)
    payload = response.json()
    assert_true(payload["count"] == 1, "configs/import 解析数量错误")
    assert_true(payload["preview"][0]["turns_count"] == 1, "configs/import 预览错误")


def test_score_template_loader():
    system_prompt, user_template = load_scoring_template(
        Path(r"E:\提效工具\promptfoo-pipeline\scoring_prompts\长文模式")
    )
    assert_true(
        isinstance(system_prompt, str) and isinstance(user_template, str),
        "score_excel 模板加载失败",
    )


def test_single_turn_scoring_and_alias(client: TestClient):
    scoring_router._scoring = FakeScoringService()
    conv_id = db.create_conversation(
        "doubao-pro",
        {
            "prompt_file": "test_prompt.md",
            "character": {"Role_Nickname": "单轮角色", "personality": "冷静"},
            "context": {"relationship": "暧昧"},
            "modules": {},
        },
    )
    db.insert_turn_result(
        conv_id,
        {
            "turn": 1,
            "user_input": "测试输入",
            "ai_output": "测试输出",
            "model_id": "doubao-pro",
        },
    )

    response = client.post(f"/api/scoring/{conv_id}/turn/1")
    assert_true(response.status_code == 200, response.text)

    alias_response = client.get(f"/api/scoring/{conv_id}")
    assert_true(alias_response.status_code == 200, alias_response.text)
    payload = alias_response.json()
    assert_true(payload["summary"]["scored_count"] == 1, "单轮打分未落库")
    scoring_router._scoring = None


def test_rescore_all_persists_latest_scoring_prompt_version(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
):
    scoring_router._scoring = FakeScoringService()
    conv_id = create_scored_conversation("最新重打分角色", [7.2, 7.4])
    conversation = db.get_conversation(conv_id)
    config = conversation["config"]
    runtime = config.setdefault("runtime", {})
    runtime["scoring_prompt_version"] = "旧版打分提示词.md"
    runtime["scoring_model_id"] = "doubao-pro"
    db.update_conversation_config(conv_id, config)

    latest_prompt = "长文模式打分提示词_v4.0_20260421.md"
    monkeypatch.setattr(
        scoring_router,
        "_resolve_scoring_prompt_version",
        lambda prompt_version=None: latest_prompt,
    )

    response = client.post(
        f"/api/scoring/{conv_id}/rescore-all",
        json={
            "scoring_prompt_version": "latest",
            "scoring_model_id": "doubao-pro",
        },
    )
    assert_true(response.status_code == 200, response.text)

    refreshed = db.get_conversation(conv_id)
    assert_true(
        refreshed["scoring_prompt_version"] == latest_prompt,
        f"重打分未落最新提示词版本: {refreshed['scoring_prompt_version']}",
    )
    result_payload = client.get(f"/api/scoring/{conv_id}/results").json()
    assert_true(
        result_payload["meta"]["scoring_prompt_version"] == latest_prompt,
        f"打分结果元信息未更新最新提示词: {result_payload['meta']}",
    )
    scoring_router._scoring = None


def test_chat_score_route_supports_legacy_scoring_signature(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
):
    class LegacyScoringService:
        def is_available(self, model_id=None):
            return True

        def get_last_error(self):
            return ""

        async def score_turn(self, turn_data):
            return {
                "success": True,
                "scores": {"persona_fidelity": 8.6},
                "mapped_total": 8.7,
                "reasoning": f"legacy scoring ok: {turn_data.get('role_name')}",
            }

    monkeypatch.setattr(scoring_service_module, "ScoringService", LegacyScoringService)

    response = client.post(
        "/api/chat/score",
        json={
            "user_input": "测试输入",
            "ai_output": "测试输出",
            "config": {
                "prompt_file": "test_prompt.md",
                "character": {
                    "Role_Nickname": "旧签名角色",
                    "personality": "冷静",
                },
                "context": {"relationship": "暧昧"},
                "runtime": {
                    "scoring_prompt_version": "长文模式打分提示词",
                    "scoring_model_id": "doubao-pro",
                },
            },
        },
    )
    assert_true(response.status_code == 200, response.text)
    payload = response.json()
    assert_true(payload["success"] is True, str(payload))
    assert_true(payload["mapped_total"] == 8.7, str(payload))
    assert_true("legacy scoring ok" in payload["reasoning"], str(payload))


def test_auto_scoring_and_injection_depth(client: TestClient):
    scoring_router._scoring = FakeScoringService()
    conversations_router._conv_service = FakeConversationService()

    response = client.post(
        "/api/conversations",
        json={
            "model_id": "doubao-pro",
            "injection_depth": 5,
            "turns": ["第一轮"],
            "character": {
                "Role_Nickname": "自动评分角色",
                "personality": "霸道腹黑",
            },
            "context": {
                "relationship": "暧昧",
                "scene": "花园",
                "time_period": "深夜",
                "user_nickname": "小鹿",
            },
            "modules": {},
        },
    )
    assert_true(response.status_code == 200, response.text)
    conv_id = response.json()["id"]
    conversation = wait_for_turn_score(conv_id, 1)
    assert_true(
        conversation["config"]["runtime"]["injection_depth"] == 5,
        "injection_depth 未写入运行时配置",
    )
    assert_true(
        conversation["results"][0]["score_status"] == "scored",
        "自动打分未在对话完成时落库",
    )
    assert_true(
        conversation["config"]["context"]["current_scene"] == "花园",
        "前端场景字段别名未归一化",
    )

    conversations_router._conv_service = None
    scoring_router._scoring = None


def test_conversation_queue_and_websocket_notice(client: TestClient):
    release_event = threading.Event()
    original_max_concurrency = conversations_router._MAX_CONCURRENT_CONVERSATIONS
    conversations_router._MAX_CONCURRENT_CONVERSATIONS = 3
    conversations_router._conv_service = BlockingConversationService(release_event)
    conversations_router._running_conversations.clear()
    conversations_router._queued_conversations.clear()
    conversations_router._ws_connections.clear()

    payload = {
        "model_id": "doubao-pro",
        "dry_run": True,
        "turns": ["排队测试"],
        "character": {
            "Role_Nickname": "排队角色",
            "personality": "冷静",
            "gender": "男",
        },
        "context": {
            "relationship": "暧昧",
            "scene": "走廊",
            "time_period": "夜晚",
            "user_nickname": "小鹿",
        },
        "modules": {},
    }
    conv_ids = []

    try:
        for index in range(3):
            response = client.post("/api/conversations", json=payload)
            assert_true(response.status_code == 200, response.text)
            body = response.json()
            conv_ids.append(body["id"])
            assert_true(body["status"] == "pending", f"第 {index + 1} 个并发任务未立即占槽")
            assert_true(body["queue_position"] == 0, "未排队任务 queue_position 应为 0")

        queued_response = client.post("/api/conversations", json=payload)
        assert_true(queued_response.status_code == 200, queued_response.text)
        queued_body = queued_response.json()
        queued_id = queued_body["id"]
        conv_ids.append(queued_id)
        assert_true(queued_body["status"] == "queued", "第 4 个并发任务未进入排队")
        assert_true(queued_body["queue_position"] == 1, "第 4 个并发任务排队位置错误")

        queued_conversation = wait_for_conversation_status(queued_id, "queued")
        assert_true(queued_conversation["status"] == "queued", "排队状态未写入数据库")

        with client.websocket_connect(f"/ws/conversations/{queued_id}") as websocket:
            queued_message = websocket.receive_json()
            assert_true(queued_message["type"] == "queued", "WebSocket 未补发排队消息")
            assert_true(queued_message["queue_position"] == 1, "WebSocket 排队位置错误")

        release_event.set()
        for conv_id in conv_ids:
            wait_for_conversation_status(conv_id, "completed")
    finally:
        release_event.set()
        deadline = time.time() + 5.0
        while time.time() < deadline:
            if (
                not conversations_router._running_conversations
                and not conversations_router._queued_conversations
            ):
                break
            time.sleep(0.1)
        conversations_router._MAX_CONCURRENT_CONVERSATIONS = original_max_concurrency
        conversations_router._conv_service = None
        conversations_router._running_conversations.clear()
        conversations_router._queued_conversations.clear()
        conversations_router._ws_connections.clear()


def test_websocket_replays_completed_conversation_events(client: TestClient):
    conversations_router._conv_service = FakeConversationService()
    try:
        response = client.post(
            "/api/conversations",
            json={
                "model_id": "doubao-pro",
                "dry_run": True,
                "turns": ["补发事件测试"],
                "character": {
                    "Role_Nickname": "回放角色",
                    "personality": "冷静",
                    "gender": "男",
                },
                "context": {
                    "relationship": "朋友",
                    "scene": "书房",
                    "time_period": "晚上",
                    "user_nickname": "小鹿",
                },
                "modules": {},
            },
        )
        assert_true(response.status_code == 200, response.text)
        conv_id = response.json()["id"]
        wait_for_conversation_status(conv_id, "completed")

        with client.websocket_connect(f"/ws/conversations/{conv_id}") as websocket:
            replay_types = [websocket.receive_json()["type"] for _ in range(3)]
        assert_true(
            replay_types == ["turn_complete", "turn_result", "completed"],
            f"已完成会话的 WebSocket 补发顺序错误: {replay_types}",
        )
    finally:
        conversations_router._conv_service = None
        conversations_router._ws_connections.clear()


def test_conversation_websocket_pushes_turn_started(client: TestClient):
    release_event = threading.Event()
    conversations_router._conv_service = BlockingConversationService(release_event)
    conversations_router._ws_connections.clear()
    try:
        response = client.post(
            "/api/conversations",
            json={
                "model_id": "doubao-pro",
                "dry_run": True,
                "turns": ["第一轮内容"],
                "character": {
                    "Role_Nickname": "实时角色",
                    "personality": "冷静",
                    "gender": "男",
                },
                "context": {
                    "relationship": "朋友",
                    "scene": "书房",
                    "time_period": "晚上",
                    "user_nickname": "小鹿",
                },
                "modules": {},
            },
        )
        assert_true(response.status_code == 200, response.text)
        conv_id = response.json()["id"]

        with client.websocket_connect(f"/ws/conversations/{conv_id}") as websocket:
            first = websocket.receive_json()
            second = websocket.receive_json()
            assert_true(first["type"] == "started", f"运行中补发类型错误: {first}")
            assert_true(second["type"] == "task_status", f"运行中状态补发错误: {second}")

            release_event.set()
            live_messages = [websocket.receive_json() for _ in range(5)]

        turn_started = next(
            (item for item in live_messages if item.get("type") == "turn_started"),
            None,
        )
        assert_true(turn_started is not None, f"未收到 turn_started: {live_messages}")
        assert_true(turn_started["turn"] == 1, f"turn_started turn 错误: {turn_started}")
        assert_true(turn_started["total_turns"] == 1, f"turn_started total_turns 错误: {turn_started}")
        assert_true(turn_started["user_input_preview"] == "第一轮内容", f"turn_started 预览错误: {turn_started}")

        live_types = [item.get("type") for item in live_messages]
        assert_true("turn_complete" in live_types, f"未收到 turn_complete: {live_types}")
        assert_true("turn_result" in live_types, f"未收到 turn_result: {live_types}")
        assert_true("completed" in live_types, f"未收到 completed: {live_types}")
        wait_for_conversation_status(conv_id, "completed")
    finally:
        release_event.set()
        conversations_router._conv_service = None
        conversations_router._ws_connections.clear()


def test_on_turn_callback_error_does_not_fail_conversation(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
):
    conversations_router._conv_service = FakeConversationService()
    conversations_router._ws_connections.clear()
    original_persist = conversations_router._persist_conversation_runtime

    def broken_persist(*args, **kwargs):
        next_turn_index = kwargs.get("next_turn_index")
        total_turns = kwargs.get("total_turns")
        if next_turn_index is not None and total_turns is not None and next_turn_index < total_turns:
            raise RuntimeError("persist boom")
        return original_persist(*args, **kwargs)

    monkeypatch.setattr(conversations_router, "_persist_conversation_runtime", broken_persist)
    try:
        response = client.post(
            "/api/conversations",
            json={
                "model_id": "doubao-pro",
                "dry_run": True,
                "turns": ["第一轮", "第二轮"],
                "character": {
                    "Role_Nickname": "防崩角色",
                    "personality": "冷静",
                    "gender": "男",
                },
                "context": {
                    "relationship": "朋友",
                    "scene": "客厅",
                    "time_period": "晚上",
                    "user_nickname": "小鹿",
                },
                "modules": {},
            },
        )
        assert_true(response.status_code == 200, response.text)
        conv_id = response.json()["id"]
        conversation = wait_for_conversation_status(conv_id, "completed")
        assert_true(len(conversation["results"]) == 2, "on_turn 异常后未完成全部轮次")
    finally:
        monkeypatch.setattr(conversations_router, "_persist_conversation_runtime", original_persist)
        conversations_router._conv_service = None
        conversations_router._ws_connections.clear()


def test_conversation_list_contract(client: TestClient):
    conv_id = create_scored_conversation("列表契约角色", [8.6])
    list_response = client.get("/api/conversations")
    assert_true(list_response.status_code == 200, list_response.text)
    payload = list_response.json()
    assert_true("conversations" in payload, "列表接口未返回 conversations 字段")
    conversations = payload["conversations"]
    target = next((item for item in conversations if item.get("id") == conv_id), None)
    assert_true(target is not None, "列表接口未返回新创建会话")
    assert_true(target.get("model_id") == "doubao-pro", "列表项 model_id 错误")
    assert_true("prompt_version" in target, "列表项 prompt_version 缺失")
    assert_true("nickname" in target, "列表项 nickname 缺失")
    assert_true("source" in target, "列表项 source 缺失")
    detail_response = client.get(f"/api/conversations/{conv_id}")
    assert_true(detail_response.status_code == 200, detail_response.text)
    detail = detail_response.json()
    assert_true(isinstance(detail.get("config"), dict), "详情接口 config 缺失")
    runtime = detail["config"].get("runtime", {})
    assert_true(isinstance(runtime.get("model_ids", []), list), "详情接口 runtime.model_ids 缺失或类型错误")
    assert_true("model_mini" in detail, "详情接口 model_mini 缺失")
    assert_true("prompt_version" in detail, "详情接口 prompt_version 缺失")
    assert_true(detail.get("summary_interval") == 2, "详情接口未回填 summary_interval")
    assert_true(detail.get("model_ids") == ["doubao-pro"], "详情接口未回填 model_ids")


def test_model_compare_conversation_contract(client: TestClient):
    conversations_router._conv_service = FakeConversationService()
    try:
        requested_model_ids = ["doubao-pro", "deepseek-v3"]
        payload = {
            "model_id": requested_model_ids[0],
            "model_ids": requested_model_ids,
            "compare_mode": "model",
            "summary_interval": 4,
            "injection_depth": 3,
            "dry_run": True,
            "turns": ["第一轮测试"],
            "character": {
                "Role_Nickname": "模型对比角色",
                "personality": "理性沉稳",
                "gender": "男",
            },
            "context": {
                "relationship": "暧昧",
                "scene": "书房",
                "time_period": "深夜",
                "user_nickname": "小鹿",
            },
            "modules": {},
        }
        create_response = client.post("/api/conversations", json=payload)
        assert_true(create_response.status_code == 200, create_response.text)
        create_payload = create_response.json()
        assert_true(create_payload.get("compare_mode") == "model", "创建响应未返回 compare_mode")
        conversations = create_payload.get("conversations", [])
        assert_true(len(conversations) == 2, f"模型对比分支未返回 2 个子会话: {create_payload}")
        assert_true(
            create_payload.get("conversation_ids") == [item.get("id") for item in conversations],
            "conversation_ids 与 conversations 不一致",
        )

        for expected_model_id, item in zip(requested_model_ids, conversations):
            conv_id = item["id"]
            conversation = wait_for_turns(conv_id, 1)
            assert_true(conversation["model_id"] == expected_model_id, "子会话主模型写入错误")
            runtime = conversation["config"].get("runtime", {})
            assert_true(runtime.get("compare_mode") == "model", "runtime.compare_mode 未写入")
            assert_true(runtime.get("model_ids") == requested_model_ids, "runtime.model_ids 未保留完整对比集合")
            assert_true(runtime.get("active_model_id") == expected_model_id, "runtime.active_model_id 未指向当前执行模型")
            assert_true(conversation["results"][0]["model_id"] == expected_model_id, "轮次结果 model_id 错误")

            detail_response = client.get(f"/api/conversations/{conv_id}")
            assert_true(detail_response.status_code == 200, detail_response.text)
            detail = detail_response.json()
            assert_true(detail.get("compare_mode") == "model", "详情接口未回填 compare_mode")
            assert_true(detail.get("model_ids") == requested_model_ids, "详情接口未回填完整 model_ids")
            assert_true(detail.get("summary_interval") == 4, "详情接口 summary_interval 错误")
    finally:
        conversations_router._conv_service = None


def test_batch_conversation_persists_resume_runtime(client: TestClient):
    conversations_router._conv_service = FakeConversationService()
    try:
        response = client.post(
            "/api/conversations",
            json={
                "model_id": "doubao-pro",
                "dry_run": True,
                "turns": ["第一轮", "第二轮", "第三轮"],
                "character": {
                    "Role_Nickname": "恢复角色",
                    "personality": "冷静",
                    "gender": "男",
                },
                "context": {
                    "relationship": "暧昧",
                    "scene": "书房",
                    "time_period": "深夜",
                    "user_nickname": "小鹿",
                },
                "modules": {},
            },
        )
        assert_true(response.status_code == 200, response.text)
        conv_id = response.json()["id"]
        conversation = wait_for_conversation_status(conv_id, "completed")
        runtime = conversation["config"].get("runtime", {})
        assert_true(runtime.get("turns") == ["第一轮", "第二轮", "第三轮"], "批量任务未持久化原始 turns")
        assert_true(runtime.get("total_turns") == 3, "批量任务 total_turns 错误")
        assert_true(runtime.get("next_turn_index") == 3, "批量任务完成后 next_turn_index 未推进")
        assert_true(runtime.get("resume_supported") is True, "批量任务未标记为可恢复")
        assert_true(runtime.get("dry_run") is True, "批量任务未持久化 dry_run")

        detail_response = client.get(f"/api/conversations/{conv_id}")
        assert_true(detail_response.status_code == 200, detail_response.text)
        detail = detail_response.json()
        assert_true(detail.get("total_turns") == 3, "详情接口未回填 total_turns")
        assert_true(detail.get("next_turn_index") == 3, "详情接口未回填 next_turn_index")
        assert_true(detail.get("resume_supported") is True, "详情接口未回填 resume_supported")
    finally:
        conversations_router._conv_service = None


def test_resume_rejects_non_batch_and_completed_conversations(client: TestClient):
    interactive_response = client.post(
        "/api/conversations/interactive",
        json={
            "model_id": "doubao-pro",
            "character": {"Role_Nickname": "交互壳", "personality": "冷静"},
            "context": {"relationship": "朋友", "scene": "客厅"},
            "modules": {},
        },
    )
    assert_true(interactive_response.status_code == 200, interactive_response.text)
    interactive_id = interactive_response.json()["id"]

    interactive_resume = client.post(f"/api/conversations/{interactive_id}/resume")
    assert_true(interactive_resume.status_code == 400, interactive_resume.text)
    assert_true("不支持恢复" in interactive_resume.text, interactive_resume.text)

    conversations_router._conv_service = FakeConversationService()
    try:
        batch_response = client.post(
            "/api/conversations",
            json={
                "model_id": "doubao-pro",
                "dry_run": True,
                "turns": ["已完成轮次"],
                "character": {
                    "Role_Nickname": "已完成批量",
                    "personality": "冷静",
                    "gender": "男",
                },
                "context": {
                    "relationship": "朋友",
                    "scene": "书房",
                    "time_period": "晚上",
                    "user_nickname": "小鹿",
                },
                "modules": {},
            },
        )
        assert_true(batch_response.status_code == 200, batch_response.text)
        batch_id = batch_response.json()["id"]
        wait_for_conversation_status(batch_id, "completed")

        completed_resume = client.post(f"/api/conversations/{batch_id}/resume")
        assert_true(completed_resume.status_code == 400, completed_resume.text)
        assert_true("已完成" in completed_resume.text, completed_resume.text)
    finally:
        conversations_router._conv_service = None


def test_resume_conversation_only_runs_remaining_turns(client: TestClient):
    conversations_router._conv_service = FakeConversationService()
    try:
        config = {
            "prompt_file": "test_prompt.md",
            "character": {
                "Role_Nickname": "续跑角色",
                "gender": "男",
                "personality": "克制",
            },
            "context": {"relationship": "暧昧"},
            "modules": {},
            "runtime": {
                "conversation_mode": "batch",
                "turns": ["第一轮", "第二轮", "第三轮"],
                "total_turns": 3,
                "next_turn_index": 1,
                "resume_supported": True,
                "dry_run": True,
                "model_ids": ["doubao-pro"],
                "active_model_id": "doubao-pro",
                "summary_interval": 2,
            },
        }
        conv_id = db.create_conversation(
            "doubao-pro",
            config,
            model_mini="doubao-mini",
            prompt_version="test_prompt.md",
        )
        db.insert_turn_result(
            conv_id,
            {
                "turn": 1,
                "user_input": "第一轮",
                "ai_output": "已完成第一轮",
                "model_id": "doubao-pro",
            },
        )
        db.update_conversation_status(conv_id, "interrupted")

        response = client.post(f"/api/conversations/{conv_id}/resume")
        assert_true(response.status_code == 200, response.text)
        payload = response.json()
        assert_true(payload["turns_count"] == 3, "resume 响应 turns_count 错误")

        conversation = wait_for_conversation_status(conv_id, "completed")
        assert_true(
            [item["turn"] for item in conversation["results"]] == [1, 2, 3],
            f"resume 后轮次编号错误: {conversation['results']}",
        )
        assert_true(
            [item["user_input"] for item in conversation["results"]] == ["第一轮", "第二轮", "第三轮"],
            "resume 未只续跑剩余轮次",
        )
        runtime = conversation["config"].get("runtime", {})
        assert_true(runtime.get("next_turn_index") == 3, "resume 完成后 next_turn_index 未更新")
    finally:
        conversations_router._conv_service = None


def test_reconcile_conversation_runtime_state_normalizes_stale_statuses(client: TestClient):
    batch_config = {
        "prompt_file": "test_prompt.md",
        "character": {"Role_Nickname": "中断批量", "gender": "男", "personality": "冷静"},
        "context": {"relationship": "朋友"},
        "modules": {},
        "runtime": {
            "conversation_mode": "batch",
            "turns": ["第一轮", "第二轮"],
            "total_turns": 2,
            "next_turn_index": 0,
            "resume_supported": True,
            "model_ids": ["doubao-pro"],
            "active_model_id": "doubao-pro",
            "summary_interval": 2,
        },
    }
    batch_id = db.create_conversation("doubao-pro", batch_config, prompt_version="test_prompt.md")
    db.insert_turn_result(
        batch_id,
        {
            "turn": 1,
            "user_input": "第一轮",
            "ai_output": "已完成第一轮",
            "model_id": "doubao-pro",
        },
    )
    db.update_conversation_status(batch_id, "running")

    interactive_config = {
        "prompt_file": "test_prompt.md",
        "character": {"Role_Nickname": "空壳交互", "gender": "男", "personality": "冷静"},
        "context": {"relationship": "朋友"},
        "modules": {},
        "runtime": {
            "conversation_mode": "interactive",
            "model_ids": ["doubao-pro"],
            "active_model_id": "doubao-pro",
        },
    }
    interactive_id = db.create_conversation(
        "doubao-pro",
        interactive_config,
        prompt_version="test_prompt.md",
    )
    db.update_conversation_status(interactive_id, "running")

    asyncio.run(conversations_router.reconcile_conversation_runtime_state())

    batch_conversation = db.get_conversation(batch_id)
    interactive_conversation = db.get_conversation(interactive_id)
    assert_true(batch_conversation["status"] == "interrupted", "可恢复批量任务未转为 interrupted")
    assert_true(
        batch_conversation["config"]["runtime"].get("next_turn_index") == 1,
        "状态收敛未按已完成轮次推进 next_turn_index",
    )
    assert_true(interactive_conversation["status"] == "pending", "空壳 interactive 会话未转回 pending")


def test_conversation_control_pause_resume_and_event_log(client: TestClient):
    pause_gate = threading.Event()
    controlled_service = ControlledConversationService(pause_gate=pause_gate)
    conversations_router._conv_service = controlled_service
    try:
        response = client.post(
            "/api/conversations",
            json={
                "model_id": "doubao-pro",
                "dry_run": True,
                "turns": ["第一轮", "第二轮"],
                "character": {"Role_Nickname": "可暂停任务", "personality": "克制", "gender": "男"},
                "context": {"relationship": "朋友"},
                "modules": {},
            },
        )
        assert_true(response.status_code == 200, response.text)
        conv_id = response.json()["id"]

        assert_true(controlled_service.after_first_turn.wait(timeout=3), "第一轮未完成")

        pause_response = client.post(
            f"/api/conversations/{conv_id}/control",
            json={"action": "pause"},
        )
        assert_true(pause_response.status_code == 200, pause_response.text)
        assert_true(pause_response.json()["status"] == "paused", pause_response.text)

        pause_gate.set()
        wait_for_condition(
            lambda: (db.get_conversation(conv_id) or {}).get("status") == "paused",
            timeout=3.0,
            message="暂停后状态未进入 paused",
        )
        paused_conv = db.get_conversation(conv_id)
        assert_true(len(paused_conv.get("results", [])) == 1, "暂停后不应继续执行下一轮")

        events_response = client.get(f"/api/conversations/{conv_id}/events")
        assert_true(events_response.status_code == 200, events_response.text)
        event_types = [item["event_type"] for item in events_response.json()["events"]]
        assert_true("paused" in event_types, f"暂停事件未记日志: {event_types}")

        resume_response = client.post(
            f"/api/conversations/{conv_id}/control",
            json={"action": "resume"},
        )
        assert_true(resume_response.status_code == 200, resume_response.text)
        conversation = wait_for_conversation_status(conv_id, "completed")
        assert_true(len(conversation.get("results", [])) == 2, "恢复后未完成剩余轮次")
    finally:
        conversations_router._conv_service = None


def test_conversation_control_cancel_can_resume_remaining_turns(client: TestClient):
    pause_gate = threading.Event()
    controlled_service = ControlledConversationService(pause_gate=pause_gate)
    conversations_router._conv_service = controlled_service
    try:
        response = client.post(
            "/api/conversations",
            json={
                "model_id": "doubao-pro",
                "dry_run": True,
                "turns": ["第一轮", "第二轮"],
                "character": {"Role_Nickname": "可取消任务", "personality": "冷静", "gender": "男"},
                "context": {"relationship": "朋友"},
                "modules": {},
            },
        )
        assert_true(response.status_code == 200, response.text)
        conv_id = response.json()["id"]

        assert_true(controlled_service.after_first_turn.wait(timeout=3), "第一轮未完成")
        cancel_response = client.post(
            f"/api/conversations/{conv_id}/control",
            json={"action": "cancel"},
        )
        assert_true(cancel_response.status_code == 200, cancel_response.text)
        pause_gate.set()

        cancelled_conv = wait_for_conversation_status(conv_id, "cancelled")
        assert_true(len(cancelled_conv.get("results", [])) == 1, "取消后不应继续执行下一轮")

        events_response = client.get(f"/api/conversations/{conv_id}/events")
        assert_true(events_response.status_code == 200, events_response.text)
        event_types = [item["event_type"] for item in events_response.json()["events"]]
        assert_true("cancelled" in event_types, f"取消事件未记日志: {event_types}")

        resume_response = client.post(f"/api/conversations/{conv_id}/resume")
        assert_true(resume_response.status_code == 200, resume_response.text)
        resumed_conv = wait_for_conversation_status(conv_id, "completed")
        assert_true(
            [item["turn"] for item in resumed_conv.get("results", [])] == [1, 2],
            f"取消后恢复未补跑剩余轮次: {resumed_conv.get('results', [])}",
        )
    finally:
        conversations_router._conv_service = None


def test_active_orchestration_route_returns_latest_recoverable_run(client: TestClient):
    batch_run = create_test_orchestration_run(kind="batch", status="running", item_status="running", title="批量活动任务")
    compare_run = create_test_orchestration_run(kind="compare", status="paused", item_status="paused", title="对比活动任务")
    create_test_orchestration_run(kind="batch", status="completed", item_status="completed", title="批量已完成任务")

    batch_response = client.get("/api/orchestrations/active", params={"kind": "batch"})
    assert_true(batch_response.status_code == 200, batch_response.text)
    batch_payload = batch_response.json()["run"]
    assert_true(batch_payload["id"] == batch_run["id"], batch_payload)
    assert_true(batch_payload["status"] == "running", batch_payload)
    assert_true(batch_payload["groups"][0]["items"][0]["status"] == "running", batch_payload)

    compare_response = client.get("/api/orchestrations/active", params={"kind": "compare"})
    assert_true(compare_response.status_code == 200, compare_response.text)
    compare_payload = compare_response.json()["run"]
    assert_true(compare_payload["id"] == compare_run["id"], compare_payload)
    assert_true(compare_payload["status"] == "paused", compare_payload)
    assert_true(compare_payload["groups"][0]["items"][0]["status"] == "paused", compare_payload)


def test_active_orchestration_route_returns_cancelling_run(client: TestClient):
    run = create_test_orchestration_run(kind="batch", status="cancelling", item_status="running", title="批量停止中任务")

    response = client.get("/api/orchestrations/active", params={"kind": "batch"})
    assert_true(response.status_code == 200, response.text)
    payload = response.json()["run"]
    assert_true(payload["id"] == run["id"], payload)
    assert_true(payload["status"] == "cancelling", payload)


def test_latest_orchestration_route_returns_completed_run_when_no_active(client: TestClient):
    run = create_test_orchestration_run(kind="compare", status="completed", item_status="completed", title="对比已完成任务")

    response = client.get("/api/orchestrations/latest", params={"kind": "compare"})
    assert_true(response.status_code == 200, response.text)
    payload = response.json()["run"]
    assert_true(payload["id"] == run["id"], payload)
    assert_true(payload["status"] == "completed", payload)
    assert_true(payload["groups"][0]["items"][0]["status"] == "completed", payload)


def test_orchestration_config_snapshot_is_public_and_listed(client: TestClient):
    response = client.post(
        "/api/orchestrations",
        json={
            "kind": "batch",
            "title": "配置快照测试",
            "concurrency": 2,
            "config_snapshot": {
                "model_id": "deepseek-v4-pro",
                "thinking_effort": "max",
                "auto_scoring": True,
            },
            "groups": [
                {
                    "key": "role:1",
                    "label": "角色一",
                    "relationship": "暧昧",
                    "planned_turns": 1,
                    "items": [
                        {
                            "key": "role:1:item:1",
                            "label": "角色一",
                            "relationship": "暧昧",
                            "model_id": "deepseek-v4-pro",
                            "planned_turns": 1,
                            "payload": {"dry_run": True},
                        }
                    ],
                }
            ],
        },
    )
    assert_true(response.status_code == 200, response.text)
    payload = response.json()
    assert_true(payload["config_snapshot"]["thinking_effort"] == "max", payload)
    assert_true(payload["manifest"]["config_snapshot"]["model_id"] == "deepseek-v4-pro", payload)

    list_response = client.get("/api/orchestrations", params={"kind": "batch", "limit": 5})
    assert_true(list_response.status_code == 200, list_response.text)
    runs = list_response.json()["runs"]
    matched = [item for item in runs if item["id"] == payload["id"]]
    assert_true(bool(matched), runs)
    assert_true(matched[0]["config_snapshot"]["auto_scoring"] is True, matched[0])


def test_orchestration_control_pause_route_updates_status(client: TestClient):
    run = create_test_orchestration_run(kind="batch", status="running", item_status="running", title="可暂停批量任务")

    response = client.post(
        f"/api/orchestrations/{run['id']}/control",
        json={"action": "pause"},
    )
    assert_true(response.status_code == 200, response.text)
    payload = response.json()
    assert_true(payload["status"] == "paused", payload)
    persisted = db.get_orchestration_run(run["id"])
    assert_true((persisted or {}).get("status") == "paused", persisted)
    assert_true(
        task_control_module.get(f"orchestration:{run['id']}").status == "paused",
        "编排任务控制器未进入 paused",
    )


def test_orchestration_control_cancel_route_marks_cancelling_until_active_items_settle(client: TestClient):
    run = create_test_orchestration_run(kind="batch", status="running", item_status="running", title="可停止批量任务")

    response = client.post(
        f"/api/orchestrations/{run['id']}/control",
        json={"action": "cancel"},
    )
    assert_true(response.status_code == 200, response.text)
    payload = response.json()
    assert_true(payload["status"] == "cancelling", payload)
    persisted = db.get_orchestration_run(run["id"])
    assert_true((persisted or {}).get("status") == "cancelling", persisted)


def test_orchestration_get_run_promotes_cancelling_to_cancelled_after_terminal_settlement(client: TestClient):
    run = create_test_orchestration_run(kind="batch", status="cancelling", item_status="cancelled", title="停止中批量任务")

    response = client.get(f"/api/orchestrations/{run['id']}")
    assert_true(response.status_code == 200, response.text)
    payload = response.json()
    assert_true(payload["status"] == "cancelled", payload)


def test_compare_orchestration_runner_persists_child_state_and_advances_groups():
    pause_gate = threading.Event()
    pause_gate.set()
    controlled_service = ControlledConversationService(pause_gate=pause_gate)
    conversations_router._conv_service = controlled_service

    def make_item(role: str, relationship: str, model_id: str, turns: list[str]) -> OrchestrationItemRequest:
        return OrchestrationItemRequest(
            key=f"{role}:{model_id}",
            label=model_id,
            relationship=relationship,
            model_id=model_id,
            planned_turns=len(turns),
            payload={
                "character": {
                    "Role_Nickname": role,
                    "gender": "女",
                    "personality": "冷静",
                },
                "context": {"relationship": relationship},
                "modules": {},
                "turns": turns,
                "model_id": model_id,
                "model_ids": [model_id],
                "compare_mode": "model",
                "dry_run": True,
            },
        )

    async def scenario() -> dict:
        run = await orchestration_service_module.create_run(
            OrchestrationRunCreate(
                kind="compare",
                title="多角色顺延测试",
                concurrency=2,
                groups=[
                    OrchestrationGroupRequest(
                        key="role:1",
                        label="角色一",
                        relationship="暧昧",
                        planned_turns=2,
                        items=[
                            make_item("角色一", "暧昧", "model-a", ["第一轮", "第二轮"]),
                            make_item("角色一", "暧昧", "model-b", ["第一轮", "第二轮"]),
                        ],
                    ),
                    OrchestrationGroupRequest(
                        key="role:2",
                        label="角色二",
                        relationship="朋友",
                        planned_turns=2,
                        items=[
                            make_item("角色二", "朋友", "model-a", ["第三轮", "第四轮"]),
                            make_item("角色二", "朋友", "model-b", ["第三轮", "第四轮"]),
                        ],
                    ),
                ],
            )
        )
        deadline = time.time() + 8.0
        latest = None
        while time.time() < deadline:
            latest = await orchestration_service_module.get_run(run["id"])
            if latest and latest.get("status") == "completed":
                return latest
            await asyncio.sleep(0.1)
        raise AssertionError(f"编排任务未按预期完成: {latest}")

    try:
        latest = asyncio.run(scenario())
    finally:
        conversations_router._conv_service = None

    items = [
        item
        for group in latest.get("groups", [])
        for item in group.get("items", [])
    ]
    assert_true(len(items) == 4, latest)
    assert_true(all(item.get("status") == "completed" for item in items), latest)
    conv_ids = [str(item.get("conversation_id", "") or "").strip() for item in items]
    assert_true(all(conv_ids), f"存在未落库的子会话引用: {latest}")
    assert_true(len(set(conv_ids)) == 4, f"子会话引用重复: {conv_ids}")
    conn = db.get_connection()
    try:
        row = conn.execute("SELECT COUNT(*) AS cnt FROM conversations").fetchone()
    finally:
        conn.close()
    assert_true(int(row["cnt"] or 0) == 4, f"编排重复创建子会话: {row['cnt']}")


def test_orchestration_control_rejects_terminal_run(client: TestClient):
    run = create_test_orchestration_run(kind="compare", status="completed", item_status="completed", title="已完成对比任务")

    response = client.post(
        f"/api/orchestrations/{run['id']}/control",
        json={"action": "pause"},
    )
    assert_true(response.status_code == 400, response.text)
    payload = response.json()
    assert_true("已结束" in payload.get("detail", ""), payload)
    persisted = db.get_orchestration_run(run["id"])
    assert_true((persisted or {}).get("status") == "completed", persisted)


def test_scoring_control_cancel_clears_cache_and_records_events(client: TestClient):
    conv_id = create_scored_conversation("评分控制角色", [8.2, 7.9])
    db.save_ai_report_summary(
        target_type="conversation_scoring",
        target_id=conv_id,
        report_kind="scoring_report",
        model_id="qwen-plus",
        prompt_filename="summary_prompt.md",
        source_signature="old-signature",
        markdown="old cache",
    )
    assert_true(count_ai_report_summaries(conv_id) == 1, "测试前缓存准备失败")

    pause_gate = threading.Event()
    controlled_service = ControlledScoringService(pause_gate=pause_gate)
    scoring_router._scoring = controlled_service
    try:
        response = client.post(f"/api/scoring/{conv_id}", json={"scoring_model_id": "qwen3.6-plus"})
        assert_true(response.status_code == 200, response.text)
        assert_true(controlled_service.after_first_turn.wait(timeout=3), "评分第一轮未完成")

        cancel_response = client.post(
            f"/api/scoring/{conv_id}/control",
            json={"action": "cancel"},
        )
        assert_true(cancel_response.status_code == 200, cancel_response.text)
        pause_gate.set()

        wait_for_condition(
            lambda: count_ai_report_summaries(conv_id) == 0,
            timeout=3.0,
            message="打分控制未清空旧摘要缓存",
        )

        events_response = client.get(f"/api/conversations/{conv_id}/events?scope=scoring")
        assert_true(events_response.status_code == 200, events_response.text)
        event_types = [item["event_type"] for item in events_response.json()["events"]]
        assert_true("cancelled" in event_types, f"评分取消事件未记日志: {event_types}")
    finally:
        scoring_router._scoring = None


def test_history_context_query_and_reset_context_coherence():
    config = {
        "character": {"Role_Nickname": "上下文角色", "gender": "男", "personality": "冷静"},
        "context": {"relationship": "朋友"},
        "modules": {},
        "runtime": {"model_ids": ["doubao-pro"], "summary_interval": 2},
    }
    conv_id = db.create_conversation("doubao-pro", config)
    for turn in (1, 2):
        db.insert_turn_result(
            conv_id,
            {
                "turn": turn,
                "user_input": f"用户输入{turn}",
                "ai_output": f"AI输出{turn}",
                "dialogue_summary": "" if turn == 1 else "已有摘要",
                "model_id": "doubao-pro",
            },
        )
        db.update_turn_scores(
            conv_id,
            turn,
            {
                "persona_fidelity": 4.5,
                "narrative_immersion": 4.4,
                "emotional_tension": 4.3,
                "boundary_memory": 4.2,
                "format_compliance": 4.1,
                "context_coherence": 4.0,
                "mapped_total": 8.5,
                "reasoning": "测试写入",
                "success": True,
            },
        )

    history_context = db.get_history_context(conv_id, 3)
    assert_true("[用户] 用户输入1" in history_context, f"history_context 缺少首轮用户输入: {history_context}")
    assert_true("[AI] AI输出2" in history_context, f"history_context 缺少第二轮 AI 输出: {history_context}")

    db.reset_conversation_scores(conv_id)
    refreshed = db.get_conversation(conv_id)
    assert_true(refreshed is not None, "重置后会话丢失")
    for item in refreshed["results"]:
        assert_true(float(item.get("score_context_coherence", 0) or 0) == 0, "D6 重置失败")
        assert_true(item.get("score_status") == "unscored", "重置后状态未回到 unscored")


def test_scoring_results_route_exposes_context_coherence(client: TestClient):
    conv_id = create_scored_conversation("D6 角色", [8.6])
    response = client.get(f"/api/scoring/{conv_id}/results")
    assert_true(response.status_code == 200, response.text)
    payload = response.json()
    assert_true("context_coherence" in payload["turns"][0]["scores"], payload)
    assert_true(
        abs(float(payload["turns"][0]["scores"]["context_coherence"]) - 8.1) < 1e-6,
        f"D6 分数返回错误: {payload['turns'][0]['scores']}",
    )


def test_archive_and_server_side_history_filters(client: TestClient):
    high_score_id = create_scored_conversation("高分角色", [8.6])
    low_score_id = create_scored_conversation("低分角色", [5.1])
    db.update_conversation_status(low_score_id, "failed")

    archive_response = client.put(f"/api/conversations/{low_score_id}/archive?archived=true")
    assert_true(archive_response.status_code == 200, archive_response.text)

    default_list = client.get("/api/conversations")
    assert_true(default_list.status_code == 200, default_list.text)
    default_ids = [item["id"] for item in default_list.json()["conversations"]]
    assert_true(high_score_id in default_ids, "默认历史列表缺少未归档记录")
    assert_true(low_score_id not in default_ids, "默认历史列表不应包含已归档记录")

    filtered = client.get(
        "/api/conversations",
        params={
            "include_archived": "true",
            "model_id": "doubao-pro",
            "status": "completed",
            "min_score": "8",
            "max_score": "9",
        },
    )
    assert_true(filtered.status_code == 200, filtered.text)
    filtered_ids = [item["id"] for item in filtered.json()["conversations"]]
    assert_true(filtered_ids == [high_score_id], f"服务端筛选结果错误: {filtered_ids}")

    archived_only = client.get("/api/conversations", params={"archived": "true", "include_archived": "true"})
    assert_true(archived_only.status_code == 200, archived_only.text)
    archived_ids = [item["id"] for item in archived_only.json()["conversations"]]
    assert_true(low_score_id in archived_ids, "已归档筛选未返回目标记录")

    unarchive_response = client.put(f"/api/conversations/{low_score_id}/archive?archived=false")
    assert_true(unarchive_response.status_code == 200, unarchive_response.text)
    restored = client.get("/api/conversations")
    restored_ids = [item["id"] for item in restored.json()["conversations"]]
    assert_true(low_score_id in restored_ids, "取消归档后记录未回到默认列表")


def test_conversation_rejects_more_than_three_model_ids(client: TestClient):
    response = client.post(
        "/api/conversations",
        json={
            "model_id": "doubao-pro",
            "model_ids": [f"m{i}" for i in range(11)],
            "compare_mode": "model",
            "turns": ["第一轮"],
            "character": {
                "Role_Nickname": "模型上限角色",
                "personality": "冷静",
                "gender": "男",
            },
            "context": {
                "relationship": "朋友",
                "scene": "客厅",
                "time_period": "下午",
                "user_nickname": "小鹿",
            },
            "modules": {},
        },
    )
    assert_true(response.status_code == 400, response.text)
    assert_true("最多支持" in response.text and "个模型" in response.text, response.text)


def test_interactive_conversation_session_flow(client: TestClient):
    payload = {
        "model_id": "doubao-pro",
        "model_mini": "doubao-mini",
        "prompt_version": DEFAULT_PROMPT_FILE,
        "summary_interval": 5,
        "injection_depth": 4,
        "character": {
            "Role_Nickname": "交互式角色",
            "personality": "霸道腹黑",
        },
        "context": {
            "relationship": "暧昧",
            "scene": "咖啡店",
            "time_period": "傍晚",
            "user_nickname": "小鹿",
        },
        "modules": {
            "system_Role_acting": "请保持强势角色口吻",
        },
    }
    create_response = client.post("/api/conversations/interactive", json=payload)
    assert_true(create_response.status_code == 200, create_response.text)
    conv_id = create_response.json()["id"]

    append_response = client.post(
        f"/api/conversations/{conv_id}/turns",
        json={
            "user_input": "你在想什么？",
            "ai_output": "我在想你会不会继续追问。",
            "word_count": 16,
            "msg_count": 3,
            "input_tokens": 32,
            "output_tokens": 48,
            "latency_s": 1.2,
            "messages_snapshot": [
                {"role": "system", "content": "你是交互式角色"},
                {"role": "user", "content": "<user_input>你在想什么？</user_input>"},
                {"role": "assistant", "content": "我在想你会不会继续追问。"},
            ],
            "model_id": "doubao-pro",
        },
    )
    assert_true(append_response.status_code == 200, append_response.text)
    assert_true(append_response.json()["turn"] == 1, "交互式 turn 未从 1 开始")

    score_response = client.post(
        f"/api/conversations/{conv_id}/turns/1/scores",
        json={
            "scores": {
                "persona_fidelity": 9.1,
                "narrative_immersion": 8.9,
                "emotional_tension": 8.8,
                "boundary_memory": 9.0,
                "format_compliance": 9.2,
            },
            "mapped_total": 9.0,
            "reasoning": "交互式回写评分",
            "success": True,
        },
    )
    assert_true(score_response.status_code == 200, score_response.text)

    detail_response = client.get(f"/api/conversations/{conv_id}")
    assert_true(detail_response.status_code == 200, detail_response.text)
    detail = detail_response.json()
    assert_true(len(detail.get("results", [])) == 1, "交互式会话未写入 turn")
    assert_true(detail["results"][0].get("score_total") == 9.0, "交互式评分未写入数据库")
    assert_true(detail["config"]["runtime"]["injection_depth"] == 4, "交互式会话未写入数字注入深度")

    second_response = client.post("/api/conversations/interactive", json=payload)
    assert_true(second_response.status_code == 200, second_response.text)
    second_id = second_response.json()["id"]
    assert_true(second_id != conv_id, "新建交互式会话未生成新 id")

    list_response = client.get("/api/conversations")
    assert_true(list_response.status_code == 200, list_response.text)
    ids = {item["id"] for item in list_response.json()["conversations"]}
    assert_true(conv_id in ids and second_id in ids, "新旧交互式会话未同时保留在历史列表")

    clear_response = client.delete(f"/api/conversations/{second_id}/turns")
    assert_true(clear_response.status_code == 200, clear_response.text)
    cleared = client.get(f"/api/conversations/{second_id}").json()
    assert_true(cleared.get("results") == [], "清除上下文后 turn 未被清空")


def test_interactive_create_prewarms_dialogue_summary_with_lite_model(client: TestClient):
    create_response = client.post(
        "/api/conversations/interactive",
        json={
            "model_id": "doubao-pro",
            "model_mini": "doubao-seed-2-0-lite-260215",
            "summary_interval": 5,
            "dry_run": True,
            "character": {"Role_Nickname": "摘要预热角色", "personality": "冷静"},
            "context": {"relationship": "朋友", "scene": "客厅"},
            "modules": {},
        },
    )
    assert_true(create_response.status_code == 200, create_response.text)
    conv_id = create_response.json()["id"]

    def _summary_completed():
        conversation = db.get_conversation(conv_id) or {}
        runtime = conversation.get("config", {}).get("runtime", {})
        return (
            runtime.get("summary_job_status") == "completed"
            and bool(str(runtime.get("latest_dialogue_summary", "")).strip())
        )

    wait_for_condition(_summary_completed, timeout=5.0, message="交互会话创建后未完成摘要预热")
    conversation = db.get_conversation(conv_id) or {}
    runtime = conversation.get("config", {}).get("runtime", {})
    assert_true(conversation.get("model_mini") == "doubao-lite", "摘要模型未归一化为 doubao-lite")
    assert_true(runtime.get("summary_job_target_turn") == 0, "摘要预热目标轮次应为 0")


def test_interactive_create_loads_previous_dialogue_summary_before_warmup(client: TestClient):
    previous_summary = "=== 之前剧情摘要 ===\n上一通持久化摘要\n=== 摘要结束 ==="
    previous_config = {
        "character": {"Role_Nickname": "跨会话摘要角色", "personality": "冷静"},
        "context": {"relationship": "朋友"},
        "modules": {},
        "runtime": {"model_ids": ["doubao-pro"], "summary_interval": 5},
    }
    previous_id = db.create_conversation("doubao-pro", previous_config, model_mini="doubao-lite")
    db.insert_turn_result(
        previous_id,
        {
            "turn": 1,
            "user_input": "上一通用户输入",
            "ai_output": "上一通角色回复",
            "dialogue_summary": previous_summary,
            "model_id": "doubao-pro",
        },
    )
    db.update_conversation_status(previous_id, "completed")

    create_response = client.post(
        "/api/conversations/interactive",
        json={
            "model_id": "doubao-pro",
            "model_mini": "doubao-seed-2-0-lite-260215",
            "summary_interval": 5,
            "dry_run": True,
            "character": {"Role_Nickname": "跨会话摘要角色", "personality": "冷静"},
            "context": {"relationship": "朋友", "scene": "客厅"},
            "modules": {},
        },
    )
    assert_true(create_response.status_code == 200, create_response.text)
    conv_id = create_response.json()["id"]
    conversation = db.get_conversation(conv_id) or {}
    runtime = conversation.get("config", {}).get("runtime", {})
    assert_true(runtime.get("latest_dialogue_summary") == previous_summary, "未冷启动加载上一通摘要")
    assert_true(conversation.get("config", {}).get("dialogue_summary") == previous_summary, "上一通摘要未写入种子摘要")
    assert_true(runtime.get("summary_job_status") == "completed", "加载上一通摘要后不应再等待空历史预热")


def test_interactive_generate_surfaces_model_failure(client: TestClient, monkeypatch: pytest.MonkeyPatch):
    payload = {
        "model_id": "doubao-pro",
        "model_mini": "doubao-mini",
        "prompt_version": DEFAULT_PROMPT_FILE,
        "summary_interval": 5,
        "injection_depth": 4,
        "character": {
            "Role_Nickname": "交互式角色",
            "personality": "霸道腹黑",
        },
        "context": {
            "relationship": "暧昧",
            "scene": "咖啡店",
            "time_period": "傍晚",
            "user_nickname": "小鹿",
        },
        "modules": {},
    }
    create_response = client.post("/api/conversations/interactive", json=payload)
    assert_true(create_response.status_code == 200, create_response.text)
    conv_id = create_response.json()["id"]

    class _FailingConversationService:
        def generate_interactive_turn(self, *args, **kwargs):
            raise RuntimeError("MINIMAX_API_KEY 未配置")

    monkeypatch.setattr(
        conversations_router,
        "_get_conv_service",
        lambda: _FailingConversationService(),
    )

    response = client.post(
        f"/api/conversations/{conv_id}/generate",
        json={
            "user_input": "你好",
            "model_id": "minimax-m27",
            "web_search": False,
            "thinking_effort": "disabled",
        },
    )

    assert_true(response.status_code == 502, response.text)
    assert_true("MINIMAX_API_KEY 未配置" in response.text, response.text)


def test_config_and_variable_endpoints(client: TestClient):
    saved_config = {
        "prompt_file": "test_prompt.md",
        "character": {"Role_Nickname": "配置角色", "personality": "理性沉稳"},
        "context": {"relationship": "朋友", "current_scene": "书房"},
        "modules": {"user_Nickname": "测试用户"},
        "runtime": {"model_ids": ["doubao-pro"], "summary_interval": 3},
    }
    save_response = client.post(
        "/api/configs",
        json={
            "name": "保存的配置",
            "type": "custom_config",
            "config": saved_config,
        },
    )
    assert_true(save_response.status_code == 200, save_response.text)
    saved_config_id = save_response.json()["id"]
    assert_true(saved_config_id.startswith("cfg_"), "配置保存 ID 未使用 cfg_ 前缀")

    conv_id = db.create_conversation("doubao-pro", saved_config)
    db.insert_turn_result(
        conv_id,
        {
            "turn": 1,
            "user_input": "测试输入",
            "ai_output": "测试输出",
            "model_id": "doubao-pro",
        },
    )

    list_response = client.get("/api/configs")
    assert_true(list_response.status_code == 200, list_response.text)
    configs = list_response.json()["configs"]
    assert_true(
        any(item["id"] == saved_config_id and item["source"] == "saved_config" for item in configs),
        "保存的配置未出现在 GET /api/configs 列表",
    )

    export_response = client.get(f"/api/configs/{saved_config_id}/export")
    assert_true(export_response.status_code == 200, export_response.text)

    variables_response = client.get(
        f"/api/configs/{saved_config_id}/variables/export?format=json"
    )
    assert_true(variables_response.status_code == 200, variables_response.text)
    exported = variables_response.json()
    assert_true(
        any(item["name"] == "Role_Nickname" for item in exported["variables"]),
        "变量导出缺少 Role_Nickname",
    )

    variable_workbook = create_workbook_bytes(
        ["变量名", "变量值"],
        [["Role_Nickname", "变量导入角色"], ["relationship", "恋人"]],
    )
    import_response = client.post(
        "/api/configs/variables/import",
        files={
            "file": (
                "variables.xlsx",
                variable_workbook.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert_true(import_response.status_code == 200, import_response.text)
    import_payload = import_response.json()
    assert_true(
        import_payload["config"]["character"]["Role_Nickname"] == "变量导入角色",
        "变量导入未写入 character 分组",
    )

    preset_vars = client.get("/api/presets/霸道腹黑/variables?gender=男&relationship=暧昧")
    assert_true(preset_vars.status_code == 200, preset_vars.text)
    assert_true(
        "intimacy_boundary" in preset_vars.json(),
        "预设变量联动接口未返回关系边界",
    )


def test_model_save_and_multi_model_export(client: TestClient):
    with tempfile.TemporaryDirectory() as temp_dir:
        original_dir = models_router.MODELS_CONFIG_DIR
        try:
            models_router.MODELS_CONFIG_DIR = Path(temp_dir)
            models_router._adapter = None
            save_response = client.post(
                "/api/models",
                json={
                    "id": "unit-test-model",
                    "name": "unit-test-model",
                    "display_name": "单测模型",
                    "provider": "volcengine",
                    "api": {
                        "base_url": "https://example.com",
                        "api_key_env": "TEST_KEY",
                    },
                    "parameters": {"temperature": 0.7},
                },
            )
            assert_true(save_response.status_code == 200, save_response.text)
            assert_true(
                (Path(temp_dir) / "unit-test-model.yaml").exists(),
                "模型配置文件未写入",
            )
        finally:
            models_router.MODELS_CONFIG_DIR = original_dir
            models_router._adapter = None

    conv_a = create_scored_conversation("模型A", [8.1, 8.4])
    conv_b = create_scored_conversation("模型B", [9.0, 9.2])
    export_response = client.get(
        f"/api/scoring/multi-model/export?conv_ids={conv_a},{conv_b}"
    )
    assert_true(export_response.status_code == 200, export_response.text)

    dialogue_export = client.get(f"/api/scoring/dialogue/export?session_id={conv_a}")
    assert_true(dialogue_export.status_code == 200, dialogue_export.text)


def test_prompt_management_crud(client: TestClient, monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    prompt_dir = tmp_path / "prompts"
    prompt_dir.mkdir(parents=True, exist_ok=True)
    base_prompt = prompt_dir / "alpha_prompt.md"
    latest_prompt = prompt_dir / "beta_prompt.md"
    base_prompt.write_text("# alpha\n初始内容\n", encoding="utf-8")
    latest_prompt.write_text("# beta\n最新内容\n", encoding="utf-8")

    monkeypatch.setattr(prompts_router, "PROMPT_DIR", prompt_dir)
    monkeypatch.setattr(
        prompts_router,
        "list_prompt_files",
        lambda: sorted(prompt_dir.glob("*.md")),
    )
    monkeypatch.setattr(prompts_router, "get_latest_prompt_file", lambda: latest_prompt.name)
    monkeypatch.setattr(
        prompts_router,
        "is_main_prompt_file",
        lambda filename: filename == latest_prompt.name,
    )

    list_response = client.get("/api/prompts")
    assert_true(list_response.status_code == 200, list_response.text)
    list_payload = list_response.json()
    assert_true(list_payload["latest_filename"] == latest_prompt.name, "最新提示词识别错误")
    assert_true(len(list_payload["prompts"]) == 2, "提示词列表数量错误")

    get_response = client.get(f"/api/prompts/{base_prompt.name}")
    assert_true(get_response.status_code == 200, get_response.text)
    assert_true(get_response.json()["content"].startswith("# alpha"), "提示词查看内容错误")

    download_response = client.get(f"/api/prompts/{base_prompt.name}/download")
    assert_true(download_response.status_code == 200, download_response.text)
    assert_true("初始内容" in download_response.content.decode("utf-8"), "提示词下载内容错误")

    upload_response = client.post(
        "/api/prompts/upload",
        files={
            "file": (
                "gamma_prompt.md",
                "# gamma\n上传内容\n".encode("utf-8"),
                "text/markdown",
            )
        },
    )
    assert_true(upload_response.status_code == 200, upload_response.text)
    assert_true((prompt_dir / "gamma_prompt.md").exists(), "提示词上传未落盘")

    edit_response = client.put(
        "/api/prompts/gamma_prompt.md",
        json={"content": "# gamma\n已编辑内容\n"},
    )
    assert_true(edit_response.status_code == 200, edit_response.text)
    assert_true(
        (prompt_dir / "gamma_prompt.md").read_text(encoding="utf-8") == "# gamma\n已编辑内容\n",
        "提示词在线编辑未生效",
    )


def test_history_export_and_delete(client: TestClient):
    conv_id = create_scored_conversation("历史删除角色", [8.7, 8.9])

    list_response = client.get("/api/conversations")
    assert_true(list_response.status_code == 200, list_response.text)
    assert_true(
        any(item.get("id") == conv_id for item in list_response.json()["conversations"]),
        "历史列表未返回新创建会话",
    )

    export_response = client.get(f"/api/conversations/{conv_id}/export")
    assert_true(export_response.status_code == 200, export_response.text)
    assert_true(
        export_response.headers.get("content-type", "").startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        "历史导出未返回 Excel 文件",
    )

    delete_response = client.delete(f"/api/conversations/{conv_id}")
    assert_true(delete_response.status_code == 200, delete_response.text)

    detail_response = client.get(f"/api/conversations/{conv_id}")
    assert_true(detail_response.status_code == 404, "删除后仍可读取历史详情")


def test_dialogue_guideline_end_to_end_injection(client: TestClient, tmp_path: Path):
    """R-1: 验证 longform_dialogue_guideline 从 modules → build_variables → render_template → system prompt 的完整路径。"""
    # 1. 先验证 load_dialogue_guideline 能按 personal_type 加载到值
    guideline_value = PromptService().load_dialogue_guideline("温暖陪伴")
    assert_true(bool(guideline_value), "load_dialogue_guideline('温暖陪伴') 返回空值")

    # 2. 验证 build_longform_variable_bundle 链路包含 guideline
    from services.runtime_config import build_longform_variable_bundle
    from config import PRESET_CHARACTERS, RELATIONSHIP_PRESETS
    bundle = build_longform_variable_bundle(
        personality="温暖陪伴",
        relationship="暧昧",
        gender="男",
        preset_characters=PRESET_CHARACTERS,
        relationship_presets=RELATIONSHIP_PRESETS,
        prompt_service=PromptService(),
    )
    assert_true(
        bundle.get("longform_dialogue_guideline", "") != "",
        "build_longform_variable_bundle 未返回 longform_dialogue_guideline",
    )

    # 3. 验证 build_variables + render_template 端到端
    ps = PromptService()
    config = {
        "character": {"Role_Nickname": "测试角色", "personality": "温暖陪伴", "gender": "男"},
        "context": {"relationship": "暧昧"},
        "modules": {"longform_dialogue_guideline": guideline_value},
    }
    variables = ps.build_variables(config)
    assert_true(
        variables.get("longform_dialogue_guideline", "") != "",
        "build_variables 未传递 longform_dialogue_guideline",
    )
    template = "你是{{Role_Nickname}}。\n【对白规范】\n{{longform_dialogue_guideline}}\n"
    rendered = ps.render_template(template, variables)
    first_line = guideline_value.split("\n")[0].strip()
    probe = first_line[:8] if len(first_line) >= 8 else first_line
    assert_true(
        probe in rendered,
        f"render_template 后 system prompt 未包含 guideline 内容 (探针: {probe!r})",
    )
    assert_true(
        "{{longform_dialogue_guideline}}" not in rendered,
        "guideline 占位符未被替换",
    )


def test_quality_guard_strips_thinking_channel_tags():
    """R-7: 验证 QualityGuard 能 strip 思考通道标记 <|channel>thought...<channel|>"""
    qa = QualityGuard()

    narration_body = (
        "落地窗外的霓虹在二十九层玻璃上晕开模糊的光斑，萧璟言指尖转着钢笔。"
        "衬衫袖口挽到小臂，腕间那只古董表在台灯下泛着冷光。\n\n"
        '**"你今天怎么来这么晚？"**\n\n'
        "他把文件推到一边，拿起桌角温着的柠檬水喝了一口，喉结动了动。"
        "手机屏幕亮着，是下午部门提交的会议纪要。\n\n"
        '**"下午的渠道对接会又超时了？"**\n\n'
        "他翻了翻会议记录的最后一页，看到你标注的三个待解决问题。"
        "指尖在你名字旁边敲了敲，指腹蹭过纸面上的墨痕。\n\n"
        '**"我让司机去接你，二十分钟到你楼下。"**\n\n'
        "他站起身走到窗边，指尖擦过冰凉的玻璃。"
    )

    # 场景1：思考通道在正文前
    text_with_thinking = (
        "<|channel>thought\n"
        "用户在加班后疲惫地来找角色，我需要展示角色的关心但不能太过直白。\n"
        "根据温暖陪伴型人设，用行动暗示关心而非直说。\n"
        "<channel|>\n"
        + narration_body
    )
    result = qa.check(text_with_thinking)
    assert_true(
        "<|channel>" not in result["processed_text"],
        "思考通道开标记未被 strip",
    )
    assert_true(
        "<channel|>" not in result["processed_text"],
        "思考通道闭标记未被 strip",
    )
    assert_true(
        "用户在加班后" not in result["processed_text"],
        "思考通道内容未被 strip",
    )
    assert_true(
        "萧璟言" in result["processed_text"],
        "strip 思考通道后正文内容丢失",
    )

    # 场景2：空思考通道
    text_empty_thinking = "<|channel>thought\n<channel|>\n" + narration_body
    result2 = qa.check(text_empty_thinking)
    assert_true(
        "<|channel>" not in result2["processed_text"],
        "空思考通道标记未被 strip",
    )
    assert_true(
        "萧璟言" in result2["processed_text"],
        "strip 空思考通道后正文内容丢失",
    )


def test_live_scoring_dispatcher_serializes_turns_per_conversation():
    from services.live_scoring_dispatcher import LiveScoringDispatcher

    conv_id, config = create_unscored_conversation("串行会话", turn_count=2, runtime={"auto_scoring": True})
    events = []

    async def worker(conv_id_arg: str, turn: int, config_arg: dict | None):
        events.append(("start", conv_id_arg, turn))
        if turn == 1:
            await asyncio.sleep(0.05)
        events.append(("end", conv_id_arg, turn))
        return {"success": True}

    async def scenario():
        dispatcher = LiveScoringDispatcher(worker=worker, get_max_workers=lambda: 24)
        assert await dispatcher.enqueue(conv_id, 1, config=config)
        assert await dispatcher.enqueue(conv_id, 2, config=config)
        assert await dispatcher.wait_for_idle(timeout=2.0)

    asyncio.run(scenario())
    assert_true(
        events == [
            ("start", conv_id, 1),
            ("end", conv_id, 1),
            ("start", conv_id, 2),
            ("end", conv_id, 2),
        ],
        f"同会话 live scoring 未串行执行: {events}",
    )


def test_live_scoring_dispatcher_respects_global_pool_limit():
    from services.live_scoring_dispatcher import LiveScoringDispatcher

    conv_ids = [
        create_unscored_conversation(f"并发会话{idx}", turn_count=1, runtime={"auto_scoring": True})[0]
        for idx in range(1, 4)
    ]
    state = {"running": 0, "max_running": 0}
    lock = asyncio.Lock()

    async def worker(conv_id_arg: str, turn: int, config_arg: dict | None):
        async with lock:
            state["running"] += 1
            state["max_running"] = max(state["max_running"], state["running"])
        await asyncio.sleep(0.05)
        async with lock:
            state["running"] -= 1
        return {"success": True}

    async def scenario():
        dispatcher = LiveScoringDispatcher(worker=worker, get_max_workers=lambda: 2)
        for conv_id in conv_ids:
            assert await dispatcher.enqueue(conv_id, 1)
        assert await dispatcher.wait_for_idle(timeout=2.0)

    asyncio.run(scenario())
    assert_true(state["max_running"] == 2, f"全局并发上限未生效: {state}")


def test_live_scoring_dispatcher_respects_runtime_group_limit():
    from services.live_scoring_dispatcher import LiveScoringDispatcher

    conv_ids = [
        create_unscored_conversation(
            f"AB并发会话{idx}",
            turn_count=1,
            runtime={
                "auto_scoring": True,
                "ab_session_id": "ab-group-limit",
                "scoring_max_workers": 2,
            },
        )[0]
        for idx in range(1, 5)
    ]
    config = {
        "runtime": {
            "ab_session_id": "ab-group-limit",
            "scoring_max_workers": 2,
        }
    }
    state = {"running": 0, "max_running": 0}
    lock = asyncio.Lock()

    async def worker(conv_id_arg: str, turn: int, config_arg: dict | None):
        async with lock:
            state["running"] += 1
            state["max_running"] = max(state["max_running"], state["running"])
        await asyncio.sleep(0.05)
        async with lock:
            state["running"] -= 1
        return {"success": True}

    async def scenario():
        dispatcher = LiveScoringDispatcher(worker=worker, get_max_workers=lambda: 6)
        for conv_id in conv_ids:
            assert await dispatcher.enqueue(conv_id, 1, config=config)
        assert await dispatcher.wait_for_idle(timeout=2.0)

    asyncio.run(scenario())
    assert_true(state["max_running"] == 2, f"任务组并发上限未生效: {state}")


def test_live_scoring_dispatcher_deduplicates_duplicate_turn_submission():
    from services.live_scoring_dispatcher import LiveScoringDispatcher

    conv_id, config = create_unscored_conversation("去重会话", turn_count=1, runtime={"auto_scoring": True})
    calls = []

    async def worker(conv_id_arg: str, turn: int, config_arg: dict | None):
        calls.append((conv_id_arg, turn))
        return {"success": True}

    async def scenario():
        dispatcher = LiveScoringDispatcher(worker=worker, get_max_workers=lambda: 24)
        first = await dispatcher.enqueue(conv_id, 1, config=config)
        second = await dispatcher.enqueue(conv_id, 1, config=config)
        assert first is True
        assert second is False
        assert await dispatcher.wait_for_idle(timeout=2.0)

    asyncio.run(scenario())
    assert_true(calls == [(conv_id, 1)], f"重复提交未被去重: {calls}")


def test_live_scoring_failed_turn_uses_retry_backoff_and_marks_failed(monkeypatch):
    conv_id, config = create_unscored_conversation(
        "失败重试会话",
        turn_count=1,
        runtime={
            "auto_scoring": True,
            "scoring_model_id": "doubao-pro",
            "scoring_retry_count": 3,
        },
    )
    delays = []

    class FakeRetryScoringService:
        def __init__(self):
            self._default_retry_delays = (5.0, 15.0, 30.0)
            self._config = {"dimensions": scoring_router.DIMENSIONS}

        def is_available(self, model_id=None):
            return True

        def get_last_error(self):
            return ""

        def resolve_scoring_thinking_effort(self, *args, **kwargs):
            return "disabled"

        def _sanitize_error_message(self, message: str):
            return str(message)

    async def always_fail(*args, **kwargs):
        raise RuntimeError("live boom")

    async def fake_sleep(ctrl, delay_s):
        delays.append(delay_s)

    monkeypatch.setattr(scoring_router, "_get_scoring", lambda: FakeRetryScoringService())
    monkeypatch.setattr(scoring_router, "invoke_score_turn_compat", always_fail)
    monkeypatch.setattr(scoring_router, "_sleep_with_control", fake_sleep)

    asyncio.run(scoring_router._run_live_scoring_turn(conv_id, 1, config=config))
    latest = db.get_conversation(conv_id)
    target = next((item for item in latest.get("results", []) if item.get("turn") == 1), None)
    assert_true(delays == [5.0, 15.0], f"重试退避节奏不正确: {delays}")
    assert_true(target is not None, "失败重试后未找到目标轮次")
    assert_true(target.get("score_status") == "failed", f"失败重试后状态错误: {target}")
    assert_true("live boom" in str(target.get("score_reasoning", "")), f"失败原因未落库: {target}")


def test_live_scoring_turn_handles_uninitialized_service_config(monkeypatch):
    conv_id, config = create_unscored_conversation(
        "未初始化评分配置会话",
        turn_count=1,
        runtime={
            "auto_scoring": True,
            "scoring_model_id": "qwen-plus",
        },
    )

    class FakeScoringService:
        def __init__(self):
            self._default_retry_delays = (5.0, 15.0, 30.0)
            self._config = None

        def is_available(self, model_id=None):
            return True

        def get_last_error(self):
            return ""

        def resolve_scoring_thinking_effort(self, *args, **kwargs):
            return "disabled"

        def _sanitize_error_message(self, message: str):
            return str(message)

    async def fake_score_turn(*args, **kwargs):
        return {
            "success": True,
            "scores": {dimension: 4 for dimension in scoring_router.DIMENSIONS},
            "weighted_total": 24,
            "mapped_total": 80,
            "reasoning": "评分成功",
            "reasoning_content": "",
            "error": "",
            "model_id": "qwen-plus",
            "score_status": "scored",
        }

    monkeypatch.setattr(scoring_router, "_get_scoring", lambda: FakeScoringService())
    monkeypatch.setattr(scoring_router, "invoke_score_turn_compat", fake_score_turn)

    asyncio.run(scoring_router._run_live_scoring_turn(conv_id, 1, config=config))

    latest = db.get_conversation(conv_id)
    target = next((item for item in latest.get("results", []) if item.get("turn") == 1), None)
    assert_true(target is not None, "live scoring 后未找到目标轮次")
    assert_true(target.get("score_status") == "scored", f"未初始化 _config 时 live scoring 未成功: {target}")
    assert_true(float(target.get("score_total", 0) or 0) == 80.0, f"live scoring 分数未落库: {target}")
    scoring_events = db.get_conversation_events(conv_id, scope="scoring")
    assert_true(
        any(event.get("event_type") == "turn_scored" for event in scoring_events),
        f"live scoring 成功后未记录 scoring 事件: {scoring_events}",
    )


def test_live_scoring_uses_single_layer_retry_and_timeout(monkeypatch):
    conv_id, config = create_unscored_conversation(
        "live timeout 会话",
        turn_count=1,
        runtime={
            "auto_scoring": True,
            "scoring_model_id": "qwen-plus",
            "live_scoring_timeout_s": 17,
            "scoring_retry_count": 3,
        },
    )
    captured = {}

    class FakeScoringService:
        def __init__(self):
            self._default_retry_delays = (5.0, 15.0, 30.0)
            self._config = {"dimensions": scoring_router.DIMENSIONS}

        def is_available(self, model_id=None):
            return True

        def get_last_error(self):
            return ""

        def resolve_scoring_thinking_effort(self, *args, **kwargs):
            return "disabled"

        def _sanitize_error_message(self, message: str):
            return str(message)

    async def fake_score_turn(service, payload, **kwargs):
        captured.update(kwargs)
        return {
            "success": True,
            "scores": {dimension: 5 for dimension in scoring_router.DIMENSIONS},
            "weighted_total": 30,
            "mapped_total": 90,
            "reasoning": "评分成功",
            "reasoning_content": "",
            "error": "",
            "model_id": "qwen-plus",
            "score_status": "scored",
        }

    monkeypatch.setattr(scoring_router, "_get_scoring", lambda: FakeScoringService())
    monkeypatch.setattr(scoring_router, "invoke_score_turn_compat", fake_score_turn)

    asyncio.run(scoring_router._run_live_scoring_turn(conv_id, 1, config=config))

    assert_true(captured.get("timeout_s") == 17, f"live scoring timeout 未生效: {captured}")
    assert_true(captured.get("retry_delays") == (), f"live scoring 未禁用内层 retry: {captured}")
    assert_true(
        captured.get("provider_retry_delays") == (),
        f"live scoring 未禁用 provider retry: {captured}",
    )


def test_live_scoring_timeout_defaults_to_global_scoring_timeout(monkeypatch):
    conv_id, config = create_unscored_conversation(
        "live timeout 默认继承会话",
        turn_count=1,
        runtime={
            "auto_scoring": True,
            "scoring_model_id": "qwen-plus",
        },
    )
    captured = {}

    class FakeScoringService:
        def __init__(self):
            self._default_retry_delays = (5.0, 15.0, 30.0)
            self._config = {"dimensions": scoring_router.DIMENSIONS}

        def is_available(self, model_id=None):
            return True

        def get_last_error(self):
            return ""

        def resolve_scoring_thinking_effort(self, *args, **kwargs):
            return "disabled"

        def _sanitize_error_message(self, message: str):
            return str(message)

    async def fake_score_turn(service, payload, **kwargs):
        captured.update(kwargs)
        return {
            "success": True,
            "scores": {dimension: 5 for dimension in scoring_router.DIMENSIONS},
            "weighted_total": 30,
            "mapped_total": 90,
            "reasoning": "评分成功",
            "reasoning_content": "",
            "error": "",
            "model_id": "qwen-plus",
            "score_status": "scored",
        }

    monkeypatch.setenv("SCORING_REQUEST_TIMEOUT_S", "120")
    monkeypatch.delenv(scoring_router.LIVE_SCORING_TIMEOUT_ENV, raising=False)
    monkeypatch.setattr(scoring_router, "_get_scoring", lambda: FakeScoringService())
    monkeypatch.setattr(scoring_router, "invoke_score_turn_compat", fake_score_turn)

    asyncio.run(scoring_router._run_live_scoring_turn(conv_id, 1, config=config))

    assert_true(captured.get("timeout_s") == 120.0, f"live scoring 未继承全局 timeout: {captured}")


def test_live_scoring_retry_count_zero_means_single_attempt(monkeypatch):
    conv_id, config = create_unscored_conversation(
        "live 不重试会话",
        turn_count=1,
        runtime={
            "auto_scoring": True,
            "scoring_model_id": "qwen-plus",
            "scoring_retry_count": 0,
        },
    )
    attempts = []
    delays = []

    class FakeScoringService:
        def __init__(self):
            self._default_retry_delays = (5.0, 15.0, 30.0)
            self._config = {"dimensions": scoring_router.DIMENSIONS}

        def is_available(self, model_id=None):
            return True

        def get_last_error(self):
            return ""

        def resolve_scoring_thinking_effort(self, *args, **kwargs):
            return "disabled"

        def _sanitize_error_message(self, message: str):
            return str(message)

    async def always_fail(*args, **kwargs):
        attempts.append(kwargs)
        raise RuntimeError("retry disabled")

    async def fake_sleep(ctrl, delay_s):
        delays.append(delay_s)

    monkeypatch.setattr(scoring_router, "_get_scoring", lambda: FakeScoringService())
    monkeypatch.setattr(scoring_router, "invoke_score_turn_compat", always_fail)
    monkeypatch.setattr(scoring_router, "_sleep_with_control", fake_sleep)

    asyncio.run(scoring_router._run_live_scoring_turn(conv_id, 1, config=config))

    latest = db.get_conversation(conv_id)
    target = next((item for item in latest.get("results", []) if item.get("turn") == 1), None)
    assert_true(len(attempts) == 1, f"scoring_retry_count=0 应只尝试一次: {len(attempts)}")
    assert_true(delays == [], f"scoring_retry_count=0 不应进入退避等待: {delays}")
    assert_true(target and target.get("score_status") == "failed", f"最终失败应保留手动重试入口: {target}")


def test_push_score_message_tolerates_racy_socket_registry():
    class RaceDict(dict):
        def __contains__(self, key):
            return True

        def __getitem__(self, key):
            raise KeyError(key)

    original = scoring_router._score_ws
    scoring_router._score_ws = RaceDict()
    try:
        asyncio.run(
            scoring_router._push_score_message(
                "race-conv",
                {"type": "score_progress", "conversation_id": "race-conv"},
            )
        )
    finally:
        scoring_router._score_ws = original


def test_scoring_config_route_returns_current_pool_size(client, monkeypatch):
    class FakeScoringService:
        def get_max_workers(self):
            return 9

    monkeypatch.setattr(scoring_router, "_get_scoring", lambda: FakeScoringService())

    response = client.get("/api/scoring/config")

    assert_true(response.status_code == 200, f"/api/scoring/config 未返回 200: {response.text}")
    payload = response.json()
    assert_true(payload.get("max_workers") == 9, f"/api/scoring/config 返回值错误: {payload}")


def test_ab_session_routes_create_real_conversations_and_support_active_restore(client, monkeypatch):
    def fake_generate_turn(
        service,
        *,
        conv_id: str,
        conversation: dict,
        user_input: str,
        model_id: str,
        model_mini: str,
        dry_run: bool,
        web_search: bool,
        thinking_enabled,
        thinking_effort: str,
        temperature,
        top_p,
    ):
        latest = db.get_conversation(conv_id) or conversation or {}
        turn = len(latest.get("results", []) or []) + 1
        turn_data = {
            "turn": turn,
            "user_input": user_input,
            "ai_output": f"{model_id}:{user_input}",
            "word_count": len(f"{model_id}:{user_input}"),
            "dialogue_summary": f"摘要{turn}",
            "msg_count": 2,
            "input_tokens": 10,
            "output_tokens": 20,
            "latency_s": 0.1,
            "has_deep_injection": False,
            "has_style_isolation": False,
            "has_cooldown_reinject": False,
            "token_trim_level": 0,
            "quality_retries": 0,
            "messages_snapshot": [{"role": "user", "content": user_input}],
            "request_payload_snapshot": {"turn": turn},
            "model_id": model_id,
        }
        db.insert_turn_result(conv_id, turn_data)
        return turn_data

    live_calls = []

    async def fake_enqueue(conv_id: str, turn: int, *, config=None):
        live_calls.append((conv_id, turn))
        return True

    monkeypatch.setattr(conversations_router, "_invoke_interactive_turn_generation", fake_generate_turn)
    monkeypatch.setattr(scoring_router, "enqueue_live_score_turn", fake_enqueue)

    create_response = client.post(
        "/api/ab-sessions",
        json={
            "shared_config": {"scenario": "ab"},
            "base": {
                "model_id": "doubao-pro",
                "prompt_version": DEFAULT_PROMPT_FILE,
                "character": {"Role_Nickname": "角色A"},
                "context": {"relationship": "暧昧"},
                "modules": {},
                "custom_variables": {},
                "auto_scoring": True,
            },
            "compare": {
                "model_id": "qwen-plus",
                "prompt_version": DEFAULT_PROMPT_FILE,
                "character": {"Role_Nickname": "角色A"},
                "context": {"relationship": "暧昧"},
                "modules": {},
                "custom_variables": {},
                "auto_scoring": True,
            },
        },
    )
    assert_true(create_response.status_code == 200, create_response.text)
    session = create_response.json()
    base_conv = db.get_conversation(session["base_conversation_id"])
    compare_conv = db.get_conversation(session["compare_conversation_id"])
    assert_true(base_conv is not None and compare_conv is not None, f"A/B 会话未创建真实 conversation: {session}")
    assert_true(base_conv["config"]["runtime"].get("ab_session_id") == session["id"], f"base runtime 缺少 ab_session_id: {base_conv}")
    assert_true(compare_conv["config"]["runtime"].get("ab_variant") == "compare", f"compare runtime 缺少 ab_variant: {compare_conv}")

    turn_response = client.post(
        f"/api/ab-sessions/{session['id']}/turns",
        json={"user_input": "继续推进"},
    )
    assert_true(turn_response.status_code == 200, turn_response.text)
    wait_for_turns(session["base_conversation_id"], 1)
    wait_for_turns(session["compare_conversation_id"], 1)

    active_response = client.get("/api/ab-sessions/active")
    assert_true(active_response.status_code == 200, active_response.text)
    active_payload = active_response.json()
    assert_true(active_payload["id"] == session["id"], f"active A/B 实验恢复失败: {active_payload}")
    assert_true(
        sorted(live_calls) == sorted(
            [
                (session["base_conversation_id"], 1),
                (session["compare_conversation_id"], 1),
            ]
        ),
        f"A/B 两侧未独立触发 live scoring: {live_calls}",
    )


def test_ab_session_completed_session_rejects_new_turns(client: TestClient):
    create_response = client.post(
        "/api/ab-sessions",
        json={
            "shared_config": {"scenario": "ab"},
            "base": {
                "model_id": "doubao-pro",
                "prompt_version": DEFAULT_PROMPT_FILE,
                "character": {"Role_Nickname": "角色A"},
                "context": {"relationship": "暧昧"},
                "modules": {},
                "custom_variables": {},
                "auto_scoring": True,
            },
            "compare": {
                "model_id": "qwen-plus",
                "prompt_version": DEFAULT_PROMPT_FILE,
                "character": {"Role_Nickname": "角色A"},
                "context": {"relationship": "暧昧"},
                "modules": {},
                "custom_variables": {},
                "auto_scoring": True,
            },
        },
    )
    assert_true(create_response.status_code == 200, create_response.text)
    session = create_response.json()
    db.update_ab_session(session["id"], status="completed")

    response = client.post(
        f"/api/ab-sessions/{session['id']}/turns",
        json={"user_input": "继续推进"},
    )

    assert_true(response.status_code == 409, response.text)
    assert_true("已结束" in response.text, response.text)


def test_ab_session_callback_does_not_revive_completed_session(client: TestClient, monkeypatch):
    create_response = client.post(
        "/api/ab-sessions",
        json={
            "shared_config": {"scenario": "ab"},
            "base": {
                "model_id": "doubao-pro",
                "prompt_version": DEFAULT_PROMPT_FILE,
                "character": {"Role_Nickname": "角色A"},
                "context": {"relationship": "暧昧"},
                "modules": {},
                "custom_variables": {},
                "auto_scoring": True,
            },
            "compare": {
                "model_id": "qwen-plus",
                "prompt_version": DEFAULT_PROMPT_FILE,
                "character": {"Role_Nickname": "角色A"},
                "context": {"relationship": "暧昧"},
                "modules": {},
                "custom_variables": {},
                "auto_scoring": True,
            },
        },
    )
    assert_true(create_response.status_code == 200, create_response.text)
    session = create_response.json()

    async def fake_generate_interactive_turn(conv_id: str, payload):
        return {"success": True, "conversation_id": conv_id, "turn": 1}

    monkeypatch.setattr(
        ab_sessions_router.conversations_router,
        "generate_interactive_turn",
        fake_generate_interactive_turn,
    )
    db.update_ab_session(session["id"], status="completed", current_turn=1)

    asyncio.run(
        ab_sessions_router._launch_ab_generation(
            session["id"],
            1,
            base_conversation_id=session["base_conversation_id"],
            compare_conversation_id=session["compare_conversation_id"],
            request=types.SimpleNamespace(
                user_input="继续推进",
                web_search=False,
                thinking_enabled=None,
                thinking_effort="",
                temperature=None,
                top_p=None,
            ),
        )
    )

    refreshed = db.get_ab_session(session["id"])
    assert_true((refreshed or {}).get("status") == "completed", refreshed)
    assert_true(not (refreshed or {}).get("config", {}).get("last_turn_result"), refreshed)


def test_ab_session_callback_marks_failed_when_any_side_fails(client: TestClient, monkeypatch):
    create_response = client.post(
        "/api/ab-sessions",
        json={
            "shared_config": {"scenario": "ab"},
            "base": {
                "model_id": "doubao-pro",
                "prompt_version": DEFAULT_PROMPT_FILE,
                "character": {"Role_Nickname": "角色A"},
                "context": {"relationship": "暧昧"},
                "modules": {},
                "custom_variables": {},
                "auto_scoring": True,
            },
            "compare": {
                "model_id": "qwen-plus",
                "prompt_version": DEFAULT_PROMPT_FILE,
                "character": {"Role_Nickname": "角色A"},
                "context": {"relationship": "暧昧"},
                "modules": {},
                "custom_variables": {},
                "auto_scoring": True,
            },
        },
    )
    assert_true(create_response.status_code == 200, create_response.text)
    session = create_response.json()
    db.update_ab_session(session["id"], status="running", current_turn=1)

    async def fake_generate_interactive_turn(conv_id: str, payload):
        if conv_id == session["compare_conversation_id"]:
            return {"success": False, "conversation_id": conv_id, "error": "compare failed"}
        return {"success": True, "conversation_id": conv_id, "turn": 1}

    monkeypatch.setattr(
        ab_sessions_router.conversations_router,
        "generate_interactive_turn",
        fake_generate_interactive_turn,
    )

    asyncio.run(
        ab_sessions_router._launch_ab_generation(
            session["id"],
            1,
            base_conversation_id=session["base_conversation_id"],
            compare_conversation_id=session["compare_conversation_id"],
            request=types.SimpleNamespace(
                user_input="继续推进",
                web_search=False,
                thinking_enabled=None,
                thinking_effort="",
                temperature=None,
                top_p=None,
            ),
        )
    )

    refreshed = db.get_ab_session(session["id"])
    last_turn_result = (refreshed or {}).get("config", {}).get("last_turn_result", {})
    assert_true((refreshed or {}).get("status") == "failed", refreshed)
    assert_true(last_turn_result.get("compare", {}).get("success") is False, last_turn_result)


def main():
    tests = []
    with TestClient(app) as client:
        tests.extend(
            [
                ("评分模型别名解析", test_scoring_service_resolves_model_alias_to_api_name),
                ("内置预设与摘要间隔", lambda: test_builtin_preset_and_summary_interval(client)),
                ("自定义预设复用", lambda: test_custom_preset_reuse(client)),
                ("提示词版本绑定与校验", lambda: test_prompt_version_binding_and_validation(client)),
                ("中文导出与人工打分", lambda: test_chinese_export_and_manual_score(client)),
                ("评分上传与测试文件导入", lambda: test_scoring_upload_and_test_file_import(client)),
                (
                    "提示词变量替换与单对话生成",
                    lambda: test_prompt_template_variable_replacement_and_interactive_generate(
                        client,
                        Path(tempfile.mkdtemp(prefix="prompt_exec_")),
                    ),
                ),
                ("Compare 导出", lambda: test_compare_exports(client)),
                ("Configs 导入", lambda: test_configs_import(client)),
                ("单轮打分与别名", lambda: test_single_turn_scoring_and_alias(client)),
                ("自动打分与注入深度", lambda: test_auto_scoring_and_injection_depth(client)),
                ("完成会话 WebSocket 补发", lambda: test_websocket_replays_completed_conversation_events(client)),
                ("会话列表契约", lambda: test_conversation_list_contract(client)),
                ("模型对比会话契约", lambda: test_model_compare_conversation_contract(client)),
                ("批量恢复元数据落库", lambda: test_batch_conversation_persists_resume_runtime(client)),
                ("恢复误用拦截", lambda: test_resume_rejects_non_batch_and_completed_conversations(client)),
                ("恢复只续跑剩余轮次", lambda: test_resume_conversation_only_runs_remaining_turns(client)),
                ("启动状态收敛", lambda: test_reconcile_conversation_runtime_state_normalizes_stale_statuses(client)),
                ("活动编排恢复接口", lambda: test_active_orchestration_route_returns_latest_recoverable_run(client)),
                ("编排暂停控制接口", lambda: test_orchestration_control_pause_route_updates_status(client)),
                ("模型对比数量上限", lambda: test_conversation_rejects_more_than_three_model_ids(client)),
                ("交互式会话流", lambda: test_interactive_conversation_session_flow(client)),
                ("配置与变量接口", lambda: test_config_and_variable_endpoints(client)),
                ("模型保存与多模型导出", lambda: test_model_save_and_multi_model_export(client)),
                (
                    "提示词管理 CRUD",
                    lambda: test_prompt_management_crud(
                        client,
                        pytest.MonkeyPatch(),
                        Path(tempfile.mkdtemp(prefix="prompt_crud_")),
                    ),
                ),
                ("历史导出与删除", lambda: test_history_export_and_delete(client)),
            ]
        )
        tests.append(("评分模板兼容 list", test_score_template_loader))
        tests.append((
            "对白规范端到端注入",
            lambda: test_dialogue_guideline_end_to_end_injection(
                client,
                Path(tempfile.mkdtemp(prefix="guideline_e2e_")),
            ),
        ))
        tests.append(("思考通道 strip", test_quality_guard_strips_thinking_channel_tags))

        passed = 0
        for name, test in tests:
            test()
            passed += 1
            print(f"[PASS] {name}")

        print(f"\n回归通过: {passed}/{len(tests)}")


if __name__ == "__main__":
    main()
